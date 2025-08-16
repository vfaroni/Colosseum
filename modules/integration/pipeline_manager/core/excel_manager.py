#!/usr/bin/env python3
"""
Pipeline Manager - Excel Integration System
Roman Engineering Standard: Built for 2000+ year reliability

Excel pipeline integration for automated row creation from extracted document data.
Integrates with existing Excel workflows while preserving formatting and formulas.
"""

import os
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
from datetime import datetime

# Excel processing imports
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Internal imports
from .document_processor import ExtractionResult, DocumentType

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelConfig:
    """Excel integration configuration"""
    file_path: str
    sheet_name: str = "Pipeline"
    header_row: int = 1
    data_start_row: int = 2
    backup_enabled: bool = True
    preserve_formatting: bool = True
    auto_fit_columns: bool = True

@dataclass
class ColumnMapping:
    """Mapping between extracted data and Excel columns"""
    excel_column: str  # Excel column letter (A, B, C, etc.)
    data_path: str     # Dot notation path to data (e.g., "property_details.name")
    column_name: str   # Human readable column name
    data_type: str     # Expected data type (text, number, date, percentage)
    formula: Optional[str] = None  # Excel formula if applicable
    validation: Optional[Dict[str, Any]] = None  # Validation rules

class ExcelManager:
    """Excel pipeline integration manager"""
    
    def __init__(self, config: ExcelConfig, column_mappings: Optional[List[ColumnMapping]] = None):
        """Initialize Excel manager with configuration"""
        self.config = config
        self.column_mappings = column_mappings or self._create_default_mappings()
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        
        logger.info(f"ExcelManager initialized for {config.file_path}")
    
    def _create_default_mappings(self) -> List[ColumnMapping]:
        """Create default column mappings for common pipeline fields"""
        return [
            # Basic Property Information
            ColumnMapping("A", "property_details.name", "Property Name", "text"),
            ColumnMapping("B", "property_details.address", "Address", "text"),
            ColumnMapping("C", "property_details.city", "City", "text"),
            ColumnMapping("D", "property_details.state", "State", "text"),
            ColumnMapping("E", "property_details.zip_code", "ZIP Code", "text"),
            
            # Property Metrics
            ColumnMapping("F", "property_details.unit_count", "Units", "number"),
            ColumnMapping("G", "property_details.square_footage", "Sq Ft", "number"),
            ColumnMapping("H", "property_details.year_built", "Year Built", "number"),
            ColumnMapping("I", "property_details.property_class", "Class", "text"),
            
            # Financial Metrics
            ColumnMapping("J", "financial_metrics.purchase_price", "Purchase Price", "currency"),
            ColumnMapping("K", "financial_metrics.price_per_unit", "Price/Unit", "currency", 
                         formula="=J{row}/F{row}"),
            ColumnMapping("L", "financial_metrics.net_operating_income", "NOI", "currency"),
            ColumnMapping("M", "financial_metrics.cap_rate", "Cap Rate", "percentage"),
            ColumnMapping("N", "financial_metrics.gross_rent_multiplier", "GRM", "number"),
            
            # Rent Information
            ColumnMapping("O", "rent_information.average_rent", "Avg Rent", "currency"),
            ColumnMapping("P", "rent_information.rent_per_sqft", "Rent/SF", "currency"),
            ColumnMapping("Q", "financial_metrics.occupancy_rate", "Occupancy", "percentage"),
            
            # Market Data
            ColumnMapping("R", "market_data.submarket", "Submarket", "text"),
            ColumnMapping("S", "market_data.median_household_income", "Median Income", "currency"),
            
            # Transaction Details
            ColumnMapping("T", "transaction_details.listing_broker", "Broker", "text"),
            ColumnMapping("U", "transaction_details.broker_company", "Broker Company", "text"),
            ColumnMapping("V", "transaction_details.days_on_market", "Days on Market", "number"),
            
            # Analysis Fields
            ColumnMapping("W", "_analysis.deal_quality_score", "Deal Score", "text"),
            ColumnMapping("X", "_analysis.investment_rating", "Rating", "text"),
            ColumnMapping("Y", "_metadata.extraction_date", "Extracted Date", "date"),
            ColumnMapping("Z", "_metadata.document_source", "Source Document", "text"),
        ]
    
    def load_workbook(self) -> bool:
        """Load Excel workbook for processing"""
        try:
            if os.path.exists(self.config.file_path):
                self.workbook = openpyxl.load_workbook(self.config.file_path)
                logger.info(f"Loaded existing workbook: {self.config.file_path}")
            else:
                self.workbook = Workbook()
                logger.info(f"Created new workbook: {self.config.file_path}")
            
            # Get or create worksheet
            if self.config.sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[self.config.sheet_name]
            else:
                self.worksheet = self.workbook.create_sheet(self.config.sheet_name)
                self._create_headers()
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            return False
    
    def _create_headers(self):
        """Create header row with column names"""
        if not self.worksheet:
            return
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Create headers
        for mapping in self.column_mappings:
            cell = self.worksheet[f"{mapping.excel_column}{self.config.header_row}"]
            cell.value = mapping.column_name
            
            if self.config.preserve_formatting:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = header_alignment
        
        logger.info(f"Created headers in {self.config.sheet_name}")
    
    def add_extraction_to_pipeline(self, extraction_result: ExtractionResult, 
                                  document_path: str) -> Tuple[bool, int]:
        """Add extracted data as new row in Excel pipeline"""
        if not self.workbook or not self.worksheet:
            if not self.load_workbook():
                return False, -1
        
        try:
            # Create backup if enabled
            if self.config.backup_enabled:
                self._create_backup()
            
            # Find next available row
            next_row = self._find_next_row()
            
            # Prepare data with analysis and metadata
            enhanced_data = self._enhance_extraction_data(extraction_result, document_path)
            
            # Populate row data
            self._populate_row(next_row, enhanced_data)
            
            # Apply formatting
            if self.config.preserve_formatting:
                self._apply_row_formatting(next_row)
            
            # Auto-fit columns if enabled
            if self.config.auto_fit_columns:
                self._auto_fit_columns()
            
            # Save workbook
            self.workbook.save(self.config.file_path)
            
            logger.info(f"Added extraction data to row {next_row}")
            return True, next_row
            
        except Exception as e:
            logger.error(f"Failed to add extraction to pipeline: {str(e)}")
            return False, -1
    
    def _create_backup(self):
        """Create backup of Excel file before modification"""
        backup_path = self.config.file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        try:
            if os.path.exists(self.config.file_path):
                import shutil
                shutil.copy2(self.config.file_path, backup_path)
                logger.info(f"Created backup: {backup_path}")
        except Exception as e:
            logger.warning(f"Failed to create backup: {str(e)}")
    
    def _find_next_row(self) -> int:
        """Find the next available row for data entry"""
        row = self.config.data_start_row
        
        # Find first empty row
        while self.worksheet[f"A{row}"].value is not None:
            row += 1
        
        return row
    
    def _enhance_extraction_data(self, extraction_result: ExtractionResult, 
                               document_path: str) -> Dict[str, Any]:
        """Enhance extracted data with analysis and metadata"""
        enhanced_data = extraction_result.data.copy()
        
        # Add analysis data
        enhanced_data["_analysis"] = {
            "deal_quality_score": self._calculate_deal_score(extraction_result.data),
            "investment_rating": self._calculate_investment_rating(extraction_result.data),
            "extraction_confidence": extraction_result.confidence_score
        }
        
        # Add metadata
        enhanced_data["_metadata"] = {
            "extraction_date": datetime.now().strftime("%Y-%m-%d"),
            "document_source": os.path.basename(document_path),
            "processing_time": extraction_result.processing_time,
            "document_type": extraction_result.document_type.value
        }
        
        return enhanced_data
    
    def _calculate_deal_score(self, data: Dict[str, Any]) -> str:
        """Calculate deal quality score (A-D rating)"""
        score = 0
        
        # Property age scoring
        year_built = self._get_nested_value(data, "property_details.year_built")
        if year_built and isinstance(year_built, int):
            age = datetime.now().year - year_built
            if age < 10: score += 3
            elif age < 20: score += 2
            elif age < 30: score += 1
        
        # Cap rate scoring
        cap_rate = self._get_nested_value(data, "financial_metrics.cap_rate")
        if cap_rate and isinstance(cap_rate, (int, float)):
            if cap_rate >= 0.06: score += 3  # 6%+
            elif cap_rate >= 0.05: score += 2  # 5-6%
            elif cap_rate >= 0.04: score += 1  # 4-5%
        
        # Occupancy scoring
        occupancy = self._get_nested_value(data, "financial_metrics.occupancy_rate")
        if occupancy and isinstance(occupancy, (int, float)):
            if occupancy >= 0.95: score += 3  # 95%+
            elif occupancy >= 0.90: score += 2  # 90-95%
            elif occupancy >= 0.85: score += 1  # 85-90%
        
        # Convert to letter grade
        if score >= 7: return "A"
        elif score >= 5: return "B"
        elif score >= 3: return "C"
        else: return "D"
    
    def _calculate_investment_rating(self, data: Dict[str, Any]) -> str:
        """Calculate investment attractiveness rating"""
        purchase_price = self._get_nested_value(data, "financial_metrics.purchase_price")
        noi = self._get_nested_value(data, "financial_metrics.net_operating_income")
        units = self._get_nested_value(data, "property_details.unit_count")
        
        if not all([purchase_price, noi, units]):
            return "Unknown"
        
        # Price per unit analysis
        price_per_unit = purchase_price / units if units > 0 else 0
        
        # Simple rating logic (can be enhanced)
        if price_per_unit < 100000 and noi > 0:
            return "Strong Buy"
        elif price_per_unit < 150000 and noi > 0:
            return "Buy"
        elif price_per_unit < 200000:
            return "Hold"
        else:
            return "Pass"
    
    def _populate_row(self, row: int, data: Dict[str, Any]):
        """Populate Excel row with extracted data"""
        for mapping in self.column_mappings:
            cell = self.worksheet[f"{mapping.excel_column}{row}"]
            
            # Get data value
            value = self._get_nested_value(data, mapping.data_path)
            
            # Handle formula columns
            if mapping.formula:
                formula = mapping.formula.replace("{row}", str(row))
                cell.value = formula
            elif value is not None:
                # Format value based on data type
                formatted_value = self._format_value(value, mapping.data_type)
                cell.value = formatted_value
    
    def _format_value(self, value: Any, data_type: str) -> Any:
        """Format value based on expected data type"""
        if value is None or value == "Not specified":
            return None
        
        try:
            if data_type == "currency":
                if isinstance(value, str):
                    # Remove currency symbols and commas
                    cleaned = value.replace('$', '').replace(',', '').strip()
                    return float(cleaned) if cleaned else None
                return float(value)
            
            elif data_type == "percentage":
                if isinstance(value, str) and '%' in value:
                    return float(value.replace('%', '').strip()) / 100
                return float(value) if isinstance(value, (int, float)) else None
            
            elif data_type == "number":
                return int(value) if isinstance(value, (int, float, str)) else None
            
            elif data_type == "date":
                if isinstance(value, str):
                    return datetime.strptime(value, "%Y-%m-%d").date()
                return value
            
            else:  # text
                return str(value) if value else None
                
        except (ValueError, TypeError):
            return str(value) if value else None
    
    def _apply_row_formatting(self, row: int):
        """Apply consistent formatting to data row"""
        # Basic cell styling
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for mapping in self.column_mappings:
            cell = self.worksheet[f"{mapping.excel_column}{row}"]
            cell.border = data_border
            
            # Data type specific formatting
            if mapping.data_type == "currency":
                cell.number_format = '$#,##0.00'
            elif mapping.data_type == "percentage":
                cell.number_format = '0.00%'
            elif mapping.data_type == "number":
                cell.number_format = '#,##0'
            elif mapping.data_type == "date":
                cell.number_format = 'mm/dd/yyyy'
    
    def _auto_fit_columns(self):
        """Auto-fit column widths based on content"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_nested_value(self, data: Dict[str, Any], field_path: str) -> Any:
        """Get nested dictionary value using dot notation"""
        keys = field_path.split('.')
        current = data
        
        for key in keys:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return None
        
        return current
    
    def batch_process_extractions(self, extraction_results: List[Tuple[ExtractionResult, str]]) -> List[Tuple[bool, int]]:
        """Process multiple extractions in batch"""
        results = []
        
        for extraction_result, document_path in extraction_results:
            success, row = self.add_extraction_to_pipeline(extraction_result, document_path)
            results.append((success, row))
        
        logger.info(f"Batch processed {len(extraction_results)} extractions")
        return results
    
    def export_pipeline_summary(self, output_path: str) -> bool:
        """Export pipeline summary to new Excel file"""
        try:
            if not self.workbook:
                if not self.load_workbook():
                    return False
            
            # Create summary workbook
            summary_wb = Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = "Pipeline Summary"
            
            # Copy data from main worksheet
            for row in self.worksheet.iter_rows():
                for cell in row:
                    new_cell = summary_ws.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    
                    # Copy formatting if preserve_formatting is enabled
                    if self.config.preserve_formatting:
                        if cell.font:
                            new_cell.font = cell.font
                        if cell.fill:
                            new_cell.fill = cell.fill
                        if cell.border:
                            new_cell.border = cell.border
                        if cell.alignment:
                            new_cell.alignment = cell.alignment
                        if cell.number_format:
                            new_cell.number_format = cell.number_format
            
            # Save summary
            summary_wb.save(output_path)
            logger.info(f"Exported pipeline summary to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export pipeline summary: {str(e)}")
            return False
    
    def get_pipeline_statistics(self) -> Dict[str, Any]:
        """Get statistics about the current pipeline"""
        if not self.worksheet:
            return {}
        
        stats = {
            "total_properties": 0,
            "total_value": 0,
            "average_price_per_unit": 0,
            "average_cap_rate": 0,
            "deal_quality_distribution": {"A": 0, "B": 0, "C": 0, "D": 0},
            "property_classes": {},
            "markets": {}
        }
        
        try:
            # Count data rows
            row = self.config.data_start_row
            values = []
            
            while self.worksheet[f"A{row}"].value is not None:
                # Collect data for statistics
                purchase_price = self.worksheet[f"J{row}"].value  # Purchase Price column
                cap_rate = self.worksheet[f"M{row}"].value       # Cap Rate column
                deal_score = self.worksheet[f"W{row}"].value     # Deal Score column
                prop_class = self.worksheet[f"I{row}"].value     # Property Class column
                market = self.worksheet[f"R{row}"].value         # Submarket column
                
                if purchase_price and isinstance(purchase_price, (int, float)):
                    values.append(purchase_price)
                
                if deal_score in stats["deal_quality_distribution"]:
                    stats["deal_quality_distribution"][deal_score] += 1
                
                if prop_class:
                    stats["property_classes"][prop_class] = stats["property_classes"].get(prop_class, 0) + 1
                
                if market:
                    stats["markets"][market] = stats["markets"].get(market, 0) + 1
                
                row += 1
            
            stats["total_properties"] = row - self.config.data_start_row
            stats["total_value"] = sum(values) if values else 0
            
            logger.info(f"Generated pipeline statistics for {stats['total_properties']} properties")
            
        except Exception as e:
            logger.error(f"Failed to generate pipeline statistics: {str(e)}")
        
        return stats

# Example usage
if __name__ == "__main__":
    # Example configuration
    config = ExcelConfig(
        file_path="pipeline.xlsx",
        sheet_name="Deal Pipeline",
        backup_enabled=True
    )
    
    # Initialize manager
    excel_manager = ExcelManager(config)
    
    logger.info("ExcelManager ready for use")
