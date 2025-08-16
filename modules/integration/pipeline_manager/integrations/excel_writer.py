#!/usr/bin/env python3
"""
Pipeline Manager - Excel Writer Integration
Roman Engineering Standard: Built for 2000+ year reliability

Advanced Excel file manipulation with smart formatting, formula preservation,
and bulk operations for pipeline management workflows.
"""

import os
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass
from datetime import datetime, date
import shutil

# Excel processing imports
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelWriterConfig:
    """Excel writer configuration"""
    auto_backup: bool = True
    preserve_formulas: bool = True
    auto_fit_columns: bool = True
    apply_conditional_formatting: bool = True
    create_charts: bool = False
    freeze_panes: bool = True
    protect_headers: bool = False
    max_backup_files: int = 10

@dataclass
class CellStyle:
    """Cell styling configuration"""
    font_name: str = "Calibri"
    font_size: int = 11
    font_bold: bool = False
    font_color: str = "000000"
    fill_color: Optional[str] = None
    border_style: str = "thin"
    alignment_horizontal: str = "left"
    alignment_vertical: str = "center"
    number_format: Optional[str] = None

class ExcelWriter:
    """Advanced Excel file writer with comprehensive formatting and data management"""
    
    def __init__(self, config: Optional[ExcelWriterConfig] = None):
        """Initialize Excel writer with configuration"""
        self.config = config or ExcelWriterConfig()
        self.workbook: Optional[Workbook] = None
        self.worksheets: Dict[str, Worksheet] = {}
        self.styles = self._initialize_styles()
        
        logger.info("ExcelWriter initialized with advanced formatting capabilities")
    
    def _initialize_styles(self) -> Dict[str, NamedStyle]:
        """Initialize pre-defined cell styles"""
        styles = {}
        
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        styles["header"] = header_style
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.font = Font(name="Calibri", size=11)
        data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        data_style.alignment = Alignment(horizontal='left', vertical='center')
        styles["data"] = data_style
        
        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.font = Font(name="Calibri", size=11)
        currency_style.number_format = '$#,##0.00'
        currency_style.alignment = Alignment(horizontal='right', vertical='center')
        currency_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles["currency"] = currency_style
        
        # Percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.font = Font(name="Calibri", size=11)
        percentage_style.number_format = '0.00%'
        percentage_style.alignment = Alignment(horizontal='right', vertical='center')
        percentage_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles["percentage"] = percentage_style
        
        # Number style
        number_style = NamedStyle(name="number")
        number_style.font = Font(name="Calibri", size=11)
        number_style.number_format = '#,##0'
        number_style.alignment = Alignment(horizontal='right', vertical='center')
        number_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles["number"] = number_style
        
        # Date style
        date_style = NamedStyle(name="date")
        date_style.font = Font(name="Calibri", size=11)
        date_style.number_format = 'mm/dd/yyyy'
        date_style.alignment = Alignment(horizontal='center', vertical='center')
        date_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        styles["date"] = date_style
        
        return styles
    
    def create_workbook(self, file_path: str) -> bool:
        """Create new workbook or load existing one"""
        try:
            self.file_path = Path(file_path)
            
            if self.file_path.exists():
                # Load existing workbook
                self.workbook = openpyxl.load_workbook(str(self.file_path))
                logger.info(f"Loaded existing workbook: {self.file_path}")
                
                # Create backup if enabled
                if self.config.auto_backup:
                    self._create_backup()
            else:
                # Create new workbook
                self.workbook = Workbook()
                # Remove default sheet
                if "Sheet" in [ws.title for ws in self.workbook.worksheets]:
                    self.workbook.remove(self.workbook["Sheet"])
                logger.info(f"Created new workbook: {self.file_path}")
            
            # Add styles to workbook
            for style_name, style in self.styles.items():
                if style_name not in [s.name for s in self.workbook.named_styles]:
                    self.workbook.add_named_style(style)
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to create/load workbook: {str(e)}")
            return False
    
    def create_worksheet(self, sheet_name: str, headers: List[str], 
                        column_widths: Optional[Dict[str, float]] = None) -> bool:
        """Create new worksheet with headers and formatting"""
        try:
            if not self.workbook:
                raise ValueError("Workbook not initialized")
            
            # Create or get worksheet
            if sheet_name in [ws.title for ws in self.workbook.worksheets]:
                worksheet = self.workbook[sheet_name]
                logger.info(f"Using existing worksheet: {sheet_name}")
            else:
                worksheet = self.workbook.create_sheet(sheet_name)
                logger.info(f"Created new worksheet: {sheet_name}")
            
            self.worksheets[sheet_name] = worksheet
            
            # Add headers
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.style = "header"
            
            # Set column widths
            if column_widths:
                for col_name, width in column_widths.items():
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        worksheet.column_dimensions[col_letter].width = width
            elif self.config.auto_fit_columns:
                self._auto_fit_columns(worksheet)
            
            # Freeze header row
            if self.config.freeze_panes:
                worksheet.freeze_panes = "A2"
            
            # Protect headers if enabled
            if self.config.protect_headers:
                for col_idx in range(1, len(headers) + 1):
                    worksheet.cell(row=1, column=col_idx).protection = openpyxl.styles.Protection(locked=True)
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to create worksheet '{sheet_name}': {str(e)}")
            return False
    
    def write_data_row(self, sheet_name: str, row_data: Dict[str, Any], 
                      row_number: Optional[int] = None,
                      column_mapping: Optional[Dict[str, str]] = None) -> Tuple[bool, int]:
        """Write data row to worksheet with automatic formatting"""
        try:
            if sheet_name not in self.worksheets:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = self.worksheets[sheet_name]
            
            # Find next available row if not specified
            if row_number is None:
                row_number = self._find_next_row(worksheet)
            
            # Get headers from first row
            headers = [cell.value for cell in worksheet[1] if cell.value]
            
            # Write data to cells
            for col_idx, header in enumerate(headers, 1):
                # Get data value using column mapping if provided
                data_key = column_mapping.get(header, header) if column_mapping else header
                value = row_data.get(data_key)
                
                cell = worksheet.cell(row=row_number, column=col_idx)
                
                if value is not None:
                    # Set cell value
                    cell.value = self._format_cell_value(value)
                    
                    # Apply appropriate style based on data type
                    style_name = self._determine_cell_style(value, header)
                    cell.style = style_name
                else:
                    cell.style = "data"
            
            logger.debug(f"Wrote data row {row_number} to sheet '{sheet_name}'")
            return True, row_number
            
        except Exception as e:
            logger.error(f"Failed to write data row: {str(e)}")
            return False, -1
    
    def write_bulk_data(self, sheet_name: str, data_list: List[Dict[str, Any]],
                       column_mapping: Optional[Dict[str, str]] = None,
                       start_row: Optional[int] = None) -> Tuple[bool, List[int]]:
        """Write multiple data rows efficiently"""
        try:
            if sheet_name not in self.worksheets:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = self.worksheets[sheet_name]
            
            # Find starting row if not specified
            if start_row is None:
                start_row = self._find_next_row(worksheet)
            
            written_rows = []
            current_row = start_row
            
            for data_row in data_list:
                success, row_num = self.write_data_row(
                    sheet_name, data_row, current_row, column_mapping
                )
                if success:
                    written_rows.append(row_num)
                    current_row += 1
                else:
                    logger.warning(f"Failed to write data row {current_row}")
            
            # Apply conditional formatting if enabled
            if self.config.apply_conditional_formatting:
                self._apply_conditional_formatting(worksheet, start_row, current_row - 1)
            
            logger.info(f"Bulk wrote {len(written_rows)} rows to sheet '{sheet_name}'")
            return True, written_rows
            
        except Exception as e:
            logger.error(f"Failed to write bulk data: {str(e)}")
            return False, []
    
    def update_cell(self, sheet_name: str, row: int, column: Union[int, str], 
                   value: Any, preserve_formatting: bool = True) -> bool:
        """Update single cell value"""
        try:
            if sheet_name not in self.worksheets:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = self.worksheets[sheet_name]
            
            # Convert column to index if string
            if isinstance(column, str):
                column = column_index_from_string(column)
            
            cell = worksheet.cell(row=row, column=column)
            old_style = cell.style if preserve_formatting else None
            
            cell.value = self._format_cell_value(value)
            
            if not preserve_formatting:
                # Apply appropriate style based on data type
                headers = [cell.value for cell in worksheet[1] if cell.value]
                header = headers[column - 1] if column <= len(headers) else ""
                style_name = self._determine_cell_style(value, header)
                cell.style = style_name
            
            logger.debug(f"Updated cell {get_column_letter(column)}{row} in sheet '{sheet_name}'")
            return True
            
        except Exception as e:
            logger.error(f"Failed to update cell: {str(e)}")
            return False
    
    def insert_formula(self, sheet_name: str, row: int, column: Union[int, str], 
                      formula: str) -> bool:
        """Insert Excel formula into cell"""
        try:
            if sheet_name not in self.worksheets:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = self.worksheets[sheet_name]
            
            # Convert column to index if string
            if isinstance(column, str):
                column = column_index_from_string(column)
            
            cell = worksheet.cell(row=row, column=column)
            cell.value = formula
            
            # Apply number formatting based on formula type
            if any(func in formula.upper() for func in ['SUM', 'AVERAGE']):
                if '$' in formula or 'price' in formula.lower():
                    cell.style = "currency"
                else:
                    cell.style = "number"
            elif '%' in formula or 'rate' in formula.lower():
                cell.style = "percentage"
            
            logger.debug(f"Inserted formula in cell {get_column_letter(column)}{row}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to insert formula: {str(e)}")
            return False
    
    def add_data_validation(self, sheet_name: str, cell_range: str, 
                           validation_type: str, formula: str) -> bool:
        """Add data validation to cell range"""
        try:
            if sheet_name not in self.worksheets:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = self.worksheets[sheet_name]
            
            from openpyxl.worksheet.datavalidation import DataValidation
            
            dv = DataValidation(type=validation_type, formula1=formula)
            dv.add(cell_range)
            worksheet.add_data_validation(dv)
            
            logger.debug(f"Added data validation to range {cell_range} in sheet '{sheet_name}'")
            return True
            
        except Exception as e:
            logger.error(f"Failed to add data validation: {str(e)}")
            return False
    
    def create_summary_sheet(self, source_sheet: str, summary_data: Dict[str, Any]) -> bool:
        """Create a summary worksheet with key metrics"""
        try:
            summary_sheet_name = f"{source_sheet}_Summary"
            
            # Create summary worksheet
            if not self.create_worksheet(summary_sheet_name, ["Metric", "Value"]):
                return False
            
            worksheet = self.worksheets[summary_sheet_name]
            
            # Write summary data
            row = 2
            for metric, value in summary_data.items():
                worksheet.cell(row=row, column=1, value=metric.replace('_', ' ').title()).style = "data"
                
                cell = worksheet.cell(row=row, column=2, value=self._format_cell_value(value))
                style_name = self._determine_cell_style(value, metric)
                cell.style = style_name
                
                row += 1
            
            # Auto-fit columns
            self._auto_fit_columns(worksheet)
            
            logger.info(f"Created summary sheet: {summary_sheet_name}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to create summary sheet: {str(e)}")
            return False
    
    def save_workbook(self, file_path: Optional[str] = None) -> bool:
        """Save workbook to file"""
        try:
            if not self.workbook:
                raise ValueError("Workbook not initialized")
            
            save_path = Path(file_path) if file_path else self.file_path
            
            # Ensure directory exists
            save_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Save workbook
            self.workbook.save(str(save_path))
            
            logger.info(f"Workbook saved to: {save_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to save workbook: {str(e)}")
            return False
    
    def _create_backup(self):
        """Create backup of existing Excel file"""
        try:
            if not self.file_path.exists():
                return
            
            backup_dir = self.file_path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
            backup_path = backup_dir / backup_name
            
            shutil.copy2(self.file_path, backup_path)
            
            # Clean up old backups
            self._cleanup_old_backups(backup_dir)
            
            logger.info(f"Created backup: {backup_path}")
            
        except Exception as e:
            logger.warning(f"Failed to create backup: {str(e)}")
    
    def _cleanup_old_backups(self, backup_dir: Path):
        """Remove old backup files beyond the limit"""
        try:
            backup_files = sorted(
                [f for f in backup_dir.glob("*_backup_*")],
                key=lambda x: x.stat().st_mtime,
                reverse=True
            )
            
            if len(backup_files) > self.config.max_backup_files:
                for old_backup in backup_files[self.config.max_backup_files:]:
                    old_backup.unlink()
                    logger.debug(f"Removed old backup: {old_backup}")
                    
        except Exception as e:
            logger.warning(f"Failed to cleanup old backups: {str(e)}")
    
    def _find_next_row(self, worksheet: Worksheet) -> int:
        """Find the next available row for data entry"""
        row = 2  # Start after header
        while worksheet.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    def _format_cell_value(self, value: Any) -> Any:
        """Format value for Excel cell"""
        if value is None:
            return None
        elif isinstance(value, (list, dict)):
            return str(value)
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, (date, datetime)):
            return value
        else:
            return str(value)
    
    def _determine_cell_style(self, value: Any, header: str) -> str:
        """Determine appropriate cell style based on value and header"""
        header_lower = header.lower()
        
        # Currency fields
        if any(keyword in header_lower for keyword in ['price', 'cost', 'income', 'value', 'rent', 'fee']):
            return "currency"
        
        # Percentage fields
        elif any(keyword in header_lower for keyword in ['rate', 'ratio', 'percent', 'occupancy']):
            return "percentage"
        
        # Date fields
        elif any(keyword in header_lower for keyword in ['date', 'time']) or isinstance(value, (date, datetime)):
            return "date"
        
        # Number fields
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            if any(keyword in header_lower for keyword in ['count', 'units', 'sqft', 'sq ft', 'year']):
                return "number"
            else:
                return "currency"  # Default for numbers
        
        # Default to data style
        else:
            return "data"
    
    def _auto_fit_columns(self, worksheet: Worksheet):
        """Auto-fit column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
    
    def _apply_conditional_formatting(self, worksheet: Worksheet, start_row: int, end_row: int):
        """Apply conditional formatting to data range"""
        try:
            # Get headers to identify columns for formatting
            headers = [cell.value for cell in worksheet[1] if cell.value]
            
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                range_address = f"{col_letter}{start_row}:{col_letter}{end_row}"
                header_lower = header.lower()
                
                # Apply color scale for financial metrics
                if any(keyword in header_lower for keyword in ['price', 'income', 'noi', 'value']):
                    color_scale = ColorScaleRule(
                        start_type='min', start_color='FFFF0000',  # Red for low
                        end_type='max', end_color='FF00FF00'      # Green for high
                    )
                    worksheet.conditional_formatting.add(range_address, color_scale)
                
                # Apply data bars for occupancy rates
                elif 'occupancy' in header_lower or 'rate' in header_lower:
                    data_bar = DataBarRule(
                        start_type='min', start_value=0,
                        end_type='max', end_value=1,
                        color="FF0066CC"
                    )
                    worksheet.conditional_formatting.add(range_address, data_bar)
                    
        except Exception as e:
            logger.warning(f"Failed to apply conditional formatting: {str(e)}")
    
    def get_worksheet_statistics(self, sheet_name: str) -> Dict[str, Any]:
        """Get statistics about worksheet data"""
        try:
            if sheet_name not in self.worksheets:
                return {}
            
            worksheet = self.worksheets[sheet_name]
            
            # Count rows and columns
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Count non-empty data cells (excluding header)
            data_cells = 0
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    if worksheet.cell(row=row, column=col).value is not None:
                        data_cells += 1
            
            return {
                "total_rows": max_row,
                "total_columns": max_col,
                "data_rows": max_row - 1,  # Excluding header
                "data_cells": data_cells,
                "headers": [cell.value for cell in worksheet[1] if cell.value]
            }
            
        except Exception as e:
            logger.error(f"Failed to get worksheet statistics: {str(e)}")
            return {}

# Example usage
if __name__ == "__main__":
    # Initialize writer
    config = ExcelWriterConfig(
        auto_backup=True,
        apply_conditional_formatting=True,
        auto_fit_columns=True
    )
    
    writer = ExcelWriter(config)
    
    # Create workbook
    if writer.create_workbook("test_pipeline.xlsx"):
        # Create worksheet
        headers = ["Property Name", "Address", "Units", "Purchase Price", "NOI", "Cap Rate", "Occupancy Rate"]
        column_widths = {"Property Name": 25, "Address": 30, "Purchase Price": 15}
        
        if writer.create_worksheet("Deal Pipeline", headers, column_widths):
            # Sample data
            sample_data = [
                {
                    "Property Name": "Sunset Apartments",
                    "Address": "123 Main St, Austin, TX",
                    "Units": 150,
                    "Purchase Price": 18500000,
                    "NOI": 1200000,
                    "Cap Rate": 0.065,
                    "Occupancy Rate": 0.95
                },
                {
                    "Property Name": "Oak Ridge Complex",
                    "Address": "456 Oak Ave, Dallas, TX",
                    "Units": 200,
                    "Purchase Price": 25000000,
                    "NOI": 1650000,
                    "Cap Rate": 0.066,
                    "Occupancy Rate": 0.92
                }
            ]
            
            # Write bulk data
            success, rows = writer.write_bulk_data("Deal Pipeline", sample_data)
            
            if success:
                # Create summary
                summary_data = {
                    "total_properties": len(sample_data),
                    "total_value": sum(d["Purchase Price"] for d in sample_data),
                    "average_cap_rate": sum(d["Cap Rate"] for d in sample_data) / len(sample_data),
                    "total_units": sum(d["Units"] for d in sample_data)
                }
                
                writer.create_summary_sheet("Deal Pipeline", summary_data)
                
                # Save workbook
                if writer.save_workbook():
                    logger.info("Excel file created successfully")
                    
                    # Get statistics
                    stats = writer.get_worksheet_statistics("Deal Pipeline")
                    logger.info(f"Worksheet stats: {stats}")
    
    logger.info("ExcelWriter ready for use")