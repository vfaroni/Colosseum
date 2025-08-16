import openpyxl
import os

file_path = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals/!Pipeline Summary.xlsm"

# Load the workbook
wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)

# Get the 7.16.25 sheet
sheet = wb["7.16.25"]

print("=== Rows 21 and 22 ===")
print("Row 21:")
for col in range(1, 20):  # Check first 19 columns
    cell = sheet.cell(row=21, column=col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {cell.value}")

print("\nRow 22:")
for col in range(1, 20):  # Check first 19 columns
    cell = sheet.cell(row=22, column=col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {cell.value}")

print("\n=== Searching for San Pablo Suites ===")
found_rows = []
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        if cell.value and "San Pablo" in str(cell.value):
            found_rows.append(row)
            print(f"Found San Pablo reference in Row {row}, Column {openpyxl.utils.get_column_letter(col)}: {cell.value}")
            break

if found_rows:
    for row in found_rows:
        print(f"\nFull row {row} data:")
        for c in range(1, 20):
            val = sheet.cell(row=row, column=c).value
            if val:
                print(f"  {openpyxl.utils.get_column_letter(c)}: {val}")
else:
    print("San Pablo Suites not found in the sheet")

print("\n=== Last few rows with deal data ===")
# Find the last row with actual data
last_data_row = 0
for row in range(sheet.max_row, 0, -1):
    has_data = False
    for col in range(1, 10):  # Check first 10 columns
        if sheet.cell(row=row, column=col).value:
            has_data = True
            break
    if has_data:
        last_data_row = row
        break

print(f"Last row with data: {last_data_row}")

# Show the last 5 rows with data
start_row = max(1, last_data_row - 4)
for row in range(start_row, last_data_row + 1):
    has_data = False
    for col in range(1, 20):
        if sheet.cell(row=row, column=col).value:
            has_data = True
            break
    
    if has_data:
        print(f"\nRow {row}:")
        for col in range(1, 20):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"  {col_letter}: {cell.value}")

wb.close()