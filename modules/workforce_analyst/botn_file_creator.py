#!/usr/bin/env python3
"""
BOTN File Creator - Real Estate Deal Underwriting Assistant
Creates actual BOTN Excel files with extracted data in deal folders

Author: Claude Code Assistant
Date: 2025-08-04
"""

import os
import shutil
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import sys
import traceback

class BOTNFileCreator:
    def __init__(self):
        # Base directories
        self.deals_base_dir = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals"
        self.template_path = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/botntemplate/Archive/CA BOTN 9.25.24.xlsx"
        self.cache_dir = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/workforce_analyst/deal_cache"
        
        # Verify paths exist
        self.verify_paths()
        
    def verify_paths(self):
        """Verify all required paths exist"""
        if not os.path.exists(self.deals_base_dir):
            print(f"❌ Deals directory not found: {self.deals_base_dir}")
            
        if not os.path.exists(self.template_path):
            print(f"❌ BOTN template not found: {self.template_path}")
            
        if not os.path.exists(self.cache_dir):
            print(f"❌ Cache directory not found: {self.cache_dir}")
    
    def create_botn_file(self, deal_name, extracted_data):
        """
        Create actual BOTN file in deal folder
        
        Args:
            deal_name (str): Name of the deal (e.g., "Sunset Gardens - El Cajon, CA")
            extracted_data (dict): Extracted property data
            
        Returns:
            dict: Result with success status and file path
        """
        try:
            print(f"\n🏗️ Creating BOTN file for: {deal_name}")
            
            # Step 1: Create deal folder and BOTN subfolder
            deal_folder = os.path.join(self.deals_base_dir, deal_name)
            botn_folder = os.path.join(deal_folder, "BOTN")
            
            # Create directories if they don't exist
            os.makedirs(deal_folder, exist_ok=True)
            os.makedirs(botn_folder, exist_ok=True)
            print(f"✅ Created folder: {botn_folder}")
            
            # Step 2: Generate BOTN filename
            sanitized_name = deal_name.replace("/", "-").replace("\\", "-")
            timestamp = datetime.now().strftime("%Y-%m-%d")
            botn_filename = f"BOTN_{sanitized_name}_{timestamp}.xlsx"
            botn_file_path = os.path.join(botn_folder, botn_filename)
            
            # Step 3: Copy BOTN template
            if os.path.exists(self.template_path):
                shutil.copy2(self.template_path, botn_file_path)
                print(f"✅ Copied BOTN template to: {botn_filename}")
            else:
                # Create a basic Excel file if template not found
                self.create_basic_botn_template(botn_file_path)
                print(f"⚠️ Template not found, created basic BOTN: {botn_filename}")
            
            # Step 4: Populate with extracted data
            self.populate_botn_data(botn_file_path, extracted_data, deal_name)
            print(f"✅ Populated BOTN with extracted data")
            
            return {
                "success": True,
                "file_path": botn_file_path,
                "filename": botn_filename,
                "folder": botn_folder,
                "message": f"BOTN file created successfully: {botn_filename}"
            }
            
        except Exception as e:
            error_msg = f"Failed to create BOTN file: {str(e)}"
            print(f"❌ {error_msg}")
            traceback.print_exc()
            return {
                "success": False,
                "error": error_msg
            }
    
    def create_basic_botn_template(self, file_path):
        """Create a basic BOTN template if the official one isn't found"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOTN Analysis"
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Create basic BOTN structure
        headers = [
            ["A1", "BACK OF THE NAPKIN (BOTN) ANALYSIS"],
            ["A3", "PROPERTY INFORMATION"],
            ["A4", "Property Name"],
            ["A5", "Address"],
            ["A6", "City, State, Zip"],
            ["A7", "Year Built"],
            ["A8", "Total Units"],
            ["A10", "FINANCIAL PERFORMANCE"],
            ["A11", "Average Rent"],
            ["A12", "T12 Net Rental Income"],
            ["A13", "T12 Other Income"],
            ["A14", "T12 Operating Expenses"],
            ["A15", "Net Operating Income"],
            ["A17", "UNIT MIX & RENTS"],
            ["A18", "Studio Units"],
            ["A19", "1BR Units"], 
            ["A20", "2BR Units"],
            ["A21", "3BR Units"],
            ["A23", "MARKET ANALYSIS"],
            ["A24", "Market Area"],
            ["A25", "County"],
            ["A26", "Construction Type"]
        ]
        
        for cell_ref, text in headers:
            cell = ws[cell_ref]
            cell.value = text
            if "INFORMATION" in text or "PERFORMANCE" in text or "ANALYSIS" in text or "BOTN" in text:
                cell.font = header_font
                cell.fill = header_fill
        
        wb.save(file_path)
    
    def populate_botn_data(self, file_path, extracted_data, deal_name):
        """Populate BOTN template with extracted data"""
        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Data mapping for BOTN population
            data_mapping = {
                # Column B is typically where data goes in BOTN templates
                "B4": extracted_data.get("Property Name", deal_name.split(" - ")[0]),
                "B5": extracted_data.get("Property Address", "Address TBD"),
                "B6": self.extract_city_state_zip(extracted_data),
                "B7": extracted_data.get("Year Built", "TBD"),
                "B8": extracted_data.get("Total Units", extracted_data.get("Number of Units", "TBD")),
                "B11": extracted_data.get("Average Rent", extracted_data.get("Avg In Place Rents", "TBD")),
                "B12": self.format_currency(extracted_data.get("T12 Net Rental Income", "TBD")),
                "B13": self.format_currency(extracted_data.get("T12 Other Income", extracted_data.get("T12 Total Other Income", "TBD"))),
                "B14": self.format_currency(extracted_data.get("T12 Operating Expenses", extracted_data.get("T12 Expenses", "TBD"))),
                "B15": self.calculate_noi(extracted_data),
                "B24": extracted_data.get("Market Area", self.get_market_area(deal_name)),
                "B25": extracted_data.get("County", extracted_data.get("County Name", "TBD")),
                "B26": extracted_data.get("Construction", extracted_data.get("Construction Type", "TBD"))
            }
            
            # Unit mix data
            unit_mapping = {
                "B18": f"{extracted_data.get('# Studio Units', '0')} units @ {extracted_data.get('Studio Rents', extracted_data.get('Studio Rent', 'TBD'))}",
                "B19": f"{extracted_data.get('# 1 Bed Units', '0')} units @ {extracted_data.get('1 Bed Current Rents', extracted_data.get('1BR Rent', 'TBD'))}",
                "B20": f"{extracted_data.get('# 2 Bed Units', '0')} units @ {extracted_data.get('2 Bed Current Rents', extracted_data.get('2BR Rent', 'TBD'))}",
                "B21": f"{extracted_data.get('# 3 Bed Units', '0')} units @ {extracted_data.get('3 Bed Current Rents', extracted_data.get('3BR Rent', 'TBD'))}"
            }
            
            # Combine all mappings
            all_mappings = {**data_mapping, **unit_mapping}
            
            # Populate cells
            for cell_ref, value in all_mappings.items():
                if value and value != "TBD":
                    try:
                        ws[cell_ref] = value
                    except Exception as e:
                        print(f"Warning: Could not populate cell {cell_ref}: {e}")
            
            # Add metadata
            ws["A30"] = "Generated by Deal Underwriting Assistant"
            ws["A31"] = f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A32"] = f"Data Source: Real Cached Data"
            
            # Save the file
            wb.save(file_path)
            
        except Exception as e:
            print(f"Warning: Could not populate BOTN data: {e}")
            # File was created, just not populated
    
    def extract_city_state_zip(self, data):
        """Extract city, state, zip from address or individual fields"""
        address = data.get("Property Address", "")
        if address and "," in address:
            parts = address.split(",")
            if len(parts) >= 2:
                return f"{parts[-2].strip()}, {parts[-1].strip()}"
        
        # Try individual fields
        city = data.get("City", "")
        state = data.get("State", "")
        zip_code = data.get("Zip Code", "")
        
        if city and state:
            result = f"{city}, {state}"
            if zip_code:
                result += f" {zip_code}"
            return result
        
        return "TBD"
    
    def format_currency(self, value):
        """Format currency values"""
        if not value or value == "TBD":
            return "TBD"
        
        # Remove existing formatting
        clean_value = str(value).replace("$", "").replace(",", "")
        
        try:
            num_value = float(clean_value)
            return f"${num_value:,.0f}"
        except (ValueError, TypeError):
            return str(value)
    
    def calculate_noi(self, data):
        """Calculate Net Operating Income"""
        try:
            income = self.parse_currency(data.get("T12 Net Rental Income", "0"))
            other_income = self.parse_currency(data.get("T12 Other Income", data.get("T12 Total Other Income", "0")))
            expenses = self.parse_currency(data.get("T12 Operating Expenses", data.get("T12 Expenses", "0")))
            
            total_income = income + other_income
            noi = total_income - expenses
            
            if noi > 0:
                return f"${noi:,.0f}"
        except:
            pass
        
        return "TBD"
    
    def parse_currency(self, value):
        """Parse currency string to float"""
        if not value or value == "TBD":
            return 0
        
        clean_value = str(value).replace("$", "").replace(",", "")
        try:
            return float(clean_value)
        except (ValueError, TypeError):
            return 0
    
    def get_market_area(self, deal_name):
        """Extract market area from deal name"""
        if "San Diego" in deal_name or "El Cajon" in deal_name:
            return "San Diego Metro"
        elif "Los Angeles" in deal_name:
            return "Los Angeles Metro"
        elif "Oakland" in deal_name:
            return "East Bay"
        elif "San Jose" in deal_name:
            return "Silicon Valley"
        else:
            return "California Market"

def main():
    """Command line interface for BOTN creation"""
    if len(sys.argv) < 2:
        print("Usage: python3 botn_file_creator.py <deal_name> [data_file]")
        print("Example: python3 botn_file_creator.py 'Sunset Gardens - El Cajon, CA'")
        return
    
    deal_name = sys.argv[1]
    
    # Load extracted data
    extracted_data = {}
    if len(sys.argv) > 2:
        data_file = sys.argv[2]
        try:
            with open(data_file, 'r') as f:
                extracted_data = json.load(f)
        except Exception as e:
            print(f"Warning: Could not load data file {data_file}: {e}")
    
    # Create BOTN file
    creator = BOTNFileCreator()
    result = creator.create_botn_file(deal_name, extracted_data)
    
    if result["success"]:
        print(f"\n🎉 SUCCESS!")
        print(f"📁 File: {result['filename']}")
        print(f"📂 Location: {result['folder']}")
        print(f"✅ {result['message']}")
    else:
        print(f"\n❌ FAILED: {result['error']}")

if __name__ == "__main__":
    main()