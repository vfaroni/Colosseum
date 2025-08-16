#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

# Change to the correct directory
os.chdir('/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals')

# Load the Excel file
file_path = "!Pipeline Summary.xlsm"
wb = load_workbook(file_path, data_only=True)

print("Available sheets:")
for i, sheet_name in enumerate(wb.sheetnames):
    print(f"  {i}: {sheet_name}")

# Find the sheet that matches 7.16.25 or similar pattern
target_sheet = None
for sheet_name in wb.sheetnames:
    if "7.16.25" in sheet_name or "7.16" in sheet_name:
        target_sheet = sheet_name
        break

if not target_sheet:
    # If no exact match, look for the most recent date-like sheet
    date_sheets = [s for s in wb.sheetnames if any(c.isdigit() for c in s) and "." in s]
    if date_sheets:
        target_sheet = date_sheets[-1]  # Take the last one
    else:
        # Just use the last sheet
        target_sheet = wb.sheetnames[-1]

print(f"\nUsing sheet: {target_sheet}")

if target_sheet:
    ws = wb[target_sheet]
    print(f"Sheet max_row: {ws.max_row}")
    print(f"Sheet max_column: {ws.max_column}")
    
    # Now examine rows 15-25 specifically
    print("\nRows 15-25 content:")
    for row_num in range(15, 26):  # 15 to 25 inclusive
        row_data = []
        for col_num in range(1, min(ws.max_column + 1, 10)):  # First 10 columns
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                cell_value = "BLANK"
            row_data.append(str(cell_value))
        print(f"Row {row_num}: {row_data}")
    
    # Search for '1206 Oakmead' specifically
    print("\nSearching for '1206 Oakmead':")
    found_locations = []
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value and "1206 Oakmead" in str(cell_value):
                found_locations.append((row_num, col_num, cell_value))
    
    if found_locations:
        for row, col, value in found_locations:
            print(f"  Found at row {row}, column {col}: {value}")
    else:
        print("  Not found")
    
    # Check for test deals
    print("\nSearching for test deals:")
    test_deals = ["Rising Sun", "Bayside Apartment Homes"]
    for deal in test_deals:
        found = []
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value and deal in str(cell_value):
                    found.append((row_num, col_num, cell_value))
        if found:
            print(f"  {deal}: Found at {found}")
        else:
            print(f"  {deal}: Not found")
    
    # Find the actual last row with data
    print("\nFinding last row with data:")
    last_data_row = 0
    for row_num in range(1, ws.max_row + 1):
        has_data = False
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is not None and str(cell_value).strip() != "":
                has_data = True
                break
        if has_data:
            last_data_row = row_num
    
    print(f"Last row with data: {last_data_row}")
    
    # Show the last few rows with data
    print(f"\nShowing rows {max(1, last_data_row-5)} to {last_data_row}:")
    for row_num in range(max(1, last_data_row-5), last_data_row + 1):
        row_data = []
        for col_num in range(1, min(ws.max_column + 1, 10)):  # First 10 columns
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                cell_value = "BLANK"
            row_data.append(str(cell_value))
        print(f"Row {row_num}: {row_data}")