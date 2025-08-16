import openpyxl
import os

file_path = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals/!Pipeline Summary.xlsm"

# Load the workbook
wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)

# Get the 7.16.25 sheet
sheet = wb["7.16.25"]

print("=== Column Headers (Row 1) ===")
for col in range(1, 20):  # Check first 19 columns
    cell = sheet.cell(row=1, column=col)
    if cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}: {cell.value}")

print("\n=== Looking for actual deal data rows ===")
# Find rows with actual deal data (skip headers and instructions)
deal_rows = []
for row in range(2, 35):  # Check rows 2-34
    # Look for rows that have deal names in column A and aren't instructions
    cell_a = sheet.cell(row=row, column=1).value
    if cell_a and isinstance(cell_a, str) and cell_a not in ['Input', 'Pipeline Summary', 'Instructions:']:
        # Check if it has more data suggesting it's a deal row
        has_deal_data = False
        for col in range(2, 10):
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val and str(cell_val) not in ['Input', 'Formula']:
                has_deal_data = True
                break
        if has_deal_data:
            deal_rows.append(row)

print(f"Found {len(deal_rows)} rows with deal data: {deal_rows}")

print("\n=== All deal rows ===")
for row in deal_rows:
    print(f"\nRow {row}:")
    for col in range(1, 20):
        cell = sheet.cell(row=row, column=col)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"  {col_letter}: {cell.value}")

wb.close()