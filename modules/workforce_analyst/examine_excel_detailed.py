#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

# Change to the correct directory
os.chdir('/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals')

# Load the Excel file
file_path = "!Pipeline Summary.xlsm"
wb = load_workbook(file_path, data_only=True)

# Use the 7.16.25 sheet
ws = wb["7.16.25"]

# Get column headers
print("Column headers (Row 1):")
headers = []
for col_num in range(1, min(ws.max_column + 1, 20)):  # First 20 columns
    cell_value = ws.cell(row=1, column=col_num).value
    if cell_value is None:
        cell_value = "BLANK"
    headers.append(str(cell_value))
    print(f"  Column {col_num}: {cell_value}")

print("\nDetailed view of actual data rows (showing first 10 columns):")
# Find rows with actual deal data (non-blank in column A)
for row_num in range(2, min(ws.max_row + 1, 25)):  # Start from row 2, check up to row 25
    cell_a = ws.cell(row=row_num, column=1).value
    if cell_a and str(cell_a).strip() and "Instructions" not in str(cell_a):
        row_data = []
        for col_num in range(1, 11):  # First 10 columns
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                cell_value = "BLANK"
            row_data.append(str(cell_value))
        print(f"Row {row_num}: {row_data}")

# Find the range where deals actually are
print("\nFinding deal data range:")
first_deal_row = None
last_deal_row = None

for row_num in range(2, ws.max_row + 1):
    cell_a = ws.cell(row=row_num, column=1).value
    if cell_a and str(cell_a).strip() and "Instructions" not in str(cell_a):
        if first_deal_row is None:
            first_deal_row = row_num
        last_deal_row = row_num

print(f"First deal row: {first_deal_row}")
print(f"Last deal row: {last_deal_row}")

# Check what's right after the last deal
if last_deal_row:
    print(f"\nRows immediately after last deal ({last_deal_row}):")
    for row_num in range(last_deal_row + 1, min(last_deal_row + 6, ws.max_row + 1)):
        cell_a = ws.cell(row=row_num, column=1).value
        print(f"Row {row_num}, Column A: '{cell_a}'")