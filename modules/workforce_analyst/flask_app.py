#!/usr/bin/env python3
"""
🏢 Deal Underwriting Assistant - Flask Web Application
Converts the Streamlit BOTN analysis tool to a stable web interface
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
import os
import json
from datetime import datetime
from pathlib import Path
import logging

# Import existing business logic classes
from AcquisitionAnalyst import (
    ExcelFileManager, 
    DealFinder, 
    DataExtractor, 
    load_deal_data, 
    save_deal_data, 
    get_cached_deals_info,
    sanitize_deal_name
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this'  # Change this in production
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file upload

# Initialize business logic components
excel_manager = ExcelFileManager()
base_path = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals"
deal_finder = DealFinder(base_path)

@app.route('/')
def index():
    """Main dashboard"""
    try:
        # Get cached deals for quick access
        cached_deals = get_cached_deals_info()
        
        return render_template('index.html', 
                             cached_deals=cached_deals,
                             stage='selection')
    except Exception as e:
        logger.error(f"Error loading main page: {e}")
        flash(f"Error loading dashboard: {str(e)}", 'error')
        return render_template('index.html', cached_deals=[], stage='selection')

@app.route('/deals')
def list_deals():
    """List all available deals"""
    try:
        deals = deal_finder.find_deals()
        return render_template('deals.html', deals=deals)
    except Exception as e:
        logger.error(f"Error listing deals: {e}")
        flash(f"Error loading deals: {str(e)}", 'error')
        return render_template('deals.html', deals=[])

@app.route('/select_deal/<path:deal_path>')
def select_deal(deal_path):
    """Select a deal for analysis"""
    try:
        # Store selected deal in session
        session['selected_deal'] = {
            'name': os.path.basename(deal_path),
            'path': deal_path
        }
        
        # Check if we have cached data for this deal
        deal_name = session['selected_deal']['name']
        cached_data = load_deal_data(deal_name)
        
        if cached_data:
            session['extracted_data'] = cached_data[0]
            session['extraction_date'] = cached_data[1]
            flash(f"Loaded cached data for {deal_name}", 'success')
            return redirect(url_for('show_botn'))
        else:
            return redirect(url_for('extract_data'))
            
    except Exception as e:
        logger.error(f"Error selecting deal: {e}")
        flash(f"Error selecting deal: {str(e)}", 'error')
        return redirect(url_for('list_deals'))

@app.route('/extract')
def extract_data():
    """Data extraction page"""
    selected_deal = session.get('selected_deal')
    if not selected_deal:
        flash("Please select a deal first", 'warning')
        return redirect(url_for('list_deals'))
    
    return render_template('extract.html', deal=selected_deal)

@app.route('/api/extract', methods=['POST'])
def api_extract_data():
    """API endpoint for data extraction"""
    try:
        selected_deal = session.get('selected_deal')
        if not selected_deal:
            return jsonify({'error': 'No deal selected'}), 400
        
        openai_key = request.json.get('openai_key')
        if not openai_key:
            return jsonify({'error': 'OpenAI API key required'}), 400
        
        # Initialize data extractor
        data_extractor = DataExtractor(openai_key)
        
        # Extract data from deal folder
        extracted_data = data_extractor.extract_from_folder(selected_deal['path'])
        
        if extracted_data:
            # Save to cache
            save_deal_data(selected_deal['name'], extracted_data)
            
            # Store in session
            session['extracted_data'] = extracted_data
            session['extraction_date'] = datetime.now().isoformat()
            
            return jsonify({
                'success': True,
                'data': extracted_data,
                'message': f'Successfully extracted {len(extracted_data)} fields'
            })
        else:
            return jsonify({'error': 'No data extracted'}), 400
            
    except Exception as e:
        logger.error(f"Error extracting data: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/botn')
def show_botn():
    """BOTN analysis page"""
    selected_deal = session.get('selected_deal')
    extracted_data = session.get('extracted_data')
    
    if not selected_deal:
        flash("Please select a deal first", 'warning')
        return redirect(url_for('list_deals'))
    
    if not extracted_data:
        flash("Please extract data first", 'warning')
        return redirect(url_for('extract_data'))
    
    return render_template('botn.html', 
                         deal=selected_deal, 
                         data=extracted_data,
                         extraction_date=session.get('extraction_date'))

@app.route('/api/create_botn', methods=['POST'])
def api_create_botn():
    """API endpoint to create BOTN Excel file"""
    try:
        selected_deal = session.get('selected_deal')
        extracted_data = session.get('extracted_data')
        
        if not selected_deal or not extracted_data:
            return jsonify({'error': 'Missing deal or data'}), 400
        
        # Create BOTN file
        botn_file_path = excel_manager.copy_template_to_deal_folder(
            selected_deal['name'],
            selected_deal['path']
        )
        
        if botn_file_path:
            # Update Excel file with extracted data
            success = excel_manager.update_excel_values(botn_file_path, extracted_data)
            
            if success:
                # Store file path in session
                session['botn_file_path'] = botn_file_path
                
                return jsonify({
                    'success': True,
                    'file_path': botn_file_path,
                    'file_name': os.path.basename(botn_file_path),
                    'message': 'BOTN analysis created successfully!'
                })
            else:
                return jsonify({'error': 'Failed to update Excel file'}), 500
        else:
            return jsonify({'error': 'Failed to copy template'}), 500
            
    except Exception as e:
        logger.error(f"Error creating BOTN: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/results')
def show_results():
    """Results page"""
    selected_deal = session.get('selected_deal')
    botn_file_path = session.get('botn_file_path')
    
    if not selected_deal:
        flash("Please select a deal first", 'warning')
        return redirect(url_for('list_deals'))
    
    return render_template('results.html', 
                         deal=selected_deal,
                         botn_file_path=botn_file_path,
                         file_name=os.path.basename(botn_file_path) if botn_file_path else None)

@app.route('/api/open_file')
def api_open_file():
    """Open BOTN file in Finder/Explorer"""
    try:
        botn_file_path = session.get('botn_file_path')
        if not botn_file_path or not os.path.exists(botn_file_path):
            return jsonify({'error': 'File not found'}), 404
        
        import subprocess
        import platform
        
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", "-R", botn_file_path])
        elif platform.system() == "Windows":
            subprocess.run(["explorer", "/select,", botn_file_path])
        else:  # Linux
            subprocess.run(["xdg-open", os.path.dirname(botn_file_path)])
        
        return jsonify({'success': True, 'message': 'File opened in finder'})
        
    except Exception as e:
        logger.error(f"Error opening file: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/template_test')
def api_template_test():
    """Test template access"""
    try:
        template_path = excel_manager.template_path
        
        if not os.path.exists(template_path):
            return jsonify({
                'success': False,
                'error': f'Template file not found: {template_path}'
            })
        
        # Try to load template
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        
        result = {
            'success': True,
            'template_path': template_path,
            'file_name': os.path.basename(template_path),
            'worksheets': wb.sheetnames,
            'worksheet_count': len(wb.sheetnames)
        }
        
        # Check for Inputs sheet
        if 'Inputs' in wb.sheetnames:
            inputs_sheet = wb['Inputs']
            result['inputs_sheet'] = {
                'found': True,
                'max_row': inputs_sheet.max_row,
                'max_col': inputs_sheet.max_column
            }
        else:
            result['inputs_sheet'] = {'found': False}
        
        wb.close()
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error testing template: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/reset')
def reset_session():
    """Reset session and start over"""
    session.clear()
    flash("Session reset. You can now start a new analysis.", 'info')
    return redirect(url_for('index'))

@app.errorhandler(404)
def page_not_found(e):
    """Handle 404 errors"""
    return render_template('error.html', 
                         error_code=404, 
                         error_message="Page not found"), 404

@app.errorhandler(500)
def internal_server_error(e):
    """Handle 500 errors"""
    return render_template('error.html', 
                         error_code=500, 
                         error_message="Internal server error"), 500

if __name__ == '__main__':
    # Create templates directory if it doesn't exist
    templates_dir = os.path.join(os.path.dirname(__file__), 'templates')
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    
    os.makedirs(templates_dir, exist_ok=True)
    os.makedirs(static_dir, exist_ok=True)
    
    print("🏢 Deal Underwriting Assistant - Flask Web Server")
    print("=" * 50)
    print("🚀 Starting server...")
    print("📱 Access at: http://localhost:5001")
    print("🌐 Team access: http://[your-ip]:5001")
    print("⏹️  Press Ctrl+C to stop")
    print("=" * 50)
    
    # Run in debug mode for development
    app.run(host='0.0.0.0', port=5001, debug=True)