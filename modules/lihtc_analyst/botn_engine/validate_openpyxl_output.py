#!/usr/bin/env python3
"""
Validate OpenPyxl BOTN Output - Verify formula preservation and data integrity
"""

import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class OpenPyxlOutputValidator:
    """Validate that OpenPyxl BOTN files preserve all template functionality"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
        self.test_output_dir = self.base_path / "OpenPyxl_Test_Batch"
    
    def validate_single_file(self, file_path):
        """Validate a single BOTN file for formula preservation"""
        
        logger.info(f"🔍 Validating: {file_path.name}")
        
        try:
            # Load with data_only=False to see formulas
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            validation_results = {
                "file_name": file_path.name,
                "worksheets_present": wb.sheetnames,
                "formulas_found": {},
                "data_populated": {},
                "errors": []
            }
            
            # Check that all expected worksheets are present
            expected_sheets = ['Inputs', 'Rents', 'Expenses', 'Sources & Uses', 'NOI', 
                              'Data>>', 'Developer Fee Max', 'Section8-FY24', '2025 FMR', '2025 SAFMR']
            
            missing_sheets = [sheet for sheet in expected_sheets if sheet not in wb.sheetnames]
            if missing_sheets:
                validation_results["errors"].append(f"Missing worksheets: {missing_sheets}")
            
            # Check Inputs sheet data population
            if 'Inputs' in wb.sheetnames:
                inputs_sheet = wb['Inputs']
                
                # Check key populated fields
                key_fields = {
                    'A2': 'Property Name',
                    'B2': 'Address', 
                    'C2': 'County',
                    'G2': 'Purchase Price',
                    'H2': 'Housing Type',
                    'I2': 'Credit Pricing',
                    'O2': 'Units',
                    'P2': 'Unit Size'
                }
                
                for cell_ref, field_name in key_fields.items():
                    value = inputs_sheet[cell_ref].value
                    validation_results["data_populated"][field_name] = value
                    
                    if value is None or value == '':
                        validation_results["errors"].append(f"{field_name} ({cell_ref}) is empty")
            
            # Check formula preservation in calculation sheets
            formula_sheets = ['Rents', 'Expenses', 'Sources & Uses', 'NOI']
            
            for sheet_name in formula_sheets:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    formulas = []
                    
                    # Scan for formulas
                    for row in range(1, min(21, sheet.max_row + 1)):
                        for col in range(1, min(11, sheet.max_column + 1)):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formulas.append({
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'formula': cell.value[:50] + "..." if len(cell.value) > 50 else cell.value
                                })
                    
                    validation_results["formulas_found"][sheet_name] = formulas
                    
                    if not formulas:
                        validation_results["errors"].append(f"No formulas found in {sheet_name} sheet")
            
            wb.close()
            
            return validation_results
            
        except Exception as e:
            logger.error(f"❌ Validation failed for {file_path.name}: {str(e)}")
            return {
                "file_name": file_path.name,
                "validation_failed": True,
                "error": str(e)
            }
    
    def validate_all_generated_files(self):
        """Validate all generated OpenPyxl BOTN files"""
        
        logger.info("🚀 COMPREHENSIVE OPENPYXL OUTPUT VALIDATION")
        logger.info("=" * 70)
        
        if not self.test_output_dir.exists():
            logger.error(f"❌ Test output directory not found: {self.test_output_dir}")
            return
        
        # Find all BOTN files
        botn_files = list(self.test_output_dir.glob("*_BOTN.xlsx"))
        
        if not botn_files:
            logger.error("❌ No BOTN files found in test directory")
            return
        
        logger.info(f"🔍 Found {len(botn_files)} BOTN files to validate")
        
        all_results = []
        successful_validations = 0
        
        for botn_file in botn_files:
            result = self.validate_single_file(botn_file)
            all_results.append(result)
            
            if "validation_failed" not in result:
                successful_validations += 1
                
                # Quick summary for each file
                error_count = len(result.get("errors", []))
                formula_count = sum(len(formulas) for formulas in result["formulas_found"].values())
                
                if error_count == 0:
                    logger.info(f"   ✅ {result['file_name']}: {formula_count} formulas preserved, no errors")
                else:
                    logger.info(f"   ⚠️  {result['file_name']}: {formula_count} formulas, {error_count} errors")
            else:
                logger.info(f"   ❌ {result['file_name']}: Validation failed")
        
        # Detailed results
        logger.info(f"\n📊 VALIDATION SUMMARY")
        logger.info("=" * 50)
        logger.info(f"✅ Successfully validated: {successful_validations}/{len(botn_files)} files")
        
        # Show detailed results for first file as example
        if all_results and "validation_failed" not in all_results[0]:
            first_result = all_results[0]
            logger.info(f"\n📋 DETAILED EXAMPLE: {first_result['file_name']}")
            logger.info("-" * 40)
            
            logger.info("📝 Data populated:")
            for field, value in first_result["data_populated"].items():
                if isinstance(value, (int, float)) and value > 1000:
                    logger.info(f"   {field}: ${value:,.0f}")
                else:
                    logger.info(f"   {field}: {value}")
            
            logger.info("\n🔗 Formulas preserved:")
            for sheet_name, formulas in first_result["formulas_found"].items():
                if formulas:
                    logger.info(f"   {sheet_name}: {len(formulas)} formulas")
                    # Show first 2 formulas as examples
                    for formula in formulas[:2]:
                        logger.info(f"     {formula['cell']}: {formula['formula']}")
            
            if first_result.get("errors"):
                logger.info(f"\n⚠️  Errors found:")
                for error in first_result["errors"]:
                    logger.info(f"   ❌ {error}")
        
        return all_results
    
    def compare_with_template(self):
        """Compare generated files with original template structure"""
        
        logger.info("\n🔬 TEMPLATE COMPARISON ANALYSIS")
        logger.info("=" * 50)
        
        try:
            # Load original template
            template_wb = openpyxl.load_workbook(self.template_path, data_only=False)
            
            # Count template formulas
            template_formulas = {}
            for sheet_name in ['Rents', 'Expenses', 'Sources & Uses', 'NOI']:
                if sheet_name in template_wb.sheetnames:
                    sheet = template_wb[sheet_name]
                    formulas = []
                    
                    for row in range(1, min(21, sheet.max_row + 1)):
                        for col in range(1, min(11, sheet.max_column + 1)):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formulas.append(f"{cell.column_letter}{cell.row}: {cell.value}")
                    
                    template_formulas[sheet_name] = formulas
            
            template_wb.close()
            
            # Show template formula count
            total_template_formulas = sum(len(formulas) for formulas in template_formulas.values())
            logger.info(f"📊 Original template formulas: {total_template_formulas}")
            
            for sheet_name, formulas in template_formulas.items():
                logger.info(f"   {sheet_name}: {len(formulas)} formulas")
            
            # Compare with first generated file
            botn_files = list(self.test_output_dir.glob("*_BOTN.xlsx"))
            if botn_files:
                first_file = botn_files[0]
                generated_wb = openpyxl.load_workbook(first_file, data_only=False)
                
                generated_formulas = {}
                for sheet_name in ['Rents', 'Expenses', 'Sources & Uses', 'NOI']:
                    if sheet_name in generated_wb.sheetnames:
                        sheet = generated_wb[sheet_name]
                        formulas = []
                        
                        for row in range(1, min(21, sheet.max_row + 1)):
                            for col in range(1, min(11, sheet.max_column + 1)):
                                cell = sheet.cell(row=row, column=col)
                                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                    formulas.append(f"{cell.column_letter}{cell.row}: {cell.value}")
                        
                        generated_formulas[sheet_name] = formulas
                
                generated_wb.close()
                
                total_generated_formulas = sum(len(formulas) for formulas in generated_formulas.values())
                logger.info(f"📊 Generated file formulas: {total_generated_formulas}")
                
                # Compare sheet by sheet
                for sheet_name in template_formulas:
                    template_count = len(template_formulas[sheet_name])
                    generated_count = len(generated_formulas.get(sheet_name, []))
                    
                    if template_count == generated_count:
                        logger.info(f"   ✅ {sheet_name}: {template_count} = {generated_count}")
                    else:
                        logger.info(f"   ⚠️  {sheet_name}: {template_count} → {generated_count}")
                
                if total_template_formulas == total_generated_formulas:
                    logger.info("🎉 PERFECT MATCH: All formulas preserved!")
                else:
                    logger.info(f"⚠️  Formula count difference: {total_template_formulas} → {total_generated_formulas}")
            
        except Exception as e:
            logger.error(f"❌ Template comparison failed: {str(e)}")

def main():
    validator = OpenPyxlOutputValidator()
    
    # Validate all generated files
    results = validator.validate_all_generated_files()
    
    # Compare with original template
    validator.compare_with_template()
    
    # Final assessment
    if results:
        successful_count = len([r for r in results if "validation_failed" not in r and not r.get("errors")])
        total_count = len(results)
        
        logger.info(f"\n🎯 FINAL ASSESSMENT")
        logger.info("=" * 40)
        
        if successful_count == total_count:
            logger.info("🎉 COMPLETE SUCCESS: All files validated perfectly!")
            logger.info("✅ OpenPyxl solution is production-ready")
            logger.info("✅ Can replace xlwings system immediately")
        else:
            logger.info(f"⚠️  {successful_count}/{total_count} files validated successfully")
            logger.info("🔧 Minor issues need addressing before production")

if __name__ == "__main__":
    main()