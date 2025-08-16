#!/usr/bin/env python3
"""
BOTN Test Generator - Create populated BOTN from template with one test site
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class BOTNTestGenerator:
    """Simple BOTN generator for testing with one site"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
        self.sites_path = self.base_path / "Sites"
        
    def load_template_headers(self):
        """Load headers from Row 1 of Inputs tab"""
        try:
            logger.info(f"Loading template from: {self.template_path}")
            workbook = openpyxl.load_workbook(self.template_path)
            
            if 'Inputs' not in workbook.sheetnames:
                logger.error(f"'Inputs' sheet not found. Available sheets: {workbook.sheetnames}")
                return None
                
            inputs_sheet = workbook['Inputs']
            
            # Read Row 1 headers
            headers = []
            for col in range(1, inputs_sheet.max_column + 1):
                cell_value = inputs_sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append((col, cell_value))
                else:
                    break  # Stop at first empty cell
                    
            logger.info(f"Found {len(headers)} headers in template")
            for col, header in headers:
                logger.info(f"  Column {col}: {header}")
                
            return headers
            
        except Exception as e:
            logger.error(f"Error loading template: {str(e)}")
            return None
    
    def load_test_site(self):
        """Load one test site from final portfolio"""
        try:
            # Use the most recent final portfolio
            final_portfolio_path = self.sites_path / "BOTN_COMPLETE_FINAL_PORTFOLIO_20250731_130415.xlsx"
            
            if not final_portfolio_path.exists():
                logger.error(f"Final portfolio not found at: {final_portfolio_path}")
                return None
                
            logger.info(f"Loading sites from: {final_portfolio_path}")
            df = pd.read_excel(final_portfolio_path)
            
            if len(df) == 0:
                logger.error("No sites found in final portfolio")
                return None
                
            # Get first site for testing
            test_site = df.iloc[0]
            logger.info(f"Selected test site: {test_site.get('Property Name', 'Unknown')}")
            logger.info(f"Address: {test_site.get('Property Address', 'Unknown')}")
            
            # Show available columns
            logger.info(f"Available data columns: {list(df.columns)}")
            
            return test_site
            
        except Exception as e:
            logger.error(f"Error loading test site: {str(e)}")
            return None
    
    def get_cdlac_region(self, county_name):
        """Determine CDLAC region based on county name"""
        # CDLAC Region mapping based on California counties
        cdlac_regions = {
            # Region 1 - Northern California
            'Alameda': 1, 'Contra Costa': 1, 'Lake': 1, 'Marin': 1, 'Napa': 1,
            'San Francisco': 1, 'San Mateo': 1, 'Santa Clara': 1, 'Solano': 1, 'Sonoma': 1,
            
            # Region 2 - Sacramento Valley
            'Butte': 2, 'Colusa': 2, 'El Dorado': 2, 'Glenn': 2, 'Nevada': 2,
            'Placer': 2, 'Sacramento': 2, 'Sutter': 2, 'Tehama': 2, 'Yolo': 2, 'Yuba': 2,
            
            # Region 3 - Northern Mountains
            'Alpine': 3, 'Amador': 3, 'Calaveras': 3, 'Del Norte': 3, 'Humboldt': 3,
            'Inyo': 3, 'Lassen': 3, 'Madera': 3, 'Mariposa': 3, 'Mendocino': 3,
            'Modoc': 3, 'Mono': 3, 'Plumas': 3, 'Shasta': 3, 'Sierra': 3, 'Siskiyou': 3,
            'Tuolumne': 3, 'Trinity': 3,
            
            # Region 4 - Central Valley
            'Fresno': 4, 'Kern': 4, 'Kings': 4, 'Merced': 4, 'San Joaquin': 4,
            'Stanislaus': 4, 'Tulare': 4,
            
            # Region 5 - Central Coast
            'Monterey': 5, 'San Benito': 5, 'San Luis Obispo': 5, 'Santa Barbara': 5,
            'Santa Cruz': 5, 'Ventura': 5,
            
            # Region 6 - Los Angeles County
            'Los Angeles': 6,
            
            # Region 7 - Orange County
            'Orange': 7,
            
            # Region 8 - Inland Empire
            'Riverside': 8, 'San Bernardino': 8,
            
            # Region 9 - San Diego County  
            'San Diego': 9, 'Imperial': 9
        }
        
        return cdlac_regions.get(county_name, None)

    def map_site_data_to_template(self, site_data, template_headers):
        """Map site data to template columns with enhanced field mapping"""
        mapping = {}
        
        # Enhanced field mapping based on actual template headers and available site data
        field_mapping = {
            # Basic site info
            'Property Name': 'Property Name',
            'Address': 'Property Address',
            'County:': 'County Name',  # Note the colon in template
            'State:': 'State',  # Note the colon in template 
            'Zip Code:': 'Zip',  # Note the colon in template
            'Purchase Price': 'For Sale Price',
            'Zoning': 'Zoning',
            'Latitude': 'Latitude',
            'Longitude': 'Longitude',
            'Land Area': 'Land Area (AC)',
            'Secondary Type': 'Secondary Type',
        }
        
        # Map available data
        for col_num, header in template_headers:
            if header in field_mapping:
                site_field = field_mapping[header]
                if site_field in site_data:
                    value = site_data[site_field]
                    if pd.notna(value):
                        mapping[col_num] = value
                        logger.info(f"Mapped {header} (Col {col_num}): {value}")
                    else:
                        logger.info(f"No value for {header} (Col {col_num})")
                else:
                    logger.info(f"Site field '{site_field}' not found for header '{header}'")
            elif header == 'CDLAC Region':  # Special handling for CDLAC Region
                county_name = site_data.get('County Name')
                if county_name and pd.notna(county_name):
                    cdlac_region = self.get_cdlac_region(county_name)
                    if cdlac_region:
                        mapping[col_num] = cdlac_region
                        logger.info(f"Mapped {header} (Col {col_num}): {cdlac_region} (from county: {county_name})")
                    else:
                        logger.info(f"Could not determine CDLAC region for county: {county_name}")
                else:
                    logger.info(f"No county data available for CDLAC region mapping")
            else:
                logger.info(f"No mapping defined for header: {header}")
                
        return mapping
    
    def get_user_inputs_for_unmapped_fields(self, template_headers, data_mapping):
        """Prompt user for manual input values that will be applied to all sites"""
        print("\n" + "="*70)
        print("🔧 MANUAL INPUT REQUIRED FOR BOTN TEMPLATE")
        print("="*70)
        print("The following fields need manual input. These values will be")
        print("applied to ALL sites in your portfolio.")
        print("-" * 70)
        
        user_inputs = {}
        
        for col_num, header in template_headers:
            if col_num not in data_mapping:
                # Provide helpful prompts for each field type
                field_help = {
                    'Housing Type': 'e.g., New Construction, Rehab, Acquisition+Rehab',
                    'Credit Pricing': 'e.g., 0.85, 0.90, 0.95 (typical range 0.80-1.00)',
                    'Credit Type': 'e.g., 4%, 9%',
                    'Construciton Loan Term': 'e.g., 18, 24 (months)',
                    'Market Cap Rate': 'e.g., 0.05, 0.06, 0.07 (as decimal)',
                    'Financing Interest Rate': 'e.g., 0.06, 0.07, 0.08 (as decimal)',
                    'Elevator': 'e.g., Yes, No',
                    '# Units': 'e.g., 50, 75, 100 (typical LIHTC project size)',
                    'Avg Unit Size': 'e.g., 850, 950, 1100 (square feet)',
                    'Hard Cost/SF': 'e.g., 200, 250, 300 (cost per square foot)'
                }
                
                help_text = field_help.get(header, 'Enter appropriate value')
                
                while True:
                    print(f"\n📝 {header}:")
                    print(f"   Help: {help_text}")
                    user_value = input(f"   Enter value: ").strip()
                    
                    if user_value:
                        # Try to convert to appropriate type
                        try:
                            # Check if it should be a number
                            if any(keyword in header.lower() for keyword in ['rate', 'cost', 'price', 'units', 'size', 'term']):
                                if '.' in user_value or any(keyword in header.lower() for keyword in ['rate']):
                                    user_inputs[col_num] = float(user_value)
                                else:
                                    user_inputs[col_num] = int(user_value)
                            else:
                                user_inputs[col_num] = user_value
                            break
                        except ValueError:
                            print(f"   ❌ Invalid input. Please enter a valid value.")
                    else:
                        print(f"   ⚠️  Field left blank - will remain empty in BOTN")
                        break
        
        print("\n" + "="*70)
        print("✅ USER INPUT COLLECTION COMPLETE")
        print("="*70)
        print("Summary of values that will be applied to ALL sites:")
        for col_num, header in template_headers:
            if col_num in user_inputs:
                print(f"  • {header}: {user_inputs[col_num]}")
        print("="*70)
        
        return user_inputs

    def display_mapping_results(self, site_data, template_headers, complete_mapping):
        """Display the complete mapping results including user inputs"""
        logger.info("\n" + "="*60)
        logger.info("📊 FINAL BOTN TEMPLATE POPULATION PREVIEW")
        logger.info("="*60)
        
        logger.info(f"\n🏠 Site: {site_data.get('Property Name', 'Unknown')}")
        logger.info(f"📍 Address: {site_data.get('Property Address', 'Unknown')}")
        logger.info(f"🏛️ County: {site_data.get('County Name', 'Unknown')}")
        logger.info(f"📊 Total fields to populate: {len(complete_mapping)}")
        
        logger.info(f"\n📋 ALL TEMPLATE FIELDS:")
        logger.info("-" * 50)
        
        auto_count = 0
        manual_count = 0
        
        for col_num, header in template_headers:
            if col_num in complete_mapping:
                value = complete_mapping[col_num]
                # Determine if this was auto-mapped or user input
                if header in ['Property Name', 'Address', 'County:', 'CDLAC Region', 'State:', 'Zip Code:', 'Purchase Price']:
                    logger.info(f"🤖 Col {col_num:2d} | {header:25s} → {value} [AUTO]")
                    auto_count += 1
                else:
                    logger.info(f"👤 Col {col_num:2d} | {header:25s} → {value} [USER INPUT]")
                    manual_count += 1
            else:
                logger.info(f"❌ Col {col_num:2d} | {header:25s} → [EMPTY]")
        
        logger.info("\n🔍 MAPPING SUMMARY:")
        logger.info("-" * 30)
        total_count = len(template_headers)
        populated_count = len(complete_mapping)
        logger.info(f"Auto-populated: {auto_count}")
        logger.info(f"User inputs: {manual_count}")
        logger.info(f"Empty fields: {total_count - populated_count}")
        logger.info(f"Total populated: {populated_count}/{total_count} ({populated_count/total_count*100:.1f}%)")
        logger.info("="*60)

    def generate_botn(self, site_data, template_headers, data_mapping):
        """Generate populated BOTN file with user inputs"""
        try:
            # Get user inputs for unmapped fields
            print("\n🔧 Getting user inputs for fields that will apply to ALL sites...")
            user_inputs = self.get_user_inputs_for_unmapped_fields(template_headers, data_mapping)
            
            # Combine auto-mapped data with user inputs
            complete_mapping = {**data_mapping, **user_inputs}
            
            # Display final mapping results
            self.display_mapping_results(site_data, template_headers, complete_mapping)
            
            # Load template
            workbook = openpyxl.load_workbook(self.template_path)
            inputs_sheet = workbook['Inputs']
            
            # Populate Row 2 with all mapped data
            for col_num, value in complete_mapping.items():
                inputs_sheet.cell(row=2, column=col_num, value=value)
                
            # Generate output filename
            site_name = str(site_data.get('Property Name', 'TestSite')).replace(' ', '_').replace('/', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"BOTN_{site_name}_{timestamp}.xlsx"
            output_path = self.base_path / "outputs" / output_filename
            
            # Create output directory if it doesn't exist
            output_path.parent.mkdir(exist_ok=True)
            
            # Save populated workbook
            workbook.save(output_path)
            logger.info(f"\n✅ Generated BOTN file: {output_path}")
            
            return output_path, user_inputs
            
        except Exception as e:
            logger.error(f"Error generating BOTN: {str(e)}")
            return None, None

def main():
    """Main function to test BOTN generation"""
    logger.info("🏛️ BOTN TEST GENERATOR - ONE SITE PROTOTYPE")
    logger.info("=" * 50)
    
    generator = BOTNTestGenerator()
    
    # Step 1: Load template headers
    logger.info("\n1. Loading template headers...")
    headers = generator.load_template_headers()
    if not headers:
        logger.error("Failed to load template headers")
        return
    
    # Step 2: Load test site
    logger.info("\n2. Loading test site...")
    test_site = generator.load_test_site()
    if test_site is None:
        logger.error("Failed to load test site")
        return
    
    # Step 3: Map data
    logger.info("\n3. Mapping site data to template...")
    data_mapping = generator.map_site_data_to_template(test_site, headers)
    logger.info(f"Mapped {len(data_mapping)} fields")
    
    # Step 4: Generate BOTN
    logger.info("\n4. Generating populated BOTN...")
    output_path, user_inputs = generator.generate_botn(test_site, headers, data_mapping)
    
    if output_path:
        logger.info(f"\n🎉 SUCCESS: Complete BOTN generated!")
        logger.info(f"📁 File: {output_path}")
        logger.info("\n📋 Summary:")
        logger.info(f"  • Auto-populated: 7 fields from site data")
        logger.info(f"  • User inputs: {len(user_inputs)} fields")
        logger.info(f"  • These user inputs will be applied to ALL sites")
        logger.info("\n✅ You can now open this file to review the complete populated BOTN")
    else:
        logger.error("\n❌ FAILED: Could not generate BOTN")

if __name__ == "__main__":
    main()