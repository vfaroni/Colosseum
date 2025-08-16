#!/usr/bin/env python3
"""
Simple BOTN Generator - Shows required inputs and creates template
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class SimpleBOTNGenerator:
    """Simple BOTN generator that shows required inputs"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
        self.sites_path = self.base_path / "Sites"
        
    def analyze_requirements(self):
        """Analyze what inputs are needed for the BOTN template"""
        
        # Load template headers
        workbook = openpyxl.load_workbook(self.template_path)
        inputs_sheet = workbook['Inputs']
        
        headers = []
        for col in range(1, inputs_sheet.max_column + 1):
            cell_value = inputs_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append((col, cell_value))
            else:
                break
        
        # Load test site
        final_portfolio_path = self.sites_path / "BOTN_COMPLETE_FINAL_PORTFOLIO_20250731_130415.xlsx"
        df = pd.read_excel(final_portfolio_path)
        test_site = df.iloc[0]
        
        # Analyze what can be auto-populated
        auto_fields = {
            'Property Name': test_site.get('Property Name'),
            'Address': test_site.get('Property Address'),
            'County:': test_site.get('County Name'),
            'CDLAC Region': self.get_cdlac_region(test_site.get('County Name')),
            'State:': test_site.get('State'),
            'Zip Code:': test_site.get('Zip'),
            'Purchase Price': test_site.get('For Sale Price'),
        }
        
        print("🏛️ BOTN TEMPLATE ANALYSIS")
        print("="*60)
        print(f"📊 Test Site: {test_site.get('Property Name')}")
        print(f"📍 Address: {test_site.get('Property Address')}")
        print(f"🏛️ County: {test_site.get('County Name')} (CDLAC Region: {auto_fields['CDLAC Region']})")
        print(f"💰 Purchase Price: ${test_site.get('For Sale Price'):,.0f}")
        
        print(f"\n✅ FIELDS THAT WILL BE AUTO-POPULATED ({len(auto_fields)}/17):")
        print("-" * 50)
        for field, value in auto_fields.items():
            if value is not None:
                print(f"  • {field}: {value}")
        
        print(f"\n⚠️  FIELDS REQUIRING MANUAL INPUT (10/17):")
        print("-" * 50)
        manual_fields = [
            ('Housing Type', 'e.g., New Construction, Rehab, Acquisition+Rehab'),
            ('Credit Pricing', 'e.g., 0.90 (typical range 0.80-1.00)'),
            ('Credit Type', 'e.g., 4%, 9%'),
            ('Construction Loan Term', 'e.g., 18, 24 (months)'),
            ('Market Cap Rate', 'e.g., 0.06 (as decimal)'),
            ('Financing Interest Rate', 'e.g., 0.07 (as decimal)'),
            ('Elevator', 'e.g., Yes, No'),
            ('# Units', 'e.g., 75 (typical LIHTC project size)'),
            ('Avg Unit Size', 'e.g., 950 (square feet)'),
            ('Hard Cost/SF', 'e.g., 250 (cost per square foot)')
        ]
        
        for field, help_text in manual_fields:
            print(f"  • {field}: {help_text}")
        
        print(f"\n📋 NEXT STEPS:")
        print("Please provide the values for the 10 manual input fields.")
        print("These values will be applied to ALL sites in your portfolio.")
        print("="*60)
        
        return headers, test_site, auto_fields, manual_fields
    
    def get_cdlac_region(self, county_name):
        """Determine CDLAC region based on county name"""
        cdlac_regions = {
            # Region 1 - Northern California
            'Alameda': 1, 'Contra Costa': 1, 'Lake': 1, 'Marin': 1, 'Napa': 1,
            'San Francisco': 1, 'San Mateo': 1, 'Santa Clara': 1, 'Solano': 1, 'Sonoma': 1,
            # Region 2 - Sacramento Valley
            'Butte': 2, 'Colusa': 2, 'El Dorado': 2, 'Glenn': 2, 'Nevada': 2,
            'Placer': 2, 'Sacramento': 2, 'Sutter': 2, 'Tehama': 2, 'Yolo': 2, 'Yuba': 2,
            # Region 3 - Northern Mountains
            'Alpine': 3, 'Amador': 3, 'Calaveras': 3, 'Del Norte': 3, 'Humboldt': 3,
            'Inyo': 3, 'Lassen': 3, 'Madera': 3, 'Mariposa': 3, 'Mendocino': 3,
            'Modoc': 3, 'Mono': 3, 'Plumas': 3, 'Shasta': 3, 'Sierra': 3, 'Siskiyou': 3,
            'Tuolumne': 3, 'Trinity': 3,
            # Region 4 - Central Valley
            'Fresno': 4, 'Kern': 4, 'Kings': 4, 'Merced': 4, 'San Joaquin': 4,
            'Stanislaus': 4, 'Tulare': 4,
            # Region 5 - Central Coast
            'Monterey': 5, 'San Benito': 5, 'San Luis Obispo': 5, 'Santa Barbara': 5,
            'Santa Cruz': 5, 'Ventura': 5,
            # Region 6 - Los Angeles County
            'Los Angeles': 6,
            # Region 7 - Orange County
            'Orange': 7,
            # Region 8 - Inland Empire
            'Riverside': 8, 'San Bernardino': 8,
            # Region 9 - San Diego County  
            'San Diego': 9, 'Imperial': 9
        }
        return cdlac_regions.get(county_name, None)

def main():
    generator = SimpleBOTNGenerator()
    headers, test_site, auto_fields, manual_fields = generator.analyze_requirements()

if __name__ == "__main__":
    main()