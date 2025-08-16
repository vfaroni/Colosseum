#!/usr/bin/env python3
"""
Inspect County Dropdown - Check what county values are in the template dropdown
"""

import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def inspect_county_dropdown():
    """Inspect the county dropdown options in the template"""
    
    base_path = Path(__file__).parent
    template_path = base_path / "botntemplate" / "CABOTNTemplate.xlsx"
    
    try:
        logger.info("🔍 Inspecting county dropdown in template...")
        workbook = openpyxl.load_workbook(template_path, data_only=False)
        inputs_sheet = workbook['Inputs']
        
        # Check for data validations on the Inputs sheet
        logger.info("\n📋 DATA VALIDATIONS FOUND:")
        if hasattr(inputs_sheet, 'data_validations'):
            for dv in inputs_sheet.data_validations.dataValidation:
                logger.info(f"Range: {dv.sqref}")
                if hasattr(dv, 'formula1') and dv.formula1:
                    logger.info(f"Formula: {dv.formula1}")
                if hasattr(dv, 'formula2') and dv.formula2:
                    logger.info(f"Formula2: {dv.formula2}")
                logger.info("---")
        
        # Also check if there's a hidden sheet with dropdown values
        logger.info(f"\n📄 ALL SHEETS IN TEMPLATE:")
        for sheet_name in workbook.sheetnames:
            logger.info(f"  • {sheet_name}")
            
        # Check if there's a Data sheet with county values
        if 'Data>>' in workbook.sheetnames:
            data_sheet = workbook['Data>>']
            logger.info(f"\n🔍 CHECKING Data>> SHEET:")
            
            # Look for county-related data in first few columns
            for row in range(1, 11):
                for col in range(1, 6):
                    cell = data_sheet.cell(row=row, column=col)
                    if cell.value and 'county' in str(cell.value).lower():
                        logger.info(f"Found at {cell.coordinate}: {cell.value}")
                        
                        # Check the column below for county values
                        logger.info(f"Values in column {cell.column_letter}:")
                        for check_row in range(row + 1, min(row + 21, data_sheet.max_row + 1)):
                            check_cell = data_sheet.cell(row=check_row, column=col)
                            if check_cell.value:
                                logger.info(f"  • {check_cell.value}")
                            else:
                                break
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error inspecting template: {str(e)}")

if __name__ == "__main__":
    inspect_county_dropdown()