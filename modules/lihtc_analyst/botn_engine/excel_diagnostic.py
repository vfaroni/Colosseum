#!/usr/bin/env python3
"""
Excel Diagnostic Tool - Analyze template structure and test different approaches
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelDiagnostic:
    """Diagnose Excel template issues"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
    
    def analyze_template_structure(self):
        """Analyze the Excel template structure to understand potential issues"""
        
        logger.info("🔍 ANALYZING EXCEL TEMPLATE STRUCTURE")
        logger.info("="*60)
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(self.template_path, data_only=False)
            
            logger.info(f"📊 Workbook: {self.template_path.name}")
            logger.info(f"📄 Worksheets: {workbook.sheetnames}")
            
            # Focus on Inputs sheet
            if 'Inputs' in workbook.sheetnames:
                inputs_sheet = workbook['Inputs']
                
                logger.info(f"\n📋 INPUTS SHEET ANALYSIS:")
                logger.info("-" * 40)
                logger.info(f"Max Row: {inputs_sheet.max_row}")
                logger.info(f"Max Column: {inputs_sheet.max_column}")
                
                # Check Row 1 (headers)
                logger.info(f"\n🏷️  ROW 1 HEADERS:")
                headers = []
                for col in range(1, inputs_sheet.max_column + 1):
                    cell = inputs_sheet.cell(row=1, column=col)
                    if cell.value:
                        headers.append(f"Col {col}: {cell.value}")
                        # Check for data validation
                        if cell.data_type:
                            logger.info(f"  Col {col} Data Type: {cell.data_type}")
                
                for header in headers:
                    logger.info(f"  {header}")
                
                # Check Row 2 (input values)
                logger.info(f"\n📝 ROW 2 CURRENT VALUES:")
                for col in range(1, inputs_sheet.max_column + 1):
                    cell = inputs_sheet.cell(row=2, column=col)
                    if cell.value is not None:
                        logger.info(f"  Col {col}: {cell.value} (type: {type(cell.value)})")
                
                # Check for data validation rules
                logger.info(f"\n🔒 DATA VALIDATION RULES:")
                validation_found = False
                for col in range(1, inputs_sheet.max_column + 1):
                    cell = inputs_sheet.cell(row=2, column=col)
                    if hasattr(cell, 'data_validation') and cell.data_validation:
                        logger.info(f"  Col {col}: Has data validation")
                        validation_found = True
                
                if not validation_found:
                    logger.info("  No data validation rules found")
                
                # Check for conditional formatting
                logger.info(f"\n🎨 CONDITIONAL FORMATTING:")
                if inputs_sheet.conditional_formatting:
                    logger.info(f"  Found {len(inputs_sheet.conditional_formatting)} conditional formatting rules")
                else:
                    logger.info("  No conditional formatting found")
                
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
    
    def test_safe_modification_approach(self):
        """Test a safer approach to modifying the Excel file"""
        
        logger.info(f"\n🧪 TESTING SAFE MODIFICATION APPROACHES")
        logger.info("="*60)
        
        try:
            # Approach 1: Copy template first, then modify
            logger.info("📋 Approach 1: Copy template first")
            
            import shutil
            temp_file = self.base_path / "temp_botn_test.xlsx"
            
            # Copy template
            shutil.copy2(self.template_path, temp_file)
            
            # Load the copy
            workbook = openpyxl.load_workbook(temp_file)
            inputs_sheet = workbook['Inputs']
            
            # Test data that should work
            test_data = {
                1: "Test Property Name",
                2: "123 Test Address", 
                3: "Test County",
                4: "Los Angeles County",
                5: "CA",
                6: "90210",
                7: 1000000,  # Number
                8: "New Construction",  # Text
                9: 0.8,      # Decimal
                10: "4%",    # Text with %
                11: 36,      # Integer
                12: 0.05,    # Decimal
                13: 0.06,    # Decimal
                14: "No",    # Text
                15: 100,     # Integer
                16: 900,     # Integer
                17: 250      # Integer
            }
            
            # Apply test data with careful type handling
            for col_num, value in test_data.items():
                cell = inputs_sheet.cell(row=2, column=col_num)
                
                # Preserve number formatting for numeric cells
                if isinstance(value, (int, float)):
                    cell.value = value
                    if col_num in [9, 12, 13]:  # Rate fields
                        cell.number_format = '0.00%' if value < 1 else '0.00'
                    elif col_num == 7:  # Purchase price
                        cell.number_format = '$#,##0'
                else:
                    cell.value = value
            
            # Save with a test name
            test_output = self.base_path / "outputs" / "TEST_BOTN_SAFE_APPROACH.xlsx"
            workbook.save(test_output)
            workbook.close()
            
            # Clean up temp file
            temp_file.unlink()
            
            logger.info(f"✅ Safe approach test file created: {test_output}")
            logger.info("   Try opening this file to see if the error persists")
            
        except Exception as e:
            logger.error(f"Error in safe modification test: {str(e)}")
    
    def suggest_solutions(self):
        """Suggest solutions for the Excel corruption issue"""
        
        logger.info(f"\n💡 SUGGESTED SOLUTIONS")
        logger.info("="*60)
        
        solutions = [
            "1. **Data Type Preservation**: Ensure numeric values maintain proper Excel formatting",
            "2. **Template Copying**: Copy template first, then modify (preserves structure better)",
            "3. **Number Formatting**: Apply proper Excel number formats for rates, currency, etc.",
            "4. **Alternative Library**: Consider using xlwings instead of openpyxl for better compatibility",
            "5. **Validation Handling**: Remove/update data validation rules that conflict with new data",
            "6. **Formula Preservation**: Ensure formulas in other sheets still reference correct cells"
        ]
        
        for solution in solutions:
            logger.info(f"  {solution}")
        
        logger.info(f"\n🔧 IMMEDIATE NEXT STEPS:")
        logger.info("1. Test the 'TEST_BOTN_SAFE_APPROACH.xlsx' file created above")
        logger.info("2. If it opens without errors, we'll use this safer approach")
        logger.info("3. If errors persist, we'll try xlwings or other alternatives")

def main():
    diagnostic = ExcelDiagnostic()
    diagnostic.analyze_template_structure()
    diagnostic.test_safe_modification_approach()
    diagnostic.suggest_solutions()

if __name__ == "__main__":
    main()