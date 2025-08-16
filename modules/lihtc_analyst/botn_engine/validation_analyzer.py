#!/usr/bin/env python3
"""
Data Validation Analyzer - Find and preserve Excel data validations
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ValidationAnalyzer:
    """Analyze and preserve Excel data validations"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
    
    def deep_analyze_validations(self):
        """Deep analysis of data validations in the template"""
        
        logger.info("🔍 DEEP DATA VALIDATION ANALYSIS")
        logger.info("="*60)
        
        try:
            # Load with data_only=False to preserve formulas and validations
            workbook = openpyxl.load_workbook(self.template_path, data_only=False, keep_vba=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"\n📄 Sheet: {sheet_name}")
                
                # Check worksheet-level data validations
                if hasattr(sheet, 'data_validations') and sheet.data_validations:
                    logger.info(f"  📋 Found {len(sheet.data_validations.dataValidation)} data validations")
                    
                    for i, dv in enumerate(sheet.data_validations.dataValidation):
                        logger.info(f"    Validation {i+1}:")
                        logger.info(f"      Type: {dv.type}")
                        logger.info(f"      Formula1: {dv.formula1}")
                        logger.info(f"      Formula2: {dv.formula2}")
                        logger.info(f"      Ranges: {dv.sqref}")
                        if dv.formula1:
                            logger.info(f"      List values: {dv.formula1}")
                else:
                    logger.info(f"  ❌ No data validations found")
                
                # Check for specific cells with validations in Inputs sheet
                if sheet_name == 'Inputs':
                    logger.info(f"\n  🔍 Checking individual cells in Row 2:")
                    for col in range(1, 18):  # Check first 17 columns
                        cell = sheet.cell(row=2, column=col)
                        if hasattr(cell, 'data_validation') and cell.data_validation:
                            logger.info(f"    Col {col}: Has individual validation")
                        
                        # Check if cell is part of any validation range
                        cell_coord = cell.coordinate
                        for dv in sheet.data_validations.dataValidation:
                            if cell_coord in dv.cells:
                                logger.info(f"    Col {col} ({cell_coord}): Part of validation range {dv.sqref}")
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error analyzing validations: {str(e)}")
    
    def test_xlsxwriter_approach(self):
        """Test using xlsxwriter which better handles Excel features"""
        
        logger.info(f"\n🧪 TESTING XLSXWRITER APPROACH")
        logger.info("="*50)
        
        try:
            import xlsxwriter
            
            # Create new workbook with xlsxwriter
            output_path = self.base_path / "outputs" / "TEST_XLSXWRITER_BOTN.xlsx"
            workbook = xlsxwriter.Workbook(str(output_path))
            
            # Create Inputs worksheet
            inputs_sheet = workbook.add_worksheet('Inputs')
            
            # Add headers
            headers = [
                'Property Name', 'Address', 'County:', 'CDLAC Region', 'State:', 'Zip Code:', 
                'Purchase Price', 'Housing Type', 'Credit Pricing', 'Credit Type', 
                'Construction Loan Term', 'Market Cap Rate', 'Financing Interest Rate', 
                'Elevator', '# Units', 'Avg Unit Size', 'Hard Cost/SF'
            ]
            
            for col, header in enumerate(headers):
                inputs_sheet.write(0, col, header)
            
            # Add test data
            test_data = [
                "Test Property", "123 Test St", "Los Angeles", "Los Angeles County", "CA", "90210",
                1000000, "New Construction", 0.8, "4%", 36, 0.05, 0.06, "Non-Elevator", 100, 900, 250
            ]
            
            for col, value in enumerate(test_data):
                inputs_sheet.write(1, col, value)
            
            # Add data validation for CDLAC Region (Column D, row 2)
            cdlac_options = [
                'Northern Region', 'Capital Region', 'Central Valley Region', 
                'Central Coast Region', 'East Bay Region', 'South & West Bay Region',
                'Inland Empire Region', 'Los Angeles County', 'Orange County',
                'San Diego County', 'San Francisco County'
            ]
            
            inputs_sheet.data_validation('D2', {
                'validate': 'list',
                'source': cdlac_options,
                'dropdown': True
            })
            
            # Add validation for Housing Type
            housing_options = ['New Construction', 'Rehab', 'Acquisition+Rehab']
            inputs_sheet.data_validation('H2', {
                'validate': 'list', 
                'source': housing_options,
                'dropdown': True
            })
            
            # Add validation for Elevator
            elevator_options = ['Elevator', 'Non-Elevator']
            inputs_sheet.data_validation('N2', {
                'validate': 'list',
                'source': elevator_options, 
                'dropdown': True
            })
            
            workbook.close()
            
            logger.info(f"✅ XlsxWriter test file created: {output_path}")
            logger.info("   This should open without corruption and preserve dropdowns")
            
        except ImportError:
            logger.error("xlsxwriter not installed. Install with: pip install xlsxwriter")
        except Exception as e:
            logger.error(f"Error with xlsxwriter approach: {str(e)}")
    
    def suggest_final_solution(self):
        """Suggest the best approach based on analysis"""
        
        logger.info(f"\n💡 RECOMMENDED SOLUTION")
        logger.info("="*50)
        
        logger.info("Based on the analysis, here are the best approaches:")
        logger.info("")
        logger.info("1. **XlsxWriter Approach (RECOMMENDED)**:")
        logger.info("   - Create new Excel file from scratch")
        logger.info("   - Add data validations properly") 
        logger.info("   - Better compatibility with Excel")
        logger.info("   - No corruption issues")
        logger.info("")
        logger.info("2. **Manual Template Recreation**:")
        logger.info("   - Recreate template with proper validations")
        logger.info("   - Use openpyxl with manual validation setup")
        logger.info("")
        logger.info("3. **Hybrid Approach**:")
        logger.info("   - Use xlsxwriter for data input sheet")
        logger.info("   - Copy other sheets from original template")

def main():
    analyzer = ValidationAnalyzer()
    analyzer.deep_analyze_validations()
    analyzer.test_xlsxwriter_approach()
    analyzer.suggest_final_solution()

if __name__ == "__main__":
    main()