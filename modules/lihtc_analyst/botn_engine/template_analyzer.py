#!/usr/bin/env python3
"""
Template Analyzer - Analyze the original template to understand the calculation structure
"""

import pandas as pd
import openpyxl
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TemplateAnalyzer:
    """Analyze original template structure and formulas"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
    
    def analyze_all_sheets(self):
        """Analyze all sheets in the template to understand the calculation flow"""
        
        logger.info("🔍 ANALYZING TEMPLATE CALCULATION STRUCTURE")
        logger.info("="*70)
        
        try:
            # Load with data_only=False to see formulas
            workbook = openpyxl.load_workbook(self.template_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"\n📄 SHEET: {sheet_name}")
                logger.info("-" * 40)
                
                sheet = workbook[sheet_name]
                
                # Find cells with formulas that reference Inputs
                formula_cells = []
                for row in range(1, min(sheet.max_row + 1, 21)):  # Check first 20 rows
                    for col in range(1, min(sheet.max_column + 1, 11)):  # Check first 10 columns
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            if 'Inputs' in cell.value:
                                formula_cells.append({
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'formula': cell.value,
                                    'description': self.describe_cell_content(sheet, row, col)
                                })
                
                if formula_cells:
                    logger.info(f"🔗 Found {len(formula_cells)} cells referencing Inputs:")
                    for fc in formula_cells[:5]:  # Show first 5
                        logger.info(f"  {fc['cell']}: {fc['formula'][:50]}...")
                        if fc['description']:
                            logger.info(f"    Context: {fc['description']}")
                else:
                    # Look for any significant data or structure
                    self.analyze_sheet_content(sheet, sheet_name)
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
    
    def describe_cell_content(self, sheet, row, col):
        """Try to understand what a cell represents based on nearby content"""
        try:
            # Check cell to the left (might be a label)
            if col > 1:
                left_cell = sheet.cell(row=row, column=col-1)
                if left_cell.value and isinstance(left_cell.value, str):
                    return left_cell.value
            
            # Check cell above (might be a header)
            if row > 1:
                above_cell = sheet.cell(row=row-1, column=col)
                if above_cell.value and isinstance(above_cell.value, str):
                    return f"Under: {above_cell.value}"
            
            return None
        except:
            return None
    
    def analyze_sheet_content(self, sheet, sheet_name):
        """Analyze general content of a sheet"""
        
        # Look for headers or important text in first few rows
        important_content = []
        for row in range(1, 6):
            for col in range(1, 6):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and len(cell.value) > 2:
                    important_content.append(f"{cell.coordinate}: {cell.value}")
        
        if important_content:
            logger.info("📋 Key content found:")
            for content in important_content[:3]:  # Show first 3
                logger.info(f"  {content}")
        else:
            logger.info("❌ No significant content found")
    
    def suggest_simple_recreation_approach(self):
        """Suggest how to recreate essential functionality with xlsxwriter"""
        
        logger.info(f"\n💡 RECREATION STRATEGY")
        logger.info("="*50)
        
        logger.info("Based on typical BOTN structure, we can recreate with xlsxwriter:")
        logger.info("")
        logger.info("1. **Pure XlsxWriter Approach**:")
        logger.info("   - Inputs tab: Clean data entry with dropdowns")
        logger.info("   - Basic calculations: Essential LIHTC formulas") 
        logger.info("   - Simple structure: Avoid complex Excel features")
        logger.info("")
        logger.info("2. **Template Bypass Approach**:")
        logger.info("   - Don't touch original template at all")
        logger.info("   - Create completely new workbook")
        logger.info("   - Implement key BOTN calculations manually")
        logger.info("")
        logger.info("3. **CSV + Manual Approach**:")
        logger.info("   - Export data to CSV")
        logger.info("   - User manually pastes into original template")
        logger.info("   - Preserves all original functionality")

def main():
    analyzer = TemplateAnalyzer()
    analyzer.analyze_all_sheets()
    analyzer.suggest_simple_recreation_approach()

if __name__ == "__main__":
    main()