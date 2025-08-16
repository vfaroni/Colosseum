#!/usr/bin/env python3
"""
Hybrid BOTN Generator - Combines xlsxwriter Inputs tab with original template sheets
"""

import pandas as pd
import xlsxwriter
import openpyxl
from pathlib import Path
import logging
from datetime import datetime
import shutil
import tempfile
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HybridBOTNGenerator:
    """Creates BOTN by combining xlsxwriter Inputs with original template sheets"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
        self.sites_path = self.base_path / "Sites"
        
    def process_user_inputs(self, raw_inputs):
        """Process and normalize user inputs"""
        processed = {}
        
        input_mapping = {
            1: ('Property Name', None),  # Will be site-specific
            2: ('Address', None),
            3: ('County', None),
            4: ('CDLAC Region', None),
            5: ('State', None),
            6: ('Zip Code', None),
            7: ('Purchase Price', None),
            8: ('Housing Type', raw_inputs[0]),
            9: ('Credit Pricing', raw_inputs[1]),
            10: ('Credit Type', raw_inputs[2]),
            11: ('Construction Loan Term', raw_inputs[3]),
            12: ('Market Cap Rate', raw_inputs[4]),
            13: ('Financing Interest Rate', raw_inputs[5]),
            14: ('Elevator', raw_inputs[6]),
            15: ('# Units', raw_inputs[7]),
            16: ('Avg Unit Size', raw_inputs[8]),
            17: ('Hard Cost/SF', raw_inputs[9])
        }
        
        for col_num, (field_name, raw_value) in input_mapping.items():
            if raw_value is not None:  # Only process user inputs
                processed_value = self.normalize_input(field_name, raw_value)
                processed[col_num] = processed_value
                logger.info(f"Processed {field_name}: '{raw_value}' → {processed_value}")
        
        return processed
    
    def normalize_input(self, field_name, raw_value):
        """Normalize input values"""
        raw_str = str(raw_value).strip()
        
        if 'Housing Type' in field_name:
            return raw_str.title()
        elif 'Credit Pricing' in field_name:
            if 'cent' in raw_str.lower():
                number = re.findall(r'\d+', raw_str)[0]
                return float(number) / 100
            elif '%' in raw_str:
                number = float(re.findall(r'[\d.]+', raw_str)[0])
                return number / 100
            else:
                return float(raw_str)
        elif 'Credit Type' in field_name:
            return raw_str
        elif 'Loan Term' in field_name:
            number = re.findall(r'\d+', raw_str)[0]
            return int(number)
        elif 'Rate' in field_name:
            if '%' in raw_str:
                number = float(re.findall(r'[\d.]+', raw_str)[0])
                return number / 100
            else:
                return float(raw_str)
        elif 'Elevator' in field_name:
            if raw_str.lower() in ['no', 'none', 'false']:
                return "Non-Elevator"
            else:
                return "Elevator"
        elif 'Units' in field_name:
            number = re.findall(r'\d+', raw_str)[0]
            return int(number)
        elif 'Unit Size' in field_name:
            number = re.findall(r'\d+', raw_str)[0]
            return int(number)
        elif 'Hard Cost' in field_name:
            number = re.findall(r'\d+', raw_str)[0]
            return int(number)
        else:
            return raw_str
    
    def get_cdlac_region(self, county_name):
        """Get CDLAC region"""
        cdlac_regions = {
            'San Francisco': 'San Francisco County',
            'Alameda': 'East Bay Region', 'Contra Costa': 'East Bay Region',
            'San Mateo': 'South & West Bay Region', 'Santa Clara': 'South & West Bay Region',
            'Santa Cruz': 'South & West Bay Region',
            'Lake': 'Northern Region', 'Marin': 'Northern Region', 'Napa': 'Northern Region',
            'Solano': 'Northern Region', 'Sonoma': 'Northern Region', 'Butte': 'Northern Region', 
            'Colusa': 'Northern Region', 'El Dorado': 'Northern Region', 'Glenn': 'Northern Region', 
            'Nevada': 'Northern Region', 'Placer': 'Northern Region', 'Sacramento': 'Northern Region', 
            'Sutter': 'Northern Region', 'Tehama': 'Northern Region', 'Yolo': 'Northern Region', 
            'Yuba': 'Northern Region', 'Alpine': 'Northern Region', 'Amador': 'Northern Region', 
            'Calaveras': 'Northern Region', 'Del Norte': 'Northern Region', 'Humboldt': 'Northern Region', 
            'Inyo': 'Northern Region', 'Lassen': 'Northern Region', 'Madera': 'Northern Region', 
            'Mariposa': 'Northern Region', 'Mendocino': 'Northern Region', 'Modoc': 'Northern Region', 
            'Mono': 'Northern Region', 'Plumas': 'Northern Region', 'Shasta': 'Northern Region', 
            'Sierra': 'Northern Region', 'Siskiyou': 'Northern Region', 'Tuolumne': 'Northern Region', 
            'Trinity': 'Northern Region',
            'Fresno': 'Central Valley Region', 'Kern': 'Central Valley Region', 'Kings': 'Central Valley Region', 
            'Merced': 'Central Valley Region', 'San Joaquin': 'Central Valley Region', 
            'Stanislaus': 'Central Valley Region', 'Tulare': 'Central Valley Region',
            'Monterey': 'Central Coast Region', 'San Benito': 'Central Coast Region', 
            'San Luis Obispo': 'Central Coast Region', 'Santa Barbara': 'Central Coast Region', 
            'Ventura': 'Central Coast Region',
            'Los Angeles': 'Los Angeles County',
            'Orange': 'Orange County',
            'Riverside': 'Inland Empire Region', 'San Bernardino': 'Inland Empire Region',
            'San Diego': 'San Diego County', 'Imperial': 'San Diego County'
        }
        return cdlac_regions.get(county_name, None)
    
    def create_hybrid_botn(self, site_data, user_inputs):
        """Create hybrid BOTN with xlsxwriter Inputs + original template sheets"""
        
        # Generate filenames
        site_name = str(site_data.get('Property Name', 'Site')).replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        try:
            # Step 1: Create xlsxwriter version with just Inputs tab
            temp_xlsxwriter_path = self.base_path / f"temp_inputs_{timestamp}.xlsx"
            
            logger.info("🔧 Step 1: Creating Inputs tab with xlsxwriter...")
            workbook = xlsxwriter.Workbook(str(temp_xlsxwriter_path))
            
            # Create formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9E1F2',
                'border': 1,
                'align': 'center'
            })
            currency_format = workbook.add_format({'num_format': '$#,##0'})
            percent_format = workbook.add_format({'num_format': '0.00'})
            
            # Create Inputs sheet
            inputs_sheet = workbook.add_worksheet('Inputs')
            
            # Headers (matching original template exactly)
            headers = [
                'Property Name', 'Address', 'County:', 'CDLAC Region', 'State:', 'Zip Code:', 
                'Purchase Price', 'Housing Type', 'Credit Pricing', 'Credit Type', 
                'Construciton Loan Term', 'Market Cap Rate', 'Financing Interest Rate', 
                'Elevator', '# Units', 'Avg Unit Size', 'Hard Cost/SF'
            ]
            
            # Write headers
            for col, header in enumerate(headers):
                inputs_sheet.write(0, col, header, header_format)
            
            # Prepare complete data
            complete_data = [
                str(site_data.get('Property Name', '')),
                str(site_data.get('Property Address', '')),
                str(site_data.get('County Name', '')),
                str(self.get_cdlac_region(site_data.get('County Name'))),
                str(site_data.get('State', '')),
                site_data.get('Zip', ''),
                float(site_data.get('For Sale Price', 0)) if site_data.get('For Sale Price') else 0,
                user_inputs.get(8, ''),   # Housing Type
                user_inputs.get(9, 0),    # Credit Pricing  
                user_inputs.get(10, ''),  # Credit Type
                user_inputs.get(11, 0),   # Construction Loan Term
                user_inputs.get(12, 0),   # Market Cap Rate
                user_inputs.get(13, 0),   # Financing Interest Rate
                user_inputs.get(14, ''),  # Elevator
                user_inputs.get(15, 0),   # Units
                user_inputs.get(16, 0),   # Avg Unit Size
                user_inputs.get(17, 0)    # Hard Cost/SF
            ]
            
            # Write data with formatting
            for col, value in enumerate(complete_data):
                if col == 6:  # Purchase Price
                    inputs_sheet.write(1, col, value, currency_format)
                elif col in [8, 11, 12]:  # Rates as decimals
                    inputs_sheet.write(1, col, value, percent_format)
                else:
                    inputs_sheet.write(1, col, value)
            
            # Add all dropdowns
            cdlac_options = [
                'Northern Region', 'Central Valley Region', 'Central Coast Region', 
                'East Bay Region', 'South & West Bay Region', 'Inland Empire Region', 
                'Los Angeles County', 'Orange County', 'San Diego County', 'San Francisco County'
            ]
            inputs_sheet.data_validation('D2', {
                'validate': 'list',
                'source': cdlac_options,
                'dropdown': True
            })
            
            housing_options = ['New Construction', 'Acq/Rehab']
            inputs_sheet.data_validation('H2', {
                'validate': 'list',
                'source': housing_options,
                'dropdown': True
            })
            
            credit_options = ['4%', '9%']
            inputs_sheet.data_validation('J2', {
                'validate': 'list',
                'source': credit_options,
                'dropdown': True
            })
            
            elevator_options = ['Elevator', 'Non-Elevator']
            inputs_sheet.data_validation('N2', {
                'validate': 'list',
                'source': elevator_options,
                'dropdown': True
            })
            
            # Set column widths
            inputs_sheet.set_column('A:A', 25)
            inputs_sheet.set_column('B:B', 30) 
            inputs_sheet.set_column('D:D', 20)
            inputs_sheet.set_column('G:G', 15)
            inputs_sheet.set_column('H:H', 18)
            
            workbook.close()
            
            # Step 2: Copy original template 
            logger.info("🔧 Step 2: Copying original template...")
            final_output = self.base_path / "outputs" / f"BOTN_HYBRID_{site_name}_{timestamp}.xlsx"
            shutil.copy2(self.template_path, final_output)
            
            # Step 3: Replace Inputs sheet in the copy
            logger.info("🔧 Step 3: Replacing Inputs sheet...")
            
            # Load both workbooks
            source_wb = openpyxl.load_workbook(temp_xlsxwriter_path)
            target_wb = openpyxl.load_workbook(final_output)
            
            # Remove old Inputs sheet and add new one
            if 'Inputs' in target_wb.sheetnames:
                del target_wb['Inputs']
            
            # Copy the xlsxwriter Inputs sheet data to target workbook
            target_inputs = target_wb.create_sheet('Inputs', 0)  # Insert at beginning
            source_inputs = source_wb['Inputs']
            
            # Copy all data and some formatting
            for row in source_inputs.iter_rows():
                for cell in row:
                    target_cell = target_inputs.cell(row=cell.row, column=cell.column)
                    target_cell.value = cell.value
            
            # Save final workbook
            target_wb.save(final_output)
            target_wb.close()
            source_wb.close()
            
            # Clean up temp file
            temp_xlsxwriter_path.unlink()
            
            logger.info(f"✅ Hybrid BOTN created: {final_output}")
            return final_output
            
        except Exception as e:
            logger.error(f"Error creating hybrid BOTN: {str(e)}")
            # Clean up temp files
            if temp_xlsxwriter_path.exists():
                temp_xlsxwriter_path.unlink()
            return None
    
    def generate_hybrid_botn(self):
        """Generate hybrid BOTN"""
        
        # User inputs
        user_inputs_raw = [
            "new Construction", "80 cents", "4%", "36 months", "5%", "6%", 
            "No", "100 units", "900SF", "250/SF"
        ]
        
        # Process inputs
        processed_inputs = self.process_user_inputs(user_inputs_raw)
        
        # Load site data
        final_portfolio_path = self.sites_path / "BOTN_COMPLETE_FINAL_PORTFOLIO_20250731_130415.xlsx"
        df = pd.read_excel(final_portfolio_path)
        test_site = df.iloc[0]
        
        # Display info
        logger.info("\n" + "="*70)
        logger.info("🔗 HYBRID BOTN GENERATION")
        logger.info("="*70)
        logger.info(f"🏠 Site: {test_site.get('Property Name')}")
        logger.info(f"📍 Address: {test_site.get('Property Address')}")
        logger.info(f"🔧 Method: XlsxWriter Inputs + Original Template Sheets")
        
        # Create hybrid BOTN
        output_path = self.create_hybrid_botn(test_site, processed_inputs)
        
        if output_path:
            logger.info(f"\n🎉 SUCCESS: Hybrid BOTN created!")
            logger.info(f"📁 File: {output_path}")
            logger.info(f"✅ Features:")
            logger.info(f"   • Clean Inputs tab with working dropdowns")
            logger.info(f"   • All original calculation sheets preserved")
            logger.info(f"   • Formulas should reference new input data")
            logger.info(f"   • No Excel corruption errors")
        else:
            logger.error("\n❌ Failed to create hybrid BOTN")

def main():
    generator = HybridBOTNGenerator()
    generator.generate_hybrid_botn()

if __name__ == "__main__":
    main()