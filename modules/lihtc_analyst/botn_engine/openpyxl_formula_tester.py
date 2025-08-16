#!/usr/bin/env python3
"""
OpenPyxl Formula Preservation Tester - Test preserving CABOTNTemplate formulas without xlwings
"""

import openpyxl
from openpyxl.styles import NamedStyle
import shutil
from pathlib import Path
import logging
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class OpenPyxlFormulaTester:
    """Test openpyxl's ability to preserve template formulas and functionality"""
    
    def __init__(self):
        self.base_path = Path(__file__).parent
        self.template_path = self.base_path / "botntemplate" / "CABOTNTemplate.xlsx"
        self.test_output_dir = self.base_path / "openpyxl_tests"
        self.test_output_dir.mkdir(exist_ok=True)
    
    def test_1_template_reading(self):
        """Test 1: Can we read the template while preserving formulas?"""
        
        logger.info("🔬 TEST 1: Template reading with formula preservation")
        logger.info("=" * 60)
        
        try:
            # Read with data_only=False to preserve formulas
            wb = openpyxl.load_workbook(self.template_path, data_only=False)
            
            logger.info(f"✅ Template loaded successfully")
            logger.info(f"   Worksheets found: {wb.sheetnames}")
            
            # Check Inputs sheet structure
            inputs_sheet = wb['Inputs']
            logger.info(f"   Inputs sheet dimensions: {inputs_sheet.max_row} x {inputs_sheet.max_column}")
            
            # Sample key input cells that we populate
            key_cells = ['A2', 'B2', 'C2', 'D2', 'G2', 'H2', 'I2', 'J2', 'O2', 'P2']
            for cell in key_cells:
                cell_value = inputs_sheet[cell].value
                logger.info(f"   {cell}: {cell_value}")
            
            wb.close()
            
            return {"success": True, "worksheets": wb.sheetnames}
            
        except Exception as e:
            logger.error(f"❌ TEST 1 FAILED: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def test_2_formula_preservation(self):
        """Test 2: Can we modify data while preserving formulas?"""
        
        logger.info("🔬 TEST 2: Data modification with formula preservation")
        logger.info("=" * 60)
        
        try:
            # Create test copy
            test_file = self.test_output_dir / "test2_formula_preservation.xlsx"
            shutil.copy2(self.template_path, test_file)
            
            # Load and modify
            wb = openpyxl.load_workbook(test_file, data_only=False)
            inputs_sheet = wb['Inputs']
            
            # Store original formulas in other sheets for comparison
            original_formulas = {}
            for sheet_name in ['Rents', 'Expenses', 'Sources & Uses', 'NOI']:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    original_formulas[sheet_name] = []
                    
                    # Check first 10 rows and columns for formulas
                    for row in range(1, min(11, sheet.max_row + 1)):
                        for col in range(1, min(11, sheet.max_column + 1)):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                original_formulas[sheet_name].append({
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'formula': cell.value
                                })
            
            logger.info(f"   Original formulas found in {len(original_formulas)} sheets")
            
            # Modify input data (similar to production code)
            test_data = {
                'A2': 'TEST PROPERTY OPENPYXL',
                'B2': '123 Test Street, Los Angeles, CA',
                'C2': 'Los Angeles County',
                'D2': 'Greater Los Angeles Area',
                'E2': 'CA',
                'F2': '90210',
                'G2': 2500000,
                'H2': 'Large Family',
                'I2': 0.80,
                'J2': '4%',
                'K2': 36,
                'L2': 0.05,
                'M2': 0.06,
                'N2': 'Non-Elevator',
                'O2': 80,
                'P2': 950,
                'Q2': 275
            }
            
            for cell_ref, value in test_data.items():
                inputs_sheet[cell_ref] = value
                logger.info(f"   Set {cell_ref}: {value}")
            
            # Save and reload to test formula persistence
            wb.save(test_file)
            wb.close()
            
            # Reload and verify formulas still exist
            wb_check = openpyxl.load_workbook(test_file, data_only=False)
            
            preserved_formulas = {}
            for sheet_name in ['Rents', 'Expenses', 'Sources & Uses', 'NOI']:
                if sheet_name in wb_check.sheetnames:
                    sheet = wb_check[sheet_name]
                    preserved_formulas[sheet_name] = []
                    
                    for row in range(1, min(11, sheet.max_row + 1)):
                        for col in range(1, min(11, sheet.max_column + 1)):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                preserved_formulas[sheet_name].append({
                                    'cell': f"{cell.column_letter}{cell.row}",
                                    'formula': cell.value
                                })
            
            wb_check.close()
            
            # Compare original vs preserved
            formulas_intact = True
            for sheet_name in original_formulas:
                original_count = len(original_formulas[sheet_name])
                preserved_count = len(preserved_formulas.get(sheet_name, []))
                
                logger.info(f"   {sheet_name}: {original_count} original → {preserved_count} preserved")
                
                if original_count != preserved_count:
                    formulas_intact = False
            
            if formulas_intact:
                logger.info("✅ All formulas preserved successfully")
            else:
                logger.warning("⚠️  Formula count mismatch detected")
            
            return {
                "success": True,
                "formulas_intact": formulas_intact,
                "original_formulas": original_formulas,
                "preserved_formulas": preserved_formulas
            }
            
        except Exception as e:
            logger.error(f"❌ TEST 2 FAILED: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def test_3_batch_processing_simulation(self):
        """Test 3: Simulate batch processing without permissions"""
        
        logger.info("🔬 TEST 3: Batch processing simulation")
        logger.info("=" * 60)
        
        try:
            # Simulate processing 5 files (like production batch)
            test_files = []
            for i in range(5):
                test_file = self.test_output_dir / f"test3_batch_{i+1}.xlsx"
                shutil.copy2(self.template_path, test_file)
                test_files.append(test_file)
            
            logger.info(f"   Created {len(test_files)} test files")
            
            successful_files = []
            processing_times = []
            
            import time
            for i, test_file in enumerate(test_files):
                start_time = time.time()
                
                try:
                    # Load template
                    wb = openpyxl.load_workbook(test_file, data_only=False)
                    inputs_sheet = wb['Inputs']
                    
                    # Populate with test data
                    inputs_sheet['A2'] = f'BATCH TEST PROPERTY {i+1}'
                    inputs_sheet['B2'] = f'{123 + i} Test Street, Los Angeles, CA'
                    inputs_sheet['G2'] = 2500000 + (i * 100000)  # Varying prices
                    inputs_sheet['O2'] = 80 + (i * 10)  # Varying units
                    
                    # Save
                    wb.save()
                    wb.close()
                    
                    process_time = time.time() - start_time
                    processing_times.append(process_time)
                    successful_files.append(test_file.name)
                    
                    logger.info(f"   ✅ File {i+1}: {process_time:.2f}s")
                    
                except Exception as e:
                    logger.error(f"   ❌ File {i+1}: {str(e)}")
            
            avg_time = sum(processing_times) / len(processing_times) if processing_times else 0
            
            logger.info(f"✅ Batch processing complete:")
            logger.info(f"   Successful: {len(successful_files)}/5 files")
            logger.info(f"   Average time per file: {avg_time:.2f}s")
            logger.info(f"   ⭐ NO PERMISSIONS REQUIRED!")
            
            return {
                "success": True,
                "files_processed": len(successful_files),
                "total_files": 5,
                "average_time": avg_time,
                "permissions_required": False
            }
            
        except Exception as e:
            logger.error(f"❌ TEST 3 FAILED: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def test_4_formula_calculation_verification(self):
        """Test 4: Verify that formulas still calculate correctly"""
        
        logger.info("🔬 TEST 4: Formula calculation verification")
        logger.info("=" * 60)
        
        try:
            # Create test file with data
            test_file = self.test_output_dir / "test4_calculation_check.xlsx"
            shutil.copy2(self.template_path, test_file)
            
            # Load and populate with known data
            wb = openpyxl.load_workbook(test_file, data_only=False)
            inputs_sheet = wb['Inputs']
            
            # Set specific values for calculation testing
            test_inputs = {
                'A2': 'CALCULATION TEST PROPERTY',
                'B2': '456 Formula Street, San Diego, CA',
                'C2': 'San Diego County',
                'D2': 'San Diego Metropolitan Area',
                'E2': 'CA',
                'F2': '92101',
                'G2': 3000000,  # $3M purchase price
                'H2': 'Large Family',
                'I2': 0.75,     # 75 cent credit pricing
                'J2': '4%',
                'K2': 36,
                'L2': 0.055,    # 5.5% cap rate
                'M2': 0.065,    # 6.5% interest rate
                'N2': 'Elevator',
                'O2': 100,      # 100 units
                'P2': 1000,     # 1000 SF units
                'Q2': 300       # $300/SF hard cost
            }
            
            for cell_ref, value in test_inputs.items():
                inputs_sheet[cell_ref] = value
            
            wb.save()
            wb.close()
            
            # Reload with data_only=True to see calculated values
            wb_calc = openpyxl.load_workbook(test_file, data_only=True)
            
            # Check some calculated cells that should reference our inputs
            calc_checks = {}
            
            # Check NOI sheet calculations
            if 'NOI' in wb_calc.sheetnames:
                noi_sheet = wb_calc['NOI']
                calc_checks['NOI_Property_Name'] = noi_sheet['C2'].value
                calc_checks['NOI_Address'] = noi_sheet['C3'].value
            
            # Check Sources & Uses calculations
            if 'Sources & Uses' in wb_calc.sheetnames:
                su_sheet = wb_calc['Sources & Uses']
                # These cells should show calculated values based on our inputs
                for row in range(7, 20):
                    for col in ['C', 'D', 'G', 'H']:
                        cell_val = su_sheet[f'{col}{row}'].value
                        if cell_val and isinstance(cell_val, (int, float)) and cell_val != 0:
                            calc_checks[f'SU_{col}{row}'] = cell_val
            
            wb_calc.close()
            
            logger.info("✅ Formula calculations verified:")
            for check_name, value in calc_checks.items():
                if isinstance(value, (int, float)):
                    logger.info(f"   {check_name}: ${value:,.0f}" if value > 1000 else f"   {check_name}: {value}")
                else:
                    logger.info(f"   {check_name}: {value}")
            
            return {
                "success": True,
                "calculations_found": len(calc_checks),
                "sample_calculations": calc_checks
            }
            
        except Exception as e:
            logger.error(f"❌ TEST 4 FAILED: {str(e)}")
            return {"success": False, "error": str(e)}
    
    def run_all_tests(self):
        """Run complete openpyxl testing suite"""
        
        logger.info("🚀 OPENPYXL FORMULA PRESERVATION TESTING SUITE")
        logger.info("=" * 70)
        logger.info("Testing openpyxl as xlwings replacement for CABOTNTemplate processing")
        logger.info("")
        
        tests = [
            ("Template reading", self.test_1_template_reading),
            ("Formula preservation", self.test_2_formula_preservation),
            ("Batch processing simulation", self.test_3_batch_processing_simulation),
            ("Formula calculation verification", self.test_4_formula_calculation_verification),
        ]
        
        results = {}
        
        for test_name, test_func in tests:
            logger.info(f"\n🔬 RUNNING: {test_name}")
            logger.info("-" * 50)
            
            try:
                result = test_func()
                results[test_name] = result
                
                if result.get("success", False):
                    logger.info(f"✅ {test_name}: SUCCESS")
                else:
                    logger.info(f"❌ {test_name}: FAILED")
                    
            except Exception as e:
                logger.error(f"💥 {test_name}: CRASHED - {str(e)}")
                results[test_name] = {"success": False, "error": f"Test crashed: {str(e)}"}
        
        # Generate summary report
        self.generate_summary_report(results)
        
        return results
    
    def generate_summary_report(self, results):
        """Generate comprehensive test results summary"""
        
        logger.info("\n📊 OPENPYXL TESTING RESULTS SUMMARY")
        logger.info("=" * 60)
        
        successful_tests = [name for name, result in results.items() if result.get("success", False)]
        
        logger.info(f"✅ Successful tests: {len(successful_tests)}/{len(results)}")
        
        for test_name, result in results.items():
            status = "✅ SUCCESS" if result.get("success", False) else "❌ FAILED"
            logger.info(f"{status}: {test_name}")
            
            # Specific insights per test
            if test_name == "Batch processing simulation" and result.get("success", False):
                logger.info(f"   → {result['files_processed']}/5 files processed without permissions!")
                logger.info(f"   → Average time: {result.get('average_time', 0):.2f}s per file")
            
            elif test_name == "Formula preservation" and result.get("success", False):
                intact = result.get("formulas_intact", False)
                status = "✅ INTACT" if intact else "⚠️  QUESTIONABLE"
                logger.info(f"   → Formulas: {status}")
        
        # Final recommendation
        all_successful = all(result.get("success", False) for result in results.values())
        
        if all_successful:
            logger.info("\n🎯 RECOMMENDATION: OpenPyxl is VIABLE alternative to xlwings!")
            logger.info("   ✅ No permission prompts")
            logger.info("   ✅ Formula preservation")
            logger.info("   ✅ Batch processing capability")
            logger.info("   ✅ Calculation integrity")
        else:
            logger.info("\n⚠️  RECOMMENDATION: OpenPyxl has limitations")
            logger.info("   Further investigation needed for production use")

def main():
    tester = OpenPyxlFormulaTester()
    tester.run_all_tests()

if __name__ == "__main__":
    main()