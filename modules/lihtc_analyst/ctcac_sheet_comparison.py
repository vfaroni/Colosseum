#!/usr/bin/env python3
"""
CTCAC Sheet-by-Sheet Comparison Tool
Analyze exactly where the 4,839 extra cells in 2024 vs 2025 files come from
"""

import openpyxl
from pathlib import Path
from typing import Dict, List

def analyze_sheet_differences(file_2024: Path, file_2025: Path):
    """Compare sheet-by-sheet cell counts between 2024 and 2025 files"""
    
    print(f"🔍 ANALYZING SHEET DIFFERENCES")
    print(f"2024 File: {file_2024.name}")
    print(f"2025 File: {file_2025.name}")
    print("=" * 60)
    
    # Load both files
    wb_2024 = openpyxl.load_workbook(file_2024, read_only=True)
    wb_2025 = openpyxl.load_workbook(file_2025, read_only=True)
    
    # Smart range detection function (from V1.8)
    def get_smart_range(sheet, sheet_name: str) -> tuple:
        """Apply same smart range logic as V1.8"""
        excel_max_row = sheet.max_row or 1
        excel_max_col = sheet.max_column or 1
        
        SAFE_MAX_ROWS = 500
        SAFE_MAX_COLS = 50
        
        if excel_max_col > 1000 or excel_max_row > 5000:
            # Use smart detection for corrupted ranges
            max_search_row = min(excel_max_row, 2000)
            max_search_col = min(excel_max_col, 100)
            
            last_data_row = 1
            last_data_col = 1
            
            # Quick scan for actual boundaries
            for row_num in range(min(max_search_row, 500), 0, -1):
                found_data = False
                for col_num in range(1, min(max_search_col + 1, 51)):
                    try:
                        cell = sheet.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            cell_text = str(cell.value).strip()
                            if cell_text and len(cell_text) > 0:
                                last_data_row = row_num
                                found_data = True
                                break
                    except:
                        continue
                if found_data:
                    break
            
            for col_num in range(1, min(max_search_col + 1, 51)):
                found_data = False
                for row_num in range(1, min(last_data_row + 1, 501)):
                    try:
                        cell = sheet.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            cell_text = str(cell.value).strip()
                            if cell_text and len(cell_text) > 0:
                                last_data_col = col_num
                                found_data = True
                                break
                    except:
                        continue
                if not found_data and col_num > 20:
                    break
            
            safe_row = min(last_data_row + 10, SAFE_MAX_ROWS)
            safe_col = min(last_data_col + 5, SAFE_MAX_COLS)
            return safe_row, safe_col
        
        return min(excel_max_row, SAFE_MAX_ROWS), min(excel_max_col, SAFE_MAX_COLS)
    
    # Analyze each sheet
    sheet_comparison = {}
    total_2024 = 0
    total_2025 = 0
    
    # Get all unique sheet names
    sheets_2024 = {sheet.title for sheet in wb_2024.worksheets}
    sheets_2025 = {sheet.title for sheet in wb_2025.worksheets}
    all_sheets = sheets_2024.union(sheets_2025)
    
    print(f"📊 SHEET-BY-SHEET ANALYSIS:")
    print(f"{'Sheet Name':<35} {'2024 Cells':<12} {'2025 Cells':<12} {'Difference':<12}")
    print("-" * 75)
    
    for sheet_name in sorted(all_sheets):
        cells_2024 = 0
        cells_2025 = 0
        
        # Calculate 2024 cells
        if sheet_name in sheets_2024:
            sheet_2024 = wb_2024[sheet_name]
            smart_row, smart_col = get_smart_range(sheet_2024, sheet_name)
            cells_2024 = smart_row * smart_col
        
        # Calculate 2025 cells
        if sheet_name in sheets_2025:
            sheet_2025 = wb_2025[sheet_name]
            smart_row, smart_col = get_smart_range(sheet_2025, sheet_name)
            cells_2025 = smart_row * smart_col
        
        difference = cells_2024 - cells_2025
        sheet_comparison[sheet_name] = {
            '2024': cells_2024,
            '2025': cells_2025,
            'difference': difference
        }
        
        total_2024 += cells_2024
        total_2025 += cells_2025
        
        # Format output with color coding
        status = ""
        if difference > 1000:
            status = "🔴 MAJOR DIFF"
        elif difference > 100:
            status = "🟡 DIFF"
        elif difference < -100:
            status = "🔵 2025 LARGER"
        
        print(f"{sheet_name:<35} {cells_2024:<12,} {cells_2025:<12,} {difference:<12,} {status}")
    
    print("-" * 75)
    print(f"{'TOTALS':<35} {total_2024:<12,} {total_2025:<12,} {total_2024-total_2025:<12,}")
    
    # Identify the biggest contributors
    print(f"\n🎯 TOP CONTRIBUTORS TO DIFFERENCE:")
    sorted_diffs = sorted(sheet_comparison.items(), key=lambda x: abs(x[1]['difference']), reverse=True)
    
    for sheet_name, data in sorted_diffs[:5]:
        if abs(data['difference']) > 50:
            print(f"  {sheet_name}: {data['difference']:+,} cells")
    
    wb_2024.close()
    wb_2025.close()
    
    return sheet_comparison

def main():
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    
    # Use the same files from validation
    file_2024 = data_dir / "2024_4pct_R1_24-409.xlsx"
    file_2025 = data_dir / "2025_4pct_R1_25-464.xlsx"
    
    if file_2024.exists() and file_2025.exists():
        comparison = analyze_sheet_differences(file_2024, file_2025)
    else:
        print(f"❌ Files not found:")
        print(f"2024: {file_2024.exists()}")
        print(f"2025: {file_2025.exists()}")

if __name__ == "__main__":
    main()