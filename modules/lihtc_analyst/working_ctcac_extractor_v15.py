#!/usr/bin/env python3
"""
WORKING CTCAC EXTRACTOR V1.5
Back to what actually works - V1.0 approach with 9 tabs instead of 6
No fancy pattern matching - just extract the actual data like V1.0 did
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import json
import logging
from typing import Dict, List, Any, Optional, Tuple

class WorkingCTCACExtractorV15:
    """
    Back to the V1.0 approach that actually worked
    Extract REAL DATA from Excel sheets and put it in new Excel files
    """
    
    def __init__(self, log_level=logging.INFO):
        self.setup_logging(log_level)
        
        # The 9 critical sheets we want (expanded from V1.0's 6)
        self.critical_sheets = [
            'Application',
            'Sources and Uses Budget', 
            'Basis & Credits',
            'Sources and Basis Breakdown',
            'Points System',
            'Tie Breaker',
            'CalHFA Addendum',          # New
            'Subsidy Contract Calculation',  # New  
            '15 Year Pro Forma'         # New
        ]
        
        # Stats tracking like V1.0
        self.stats = {
            'files_processed': 0,
            'sheets_extracted': 0,
            'cells_extracted': 0,
            'formulas_preserved': 0,
            'errors': []
        }
        
        self.logger.info("🚀 WORKING CTCAC EXTRACTOR V1.5 - Back to what works!")
    
    def setup_logging(self, log_level):
        """Setup logging exactly like V1.0"""
        log_dir = Path("extraction_logs_v15")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f"ctcac_extraction_v15_{timestamp}.log"
        
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.log_file = log_file
    
    def extract_3_applications(self, data_dir: Path, output_dir: Path) -> Dict[str, Any]:
        """
        Extract the same 3 applications but with 9 tabs instead of 6
        Using the V1.0 approach that actually worked
        """
        self.logger.info(f"="*80)
        self.logger.info(f"🎯 V1.5 EXTRACTION - 9 TABS FROM 3 APPLICATIONS")
        self.logger.info(f"Back to the approach that actually worked!")
        self.logger.info(f"="*80)
        
        # Create output directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # The same 3 files from our tests
        target_files = [
            "2025_4pct_R1_25-464.xlsx",
            "2025_4pct_R1_25-433.xlsx", 
            "2025_4pct_R1_25-425.xlsx"
        ]
        
        master_data = {}
        
        for idx, filename in enumerate(target_files, 1):
            file_path = data_dir / filename
            if not file_path.exists():
                self.logger.error(f"❌ File not found: {filename}")
                continue
                
            self.logger.info(f"\\n{'='*60}")
            self.logger.info(f"Processing [{idx}/3]: {filename}")
            self.logger.info(f"{'='*60}")
            
            try:
                # Extract using V1.0 approach
                app_data = self._extract_single_application_v10_style(file_path)
                master_data[filename] = app_data
                
                # Create individual Excel export like V1.0 did
                self._create_working_excel_export(app_data, output_dir, filename)
                
                self.stats['files_processed'] += 1
                
            except Exception as e:
                self.logger.error(f"❌ Error processing {filename}: {e}")
                self.stats['errors'].append({'file': filename, 'error': str(e)})
        
        # Create master comparison like V1.0
        self._create_master_comparison_excel(master_data, output_dir)
        
        # Generate report
        report = self._generate_extraction_report(master_data)
        
        self.logger.info(f"\\n{'='*80}")
        self.logger.info(f"✅ V1.5 EXTRACTION COMPLETE!")
        self.logger.info(f"Files Processed: {self.stats['files_processed']}")
        self.logger.info(f"Sheets Extracted: {self.stats['sheets_extracted']}")
        self.logger.info(f"Cells Extracted: {self.stats['cells_extracted']}")
        self.logger.info(f"Formulas Preserved: {self.stats['formulas_preserved']}")
        self.logger.info(f"{'='*80}")
        
        return report
    
    def _extract_single_application_v10_style(self, file_path: Path) -> Dict[str, Any]:
        """Extract using the V1.0 approach that actually worked"""
        app_data = {
            'file_name': file_path.name,
            'file_path': str(file_path),
            'extraction_timestamp': datetime.now().isoformat(),
            'sheets': {},
            'metadata': {}
        }
        
        # Load exactly like V1.0
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Metadata
        app_data['metadata'] = {
            'total_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'file_size_kb': file_path.stat().st_size / 1024
        }
        
        # Process each of our 9 critical sheets
        for sheet_name in self.critical_sheets:
            if sheet_name in wb.sheetnames:
                self.logger.info(f"  📋 Extracting: {sheet_name}")
                sheet_data = self._extract_sheet_v10_style(
                    wb[sheet_name],
                    wb_values[sheet_name],
                    sheet_name
                )
                app_data['sheets'][sheet_name] = sheet_data
                self.stats['sheets_extracted'] += 1
            else:
                self.logger.warning(f"  ⚠️ Sheet not found: {sheet_name}")
        
        wb.close()
        wb_values.close()
        
        return app_data
    
    def _extract_sheet_v10_style(self, ws_formulas, ws_values, sheet_name: str) -> Dict:
        """Extract sheet data exactly like V1.0 did - this is what worked!"""
        sheet_data = {
            'cells': {},
            'formulas': {},
            'values': {},
            'merged_cells': [],
            'dimensions': {
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            }
        }
        
        # Extract merged cells like V1.0
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract all cells with data - V1.0 approach
        for row in range(1, min(ws_formulas.max_row + 1, 1000)):
            for col in range(1, min(ws_formulas.max_column + 1, 50)):
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_formula.value is not None or cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__
                    }
                    
                    # Check for formula like V1.0
                    if hasattr(cell_formula, '_value') and isinstance(cell_formula._value, str) and cell_formula._value.startswith('='):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                        self.stats['formulas_preserved'] += 1
                    
                    sheet_data['cells'][cell_ref] = cell_info
                    sheet_data['values'][cell_ref] = cell_value.value
                    self.stats['cells_extracted'] += 1
        
        return sheet_data
    
    def _create_working_excel_export(self, app_data: Dict, output_dir: Path, source_name: str):
        """Create Excel export exactly like V1.0 - this worked!"""
        base_name = source_name.replace('.xlsx', '')
        output_file = output_dir / f"WORKING_V15_{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add metadata sheet first like V1.0
        self._add_metadata_sheet_v10(wb, app_data)
        
        # Add each extracted sheet with ORIGINAL NAME like V1.0
        for sheet_name, sheet_data in app_data['sheets'].items():
            self.logger.info(f"    ✍️ Writing sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name[:31])
            
            # Write the actual data like V1.0
            self._write_sheet_data_v10(ws, sheet_data)
        
        # Save the workbook
        wb.save(output_file)
        self.logger.info(f"  💾 Saved: {output_file.name}")
        
        return output_file
    
    def _write_sheet_data_v10(self, ws, sheet_data: Dict):
        """Write sheet data exactly like V1.0 - put actual cell values in actual positions"""
        # Write cells with their original positions - V1.0 approach
        if 'cells' in sheet_data:
            for cell_ref, cell_info in sheet_data['cells'].items():
                row = cell_info['row']
                col = cell_info['col']
                
                # Write the actual value to the actual position
                ws.cell(row=row, column=col, value=cell_info['value'])
                
                # Add formula as comment if present like V1.0
                if cell_info.get('formula'):
                    cell = ws.cell(row=row, column=col)
                    try:
                        cell.comment = openpyxl.comments.Comment(
                            f"Formula: {cell_info['formula']}", 
                            "CTCAC Extractor V1.5"
                        )
                    except:
                        pass  # Skip if comment fails
        
        # Apply merged cells like V1.0
        if 'merged_cells' in sheet_data:
            for merged_range in sheet_data['merged_cells']:
                try:
                    ws.merge_cells(merged_range)
                except:
                    pass
    
    def _add_metadata_sheet_v10(self, wb: Workbook, app_data: Dict):
        """Add metadata sheet like V1.0"""
        ws = wb.create_sheet(title="METADATA", index=0)
        
        # Header styling
        header_font = Font(bold=True, size=14)
        
        # Title
        ws['A1'] = "CTCAC EXTRACTION METADATA V1.5"
        ws['A1'].font = header_font
        
        # File info
        ws['A3'] = "Source File:"
        ws['B3'] = app_data['file_name']
        ws['A4'] = "Extraction Time:"
        ws['B4'] = app_data['extraction_timestamp']
        ws['A5'] = "Sheets Extracted:"
        ws['B5'] = len(app_data['sheets'])
        
        # Sheet list
        ws['A7'] = "EXTRACTED SHEETS:"
        ws['A7'].font = Font(bold=True)
        
        row = 8
        for sheet_name in app_data['sheets'].keys():
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = f"{len(app_data['sheets'][sheet_name].get('cells', {}))} cells"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_master_comparison_excel(self, master_data: Dict, output_dir: Path):
        """Create master comparison like V1.0"""
        output_file = output_dir / f"MASTER_V15_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        ws = wb.create_sheet(title="COMPARISON")
        
        # Headers
        headers = ['Application', 'Sheets Found', 'Total Cells', 'Total Formulas']
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        
        # Data rows
        row_num = 2
        for file_name, app_data in master_data.items():
            total_cells = sum(len(sheet.get('cells', {})) for sheet in app_data['sheets'].values())
            total_formulas = sum(len(sheet.get('formulas', {})) for sheet in app_data['sheets'].values())
            
            ws.cell(row=row_num, column=1, value=file_name)
            ws.cell(row=row_num, column=2, value=len(app_data['sheets']))
            ws.cell(row=row_num, column=3, value=total_cells)
            ws.cell(row=row_num, column=4, value=total_formulas)
            
            row_num += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        self.logger.info(f"\\n📊 Master Comparison saved: {output_file.name}")
    
    def _generate_extraction_report(self, master_data: Dict) -> Dict:
        """Generate report like V1.0"""
        report = {
            'timestamp': datetime.now().isoformat(),
            'statistics': self.stats,
            'files': {},
            'summary': {
                'total_files': len(master_data),
                'total_sheets': sum(len(d.get('sheets', {})) for d in master_data.values()),
                'sheets_found_per_app': {},
            }
        }
        
        # Per-file breakdown
        for file_name, app_data in master_data.items():
            sheets_found = list(app_data.get('sheets', {}).keys())
            report['files'][file_name] = {
                'sheets_extracted': sheets_found,
                'extraction_time': app_data.get('extraction_timestamp')
            }
            report['summary']['sheets_found_per_app'][file_name] = len(sheets_found)
        
        # Save JSON report
        json_file = Path("extraction_logs_v15") / f"extraction_report_v15_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_file, 'w') as f:
            json.dump(report, f, indent=2, default=str)
        
        self.logger.info(f"📄 Report saved: {json_file}")
        
        return report


def main():
    """Run the working V1.5 extractor - back to what actually works!"""
    print("="*80)
    print("🚀 WORKING CTCAC EXTRACTOR V1.5")
    print("   Back to the V1.0 approach that actually worked")
    print("   Now with 9 tabs instead of 6")
    print("="*80)
    
    # Setup paths
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/working_extractions_v15")
    
    # Initialize the extractor
    extractor = WorkingCTCACExtractorV15()
    
    # Extract the 3 applications
    report = extractor.extract_3_applications(
        data_dir=data_dir,
        output_dir=output_dir
    )
    
    print("\\n" + "="*80)
    print("✅ V1.5 EXTRACTION COMPLETE!")
    print("="*80)
    print(f"Files Processed: {report['summary']['total_files']}")
    print(f"Total Sheets: {report['summary']['total_sheets']}")
    
    print("\\n📋 SHEETS FOUND PER APPLICATION:")
    for app_name, sheet_count in report['summary']['sheets_found_per_app'].items():
        print(f"   {app_name}: {sheet_count} sheets")
    
    print("\\n🎯 This should have ACTUAL DATA in the Excel files!")
    print("   Each sheet should contain the real cell values from the source")


if __name__ == "__main__":
    main()