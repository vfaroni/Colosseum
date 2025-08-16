#!/usr/bin/env python3
"""
CONSERVATIVE CTCAC EXTRACTOR V1.6
Built on V1.5 foundation with intelligent cleanup and comprehensive logging
Conservative approach: When in doubt, preserve data
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import json
import logging
import re
from typing import Dict, List, Any, Optional, Tuple

class ConservativeCleanupEngine:
    """Conservative text analysis and cleanup engine"""
    
    def __init__(self):
        self.removal_log = []
        
        # Conservative removal patterns - only obvious boilerplate
        self.safe_removal_patterns = [
            # Long instructional text (>100 chars)
            r'^.{100,}$',  # Very long single-cell text
            # Application submission instructions
            r'PLEASE INCLUDE APPLICATION FEE WITH APPLICATION SUBMISSION',
            r'Submit this application with all required attachments',
            # Form completion guidance
            r'Complete all yellow highlighted cells',
            r'All yellow cells in this addendum must be completed',
            # Legal disclaimers over 50 chars
            r'^(The following|This application|By submitting).{50,}$'
        ]
        
        # Patterns to NEVER remove
        self.preserve_patterns = [
            r'^\d+$',  # Any numeric value
            r'^\$\d+',  # Dollar amounts
            r'%$',     # Percentages
            r'SECTION \d+',  # Section dividers
            r'Version$',     # Version stamps
            r'PROJECT NAME',  # Field labels
            r'TOTAL',        # Totals
            r'COST',         # Cost labels
            r'^[A-Z]{1,5}\d+$',  # Cell references
            r'^\s*$'         # Blank cells (preserve for standardization)
        ]
    
    def analyze_cell_for_removal(self, cell_value: Any, cell_ref: str, sheet_name: str, context: Dict) -> Tuple[bool, str]:
        """
        Conservative analysis - only remove obvious boilerplate
        Returns: (should_remove, reason)
        """
        if cell_value is None:
            return False, "Blank cell preserved for standardization"
        
        text = str(cell_value).strip()
        
        # Always preserve certain patterns
        for pattern in self.preserve_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return False, f"Preserved due to pattern: {pattern}"
        
        # Check for safe removal patterns (very conservative)
        for pattern in self.safe_removal_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                # Extra safety check - don't remove if it looks like data
                if self._contains_variable_data(text):
                    return False, "Contains variable data - preserved"
                return True, f"Safe removal pattern: {pattern}"
        
        # Default to preservation
        return False, "Conservative default - preserved"
    
    def _contains_variable_data(self, text: str) -> bool:
        """Check if text might contain variable/meaningful data"""
        # Look for signs this might be data rather than boilerplate
        data_indicators = [
            r'\d+',           # Contains numbers
            r'\$',            # Contains dollar signs
            r'%',             # Contains percentages
            r'[A-Z]{2,}\s+[A-Z]',  # Looks like names/entities
            r'^\s*[A-Z][a-z]+',    # Starts with proper noun
        ]
        
        for indicator in data_indicators:
            if re.search(indicator, text):
                return True
        return False
    
    def log_removal(self, sheet_name: str, cell_ref: str, removed_text: str, reason: str, context: Dict):
        """Log removal for validation and traceability"""
        self.removal_log.append({
            'sheet': sheet_name,
            'cell_reference': cell_ref,
            'removed_text': removed_text[:200] + ('...' if len(removed_text) > 200 else ''),
            'full_text_length': len(removed_text),
            'reason': reason,
            'context_before': context.get('before', ''),
            'context_after': context.get('after', ''),
            'timestamp': datetime.now().isoformat()
        })

class ConservativeCTCACExtractorV16:
    """
    V1.6: Conservative cleanup enhancement of V1.5
    Preserves all meaningful data while improving presentation
    """
    
    def __init__(self, log_level=logging.INFO, enable_cleanup=True):
        self.setup_logging(log_level)
        self.enable_cleanup = enable_cleanup
        self.cleanup_engine = ConservativeCleanupEngine()
        
        # Same 9 critical sheets as V1.5
        self.critical_sheets = [
            'Application',
            'Sources and Uses Budget', 
            'Basis & Credits',
            'Sources and Basis Breakdown',
            'Points System',
            'Tie Breaker',
            'CalHFA Addendum',
            'Subsidy Contract Calculation',
            '15 Year Pro Forma'
        ]
        
        # Enhanced stats tracking
        self.stats = {
            'files_processed': 0,
            'sheets_extracted': 0,
            'cells_extracted': 0,
            'cells_cleaned': 0,
            'formulas_preserved': 0,
            'errors': []
        }
        
        self.logger.info("🚀 CONSERVATIVE CTCAC EXTRACTOR V1.6 - Enhanced with cleanup!")
    
    def setup_logging(self, log_level):
        """Enhanced logging for cleanup operations"""
        log_dir = Path("extraction_logs_v16")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f"ctcac_extraction_v16_{timestamp}.log"
        
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.log_file = log_file
    
    def extract_3_applications_with_cleanup(self, data_dir: Path, output_dir: Path) -> Dict[str, Any]:
        """
        Enhanced extraction with conservative cleanup
        Same core logic as V1.5 but with presentation improvements
        """
        self.logger.info(f"="*80)
        self.logger.info(f"🧹 V1.6 CONSERVATIVE EXTRACTION - CLEANUP ENABLED: {self.enable_cleanup}")
        self.logger.info(f"Foundation: V1.5 proven extraction with presentation enhancement")
        self.logger.info(f"="*80)
        
        # Create output directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Same 3 test files as V1.5
        target_files = [
            "2025_4pct_R1_25-464.xlsx",
            "2025_4pct_R1_25-433.xlsx", 
            "2025_4pct_R1_25-425.xlsx"
        ]
        
        master_data = {}
        
        for idx, filename in enumerate(target_files, 1):
            file_path = data_dir / filename
            if not file_path.exists():
                self.logger.error(f"❌ File not found: {filename}")
                continue
                
            self.logger.info(f"\\n{'='*60}")
            self.logger.info(f"Processing [{idx}/3]: {filename}")
            self.logger.info(f"{'='*60}")
            
            try:
                # Extract using enhanced V1.6 approach
                app_data = self._extract_single_application_v16_style(file_path)
                master_data[filename] = app_data
                
                # Create individual Excel export with cleanup
                self._create_conservative_excel_export(app_data, output_dir, filename)
                
                self.stats['files_processed'] += 1
                
            except Exception as e:
                self.logger.error(f"❌ Error processing {filename}: {e}")
                self.stats['errors'].append({'file': filename, 'error': str(e)})
        
        # Create master comparison
        self._create_master_comparison_excel(master_data, output_dir)
        
        # Generate removal log if cleanup enabled
        if self.enable_cleanup and self.cleanup_engine.removal_log:
            self._create_removal_log_excel(output_dir)
        
        # Generate comprehensive report
        report = self._generate_enhanced_extraction_report(master_data)
        
        self.logger.info(f"\\n{'='*80}")
        self.logger.info(f"✅ V1.6 CONSERVATIVE EXTRACTION COMPLETE!")
        self.logger.info(f"Files Processed: {self.stats['files_processed']}")
        self.logger.info(f"Sheets Extracted: {self.stats['sheets_extracted']}")
        self.logger.info(f"Cells Extracted: {self.stats['cells_extracted']}")
        self.logger.info(f"Cells Cleaned: {self.stats['cells_cleaned']}")
        self.logger.info(f"Formulas Preserved: {self.stats['formulas_preserved']}")
        if self.enable_cleanup:
            self.logger.info(f"Removals Logged: {len(self.cleanup_engine.removal_log)}")
        self.logger.info(f"{'='*80}")
        
        return report
    
    def _extract_single_application_v16_style(self, file_path: Path) -> Dict[str, Any]:
        """Enhanced extraction with conservative cleanup"""
        app_data = {
            'file_name': file_path.name,
            'file_path': str(file_path),
            'extraction_timestamp': datetime.now().isoformat(),
            'cleanup_enabled': self.enable_cleanup,
            'sheets': {},
            'metadata': {}
        }
        
        # Load workbooks (same as V1.5)
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Metadata
        app_data['metadata'] = {
            'total_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'file_size_kb': file_path.stat().st_size / 1024,
            'v16_enhancements': 'Conservative cleanup with comprehensive logging'
        }
        
        # Process each critical sheet with optional cleanup
        for sheet_name in self.critical_sheets:
            if sheet_name in wb.sheetnames:
                self.logger.info(f"  📋 Extracting: {sheet_name}")
                sheet_data = self._extract_sheet_v16_style(
                    wb[sheet_name],
                    wb_values[sheet_name],
                    sheet_name
                )
                app_data['sheets'][sheet_name] = sheet_data
                self.stats['sheets_extracted'] += 1
            else:
                self.logger.warning(f"  ⚠️ Sheet not found: {sheet_name}")
        
        wb.close()
        wb_values.close()
        
        return app_data
    
    def _extract_sheet_v16_style(self, ws_formulas, ws_values, sheet_name: str) -> Dict:
        """Enhanced sheet extraction with conservative cleanup"""
        sheet_data = {
            'cells': {},
            'formulas': {},
            'values': {},
            'merged_cells': [],
            'removed_cells': {},  # Track cleaned cells
            'dimensions': {
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            },
            'cleanup_stats': {
                'cells_analyzed': 0,
                'cells_removed': 0,
                'cells_preserved': 0
            }
        }
        
        # Extract merged cells (same as V1.5)
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract all cells with optional cleanup
        for row in range(1, min(ws_formulas.max_row + 1, 1000)):
            for col in range(1, min(ws_formulas.max_column + 1, 50)):
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_formula.value is not None or cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    # Get context for cleanup analysis
                    context = self._get_cell_context(ws_values, row, col)
                    
                    # Analyze for potential cleanup
                    should_remove = False
                    removal_reason = ""
                    
                    if self.enable_cleanup and cell_value.value is not None:
                        should_remove, removal_reason = self.cleanup_engine.analyze_cell_for_removal(
                            cell_value.value, cell_ref, sheet_name, context
                        )
                        sheet_data['cleanup_stats']['cells_analyzed'] += 1
                    
                    if should_remove:
                        # Log the removal
                        self.cleanup_engine.log_removal(
                            sheet_name, cell_ref, str(cell_value.value), 
                            removal_reason, context
                        )
                        sheet_data['removed_cells'][cell_ref] = {
                            'original_value': cell_value.value,
                            'reason': removal_reason
                        }
                        sheet_data['cleanup_stats']['cells_removed'] += 1
                        self.stats['cells_cleaned'] += 1
                        continue  # Skip adding to main cells dict
                    
                    # Preserve the cell (same logic as V1.5)
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__,
                        'cleaned': False
                    }
                    
                    # Check for formula (same as V1.5)
                    if hasattr(cell_formula, '_value') and isinstance(cell_formula._value, str) and cell_formula._value.startswith('='):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                        self.stats['formulas_preserved'] += 1
                    
                    sheet_data['cells'][cell_ref] = cell_info
                    sheet_data['values'][cell_ref] = cell_value.value
                    sheet_data['cleanup_stats']['cells_preserved'] += 1
                    self.stats['cells_extracted'] += 1
        
        return sheet_data
    
    def _get_cell_context(self, ws, row: int, col: int, radius: int = 2) -> Dict:
        """Get surrounding cell context for cleanup analysis"""
        context = {
            'before': '',
            'after': '',
            'above': '',
            'below': ''
        }
        
        try:
            # Get adjacent cells for context
            if col > 1:
                before_cell = ws.cell(row=row, column=col-1)
                context['before'] = str(before_cell.value) if before_cell.value else ''
            
            if col < ws.max_column:
                after_cell = ws.cell(row=row, column=col+1)
                context['after'] = str(after_cell.value) if after_cell.value else ''
                
            if row > 1:
                above_cell = ws.cell(row=row-1, column=col)
                context['above'] = str(above_cell.value) if above_cell.value else ''
                
            if row < ws.max_row:
                below_cell = ws.cell(row=row+1, column=col)
                context['below'] = str(below_cell.value) if below_cell.value else ''
                
        except Exception:
            pass  # Context is optional, continue if error
            
        return context
    
    def _create_conservative_excel_export(self, app_data: Dict, output_dir: Path, source_name: str):
        """Create cleaned Excel export with validation info"""
        base_name = source_name.replace('.xlsx', '')
        output_file = output_dir / f"CONSERVATIVE_V16_{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add enhanced metadata sheet
        self._add_enhanced_metadata_sheet(wb, app_data)
        
        # Add each extracted sheet (same as V1.5 but with cleanup info)
        for sheet_name, sheet_data in app_data['sheets'].items():
            self.logger.info(f"    ✍️ Writing sheet: {sheet_name}")
            if self.enable_cleanup:
                self.logger.info(f"      🧹 Cleanup: {sheet_data['cleanup_stats']['cells_removed']} removed, {sheet_data['cleanup_stats']['cells_preserved']} preserved")
            
            ws = wb.create_sheet(title=sheet_name[:31])
            
            # Write the cleaned data
            self._write_cleaned_sheet_data(ws, sheet_data)
        
        # Save the workbook
        wb.save(output_file)
        self.logger.info(f"  💾 Saved: {output_file.name}")
        
        return output_file
    
    def _write_cleaned_sheet_data(self, ws, sheet_data: Dict):
        """Write cleaned sheet data with preservation of structure"""
        # Write preserved cells (same as V1.5)
        if 'cells' in sheet_data:
            for cell_ref, cell_info in sheet_data['cells'].items():
                row = cell_info['row']
                col = cell_info['col']
                
                # Write the actual value to the actual position
                ws.cell(row=row, column=col, value=cell_info['value'])
                
                # Add formula as comment if present
                if cell_info.get('formula'):
                    cell = ws.cell(row=row, column=col)
                    try:
                        cell.comment = openpyxl.comments.Comment(
                            f"Formula: {cell_info['formula']}", 
                            "CTCAC Extractor V1.6"
                        )
                    except:
                        pass  # Skip if comment fails
        
        # Apply merged cells (same as V1.5)
        if 'merged_cells' in sheet_data:
            for merged_range in sheet_data['merged_cells']:
                try:
                    ws.merge_cells(merged_range)
                except:
                    pass
    
    def _add_enhanced_metadata_sheet(self, wb: Workbook, app_data: Dict):
        """Enhanced metadata with cleanup information"""
        ws = wb.create_sheet(title="METADATA_V16", index=0)
        
        # Header styling
        header_font = Font(bold=True, size=14)
        
        # Title
        ws['A1'] = "CTCAC EXTRACTION METADATA V1.6 - CONSERVATIVE CLEANUP"
        ws['A1'].font = header_font
        
        # File info
        ws['A3'] = "Source File:"
        ws['B3'] = app_data['file_name']
        ws['A4'] = "Extraction Time:"
        ws['B4'] = app_data['extraction_timestamp']
        ws['A5'] = "Cleanup Enabled:"
        ws['B5'] = "Yes" if app_data['cleanup_enabled'] else "No"
        ws['A6'] = "Sheets Extracted:"
        ws['B6'] = len(app_data['sheets'])
        
        # Cleanup summary
        if app_data['cleanup_enabled']:
            ws['A8'] = "CLEANUP SUMMARY:"
            ws['A8'].font = Font(bold=True)
            
            total_removed = sum(sheet['cleanup_stats']['cells_removed'] for sheet in app_data['sheets'].values())
            total_preserved = sum(sheet['cleanup_stats']['cells_preserved'] for sheet in app_data['sheets'].values())
            
            ws['A9'] = "Total Cells Removed:"
            ws['B9'] = total_removed
            ws['A10'] = "Total Cells Preserved:"
            ws['B10'] = total_preserved
            ws['A11'] = "Preservation Rate:"
            ws['B11'] = f"{(total_preserved / (total_preserved + total_removed) * 100):.1f}%" if (total_preserved + total_removed) > 0 else "N/A"
        
        # Sheet breakdown
        row = 13
        ws[f'A{row}'] = "EXTRACTED SHEETS:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        for sheet_name, sheet_data in app_data['sheets'].items():
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = f"{len(sheet_data.get('cells', {}))} cells"
            if app_data['cleanup_enabled']:
                ws[f'C{row}'] = f"{sheet_data['cleanup_stats']['cells_removed']} removed"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
    
    def _create_removal_log_excel(self, output_dir: Path):
        """Create comprehensive removal log for validation"""
        if not self.cleanup_engine.removal_log:
            return
        
        log_file = output_dir / f"REMOVED_TEXT_LOG_V16_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Removal Log"
        
        # Headers
        headers = ['Sheet', 'Cell Reference', 'Removed Text', 'Text Length', 'Reason', 'Context Before', 'Context After', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, entry in enumerate(self.cleanup_engine.removal_log, 2):
            ws.cell(row=row, column=1, value=entry['sheet'])
            ws.cell(row=row, column=2, value=entry['cell_reference'])
            ws.cell(row=row, column=3, value=entry['removed_text'])
            ws.cell(row=row, column=4, value=entry['full_text_length'])
            ws.cell(row=row, column=5, value=entry['reason'])
            ws.cell(row=row, column=6, value=entry.get('context_before', ''))
            ws.cell(row=row, column=7, value=entry.get('context_after', ''))
            ws.cell(row=row, column=8, value=entry['timestamp'])
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(log_file)
        self.logger.info(f"📄 Removal log saved: {log_file.name}")
    
    def _create_master_comparison_excel(self, master_data: Dict, output_dir: Path):
        """Enhanced master comparison with cleanup stats"""
        output_file = output_dir / f"MASTER_V16_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        ws = wb.create_sheet(title="COMPARISON")
        
        # Enhanced headers
        headers = ['Application', 'Sheets Found', 'Total Cells', 'Cells Cleaned', 'Total Formulas', 'Cleanup Rate']
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        
        # Data rows with cleanup stats
        row_num = 2
        for file_name, app_data in master_data.items():
            total_cells = sum(len(sheet.get('cells', {})) for sheet in app_data['sheets'].values())
            total_formulas = sum(len(sheet.get('formulas', {})) for sheet in app_data['sheets'].values())
            total_cleaned = sum(sheet.get('cleanup_stats', {}).get('cells_removed', 0) for sheet in app_data['sheets'].values())
            cleanup_rate = f"{(total_cleaned / (total_cells + total_cleaned) * 100):.1f}%" if (total_cells + total_cleaned) > 0 else "0%"
            
            ws.cell(row=row_num, column=1, value=file_name)
            ws.cell(row=row_num, column=2, value=len(app_data['sheets']))
            ws.cell(row=row_num, column=3, value=total_cells)
            ws.cell(row=row_num, column=4, value=total_cleaned)
            ws.cell(row=row_num, column=5, value=total_formulas)
            ws.cell(row=row_num, column=6, value=cleanup_rate)
            
            row_num += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        self.logger.info(f"\\n📊 Master Comparison saved: {output_file.name}")
    
    def _generate_enhanced_extraction_report(self, master_data: Dict) -> Dict:
        """Enhanced report with cleanup statistics"""
        report = {
            'timestamp': datetime.now().isoformat(),
            'version': 'V1.6 Conservative Cleanup',
            'cleanup_enabled': self.enable_cleanup,
            'statistics': self.stats,
            'cleanup_summary': {
                'total_removals': len(self.cleanup_engine.removal_log),
                'removal_rate': f"{(self.stats['cells_cleaned'] / max(self.stats['cells_extracted'] + self.stats['cells_cleaned'], 1) * 100):.1f}%",
                'preservation_rate': f"{(self.stats['cells_extracted'] / max(self.stats['cells_extracted'] + self.stats['cells_cleaned'], 1) * 100):.1f}%"
            },
            'files': {},
            'summary': {
                'total_files': len(master_data),
                'total_sheets': sum(len(d.get('sheets', {})) for d in master_data.values()),
                'sheets_found_per_app': {},
            }
        }
        
        # Per-file breakdown with cleanup stats
        for file_name, app_data in master_data.items():
            sheets_found = list(app_data.get('sheets', {}).keys())
            cleanup_stats = {}
            if self.enable_cleanup:
                cleanup_stats = {
                    'total_removed': sum(sheet.get('cleanup_stats', {}).get('cells_removed', 0) for sheet in app_data['sheets'].values()),
                    'total_preserved': sum(sheet.get('cleanup_stats', {}).get('cells_preserved', 0) for sheet in app_data['sheets'].values())
                }
            
            report['files'][file_name] = {
                'sheets_extracted': sheets_found,
                'extraction_time': app_data.get('extraction_timestamp'),
                'cleanup_enabled': app_data.get('cleanup_enabled', False),
                'cleanup_stats': cleanup_stats
            }
            report['summary']['sheets_found_per_app'][file_name] = len(sheets_found)
        
        # Save enhanced JSON report
        json_file = Path("extraction_logs_v16") / f"extraction_report_v16_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_file, 'w') as f:
            json.dump(report, f, indent=2, default=str)
        
        self.logger.info(f"📄 Enhanced report saved: {json_file}")
        
        return report


def main():
    """Run the conservative V1.6 extractor with cleanup"""
    print("="*80)
    print("🧹 CONSERVATIVE CTCAC EXTRACTOR V1.6")
    print("   Built on V1.5 foundation with intelligent cleanup")
    print("   Conservative approach: When in doubt, preserve data")
    print("="*80)
    
    # Setup paths
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/conservative_extractions_v16")
    
    # Initialize the enhanced extractor
    extractor = ConservativeCTCACExtractorV16(enable_cleanup=True)
    
    # Extract with cleanup
    report = extractor.extract_3_applications_with_cleanup(
        data_dir=data_dir,
        output_dir=output_dir
    )
    
    print("\\n" + "="*80)
    print("✅ V1.6 CONSERVATIVE EXTRACTION COMPLETE!")
    print("="*80)
    print(f"Files Processed: {report['summary']['total_files']}")
    print(f"Total Sheets: {report['summary']['total_sheets']}")
    print(f"Cleanup Enabled: {report['cleanup_enabled']}")
    
    if report['cleanup_enabled']:
        print(f"Total Removals: {report['cleanup_summary']['total_removals']}")
        print(f"Preservation Rate: {report['cleanup_summary']['preservation_rate']}")
    
    print("\\n📋 SHEETS FOUND PER APPLICATION:")
    for app_name, sheet_count in report['summary']['sheets_found_per_app'].items():
        print(f"   {app_name}: {sheet_count} sheets")
    
    print("\\n🎯 V1.6 delivers clean, validated Excel files with comprehensive logging!")
    print("   Conservative cleanup preserves all meaningful data")
    print("   Ready for Step 2: Excel → JSON conversion")


if __name__ == "__main__":
    main()