#!/usr/bin/env python3
"""
CTCAC LIHTC Application Data Extractor - Precision Version
Extracts specific data from exact cell locations in CTCAC applications

Based on detailed analysis of actual CTCAC application templates from 2023-2025
"""

import pandas as pd
import json
import os
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from typing import Dict, List, Any, Optional
import re
import sys

class CTCACPrecisionExtractor:
    """
    Precision extractor that targets exact cell locations in CTCAC applications
    """
    
    def __init__(self, input_path: str, output_path: str, log_path: str):
        """Initialize the extractor with file paths"""
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        self.log_path = Path(log_path)
        
        # Create directories if they don't exist
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.log_path.mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        self.setup_logging()
        
        # Define exact sheet names we're targeting
        self.target_sheets = {
            'application': 'Application',
            'sources_uses': 'Sources and Uses Budget',
            'points_system': 'Points System',
            'basis_credits': 'Basis & Credits',
            'tie_breaker': ['Tie Breaker', 'Final Tie Breaker', 'Disaster Credit Tie Breaker']
        }
    
    def setup_logging(self):
        """Set up logging configuration"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_path / f"ctcac_precision_extraction_{timestamp}.log"
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.logger.propagate = False
        
        self.logger.info(f"Precision extraction logging initialized. Log file: {log_file}")
    
    def parse_filename(self, filename: str) -> Dict[str, str]:
        """Parse filename to extract metadata"""
        base_name = Path(filename).stem
        
        parsed = {
            'year': 'unknown',
            'credit_type': 'unknown',
            'round': 'unknown',
            'project_id': 'unknown',
            'original_filename': filename
        }
        
        try:
            # Pattern for both formats: YYYY_XpctXX_RX_YY-XXX and YYYY_XpctXX_RX_CAYYXXX
            pattern = r'(\d{4})_(\d+pct)_([rR]\d+)_(?:CA)?(.+)'
            match = re.match(pattern, base_name)
            
            if match:
                parsed['year'] = match.group(1)
                parsed['credit_type'] = match.group(2)
                parsed['round'] = match.group(3).upper()
                parsed['project_id'] = match.group(4)
        
        except Exception as e:
            self.logger.warning(f"Error parsing filename {filename}: {e}")
        
        return parsed
    
    def create_output_filename(self, parsed_info: Dict[str, str]) -> str:
        """Create standardized output filename"""
        return f"{parsed_info['year']}_{parsed_info['credit_type']}_{parsed_info['round']}_CA-{parsed_info['project_id']}.json"
    
    def get_cell_value(self, sheet, cell_ref: str):
        """Safely get cell value from sheet"""
        try:
            cell = sheet[cell_ref]
            return cell.value if cell and cell.value is not None else None
        except:
            return None
    
    def find_sheet_name(self, workbook, target_names: List[str]) -> Optional[str]:
        """Find the actual sheet name from a list of possibilities"""
        if isinstance(target_names, str):
            target_names = [target_names]
        
        for name in target_names:
            if name in workbook.sheetnames:
                return name
        return None
    
    def extract_project_info(self, workbook) -> Dict[str, Any]:
        """Extract project information from Application sheet - targeting exact locations"""
        try:
            sheet_name = self.find_sheet_name(workbook, ['Application'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            project_info = {
                # Basic project identification - exact cell locations based on analysis
                'ctcac_applicant': self.get_cell_value(sheet, 'H16'),  # Consistent across files
                'project_name': self.get_cell_value(sheet, 'H18'),     # Consistent across files
                'project_name_alt': self.get_cell_value(sheet, 'H17'), # For 9% applications
                
                # Additional project details we can extract
                'application_type': None,
                'state_credits_election': None,
                'project_address': None,
                'total_units': None,
                'affordable_units': None
            }
            
            # Determine application type from sheet content
            for row in range(8, 15):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and isinstance(cell_value, str):
                    if '4%' in cell_value and 'TAX-EXEMPT BONDS' in cell_value.upper():
                        project_info['application_type'] = '4% with Tax-Exempt Bonds'
                    elif '9%' in cell_value and 'COMPETITIVE' in cell_value.upper():
                        project_info['application_type'] = '9% Competitive'
            
            # Look for state credits election (around row 33-35)
            for row in range(32, 38):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and 'state credit' in str(cell_value).lower():
                    # Look for Yes/No in adjacent cells
                    for col in ['B', 'C', 'D', 'E']:
                        response = self.get_cell_value(sheet, f'{col}{row}')
                        if response and str(response).lower() in ['yes', 'no']:
                            project_info['state_credits_election'] = str(response)
                            break
            
            # Clean up None values
            project_info = {k: v for k, v in project_info.items() if v is not None}
            
            return project_info
            
        except Exception as e:
            self.logger.error(f"Error extracting project info: {e}")
            return {}
    
    def extract_sources_and_uses(self, workbook) -> Dict[str, Any]:
        """Extract detailed sources and uses from exact locations"""
        try:
            sheet_name = self.find_sheet_name(workbook, ['Sources and Uses Budget'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            sources_uses = {
                'land_costs': {},
                'construction_costs': {},
                'soft_costs': {},
                'financing_costs': {},
                'totals': {},
                'sources': {}
            }
            
            # Land costs section (rows 4-12 based on analysis)
            land_cost_items = [
                ('land_cost_value', 4),
                ('demolition', 5),
                ('legal', 6),
                ('total_land_cost', 8),
                ('existing_improvements', 9),
                ('total_acquisition', 11),
                ('total_land_acquisition', 12)
            ]
            
            for item_name, row in land_cost_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                res_cost = self.get_cell_value(sheet, f'C{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['land_costs'][item_name] = {
                        'total': float(total_cost),
                        'residential': float(res_cost) if res_cost else None
                    }
            
            # Construction costs (rehabilitation: rows 18-26, new construction: rows 29-38)
            construction_items = [
                ('rehab_structures', 18),
                ('rehab_general_requirements', 19),
                ('rehab_contractor_overhead', 20),
                ('rehab_contractor_profit', 21),
                ('rehab_insurance', 23),
                ('total_rehab_costs', 26),
                ('new_site_work', 29),
                ('new_structures', 30),
                ('new_general_requirements', 31),
                ('new_contractor_overhead', 32),
                ('new_contractor_profit', 33),
                ('new_prevailing_wages', 34),
                ('total_new_construction', 38)
            ]
            
            for item_name, row in construction_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['construction_costs'][item_name] = float(total_cost)
            
            # Soft costs (around rows 40-50)
            soft_cost_items = [
                ('architectural_design', 40),
                ('architectural_supervision', 41),
                ('total_architectural', 42),
                ('survey_engineering', 43),
                ('construction_interest', 45),
                ('origination_fee', 46)
            ]
            
            for item_name, row in soft_cost_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['soft_costs'][item_name] = float(total_cost)
            
            # Look for permanent sources in columns F onwards
            sources_header_row = 2
            for col_idx in range(5, 15):  # Columns F through O
                col_letter = chr(65 + col_idx)  # Convert to letter
                source_name = self.get_cell_value(sheet, f'{col_letter}{sources_header_row}')
                if source_name and isinstance(source_name, str):
                    # Look for source amounts in key total rows
                    total_amount = self.get_cell_value(sheet, f'{col_letter}12') or \
                                 self.get_cell_value(sheet, f'{col_letter}8') or \
                                 self.get_cell_value(sheet, f'{col_letter}26')
                    
                    if total_amount and isinstance(total_amount, (int, float)) and total_amount > 0:
                        clean_name = source_name.replace('\n', ' ').replace('\r', ' ').strip()
                        sources_uses['sources'][clean_name] = float(total_amount)
            
            return sources_uses
            
        except Exception as e:
            self.logger.error(f"Error extracting sources and uses: {e}")
            return {}
    
    def extract_unit_mix(self, workbook) -> Dict[str, Any]:
        """Extract unit mix information from Application sheet"""
        try:
            sheet_name = self.find_sheet_name(workbook, ['Application'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            unit_mix = {
                'total_units': None,
                'affordable_units': None,
                'market_rate_units': None,
                'unit_details': {}
            }
            
            # Scan for unit mix table (typically around rows 150-250 in Application sheet)
            for row in range(100, 300):
                for col in range(8):  # Check first 8 columns
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        
                        # Look for total units
                        if 'total' in cell_lower and 'unit' in cell_lower:
                            # Check adjacent cells for number
                            for adj_col in range(col + 1, min(col + 4, 8)):
                                adj_letter = chr(65 + adj_col)
                                adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                if adj_value and isinstance(adj_value, (int, float)):
                                    unit_mix['total_units'] = int(adj_value)
                                    break
                        
                        # Look for affordable units
                        elif any(term in cell_lower for term in ['affordable', 'low income', 'tax credit']):
                            for adj_col in range(col + 1, min(col + 4, 8)):
                                adj_letter = chr(65 + adj_col)
                                adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                if adj_value and isinstance(adj_value, (int, float)):
                                    unit_mix['affordable_units'] = int(adj_value)
                                    break
            
            return unit_mix
            
        except Exception as e:
            self.logger.error(f"Error extracting unit mix: {e}")
            return {}
    
    def extract_scoring_points(self, workbook) -> Dict[str, Any]:
        """Extract competitive scoring points from Points System sheet"""
        try:
            sheet_name = self.find_sheet_name(workbook, ['Points System'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            scoring = {
                'total_points_claimed': 0,
                'categories': {},
                'tie_breaker_items': []
            }
            
            # Scan the points sheet for scoring categories and values
            for row in range(1, 200):  # Points sheets can be quite long
                for col in range(6):  # Check first 6 columns
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str):
                        # Look for "Yes" responses which indicate claimed points
                        if cell_value.lower() == 'yes':
                            # Look for point values in the same row
                            for point_col in range(6):
                                point_letter = chr(65 + point_col)
                                point_value = self.get_cell_value(sheet, f'{point_letter}{row}')
                                
                                if point_value and isinstance(point_value, (int, float)) and 0 < point_value <= 50:
                                    # Look for category description
                                    category_desc = None
                                    for desc_row in range(max(1, row - 5), row + 1):
                                        desc_value = self.get_cell_value(sheet, f'A{desc_row}')
                                        if desc_value and isinstance(desc_value, str) and len(desc_value) > 10:
                                            category_desc = desc_value[:100]  # Truncate long descriptions
                                            break
                                    
                                    if category_desc:
                                        scoring['categories'][f'row_{row}'] = {
                                            'description': category_desc,
                                            'points': float(point_value),
                                            'claimed': True
                                        }
                                        scoring['total_points_claimed'] += float(point_value)
                                    break
            
            return scoring
            
        except Exception as e:
            self.logger.error(f"Error extracting scoring points: {e}")
            return {}
    
    def extract_basis_credits(self, workbook) -> Dict[str, Any]:
        """Extract basis and credit calculations"""
        try:
            sheet_name = self.find_sheet_name(workbook, ['Basis & Credits'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            basis_credits = {
                'eligible_basis': None,
                'qualified_basis': None,
                'federal_credits': None,
                'state_credits': None,
                'annual_federal_credits': None,
                'annual_state_credits': None
            }
            
            # Look for basis and credit amounts in typical locations
            key_terms = {
                'eligible_basis': ['total eligible basis', 'eligible basis'],
                'qualified_basis': ['qualified basis', 'total qualified basis'],
                'federal_credits': ['federal credit', 'annual federal credit'],
                'state_credits': ['state credit', 'annual state credit']
            }
            
            # Scan for basis and credit values
            for row in range(1, 100):
                for col in range(6):
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        
                        for key, terms in key_terms.items():
                            if any(term in cell_lower for term in terms):
                                # Look for numeric value in adjacent cells
                                for adj_col in range(col + 1, min(col + 4, 8)):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                    if adj_value and isinstance(adj_value, (int, float)) and adj_value > 0:
                                        basis_credits[key] = float(adj_value)
                                        break
                                break
            
            return basis_credits
            
        except Exception as e:
            self.logger.error(f"Error extracting basis and credits: {e}")
            return {}
    
    def extract_tie_breakers(self, workbook) -> Dict[str, Any]:
        """Extract tie breaker information"""
        try:
            sheet_name = self.find_sheet_name(workbook, self.target_sheets['tie_breaker'])
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            tie_breakers = {
                'criteria_met': [],
                'responses': {}
            }
            
            # Scan for tie breaker responses
            for row in range(1, 100):
                for col in range(6):
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str) and cell_value.lower() in ['yes', 'no']:
                        # Find description for this tie breaker
                        description = None
                        for desc_col in range(max(0, col - 3), col):
                            desc_letter = chr(65 + desc_col)
                            desc_value = self.get_cell_value(sheet, f'{desc_letter}{row}')
                            if desc_value and isinstance(desc_value, str) and len(desc_value) > 10:
                                description = desc_value[:150]
                                break
                        
                        if description:
                            tie_breakers['responses'][f'row_{row}'] = {
                                'description': description,
                                'response': cell_value.lower()
                            }
                            
                            if cell_value.lower() == 'yes':
                                tie_breakers['criteria_met'].append(description[:50])
            
            return tie_breakers
            
        except Exception as e:
            self.logger.error(f"Error extracting tie breakers: {e}")
            return {}
    
    def process_single_file(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Process a single Excel file with precision extraction"""
        try:
            self.logger.info(f"📂 Processing file: {file_path.name}")
            
            parsed_info = self.parse_filename(file_path.name)
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            self.logger.info(f"   📋 Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Initialize result with metadata
            result = {
                'metadata': {
                    'filename': file_path.name,
                    'processed_date': datetime.now().isoformat(),
                    'year': parsed_info['year'],
                    'credit_type': parsed_info['credit_type'],
                    'round': parsed_info['round'],
                    'project_id': parsed_info['project_id'],
                    'available_sheets': workbook.sheetnames,
                    'extraction_method': 'precision_targeting'
                },
                'project_info': {},
                'sources_and_uses': {},
                'unit_mix': {},
                'scoring_points': {},
                'basis_and_credits': {},
                'tie_breakers': {}
            }
            
            # Execute precision extractions
            extraction_results = {}
            
            # Project Information
            self.logger.info("   🎯 Extracting project information...")
            result['project_info'] = self.extract_project_info(workbook)
            extraction_results['project_info'] = len(result['project_info']) > 0
            
            # Sources and Uses Budget
            self.logger.info("   💰 Extracting sources and uses budget...")
            result['sources_and_uses'] = self.extract_sources_and_uses(workbook)
            extraction_results['sources_uses'] = len(result['sources_and_uses']) > 0
            
            # Unit Mix
            self.logger.info("   🏠 Extracting unit mix...")
            result['unit_mix'] = self.extract_unit_mix(workbook)
            extraction_results['unit_mix'] = len(result['unit_mix']) > 0
            
            # Scoring Points
            self.logger.info("   📊 Extracting scoring points...")
            result['scoring_points'] = self.extract_scoring_points(workbook)
            extraction_results['scoring'] = len(result['scoring_points']) > 0
            
            # Basis and Credits
            self.logger.info("   🧮 Extracting basis and credits...")
            result['basis_and_credits'] = self.extract_basis_credits(workbook)
            extraction_results['basis_credits'] = len(result['basis_and_credits']) > 0
            
            # Tie Breakers
            self.logger.info("   🎲 Extracting tie breakers...")
            result['tie_breakers'] = self.extract_tie_breakers(workbook)
            extraction_results['tie_breakers'] = len(result['tie_breakers']) > 0
            
            workbook.close()
            
            # Summary of extraction success
            successful_extractions = sum(extraction_results.values())
            total_extractions = len(extraction_results)
            
            self.logger.info(f"   ✅ Successfully extracted {successful_extractions}/{total_extractions} data sections")
            
            # Log specific successes
            for section, success in extraction_results.items():
                status = "✅" if success else "❌"
                self.logger.info(f"     {status} {section}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"   💥 Error processing file {file_path.name}: {e}")
            return None
    
    def process_files(self, limit: int = None, test_mode: bool = True):
        """Process files with precision extraction"""
        print("🎯 CTCAC Precision Extractor Starting...")
        print("=" * 60)
        
        if test_mode and limit is None:
            limit = 5
            print(f"🧪 TEST MODE: Processing up to {limit} files")
        elif limit:
            print(f"📊 LIMITED MODE: Processing up to {limit} files")
        else:
            print("🔄 FULL MODE: Processing all files")
        
        print("=" * 60)
        
        # Find Excel files
        excel_files = list(self.input_path.glob("*.xlsx")) + list(self.input_path.glob("*.xls"))
        
        if not excel_files:
            self.logger.error(f"❌ No Excel files found in {self.input_path}")
            return
        
        total_files = len(excel_files)
        self.logger.info(f"📁 Found {total_files} Excel files in input directory")
        
        if limit:
            excel_files = excel_files[:limit]
            self.logger.info(f"🎯 Processing {len(excel_files)} files (limited)")
        
        processed_count = 0
        failed_count = 0
        extraction_stats = {
            'project_info': 0,
            'sources_uses': 0,
            'unit_mix': 0,
            'scoring': 0,
            'basis_credits': 0,
            'tie_breakers': 0
        }
        
        print(f"\n📈 Processing Progress:")
        print("-" * 40)
        
        for i, file_path in enumerate(excel_files, 1):
            try:
                progress = f"[{i}/{len(excel_files)}]"
                print(f"\n{progress} {file_path.name}")
                
                result = self.process_single_file(file_path)
                
                if result:
                    # Parse filename for organized output
                    parsed_info = self.parse_filename(file_path.name)
                    json_filename = self.create_output_filename(parsed_info)
                    json_path = self.output_path / json_filename
                    
                    # Save as JSON
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
                    
                    self.logger.info(f"   💾 Saved: {json_filename}")
                    processed_count += 1
                    print(f"   ✅ SUCCESS")
                    
                    # Update extraction statistics
                    for section in extraction_stats.keys():
                        section_key = section if section in result else section.replace('_', '_and_') if section == 'sources_uses' else f"{section}_points" if section == 'scoring' else f"{section.replace('_', '_and_')}"
                        if section_key in result and result[section_key]:
                            extraction_stats[section] += 1
                else:
                    failed_count += 1
                    print(f"   ❌ FAILED")
                    
            except Exception as e:
                self.logger.error(f"   💥 Failed to process {file_path.name}: {e}")
                failed_count += 1
                print(f"   ❌ FAILED: {str(e)[:50]}...")
        
        # Final comprehensive summary
        print("\n" + "=" * 60)
        print("📊 COMPREHENSIVE PROCESSING SUMMARY")
        print("=" * 60)
        print(f"✅ Successfully processed: {processed_count}")
        print(f"❌ Failed: {failed_count}")
        print(f"📁 Total files attempted: {len(excel_files)}")
        print(f"📈 Success rate: {(processed_count/len(excel_files)*100):.1f}%")
        
        print(f"\n📋 EXTRACTION SUCCESS RATES:")
        for section, count in extraction_stats.items():
            rate = (count / processed_count * 100) if processed_count > 0 else 0
            print(f"  {section.replace('_', ' ').title()}: {count}/{processed_count} ({rate:.1f}%)")
        
        print(f"\n💾 JSON files saved to: {self.output_path}")
        print(f"📋 Log files saved to: {self.log_path}")
        
        self.logger.info(f"Precision extraction complete. Success: {processed_count}, Failed: {failed_count}")

def main():
    """Main function to run the precision extractor"""
    
    # Define file paths
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    output_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data'
    log_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/logs'
    
    # Create extractor instance
    extractor = CTCACPrecisionExtractor(input_path, output_path, log_path)
    
    # Run in test mode
    extractor.process_files(limit=5, test_mode=True)
    
    print("\n🎉 Precision extraction complete!")
    print(f"Check your JSON files in: {output_path}")
    print(f"Check your log files in: {log_path}")
    print("\n💡 To process more files:")
    print("   - Change limit=10 for more test files")
    print("   - Change test_mode=False to process all files")

if __name__ == "__main__":
    main()