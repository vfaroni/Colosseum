# extract_lihtc_key_data.py

import os
import json
import pandas as pd
from openpyxl import load_workbook
import logging
from pathlib import Path
import re
import time
from datetime import datetime
import traceback

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_extraction.log"),
        logging.StreamHandler()
    ]
)

class LIHTCKeyDataExtractor:
    """
    Extracts key data from LIHTC applications based on known fields and locations
    """
    
    def __init__(self, input_path, output_path):
        """
        Initialize the LIHTC Key Data Extractor
        
        Args:
            input_path (str): Path to the directory containing LIHTC applications
            output_path (str): Path to the base output directory for JSON files
        """
        self.input_path = Path(input_path)
        self.output_base_path = Path(output_path)
        
        # Define output paths
        self.output_paths = {
            "4p": Path(output_path) / "JSON_data" / "4p",
            "9p": Path(output_path) / "JSON_data" / "9p"
        }
        
        # Create output directories if they don't exist
        for path in self.output_paths.values():
            path.mkdir(parents=True, exist_ok=True)
        
        # Define key fields to extract - these are based on common LIHTC application structure
        # Format: {"json_path": {"sheet": "Sheet Name", "cell": "Cell Reference"}}
        self.key_fields_4p = {
            "metadata.ctcac_number": {"sheet": "Application", "cell": "B10"},
            "project.name": {"sheet": "Application", "cell": "B14"},
            "project.address": {"sheet": "Application", "cell": "B15"},
            "project.city": {"sheet": "Application", "cell": "B16"},
            "project.county": {"sheet": "Application", "cell": "G16"},
            "project.zip_code": {"sheet": "Application", "cell": "D16"},
            "project.census_tract": {"sheet": "Application", "cell": "B17"},
            "project.total_units": {"sheet": "Application", "cell": "D18"},
            "project.low_income_units": {"sheet": "Application", "cell": "F18"},
            "project.manager_units": {"sheet": "Application", "cell": "H18"},
            "project.housing_type": {"sheet": "Application", "cell": "E22"},
            "project.new_construction": {"sheet": "Application", "cell": "G20"},
            "project.rehabilitation": {"sheet": "Application", "cell": "G21"},
            "costs.total_project_cost": {"sheet": "Sources and Uses Budget", "cell": "F134"},
            "costs.total_eligible_basis": {"sheet": "Basis & Credits", "cell": "D98"},
            "costs.acquisition_costs.land": {"sheet": "Sources and Uses Budget", "cell": "F41"},
            "costs.construction_costs.total": {"sheet": "Sources and Uses Budget", "cell": "F58"},
            "costs.developer_fee.total": {"sheet": "Sources and Uses Budget", "cell": "F110"},
            "financing.tax_credits.federal_annual": {"sheet": "Basis & Credits", "cell": "G135"},
            "financing.tax_credits.credit_price": {"sheet": "Basis & Credits", "cell": "F142"},
            "development_team.developer.name": {"sheet": "Application", "cell": "B134"},
            "development_team.general_partner.name": {"sheet": "Application", "cell": "B144"},
            "timeline.estimated_construction_start": {"sheet": "Application", "cell": "B42"},
            "timeline.estimated_completion_date": {"sheet": "Application", "cell": "D42"},
            "scoring.tiebreaker_score": {"sheet": "Tie Breaker", "cell": "G69"}
        }
        
        # 9% applications are similar but might have slight differences
        self.key_fields_9p = {
            "metadata.ctcac_number": {"sheet": "Application", "cell": "B10"},
            "project.name": {"sheet": "Application", "cell": "B14"},
            "project.address": {"sheet": "Application", "cell": "B15"},
            "project.city": {"sheet": "Application", "cell": "B16"},
            "project.county": {"sheet": "Application", "cell": "G16"},
            "project.zip_code": {"sheet": "Application", "cell": "D16"},
            "project.census_tract": {"sheet": "Application", "cell": "B17"},
            "project.total_units": {"sheet": "Application", "cell": "D18"},
            "project.low_income_units": {"sheet": "Application", "cell": "F18"},
            "project.manager_units": {"sheet": "Application", "cell": "H18"},
            "project.housing_type": {"sheet": "Application", "cell": "E22"},
            "project.new_construction": {"sheet": "Application", "cell": "G20"},
            "project.rehabilitation": {"sheet": "Application", "cell": "G21"},
            "costs.total_project_cost": {"sheet": "Sources and Uses Budget", "cell": "F134"},
            "costs.total_eligible_basis": {"sheet": "Basis & Credits", "cell": "D98"},
            "costs.acquisition_costs.land": {"sheet": "Sources and Uses Budget", "cell": "F41"},
            "costs.construction_costs.total": {"sheet": "Sources and Uses Budget", "cell": "F58"},
            "costs.developer_fee.total": {"sheet": "Sources and Uses Budget", "cell": "F110"},
            "financing.tax_credits.federal_annual": {"sheet": "Basis & Credits", "cell": "G135"},
            "financing.tax_credits.credit_price": {"sheet": "Basis & Credits", "cell": "F142"},
            "development_team.developer.name": {"sheet": "Application", "cell": "B134"},
            "development_team.general_partner.name": {"sheet": "Application", "cell": "B144"},
            "timeline.estimated_construction_start": {"sheet": "Application", "cell": "B42"},
            "timeline.estimated_completion_date": {"sheet": "Application", "cell": "D42"},
            "scoring.total_score": {"sheet": "Points System", "cell": "F141"},
            "scoring.tiebreaker_score": {"sheet": "Tie Breaker", "cell": "G69"}
        }
        
        # Base schema for output JSON
        self.base_schema = self._create_base_schema()
        
        # Stats tracking
        self.stats = {
            "start_time": datetime.now().isoformat(),
            "total_files": 0,
            "successful": 0,
            "failed": 0,
            "files": {}
        }
    
    def _create_base_schema(self):
        """
        Create the base schema for application data
        
        Returns:
            dict: Base schema structure
        """
        return {
            "metadata": {
                "application_type": "",
                "application_round": "",
                "application_year": "",
                "file_name": "",
                "processing_date": "",
                "ctcac_number": "",
                "application_status": ""
            },
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "legislative_districts": {
                    "congressional": "",
                    "state_assembly": "",
                    "state_senate": ""
                },
                "geographic_region": "",
                "resource_area": "",
                "housing_type": "",
                "new_construction": None,
                "rehabilitation": None,
                "site_control": None,
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            },
            "scoring": {
                "total_score": 0,
                "general_partner_experience": 0,
                "management_company_experience": 0,
                "housing_needs": 0,
                "site_amenities": 0,
                "service_amenities": 0,
                "sustainable_building_methods": 0,
                "lowest_income": 0,
                "readiness_to_proceed": 0,
                "tiebreaker_score": 0
            },
            "costs": {
                "total_project_cost": 0,
                "total_eligible_basis": 0,
                "acquisition_costs": {
                    "land": 0,
                    "improvements": 0,
                    "total": 0
                },
                "construction_costs": {
                    "new_construction": 0,
                    "rehabilitation": 0,
                    "accessory_buildings": 0,
                    "general_requirements": 0,
                    "contractor_overhead": 0,
                    "contractor_profit": 0,
                    "prevailing_wage": None,
                    "total": 0
                },
                "soft_costs": {
                    "architectural": 0,
                    "survey_engineering": 0,
                    "construction_interest": 0,
                    "construction_loan_fees": 0,
                    "permanent_loan_fees": 0,
                    "legal_fees": 0,
                    "reserves": 0,
                    "appraisal": 0,
                    "hard_cost_contingency": 0,
                    "soft_cost_contingency": 0,
                    "total": 0
                },
                "developer_fee": {
                    "total": 0,
                    "deferred": 0
                },
                "per_unit_metrics": {
                    "total_cost_per_unit": 0,
                    "construction_cost_per_unit": 0
                }
            },
            "financing": {
                "tax_credits": {
                    "federal_annual": 0,
                    "federal_total": 0,
                    "state_annual": 0,
                    "state_total": 0,
                    "credit_price": 0,
                    "equity_amount": 0
                },
                "permanent_loans": [],
                "construction_loans": [],
                "public_funds": [],
                "debt_service_coverage": 0,
                "bond_allocation": 0
            },
            "development_team": {
                "developer": {
                    "name": "",
                    "contact": "",
                    "address": "",
                    "phone": "",
                    "email": ""
                },
                "general_partner": {
                    "name": "",
                    "type": "",
                    "ownership_percentage": 0
                },
                "limited_partner": {
                    "name": "",
                    "ownership_percentage": 0
                },
                "management_agent": {
                    "name": "",
                    "contact": ""
                }
            },
            "timeline": {
                "application_date": "",
                "estimated_construction_start": "",
                "estimated_completion_date": "",
                "placed_in_service_date": ""
            }
        }
    
    def get_application_files(self, app_type, limit=None, skip_processed=True):
        """
        Get a list of application files to process
        
        Args:
            app_type (str): Type of application ('4p' or '9p')
            limit (int, optional): Maximum number of files to return
            skip_processed (bool): Skip already processed files
            
        Returns:
            list: List of application file paths
        """
        files = []
        count = 0
        
        # Get list of already processed files
        processed_files = set()
        if skip_processed:
            for file in os.listdir(self.output_paths[app_type]):
                if file.endswith('.json'):
                    # Get original Excel filename
                    processed_files.add(file.replace('.json', '.xlsx'))
        
        # Find application files
        for file in os.listdir(self.input_path):
            if file.endswith('.xlsx') and (app_type.lower() in file.lower()):
                # Skip if already processed
                if skip_processed and file in processed_files:
                    logging.info(f"Skipping already processed file: {file}")
                    continue
                
                files.append(os.path.join(self.input_path, file))
                count += 1
                if limit and count >= limit:
                    break
        
        if not files:
            logging.warning(f"No {app_type} application files found to process")
        else:
            logging.info(f"Found {len(files)} {app_type} application files to process")
            
        return files
    
    def extract_value(self, workbook, sheet_name, cell_ref):
        """
        Extract a value from a specific cell in a workbook
        
        Args:
            workbook: The workbook object
            sheet_name: Name of the worksheet
            cell_ref: Cell reference (e.g., 'B10')
            
        Returns:
            The value in the cell, or None if not found
        """
        try:
            # Check if sheet exists
            if sheet_name not in workbook.sheetnames:
                return None
            
            # Get sheet
            sheet = workbook[sheet_name]
            
            # Get cell value
            value = sheet[cell_ref].value
            
            # Convert to appropriate type
            if isinstance(value, (pd.Timestamp, datetime)):
                return value.isoformat()
            
            return value
        except Exception as e:
            logging.debug(f"Error extracting {sheet_name}!{cell_ref}: {e}")
            return None
    
    def process_file(self, file_path, app_type, timeout=60):
        """
        Process a single LIHTC application file
        
        Args:
            file_path (str): Path to the Excel file
            app_type (str): Type of application ('4p' or '9p')
            timeout (int): Maximum processing time in seconds
            
        Returns:
            dict: Extracted data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Processing {app_type} application: {file_name}")
        
        # Create a copy of the base schema
        data = json.loads(json.dumps(self.base_schema))
        
        # Add basic metadata
        data["metadata"]["application_type"] = app_type
        data["metadata"]["file_name"] = file_name
        data["metadata"]["processing_date"] = datetime.now().isoformat()
        
        # Parse year and round from filename
        match = re.search(r'(\d{4})_\d+pct_R(\d+)', file_name)
        if match:
            data["metadata"]["application_year"] = match.group(1)
            data["metadata"]["application_round"] = match.group(2)
        
        # Parse CTCAC number from filename
        match_ctcac = re.search(r'_(\d+-\d+)', file_name)
        if match_ctcac:
            data["metadata"]["ctcac_number"] = match_ctcac.group(1)
        
        start_time = time.time()
        
        try:
            # Select the appropriate field mappings
            key_fields = self.key_fields_4p if app_type == '4p' else self.key_fields_9p
            
            # Load workbook with data_only=True to get values instead of formulas
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            # Extract values for each field
            for field_path, location in key_fields.items():
                sheet_name = location["sheet"]
                cell_ref = location["cell"]
                
                # Extract value
                value = self.extract_value(wb, sheet_name, cell_ref)
                
                # Set in data structure if not None
                if value is not None:
                    self._set_nested_value(data, field_path, value)
            
            # Calculate derived fields
            self._calculate_derived_fields(data)
            
            # Update stats
            elapsed_time = time.time() - start_time
            self.stats["total_files"] += 1
            self.stats["successful"] += 1
            self.stats["files"][file_name] = {
                "status": "success",
                "time": elapsed_time,
                "type": app_type
            }
            
            logging.info(f"Successfully processed {file_name} in {elapsed_time:.2f} seconds")
            
        except Exception as e:
            elapsed_time = time.time() - start_time
            logging.error(f"Error processing {file_name}: {e}")
            logging.error(traceback.format_exc())
            
            self.stats["total_files"] += 1
            self.stats["failed"] += 1
            self.stats["files"][file_name] = {
                "status": "error",
                "time": elapsed_time,
                "type": app_type,
                "error": str(e)
            }
            
            data["metadata"]["error"] = str(e)
        
        return data
    
    def _set_nested_value(self, data, path, value):
        """
        Set a value in a nested dictionary using dot notation
        
        Args:
            data (dict): Data dictionary
            path (str): Dot notation path (e.g., "project.name")
            value: Value to set
        """
        parts = path.split('.')
        current = data
        
        # Navigate to the correct location
        for i, part in enumerate(parts[:-1]):
            if part not in current:
                current[part] = {}
            current = current[part]
        
        # Set the value with appropriate type conversion
        last_part = parts[-1]
        
        # Handle common type conversions
        if value == "":
            # Keep empty strings as is for string fields
            if isinstance(current[last_part], str):
                current[last_part] = value
            # Convert empty strings to default values for numeric fields
            elif isinstance(current[last_part], (int, float)):
                current[last_part] = current[last_part]  # Keep default
            elif current[last_part] is None:
                current[last_part] = None  # Keep as None
        elif isinstance(current[last_part], bool) or last_part in ["new_construction", "rehabilitation"]:
            # Handle boolean conversions
            if isinstance(value, str):
                current[last_part] = value.lower() in ('yes', 'true', 'y', 't', '1', 'x')
            else:
                current[last_part] = bool(value)
        elif isinstance(current[last_part], int):
            # Convert to integer
            try:
                current[last_part] = int(float(value)) if value != "" else current[last_part]
            except (ValueError, TypeError):
                current[last_part] = current[last_part]  # Keep default on error
        elif isinstance(current[last_part], float):
            # Convert to float
            try:
                current[last_part] = float(value) if value != "" else current[last_part]
            except (ValueError, TypeError):
                current[last_part] = current[last_part]  # Keep default on error
        else:
            # For other types, just set the value
            current[last_part] = value
    
    def _calculate_derived_fields(self, data):
        """
        Calculate derived fields based on extracted data
        
        Args:
            data (dict): Application data
        """
        # Calculate federal total tax credits (10-year)
        if data["financing"]["tax_credits"]["federal_annual"]:
            data["financing"]["tax_credits"]["federal_total"] = data["financing"]["tax_credits"]["federal_annual"] * 10
        
        # Calculate cost per unit
        if data["costs"]["total_project_cost"] and data["project"]["total_units"]:
            try:
                data["costs"]["per_unit_metrics"]["total_cost_per_unit"] = data["costs"]["total_project_cost"] / data["project"]["total_units"]
            except (ZeroDivisionError, TypeError):
                pass
        
        # Calculate construction cost per unit
        if data["costs"]["construction_costs"]["total"] and data["project"]["total_units"]:
            try:
                data["costs"]["per_unit_metrics"]["construction_cost_per_unit"] = data["costs"]["construction_costs"]["total"] / data["project"]["total_units"]
            except (ZeroDivisionError, TypeError):
                pass
    
    def save_json(self, data, app_type):
        """
        Save data to a JSON file
        
        Args:
            data (dict): Application data
            app_type (str): Application type ('4p' or '9p')
            
        Returns:
            str: Path to saved file
        """
        try:
            # Generate filename
            ctcac_number = data["metadata"]["ctcac_number"]
            if ctcac_number:
                # Use CTCAC number if available
                filename = f"{ctcac_number}.json"
            else:
                # Otherwise use original filename without extension
                filename = data["metadata"]["file_name"].replace('.xlsx', '.json')
            
            output_path = self.output_paths[app_type] / filename
            
            with open(output_path, 'w') as f:
                json.dump(data, f, indent=2)
            
            logging.info(f"Saved to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error saving JSON: {e}")
            return None
    
    def process_batch(self, app_type, limit=None, skip_processed=True):
        """
        Process a batch of applications
        
        Args:
            app_type (str): Type of application ('4p' or '9p')
            limit (int, optional): Maximum number of files to process
            skip_processed (bool): Skip already processed files
            
        Returns:
            int: Number of successfully processed files
        """
        files = self.get_application_files(app_type, limit, skip_processed)
        
        if not files:
            logging.info(f"No {app_type} files to process")
            return 0
        
        logging.info(f"Starting batch processing of {len(files)} {app_type} applications")
        total_start_time = time.time()
        
        success_count = 0
        
        for i, file_path in enumerate(files):
            try:
                logging.info(f"Processing file {i+1} of {len(files)}")
                
                # Process file
                data = self.process_file(file_path, app_type)
                
                # Save result
                self.save_json(data, app_type)
                
                if self.stats["files"].get(os.path.basename(file_path), {}).get("status") == "success":
                    success_count += 1
                
            except Exception as e:
                logging.error(f"Unexpected error processing {file_path}: {e}")
                logging.error(traceback.format_exc())
        
        total_time = time.time() - total_start_time
        logging.info(f"Completed batch processing in {total_time:.2f} seconds")
        logging.info(f"Successfully processed {success_count} of {len(files)} files")
        
        return success_count
    
    def save_stats(self, filename="extraction_stats.json"):
        """Save processing statistics to a JSON file"""
        self.stats["end_time"] = datetime.now().isoformat()
        self.stats["total_time_seconds"] = (datetime.now() - datetime.fromisoformat(self.stats["start_time"])).total_seconds()
        
        with open(filename, 'w') as f:
            json.dump(self.stats, f, indent=2)
        
        logging.info(f"Saved processing stats to {filename}")

if __name__ == "__main__":
    # Parse command line arguments
    import argparse
    
    parser = argparse.ArgumentParser(description="Extract key data from LIHTC applications")
    parser.add_argument('--type', choices=['4p', '9p', 'both'], default='both',
                       help='Type of applications to process (4p, 9p, or both)')
    parser.add_argument('--limit', type=int, default=None,
                       help='Maximum number of files to process per type')
    parser.add_argument('--input', type=str, 
                       default='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data',
                       help='Input directory path')
    parser.add_argument('--output', type=str,
                       default='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data',
                       help='Output base directory path')
    
    args = parser.parse_args()
    
    # Create extractor
    extractor = LIHTCKeyDataExtractor(args.input, args.output)
    
    # Process files
    if args.type in ['4p', 'both']:
        extractor.process_batch('4p', args.limit)
    
    if args.type in ['9p', 'both']:
        extractor.process_batch('9p', args.limit)
    
    # Save stats
    extractor.save_stats()
    
    logging.info("Processing complete")