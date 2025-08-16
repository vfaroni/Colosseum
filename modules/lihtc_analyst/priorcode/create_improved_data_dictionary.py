# lihtc_application_processor.py

import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
import logging
import re
import time
from datetime import datetime
import traceback
import sys
import signal

# Set up logging with both file and console output
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_processor.log"),
        logging.StreamHandler()
    ]
)

class TimeoutException(Exception):
    """Exception raised when a function times out"""
    pass

def timeout_handler(signum, frame):
    """Handler for timeout signal"""
    raise TimeoutException("Function timed out")

class LIHTCApplicationProcessor:
    """
    Processes LIHTC applications from Excel format to structured JSON data.
    
    This processor handles both 4% and 9% applications, extracts key data,
    normalizes it, and generates structured JSON files for RAG system consumption.
    """
    
    def __init__(self, input_base_path, output_base_path):
        """
        Initialize the LIHTC Application Processor
        
        Args:
            input_base_path (str): Path to the directory containing LIHTC applications
            output_base_path (str): Path to the output directory for JSON files
        """
        self.input_base_path = Path(input_base_path)
        self.output_base_path = Path(output_base_path)
        
        # Define output paths
        self.output_paths = {
            "4p": Path(output_base_path) / "JSON_data" / "4p",
            "9p": Path(output_base_path) / "JSON_data" / "9p"
        }
        
        # Create output directories if they don't exist
        for path in self.output_paths.values():
            path.mkdir(parents=True, exist_ok=True)
        
        # Initialize schema
        self.define_schema()
        
        # Load field mappings from a JSON file if it exists, otherwise use default mappings
        self.field_mappings = self._load_field_mappings()
        
        # Stats tracking
        self.stats = {
            "total_files_processed": 0,
            "successful_files": 0,
            "failed_files": 0,
            "start_time": datetime.now(),
            "files": {}
        }
    
    def define_schema(self):
        """Define the JSON schema for LIHTC applications"""
        
        self.json_schema = {
            "metadata": {
                "application_type": "",       # 4% or 9%
                "application_round": "",      # Which funding round
                "application_year": "",       # Year of application
                "file_name": "",              # Original file name
                "processing_date": "",        # When this JSON was created
                "ctcac_number": "",           # CTCAC assigned number 
                "application_status": ""      # Status (awarded, waitlisted, rejected)
            },
            "project": {
                "name": "",                   # Project name
                "address": "",                # Physical address
                "city": "",                   # City
                "county": "",                 # County
                "zip_code": "",               # ZIP code
                "census_tract": "",           # Census tract
                "legislative_districts": {
                    "congressional": "",      # Congressional district
                    "state_assembly": "",     # State assembly district
                    "state_senate": ""        # State senate district
                },
                "geographic_region": "",      # CTCAC geographic region
                "resource_area": "",          # High/medium/low resource area designation
                "housing_type": "",           # Large family, senior, etc.
                "new_construction": None,     # True for new construction
                "rehabilitation": None,       # True for rehabilitation
                "site_control": None,         # Has site control
                "total_units": 0,             # Total units
                "low_income_units": 0,        # Low income units
                "manager_units": 0            # Manager units
            },
            "scoring": {
                "total_score": 0,             # Total application score
                "general_partner_experience": 0,    # GP experience points
                "management_company_experience": 0, # Management company points
                "housing_needs": 0,           # Housing needs points
                "site_amenities": 0,          # Site amenities points
                "service_amenities": 0,       # Service amenities points
                "sustainable_building_methods": 0,  # Sustainable building methods points 
                "lowest_income": 0,           # Lowest income points
                "readiness_to_proceed": 0,    # Readiness to proceed points
                "tiebreaker_score": 0         # Tiebreaker score
            },
            "costs": {
                "total_project_cost": 0,      # Total project cost
                "total_eligible_basis": 0,    # Total eligible basis
                "acquisition_costs": {
                    "land": 0,                # Land cost
                    "improvements": 0,        # Existing improvements cost
                    "total": 0                # Total acquisition cost
                },
                "construction_costs": {
                    "new_construction": 0,    # New construction cost
                    "rehabilitation": 0,      # Rehabilitation cost
                    "accessory_buildings": 0, # Accessory buildings cost
                    "general_requirements": 0, # General requirements
                    "contractor_overhead": 0, # Contractor overhead
                    "contractor_profit": 0,   # Contractor profit
                    "prevailing_wage": None,  # Subject to prevailing wage
                    "total": 0                # Total construction cost
                },
                "soft_costs": {
                    "architectural": 0,       # Architectural cost
                    "survey_engineering": 0,  # Survey and engineering
                    "construction_interest": 0, # Construction interest
                    "construction_loan_fees": 0, # Construction loan fees
                    "permanent_loan_fees": 0, # Permanent loan fees
                    "legal_fees": 0,          # Legal fees
                    "reserves": 0,            # Reserves
                    "appraisal": 0,           # Appraisal
                    "hard_cost_contingency": 0, # Hard cost contingency
                    "soft_cost_contingency": 0, # Soft cost contingency
                    "total": 0                # Total soft costs
                },
                "developer_fee": {
                    "total": 0,               # Total developer fee
                    "deferred": 0             # Deferred developer fee
                },
                "per_unit_metrics": {
                    "total_cost_per_unit": 0, # Total cost per unit
                    "construction_cost_per_unit": 0 # Construction cost per unit
                }
            },
            "financing": {
                "tax_credits": {
                    "federal_annual": 0,      # Annual federal credits
                    "federal_total": 0,       # Total federal credits (10 years)
                    "state_annual": 0,        # Annual state credits
                    "state_total": 0,         # Total state credits
                    "credit_price": 0,        # Credit price
                    "equity_amount": 0        # Equity from tax credits
                },
                "permanent_loans": [],        # List of permanent loans
                "construction_loans": [],     # List of construction loans
                "public_funds": [],           # List of public funds
                "debt_service_coverage": 0,   # Debt service coverage ratio
                "bond_allocation": 0          # Tax-exempt bond allocation
            },
            "rental_information": {
                "unit_mix": [],               # Array of unit types and counts
                "utility_allowances": {},     # Utility allowances by unit type
                "average_affordability": 0,   # Average affordability percentage
                "rent_levels": []             # Array of rent levels by AMI and unit type
            },
            "development_team": {
                "developer": {
                    "name": "",               # Developer name
                    "contact": "",            # Developer contact
                    "address": "",            # Developer address
                    "phone": "",              # Developer phone
                    "email": ""               # Developer email
                },
                "general_partner": {
                    "name": "",               # General partner name
                    "type": "",               # For-profit or non-profit
                    "ownership_percentage": 0 # GP ownership percentage
                },
                "limited_partner": {
                    "name": "",               # Limited partner name
                    "ownership_percentage": 0 # LP ownership percentage
                },
                "management_agent": {
                    "name": "",               # Management agent name
                    "contact": ""             # Management agent contact
                }
            },
            "set_asides": {
                "nonprofit": None,            # Nonprofit set-aside
                "rural": None,                # Rural set-aside
                "at_risk": None,              # At-risk set-aside
                "special_needs": None,        # Special needs set-aside
                "tribal": None                # Native American apportionment
            },
            "timeline": {
                "application_date": "",       # Application date
                "estimated_construction_start": "", # Estimated construction start
                "estimated_completion_date": "", # Estimated completion date
                "placed_in_service_date": ""  # Placed in service date
            }
        }
        
    def _load_field_mappings(self):
        """
        Load field mappings from a file if it exists, otherwise use default mappings
        
        Returns:
            dict: Field mappings for 4% and 9% applications
        """
        mapping_file = Path("field_mappings.json")
        
        if mapping_file.exists():
            try:
                with open(mapping_file, 'r') as f:
                    return json.load(f)
            except Exception as e:
                logging.warning(f"Could not load field mappings file: {e}")
        
        # Default field mappings if file doesn't exist
        return self._get_default_field_mappings()
    
    def _get_default_field_mappings(self):
        """
        Get default field mappings for 4% and 9% applications
        
        Returns:
            dict: Default field mappings
        """
        # This is a simplified version. In practice, you'd build a comprehensive mapping
        # based on your data dictionary analysis.
        return {
            "4p": {
                "metadata.ctcac_number": {"sheet": "Application", "cell": "B10"},
                "project.name": {"sheet": "Application", "cell": "B14"},
                "project.address": {"sheet": "Application", "cell": "B15"},
                "project.city": {"sheet": "Application", "cell": "B16"},
                "project.county": {"sheet": "Application", "cell": "G16"},
                "project.zip_code": {"sheet": "Application", "cell": "D16"},
                "project.census_tract": {"sheet": "Application", "cell": "B17"},
                "costs.total_project_cost": {"sheet": "Sources and Uses Budget", "cell": "F134"},
                "costs.total_eligible_basis": {"sheet": "Basis & Credits", "cell": "D98"},
                "financing.tax_credits.federal_annual": {"sheet": "Basis & Credits", "cell": "G135"},
                "financing.tax_credits.credit_price": {"sheet": "Basis & Credits", "cell": "F142"}
            },
            "9p": {
                "metadata.ctcac_number": {"sheet": "Application", "cell": "B10"},
                "project.name": {"sheet": "Application", "cell": "B14"},
                "project.address": {"sheet": "Application", "cell": "B15"},
                "project.city": {"sheet": "Application", "cell": "B16"},
                "project.county": {"sheet": "Application", "cell": "G16"},
                "project.zip_code": {"sheet": "Application", "cell": "D16"},
                "project.census_tract": {"sheet": "Application", "cell": "B17"},
                "costs.total_project_cost": {"sheet": "Sources and Uses Budget", "cell": "F134"},
                "costs.total_eligible_basis": {"sheet": "Basis & Credits", "cell": "D98"},
                "financing.tax_credits.federal_annual": {"sheet": "Basis & Credits", "cell": "G135"},
                "financing.tax_credits.credit_price": {"sheet": "Basis & Credits", "cell": "F142"}
            }
        }
    
    def get_application_files(self, app_type, limit=None, skip_processed=True):
        """
        Get a list of application files for processing
        
        Args:
            app_type (str): Type of application ('4p' or '9p')
            limit (int, optional): Maximum number of files to return
            skip_processed (bool): Skip already processed files
            
        Returns:
            list: List of application file paths
        """
        files = []
        count = 0
        
        # Get already processed files
        processed_files = set()
        if skip_processed:
            for file in os.listdir(self.output_paths[app_type]):
                if file.endswith('.json'):
                    processed_files.add(file.replace('.json', '.xlsx'))
        
        path = self.input_base_path
        for file in os.listdir(path):
            # Check if it's an Excel file and matches the application type
            if file.endswith('.xlsx') and (app_type.lower() in file.lower()):
                # Skip if already processed
                if skip_processed and file in processed_files:
                    logging.info(f"Skipping already processed file: {file}")
                    continue
                
                files.append(os.path.join(path, file))
                count += 1
                if limit and count >= limit:
                    break
        
        if not files:
            logging.warning(f"No {app_type} application files found in {path}")
        else:
            logging.info(f"Found {len(files)} {app_type} application files to process")
            
        return files
    
    def extract_value_safely(self, workbook, mapping, default=None, timeout=30):
        """
        Safely extract a value from an Excel file with timeout protection
        
        Args:
            workbook: The Excel workbook object
            mapping (dict): Mapping with sheet and cell information
            default: Default value if extraction fails
            timeout (int): Timeout in seconds
            
        Returns:
            The extracted value or default
        """
        # Set up the timeout
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(timeout)
        
        try:
            sheet_name = mapping.get('sheet')
            cell = mapping.get('cell')
            
            if not sheet_name or not cell:
                return default
            
            # Try to get the sheet
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                logging.warning(f"Sheet {sheet_name} not found")
                return default
            
            # Try to get the cell value
            try:
                value = sheet[cell].value
                
                # Convert numpy types to Python native types
                if isinstance(value, np.integer):
                    value = int(value)
                elif isinstance(value, np.floating):
                    value = float(value)
                elif isinstance(value, np.ndarray):
                    value = value.tolist()
                
                return value
            except Exception as e:
                logging.warning(f"Error getting cell {cell} from sheet {sheet_name}: {e}")
                return default
                
        except TimeoutException:
            logging.warning(f"Timeout occurred while extracting from {sheet_name}:{cell}")
            return default
        finally:
            # Cancel the alarm
            signal.alarm(0)
    
    def process_application(self, file_path, app_type):
        """
        Process a single LIHTC application
        
        Args:
            file_path (str): Path to the Excel file
            app_type (str): Type of application ('4p' or '9p')
            
        Returns:
            dict: Processed application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Processing {app_type} application: {file_name}")
        
        # Initialize with schema
        application_data = self._deep_copy_schema()
        
        # Add metadata
        application_data["metadata"]["application_type"] = app_type
        application_data["metadata"]["file_name"] = file_name
        application_data["metadata"]["processing_date"] = datetime.now().isoformat()
        
        # Extract application year and round from filename
        match = re.search(r'(\d{4})_\d+pct_R(\d+)', file_name)
        if match:
            application_data["metadata"]["application_year"] = match.group(1)
            application_data["metadata"]["application_round"] = match.group(2)
        
        # Track processing time
        start_time = time.time()
        
        try:
            # Load workbook but don't read all data at once
            workbook = pd.ExcelFile(file_path)
            
            # Use openpyxl for cell-by-cell access
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            # Process based on field mappings
            mappings = self.field_mappings.get(app_type, {})
            for field_path, mapping in mappings.items():
                # Extract value with timeout protection
                value = self.extract_value_safely(wb, mapping)
                
                # Set in the nested structure
                if value is not None:
                    self._set_nested_value(application_data, field_path, value)
            
            # Custom processing for complex fields
            self._process_special_fields(application_data, wb, app_type)
            
            # Calculate any derived fields
            self._calculate_derived_fields(application_data)
            
            # Update stats
            elapsed_time = time.time() - start_time
            self.stats["total_files_processed"] += 1
            self.stats["successful_files"] += 1
            self.stats["files"][file_name] = {
                "status": "success",
                "time": elapsed_time,
                "type": app_type
            }
            
            logging.info(f"Successfully processed {file_name} in {elapsed_time:.2f} seconds")
            
            return application_data
            
        except Exception as e:
            elapsed_time = time.time() - start_time
            self.stats["total_files_processed"] += 1
            self.stats["failed_files"] += 1
            self.stats["files"][file_name] = {
                "status": "failed",
                "time": elapsed_time,
                "type": app_type,
                "error": str(e)
            }
            
            logging.error(f"Error processing {file_name}: {e}")
            logging.error(traceback.format_exc())
            
            # Return a partial data structure with error info
            application_data["metadata"]["error"] = str(e)
            return application_data
    
    def _deep_copy_schema(self):
        """Make a deep copy of the schema for new applications"""
        return json.loads(json.dumps(self.json_schema))
    
    def _set_nested_value(self, data, path, value):
        """
        Set a value in a nested dictionary based on dot notation path
        
        Args:
            data (dict): The dictionary to modify
            path (str): Dot notation path (e.g., "project.name")
            value: The value to set
        """
        parts = path.split('.')
        current = data
        
        # Navigate to the right location
        for i, part in enumerate(parts[:-1]):
            if part not in current:
                current[part] = {}
            current = current[part]
        
        # Set the value
        current[parts[-1]] = value
    
    def _process_special_fields(self, data, workbook, app_type):
        """
        Process special fields that require custom logic
        
        Args:
            data (dict): The application data
            workbook: The Excel workbook
            app_type (str): Application type ('4p' or '9p')
        """
        # Example: Extracting unit mix from a complex range
        try:
            # This is just an example and would need to be customized
            # based on your specific Excel structure
            if 'Sources and Uses Budget' in workbook.sheetnames:
                sheet = workbook['Sources and Uses Budget']
                
                # Example: Extract construction costs from a range
                construction_costs = 0
                if sheet.cell(row=58, column=6).value:
                    construction_costs = sheet.cell(row=58, column=6).value
                data["costs"]["construction_costs"]["total"] = construction_costs
                
        except Exception as e:
            logging.warning(f"Error processing special fields: {e}")
    
    def _calculate_derived_fields(self, data):
        """
        Calculate derived fields based on extracted data
        
        Args:
            data (dict): The application data
        """
        # Example: Calculate per-unit metrics
        try:
            total_cost = data["costs"]["total_project_cost"]
            total_units = data["project"]["total_units"]
            
            if total_cost and total_units and total_units > 0:
                data["costs"]["per_unit_metrics"]["total_cost_per_unit"] = total_cost / total_units
        except Exception as e:
            logging.warning(f"Error calculating derived fields: {e}")
    
    def save_application_json(self, data, app_type):
        """
        Save application data to JSON file
        
        Args:
            data (dict): The application data
            app_type (str): Application type ('4p' or '9p')
            
        Returns:
            bool: Success or failure
        """
        try:
            # Create filename based on CTCAC number or original filename
            ctcac_number = data["metadata"]["ctcac_number"]
            filename = data["metadata"]["file_name"].replace('.xlsx', '.json')
            
            if ctcac_number:
                filename = f"{ctcac_number}.json"
            
            output_path = self.output_paths[app_type] / filename
            
            with open(output_path, 'w') as f:
                json.dump(data, f, indent=2)
            
            logging.info(f"Saved to {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error saving JSON: {e}")
            return False
    
    def process_batch(self, app_type, limit=None, save=True):
        """
        Process a batch of applications
        
        Args:
            app_type (str): Type of application ('4p' or '9p')
            limit (int, optional): Maximum number of files to process
            save (bool): Whether to save JSON output
            
        Returns:
            list: Processed application data
        """
        files = self.get_application_files(app_type, limit)
        results = []
        
        total_start_time = time.time()
        logging.info(f"Starting batch processing of {len(files)} {app_type} applications")
        
        for i, file_path in enumerate(files):
            logging.info(f"Processing file {i+1} of {len(files)}: {os.path.basename(file_path)}")
            
            try:
                # Process the application
                data = self.process_application(file_path, app_type)
                
                # Save if requested
                if save:
                    self.save_application_json(data, app_type)
                
                results.append(data)
                
            except Exception as e:
                logging.error(f"Error processing {file_path}: {e}")
                logging.error(traceback.format_exc())
        
        total_time = time.time() - total_start_time
        successful = self.stats["successful_files"]
        failed = self.stats["failed_files"]
        
        logging.info(f"Batch processing complete in {total_time:.2f} seconds")
        logging.info(f"Processed {len(files)} files: {successful} successful, {failed} failed")
        
        return results
    
    def save_stats(self):
        """Save processing statistics to a JSON file"""
        self.stats["end_time"] = datetime.now().isoformat()
        self.stats["total_time"] = (datetime.now() - self.stats["start_time"]).total_seconds()
        
        with open("processing_stats.json", 'w') as f:
            json.dump(self.stats, f, indent=2)
        
        logging.info(f"Saved processing stats to processing_stats.json")

if __name__ == "__main__":
    # Get command line arguments
    import argparse
    
    parser = argparse.ArgumentParser(description="Process LIHTC applications to JSON")
    parser.add_argument('--type', choices=['4p', '9p', 'both'], default='both',
                       help='Type of applications to process (4p, 9p, or both)')
    parser.add_argument('--limit', type=int, default=None,
                       help='Maximum number of files to process (default: all)')
    parser.add_argument('--input', type=str, 
                       default='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data',
                       help='Input directory path')
    parser.add_argument('--output', type=str,
                       default='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data',
                       help='Output base directory path')
    
    args = parser.parse_args()
    
    # Initialize processor
    processor = LIHTCApplicationProcessor(args.input, args.output)
    
    # Process applications
    if args.type in ['4p', 'both']:
        logging.info("Processing 4% applications...")
        processor.process_batch('4p', args.limit)
    
    if args.type in ['9p', 'both']:
        logging.info("Processing 9% applications...")
        processor.process_batch('9p', args.limit)
    
    # Save stats
    processor.save_stats()
    
    logging.info("Processing complete")