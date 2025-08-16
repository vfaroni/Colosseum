#!/usr/bin/env python3
"""
Roswell, NM Weather Data Analyzer
Adapted from KCEC Battery Point code for Roswell Industrial Air Park (USW00023009)
Downloads complete weather data from July 1, 2024 to present with temperature, wind, and precipitation
Uses EXACT same format as KCEC_rainfall_2024-01-01_to_2025-06-21_20250623_WER.xlsx
"""

import os
import json
import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class RoswellWeatherAnalyzer:
    def __init__(self, noaa_token):
        self.noaa_token = noaa_token
        self.station_id = "GHCND:USW00023009"  # Roswell Industrial Air Park
        self.base_url = "https://www.ncei.noaa.gov/cdo-web/api/v2/data"
        self.headers = {"token": noaa_token}
        
    def get_rain_category(self, inches):
        """Categorize rainfall amount - EXACT KCEC logic"""
        if inches == 0 or inches < 0.01:
            return ""  # Empty string for no rain (matching KCEC)
        elif inches < 0.25:
            return "Light"
        elif inches < 0.5:
            return "Moderate"
        elif inches < 1.0:
            return "Heavy"
        else:
            return "Very Heavy"
    
    def get_federal_holidays(self):
        """Get federal holidays for 2024 and 2025"""
        return {
            # 2024 Federal Holidays
            '2024-01-01': 'New Year\'s Day',
            '2024-01-15': 'Martin Luther King Jr. Day',
            '2024-02-19': 'Presidents\' Day',
            '2024-05-27': 'Memorial Day',
            '2024-06-19': 'Juneteenth',
            '2024-07-04': 'Independence Day',
            '2024-09-02': 'Labor Day',
            '2024-10-14': 'Columbus Day',
            '2024-11-11': 'Veterans Day',
            '2024-11-28': 'Thanksgiving Day',
            '2024-12-25': 'Christmas Day',
            
            # 2025 Federal Holidays
            '2025-01-01': 'New Year\'s Day',
            '2025-01-20': 'Martin Luther King Jr. Day',
            '2025-02-17': 'Presidents\' Day',
            '2025-05-26': 'Memorial Day',
            '2025-06-19': 'Juneteenth',
            '2025-07-04': 'Independence Day',
            '2025-09-01': 'Labor Day',
            '2025-10-13': 'Columbus Day',
            '2025-11-11': 'Veterans Day',
            '2025-11-27': 'Thanksgiving Day',
            '2025-12-25': 'Christmas Day'
        }
    
    def get_new_mexico_holidays(self):
        """Get New Mexico state holidays for 2024 and 2025"""
        return {
            # 2024 New Mexico Holidays (includes all federal plus NM-specific)
            '2024-01-01': 'New Year\'s Day',
            '2024-01-15': 'Martin Luther King Jr. Day',
            '2024-02-19': 'Presidents\' Day',
            '2024-05-27': 'Memorial Day',
            '2024-06-19': 'Juneteenth',
            '2024-07-04': 'Independence Day',
            '2024-08-09': 'Statehood Day',  # NM specific - August 9, 1912
            '2024-09-02': 'Labor Day',
            '2024-10-14': 'Columbus Day',
            '2024-11-11': 'Veterans Day',
            '2024-11-28': 'Thanksgiving Day',
            '2024-11-29': 'Day After Thanksgiving',  # NM observes this
            '2024-12-25': 'Christmas Day',
            
            # 2025 New Mexico Holidays
            '2025-01-01': 'New Year\'s Day',
            '2025-01-20': 'Martin Luther King Jr. Day',
            '2025-02-17': 'Presidents\' Day',
            '2025-05-26': 'Memorial Day',
            '2025-06-19': 'Juneteenth',
            '2025-07-04': 'Independence Day',
            '2025-08-08': 'Statehood Day',  # NM specific (observed Friday when falls on weekend)
            '2025-09-01': 'Labor Day',
            '2025-10-13': 'Columbus Day',
            '2025-11-11': 'Veterans Day',
            '2025-11-27': 'Thanksgiving Day',
            '2025-11-28': 'Day After Thanksgiving',  # NM observes this
            '2025-12-25': 'Christmas Day'
        }
    
    def download_weather_data(self, start_date="2024-07-01"):
        """Download complete weather data from start_date to present"""
        
        print(f"Downloading Roswell weather data: {self.station_id}")
        print(f"Date range: {start_date} to present")
        print("Data types: PRCP (precipitation), TMAX (high temp), TMIN (low temp), AWND (avg wind)")
        
        # Get current date
        end_date = datetime.now().strftime("%Y-%m-%d")
        
        # Split into monthly chunks to avoid API limits
        all_data = {}
        
        # Generate monthly chunks
        current_date = datetime.strptime(start_date, "%Y-%m-%d")
        today = datetime.now()
        date_chunks = []
        
        while current_date < today:
            # Get last day of current month
            if current_date.month == 12:
                next_month = current_date.replace(year=current_date.year + 1, month=1)
            else:
                next_month = current_date.replace(month=current_date.month + 1)
            
            last_day = next_month - timedelta(days=1)
            
            # Don't go beyond today
            if last_day > today:
                last_day = today
            
            chunk_start = current_date.strftime("%Y-%m-%d")
            chunk_end = last_day.strftime("%Y-%m-%d")
            date_chunks.append((chunk_start, chunk_end))
            
            current_date = next_month
        
        print(f"Will download {len(date_chunks)} monthly chunks...")
        
        for chunk_start, chunk_end in date_chunks:
            print(f"  Downloading {chunk_start} to {chunk_end}...")
            
            params = {
                "datasetid": "GHCND",
                "stationid": self.station_id,
                "datatypeid": "PRCP,TMAX,TMIN,AWND",  # All data types in one request
                "startdate": chunk_start,
                "enddate": chunk_end,
                "units": "standard",  # Returns inches and °F
                "limit": 1000
            }
            
            try:
                time.sleep(1)  # Rate limiting
                response = requests.get(self.base_url, params=params, headers=self.headers)
                response.raise_for_status()
                data = response.json()
                
                if "results" in data:
                    print(f"    Got {len(data['results'])} records")
                    
                    # Process the data by date and data type
                    for item in data["results"]:
                        date = item["date"][:10]
                        datatype = item["datatype"]
                        value = item["value"] if item["value"] else None
                        
                        if date not in all_data:
                            all_data[date] = {}
                        
                        # Store by data type
                        if datatype == "PRCP":
                            all_data[date]["Total_Inches"] = round(value, 2) if value else 0.0
                        elif datatype == "TMAX":
                            all_data[date]["High_Temp_F"] = round(value, 1) if value else None
                        elif datatype == "TMIN":
                            all_data[date]["Low_Temp_F"] = round(value, 1) if value else None
                        elif datatype == "AWND":
                            all_data[date]["Avg_Wind_MPH"] = round(value, 1) if value else None
                        
                else:
                    print(f"    No data for {chunk_start} to {chunk_end}")
                    
            except Exception as e:
                print(f"    Error downloading {chunk_start} to {chunk_end}: {e}")
        
        print(f"Total dates with weather data: {len(all_data)}")
        return all_data, start_date, end_date
    
    def create_complete_dataset(self, weather_data, start_date, end_date):
        """Create complete dataset with EXACT KCEC format (25 columns)"""
        
        print("Creating complete dataset with EXACT KCEC format...")
        
        # Get holiday dictionaries
        federal_holidays = self.get_federal_holidays()
        nm_holidays = self.get_new_mexico_holidays()
        
        # Create complete dataset
        results = []
        current_date = datetime.strptime(start_date, "%Y-%m-%d")
        last_date = datetime.strptime(end_date, "%Y-%m-%d")
        
        while current_date <= last_date:
            date_str = current_date.strftime('%Y-%m-%d')
            
            # Get weather data for this date
            day_data = weather_data.get(date_str, {})
            inches = day_data.get("Total_Inches", 0.0)
            high_temp = day_data.get("High_Temp_F")
            low_temp = day_data.get("Low_Temp_F")
            wind_speed = day_data.get("Avg_Wind_MPH")
            
            # Check holidays
            is_federal_holiday = 1 if date_str in federal_holidays else 0
            federal_holiday_name = federal_holidays.get(date_str, '')
            is_nm_holiday = 1 if date_str in nm_holidays else 0
            nm_holiday_name = nm_holidays.get(date_str, '')
            
            # EXACT KCEC FORMAT - 25 columns in exact order
            row = {
                # Columns 1-6: Date information  
                "Date": current_date.strftime('%-m/%-d/%y'),  # M/D/YY format (KCEC style)
                "Year": current_date.year,
                "Month": current_date.month,
                "Day": current_date.day,
                "Month_Name": current_date.strftime('%B'),
                "Day_of_Week": current_date.strftime('%A'),
                
                # Columns 7-10: Weather data
                "Total_Inches": inches,
                "High_Temp_F": high_temp,
                "Low_Temp_F": low_temp,
                "Avg_Wind_MPH": wind_speed,
                
                # Column 11: Category
                "Category": self.get_rain_category(inches),
                
                # Columns 12-15: Threshold indicators (EXACT KCEC names)
                "Measurable_0.01": 1 if inches >= 0.01 else 0,
                "Contract_0.10": 1 if inches >= 0.10 else 0,
                "Normal_0.25": 1 if inches >= 0.25 else 0,
                "Heavy_0.50": 1 if inches >= 0.50 else 0,
                
                # Columns 16-19: Category indicators (EXACT KCEC names)
                "Light_0.01": 1 if 0.01 <= inches < 0.25 else 0,
                "Moderate_0.25": 1 if 0.25 <= inches < 0.5 else 0,
                "Heavy_0.5": 1 if 0.5 <= inches < 1.0 else 0,
                "Very_Heavy_1.0": 1 if inches >= 1.0 else 0,
                
                # Column 20: Work suitability
                "Work_Suitable": 1 if inches < 0.5 else 0,
                
                # Column 21: Timestamp
                "Timestamp": datetime.now().isoformat(),
                
                # Columns 22-23: Federal holidays (EXACT KCEC names)
                "Federal_Holiday": is_federal_holiday,
                "Federal_Holiday_Name": federal_holiday_name,
                
                # Columns 24-25: State holidays (adapted for NM)
                "NM_State_Holiday": is_nm_holiday,
                "NM_State_Holiday_Name": nm_holiday_name,
            }
            
            results.append(row)
            current_date += timedelta(days=1)
        
        # Create DataFrame with exact column order
        columns = [
            "Date", "Year", "Month", "Day", "Month_Name", "Day_of_Week",
            "Total_Inches", "High_Temp_F", "Low_Temp_F", "Avg_Wind_MPH",
            "Category", "Measurable_0.01", "Contract_0.10", "Normal_0.25", "Heavy_0.50",
            "Light_0.01", "Moderate_0.25", "Heavy_0.5", "Very_Heavy_1.0",
            "Work_Suitable", "Timestamp", "Federal_Holiday", "Federal_Holiday_Name",
            "NM_State_Holiday", "NM_State_Holiday_Name"
        ]
        
        df = pd.DataFrame(results, columns=columns)
        return df
    
    def export_to_excel(self, df, filename):
        """Export to Excel with KCEC-style formatting"""
        
        print(f"Exporting to Excel: {filename}")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roswell Weather Data"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row (matching KCEC style)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders (matching KCEC style)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Save workbook
        wb.save(filename)
        print(f"Excel file saved successfully: {filename}")
    
    def analyze_and_export(self, start_date="2024-07-01"):
        """Complete analysis pipeline from download to Excel export"""
        
        print("Roswell, NM Weather Analysis - EXACT KCEC Format")
        print("=" * 60)
        
        # Download weather data
        weather_data, actual_start, actual_end = self.download_weather_data(start_date)
        
        if not weather_data:
            print("Failed to download weather data")
            return None
        
        # Create complete dataset
        df = self.create_complete_dataset(weather_data, actual_start, actual_end)
        
        # Generate filename (matching KCEC pattern)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"Roswell_rainfall_{actual_start}_to_{actual_end}_{timestamp}_WER.xlsx"
        
        # Export to Excel
        self.export_to_excel(df, excel_filename)
        
        # Display summary statistics
        print(f"\n" + "=" * 60)
        print("ANALYSIS SUMMARY")
        print("=" * 60)
        print(f"Station: {self.station_id} (Roswell Industrial Air Park)")
        print(f"Date Range: {actual_start} to {actual_end}")
        print(f"Total Records: {len(df)}")
        
        # Weather statistics
        total_rain = df['Total_Inches'].sum()
        measurable_days = df['Measurable_0.01'].sum()
        contract_days = df['Contract_0.10'].sum()
        normal_days = df['Normal_0.25'].sum()
        heavy_days = df['Heavy_0.50'].sum()
        very_heavy_days = df['Very_Heavy_1.0'].sum()
        
        print(f"\nPRECIPITATION SUMMARY:")
        print(f"Total rainfall: {total_rain:.2f} inches")
        print(f"Measurable days (≥0.01\"): {measurable_days} days ({measurable_days/len(df)*100:.1f}%)")
        print(f"Contract days (≥0.10\"): {contract_days} days ({contract_days/len(df)*100:.1f}%)")
        print(f"Normal days (≥0.25\"): {normal_days} days ({normal_days/len(df)*100:.1f}%)")
        print(f"Heavy days (≥0.50\"): {heavy_days} days ({heavy_days/len(df)*100:.1f}%)")
        print(f"Very heavy days (≥1.0\"): {very_heavy_days} days ({very_heavy_days/len(df)*100:.1f}%)")
        
        # Temperature statistics
        temp_coverage = df['High_Temp_F'].notna().sum()
        if temp_coverage > 0:
            print(f"\nTEMPERATURE SUMMARY:")
            print(f"Temperature coverage: {temp_coverage}/{len(df)} days ({temp_coverage/len(df)*100:.1f}%)")
            print(f"High Temperature: {df['High_Temp_F'].min():.1f}°F to {df['High_Temp_F'].max():.1f}°F")
            print(f"Low Temperature: {df['Low_Temp_F'].min():.1f}°F to {df['Low_Temp_F'].max():.1f}°F")
            print(f"Average High: {df['High_Temp_F'].mean():.1f}°F")
            print(f"Average Low: {df['Low_Temp_F'].mean():.1f}°F")
        
        # Wind statistics
        wind_coverage = df['Avg_Wind_MPH'].notna().sum()
        if wind_coverage > 0:
            print(f"\nWIND SUMMARY:")
            print(f"Wind coverage: {wind_coverage}/{len(df)} days ({wind_coverage/len(df)*100:.1f}%)")
            print(f"Wind Speed: {df['Avg_Wind_MPH'].min():.1f} to {df['Avg_Wind_MPH'].max():.1f} mph")
            print(f"Average Wind: {df['Avg_Wind_MPH'].mean():.1f} mph")
        
        # Holiday statistics
        federal_holidays = df['Federal_Holiday'].sum()
        nm_holidays = df['NM_State_Holiday'].sum()
        
        print(f"\nHOLIDAY SUMMARY:")
        print(f"Federal holidays: {federal_holidays}")
        print(f"New Mexico state holidays: {nm_holidays}")
        
        # Rain on holidays
        holiday_data = df[df['Federal_Holiday'] == 1]
        if len(holiday_data) > 0:
            rainy_holidays = holiday_data['Contract_0.10'].sum()
            print(f"Federal holidays with ≥0.10\" rain: {rainy_holidays}/{len(holiday_data)} ({rainy_holidays/len(holiday_data)*100:.1f}%)")
        
        print(f"\nFILE EXPORTED: {excel_filename}")
        print(f"Format: EXACT match to KCEC structure (25 columns)")
        print("=" * 60)
        
        return excel_filename, df

def main():
    """Main function to run Roswell weather analysis"""
    
    # NOAA API token - using the same token from KCEC config
    NOAA_TOKEN = "oaLvXPjjAWoSCizEBvNoHPNhATmdDmQA"
    
    # Create analyzer
    analyzer = RoswellWeatherAnalyzer(NOAA_TOKEN)
    
    # Run complete analysis
    result = analyzer.analyze_and_export(start_date="2024-07-01")
    
    if result:
        filename, df = result
        print(f"\nSuccess! Weather analysis complete.")
        print(f"Excel file ready: {filename}")
        print(f"Format matches KCEC exactly with NM holidays instead of CA")
    else:
        print("Analysis failed.")

if __name__ == "__main__":
    main()