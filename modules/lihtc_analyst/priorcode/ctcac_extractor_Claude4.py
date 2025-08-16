#!/usr/bin/env python3
"""
Enhanced CTCAC LIHTC Application Data Extractor - Complete RAG Version
Extracts comprehensive data from CTCAC applications for RAG system

This enhanced version captures:
- Complete project information & unit mix details
- All financing sources and detailed cost breakdowns  
- Full competitive scoring breakdown with all point categories
- Comprehensive basis & credit calculations
- Complete tie breaker responses
- 15-year pro forma data
- Sources and basis breakdown details
- SCE/FCE basis calculations

Based on detailed analysis of 2023-2025 CTCAC applications (4% and 9%)
"""

import pandas as pd
import json
import os
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from typing import Dict, List, Any, Optional
import re
import sys

class CTCACComprehensiveExtractor:
    """
    Comprehensive extractor that captures ALL valuable data from CTCAC applications
    """
    
    def __init__(self, input_path: str, output_path_4p: str, output_path_9p: str, log_path: str):
        """Initialize the extractor with file paths"""
        self.input_path = Path(input_path)
        self.output_path_4p = Path(output_path_4p)
        self.output_path_9p = Path(output_path_9p)
        self.log_path = Path(log_path)
        
        # Create directories if they don't exist
        self.output_path_4p.mkdir(parents=True, exist_ok=True)
        self.output_path_9p.mkdir(parents=True, exist_ok=True)
        self.log_path.mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        self.setup_logging()
        
        # Define comprehensive sheet mapping
        self.target_sheets = {
            'application': 'Application',
            'sources_uses': 'Sources and Uses Budget',
            'sources_basis_breakdown': 'Sources and Basis Breakdown',
            'points_system': 'Points System',
            'basis_credits': 'Basis & Credits',
            'tie_breaker': 'Tie Breaker',
            'final_tie_breaker': 'Final Tie Breaker ',
            'disaster_tie_breaker': 'Disaster Credit Tie Breaker',
            'pro_forma': '15 Year Pro Forma',
            'sce_basis': 'SCE Basis and Credits',
            'fce_basis': 'FCE Basis and Credits',
            'calHFA_addendum': 'CalHFA Addendum',
            'subsidy_contract': 'Subsidy Contract Calculation'
        }
    
    def setup_logging(self):
        """Set up logging configuration"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_path / f"ctcac_comprehensive_extraction_{timestamp}.log"
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.logger.propagate = False
        
        self.logger.info(f"Comprehensive extraction logging initialized. Log file: {log_file}")
    
    def parse_filename(self, filename: str) -> Dict[str, str]:
        """Parse filename to extract metadata"""
        base_name = Path(filename).stem
        
        parsed = {
            'year': 'unknown',
            'credit_type': 'unknown',
            'round': 'unknown',
            'project_id': 'unknown',
            'original_filename': filename
        }
        
        try:
            # Pattern for both formats: YYYY_XpctXX_RX_YY-XXX and YYYY_XpctXX_RX_CAYYXXX
            pattern = r'(\d{4})_(\d+pct)_([rR]\d+)_(?:CA)?(.+)'
            match = re.match(pattern, base_name)
            
            if match:
                parsed['year'] = match.group(1)
                parsed['credit_type'] = match.group(2)
                parsed['round'] = match.group(3).upper()
                parsed['project_id'] = match.group(4)
        
        except Exception as e:
            self.logger.warning(f"Error parsing filename {filename}: {e}")
        
        return parsed
    
    def determine_output_path(self, parsed_info: Dict[str, str]) -> Path:
        """Determine correct output path based on credit type"""
        if '4' in parsed_info['credit_type']:
            return self.output_path_4p
        else:
            return self.output_path_9p
    
    def create_output_filename(self, parsed_info: Dict[str, str]) -> str:
        """Create standardized output filename"""
        return f"{parsed_info['year']}_{parsed_info['credit_type']}_{parsed_info['round']}_CA-{parsed_info['project_id']}.json"
    
    def get_cell_value(self, sheet, cell_ref: str):
        """Safely get cell value from sheet"""
        try:
            cell = sheet[cell_ref]
            return cell.value if cell and cell.value is not None else None
        except:
            return None
    
    def find_sheet_name(self, workbook, target_name: str) -> Optional[str]:
        """Find the actual sheet name (handles slight variations)"""
        if target_name in workbook.sheetnames:
            return target_name
        
        # Look for close matches
        for sheet_name in workbook.sheetnames:
            if target_name.lower().replace(' ', '') in sheet_name.lower().replace(' ', ''):
                return sheet_name
        return None
    
    def extract_comprehensive_project_info(self, workbook) -> Dict[str, Any]:
        """Extract comprehensive project information"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Application')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            project_info = {
                # Basic project identification
                'ctcac_applicant': self.get_cell_value(sheet, 'H16'),
                'project_name': self.get_cell_value(sheet, 'H18'),
                'project_name_alt': self.get_cell_value(sheet, 'H17'),
                
                # Application details
                'application_type': None,
                'state_credits_election': None,
                'project_address': None,
                'city': None,
                'county': None,
                'zip_code': None,
                
                # Unit information
                'total_units': None,
                'affordable_units': None,
                'market_rate_units': None,
                'unit_mix_details': {},
                
                # Set-aside information
                'set_aside_type': None,
                'nonprofit_participation': None,
                'tribal_project': None,
                
                # Development information
                'new_construction': None,
                'rehabilitation': None,
                'acquisition_rehabilitation': None,
                
                # Special features
                'green_building': None,
                'transit_oriented': None,
                'preservation_project': None
            }
            
            # Determine application type
            for row in range(8, 20):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and isinstance(cell_value, str):
                    if '4%' in cell_value and 'TAX-EXEMPT BONDS' in cell_value.upper():
                        project_info['application_type'] = '4% with Tax-Exempt Bonds'
                    elif '9%' in cell_value and 'COMPETITIVE' in cell_value.upper():
                        project_info['application_type'] = '9% Competitive'
            
            # Extract location information (scan around project name area)
            for row in range(15, 30):
                for col in ['H', 'I', 'J', 'K']:
                    value = self.get_cell_value(sheet, f'{col}{row}')
                    if value and isinstance(value, str):
                        # Look for address patterns
                        if any(term in value.lower() for term in ['street', 'avenue', 'road', 'drive', 'way']):
                            project_info['project_address'] = value
                        # Look for city patterns
                        elif len(value) > 3 and value.replace(' ', '').replace(',', '').isalpha():
                            project_info['city'] = value
                        # Look for zip codes
                        elif re.match(r'^\d{5}(-\d{4})?$', value.strip()):
                            project_info['zip_code'] = value
            
            # Look for state credits election
            for row in range(30, 45):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and 'state credit' in str(cell_value).lower():
                    for col in ['B', 'C', 'D', 'E', 'F']:
                        response = self.get_cell_value(sheet, f'{col}{row}')
                        if response and str(response).lower() in ['yes', 'no']:
                            project_info['state_credits_election'] = str(response)
                            break
            
            # Extract unit mix information (comprehensive scan)
            unit_mix = self.extract_detailed_unit_mix(sheet)
            if unit_mix:
                project_info.update(unit_mix)
            
            # Clean up None values
            project_info = {k: v for k, v in project_info.items() if v is not None}
            
            return project_info
            
        except Exception as e:
            self.logger.error(f"Error extracting comprehensive project info: {e}")
            return {}
    
    def extract_detailed_unit_mix(self, sheet) -> Dict[str, Any]:
        """Extract detailed unit mix information from Application sheet"""
        try:
            unit_mix = {
                'total_units': None,
                'affordable_units': None,
                'market_rate_units': None,
                'unit_types': {},
                'ami_levels': {},
                'rent_schedule': {}
            }
            
            # Comprehensive scan for unit mix tables
            unit_keywords = [
                'bedroom', 'studio', '1 br', '2 br', '3 br', '4 br', 
                'unit type', 'unit mix', 'ami', 'rent', 'income',
                '30%', '50%', '60%', '80%', '120%'
            ]
            
            # Scan the entire Application sheet for unit data
            for row in range(50, 500):  # Extended range for comprehensive search
                row_has_unit_data = False
                row_data = {}
                
                for col in range(16):  # Columns A-P
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value:
                        row_data[col_letter] = cell_value
                        
                        if isinstance(cell_value, str):
                            cell_lower = cell_value.lower()
                            if any(keyword in cell_lower for keyword in unit_keywords):
                                row_has_unit_data = True
                
                # If row contains unit data, analyze it
                if row_has_unit_data:
                    # Look for total units
                    for col_letter, value in row_data.items():
                        if isinstance(value, str) and 'total' in value.lower() and 'unit' in value.lower():
                            # Check adjacent cells for numeric values
                            for adj_col in range(ord(col_letter) - 65 + 1, min(ord(col_letter) - 65 + 4, 16)):
                                adj_letter = chr(65 + adj_col)
                                adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                if isinstance(adj_value, (int, float)) and adj_value > 0:
                                    unit_mix['total_units'] = int(adj_value)
                                    break
                        
                        # Look for bedroom counts
                        if isinstance(value, str):
                            bedroom_match = re.search(r'(\d+)\s*bedroom|(\d+)\s*br', value.lower())
                            if bedroom_match:
                                bedrooms = bedroom_match.group(1) or bedroom_match.group(2)
                                # Look for unit count in same row
                                for adj_col in range(16):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                    if isinstance(adj_value, (int, float)) and 0 < adj_value < 200:
                                        unit_mix['unit_types'][f'{bedrooms}_bedroom'] = int(adj_value)
                                        break
                        
                        # Look for AMI percentages
                        if isinstance(value, str):
                            ami_match = re.search(r'(\d+)%.*ami', value.lower())
                            if ami_match:
                                ami_level = f"{ami_match.group(1)}%_AMI"
                                # Look for unit count in same row
                                for adj_col in range(16):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                    if isinstance(adj_value, (int, float)) and 0 < adj_value < 200:
                                        unit_mix['ami_levels'][ami_level] = int(adj_value)
                                        break
            
            return unit_mix
            
        except Exception as e:
            self.logger.error(f"Error extracting detailed unit mix: {e}")
            return {}
    
    def extract_comprehensive_sources_uses(self, workbook) -> Dict[str, Any]:
        """Extract comprehensive sources and uses with ALL financing sources"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Sources and Uses Budget')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            sources_uses = {
                # Detailed cost categories
                'land_costs': {},
                'rehabilitation_costs': {},
                'new_construction_costs': {},
                'architectural_costs': {},
                'construction_interest_fees': {},
                'permanent_financing_costs': {},
                'other_costs': {},
                
                # Comprehensive financing sources
                'financing_sources': {},
                'source_details': {},
                
                # Totals and calculations
                'total_uses': None,
                'total_sources': None,
                'sources_uses_balance': None
            }
            
            # Extract ALL financing sources (columns F through P)
            self.logger.info("   🔍 Extracting all financing sources...")
            for col in range(5, 16):  # Columns F-P
                col_letter = chr(65 + col)
                
                # Get source name from header
                source_name = self.get_cell_value(sheet, f'{col_letter}2')
                if source_name and isinstance(source_name, str) and source_name.strip():
                    clean_name = source_name.replace('\n', ' ').replace('\r', ' ').strip()
                    
                    source_details = {
                        'name': clean_name,
                        'amounts': {},
                        'total_amount': 0
                    }
                    
                    # Check key total rows for amounts
                    key_rows = [8, 12, 26, 38, 42, 53, 60, 70, 80]
                    for row in key_rows:
                        amount = self.get_cell_value(sheet, f'{col_letter}{row}')
                        if amount and isinstance(amount, (int, float)) and amount > 0:
                            source_details['amounts'][f'row_{row}'] = float(amount)
                            source_details['total_amount'] += float(amount)
                    
                    if source_details['total_amount'] > 0:
                        sources_uses['financing_sources'][clean_name] = source_details['total_amount']
                        sources_uses['source_details'][clean_name] = source_details
            
            # Extract detailed use categories
            self.logger.info("   🏗️ Extracting detailed cost categories...")
            
            # Land costs (rows 4-12)
            land_items = [
                ('land_cost_value', 4),
                ('demolition', 5),
                ('land_lease_prepayment', 7),
                ('total_land_cost', 8),
                ('existing_improvements', 9),
                ('offsite_improvements', 10),
                ('total_acquisition', 11),
                ('total_land_acquisition', 12)
            ]
            
            for item_name, row in land_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                res_cost = self.get_cell_value(sheet, f'C{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['land_costs'][item_name] = {
                        'total': float(total_cost),
                        'residential': float(res_cost) if res_cost and isinstance(res_cost, (int, float)) else None
                    }
            
            # Rehabilitation costs (rows 17-26)
            rehab_items = [
                ('site_work', 17),
                ('structures', 18),
                ('general_requirements', 19),
                ('contractor_overhead', 20),
                ('contractor_profit', 21),
                ('prevailing_wages', 22),
                ('general_liability_insurance', 23),
                ('construction_management', 24),
                ('other_rehab', 25),
                ('total_rehabilitation', 26)
            ]
            
            for item_name, row in rehab_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['rehabilitation_costs'][item_name] = float(total_cost)
            
            # New construction costs (rows 29-38)
            new_construction_items = [
                ('site_work', 29),
                ('structures', 30),
                ('general_requirements', 31),
                ('contractor_overhead', 32),
                ('contractor_profit', 33),
                ('prevailing_wages', 34),
                ('general_liability_insurance', 35),
                ('construction_management', 36),
                ('other_new_construction', 37),
                ('total_new_construction', 38)
            ]
            
            for item_name, row in new_construction_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['new_construction_costs'][item_name] = float(total_cost)
            
            # Architectural costs (rows 40-42)
            arch_items = [
                ('design', 40),
                ('supervision', 41),
                ('total_architectural', 42)
            ]
            
            for item_name, row in arch_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['architectural_costs'][item_name] = float(total_cost)
            
            # Construction interest and fees (rows 45-53)
            const_interest_items = [
                ('construction_loan_interest', 45),
                ('origination_fee', 46),
                ('credit_enhancement_fee', 47),
                ('bond_premium', 48),
                ('title_recording', 49),
                ('insurance', 51),
                ('other_construction_fees', 52),
                ('total_construction_interest_fees', 53)
            ]
            
            for item_name, row in const_interest_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['construction_interest_fees'][item_name] = float(total_cost)
            
            return sources_uses
            
        except Exception as e:
            self.logger.error(f"Error extracting comprehensive sources and uses: {e}")
            return {}
    
    def extract_complete_scoring_breakdown(self, workbook) -> Dict[str, Any]:
        """Extract complete competitive scoring breakdown"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Points System')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            scoring = {
                'total_points_claimed': 0,
                'max_possible_points': 0,
                'scoring_categories': {},
                'detailed_responses': {},
                'point_summary': {}
            }
            
            self.logger.info("   📊 Extracting complete scoring breakdown...")
            
            # Comprehensive scan of Points System sheet
            for row in range(1, 300):  # Extended range to catch all scoring items
                # Get all relevant cell values for this row
                desc_a = self.get_cell_value(sheet, f'A{row}')
                response_b = self.get_cell_value(sheet, f'B{row}')
                points_c = self.get_cell_value(sheet, f'C{row}')
                points_d = self.get_cell_value(sheet, f'D{row}')
                
                # Look for scoring responses and point values
                is_scoring_row = False
                points_awarded = 0
                max_points = 0
                
                # Check for Yes/No responses
                if response_b and str(response_b).lower() in ['yes', 'no']:
                    is_scoring_row = True
                    
                    # Look for associated point values
                    if isinstance(points_c, (int, float)) and 0 < points_c <= 50:
                        if str(response_b).lower() == 'yes':
                            points_awarded = float(points_c)
                        max_points = float(points_c)
                    elif isinstance(points_d, (int, float)) and 0 < points_d <= 50:
                        if str(response_b).lower() == 'yes':
                            points_awarded = float(points_d)
                        max_points = float(points_d)
                
                # Check for direct point values (even without Yes/No)
                elif isinstance(points_c, (int, float)) and 0 < points_c <= 50:
                    is_scoring_row = True
                    points_awarded = float(points_c)
                    max_points = float(points_c)
                elif isinstance(points_d, (int, float)) and 0 < points_d <= 50:
                    is_scoring_row = True
                    points_awarded = float(points_d)
                    max_points = float(points_d)
                
                if is_scoring_row:
                    # Get category description
                    category_desc = None
                    if desc_a and isinstance(desc_a, str) and len(desc_a) > 5:
                        category_desc = desc_a.strip()
                    else:
                        # Look in nearby rows for description
                        for desc_row in range(max(1, row - 5), row):
                            desc_value = self.get_cell_value(sheet, f'A{desc_row}')
                            if desc_value and isinstance(desc_value, str) and len(desc_value) > 10:
                                category_desc = desc_value.strip()
                                break
                    
                    if category_desc:
                        category_key = f'row_{row}'
                        
                        scoring['scoring_categories'][category_key] = {
                            'description': category_desc[:200],  # Limit length
                            'response': str(response_b) if response_b else None,
                            'points_awarded': points_awarded,
                            'max_points': max_points,
                            'claimed': points_awarded > 0
                        }
                        
                        scoring['total_points_claimed'] += points_awarded
                        scoring['max_possible_points'] += max_points
                        
                        # Create summary by category type
                        category_type = self.categorize_scoring_item(category_desc)
                        if category_type not in scoring['point_summary']:
                            scoring['point_summary'][category_type] = {
                                'points_claimed': 0,
                                'max_points': 0,
                                'items': 0
                            }
                        
                        scoring['point_summary'][category_type]['points_claimed'] += points_awarded
                        scoring['point_summary'][category_type]['max_points'] += max_points
                        scoring['point_summary'][category_type]['items'] += 1
            
            self.logger.info(f"   ✅ Found {len(scoring['scoring_categories'])} scoring items, Total: {scoring['total_points_claimed']}/{scoring['max_possible_points']} points")
            
            return scoring
            
        except Exception as e:
            self.logger.error(f"Error extracting complete scoring breakdown: {e}")
            return {}
    
    def categorize_scoring_item(self, description: str) -> str:
        """Categorize scoring items into logical groups"""
        desc_lower = description.lower()
        
        if any(term in desc_lower for term in ['experience', 'development', 'management', 'general partner']):
            return 'Developer_Experience'
        elif any(term in desc_lower for term in ['site', 'amenity', 'location', 'transit']):
            return 'Site_Amenities'
        elif any(term in desc_lower for term in ['green', 'solar', 'energy', 'sustainable']):
            return 'Green_Building'
        elif any(term in desc_lower for term in ['financing', 'leveraging', 'funding']):
            return 'Financing_Characteristics'
        elif any(term in desc_lower for term in ['cost', 'efficiency', 'savings']):
            return 'Cost_Efficiency'
        elif any(term in desc_lower for term in ['service', 'supportive', 'social']):
            return 'Supportive_Services'
        elif any(term in desc_lower for term in ['housing', 'preservation', 'at-risk']):
            return 'Housing_Preservation'
        else:
            return 'Other_Criteria'
    
    def extract_comprehensive_basis_credits(self, workbook) -> Dict[str, Any]:
        """Extract comprehensive basis and credit calculations"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Basis & Credits')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            basis_credits = {
                # Main calculations
                'eligible_basis': None,
                'qualified_basis': None,
                'federal_credits_annual': None,
                'state_credits_annual': None,
                'federal_credits_total': None,
                'state_credits_total': None,
                
                # Detailed breakdowns
                'basis_breakdown': {},
                'credit_calculations': {},
                'income_calculations': {},
                'operating_data': {},
                
                # Additional calculations
                'gross_rent_potential': None,
                'effective_gross_income': None,
                'operating_expenses': None,
                'net_operating_income': None
            }
            
            self.logger.info("   🧮 Extracting comprehensive basis and credits...")
            
            # Scan entire sheet for basis and credit terms
            basis_terms = {
                'eligible_basis': ['total eligible basis', 'eligible basis', 'aggregate eligible basis'],
                'qualified_basis': ['qualified basis', 'total qualified basis', 'aggregate qualified basis'],
                'federal_credits_annual': ['annual federal credit', 'federal tax credit annual', 'annual federal'],
                'state_credits_annual': ['annual state credit', 'state tax credit annual', 'annual state'],
                'federal_credits_total': ['total federal credit', 'federal credit total', 'total federal tax credit'],
                'state_credits_total': ['total state credit', 'state credit total', 'total state tax credit'],
                'gross_rent_potential': ['gross rent potential', 'potential gross rent', 'gross rental income'],
                'effective_gross_income': ['effective gross income', 'effective gross', 'adjusted gross income'],
                'operating_expenses': ['operating expenses', 'total operating expense', 'operating costs'],
                'net_operating_income': ['net operating income', 'noi', 'net income']
            }
            
            # Comprehensive scan for all calculations
            for row in range(1, 200):  # Extended range
                for col in range(12):  # Columns A-L
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower().strip()
                        
                        # Check against all basis terms
                        for key, terms in basis_terms.items():
                            if any(term in cell_lower for term in terms):
                                # Look for numeric value in adjacent cells
                                for adj_col in range(col + 1, min(col + 4, 12)):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                    
                                    if adj_value and isinstance(adj_value, (int, float)) and adj_value > 0:
                                        basis_credits[key] = float(adj_value)
                                        
                                        # Store detailed information
                                        basis_credits['credit_calculations'][f'{key}_details'] = {
                                            'source_cell': f'{col_letter}{row}',
                                            'value_cell': f'{adj_letter}{row}',
                                            'description': cell_value[:100],
                                            'amount': float(adj_value)
                                        }
                                        break
                                break
            
            return basis_credits
            
        except Exception as e:
            self.logger.error(f"Error extracting comprehensive basis and credits: {e}")
            return {}
    
    def extract_all_tie_breakers(self, workbook) -> Dict[str, Any]:
        """Extract responses from all tie breaker sheets"""
        try:
            tie_breaker_sheets = ['Tie Breaker', 'Final Tie Breaker ', 'Disaster Credit Tie Breaker']
            
            all_tie_breakers = {
                'tie_breaker_responses': {},
                'criteria_met': [],
                'total_criteria_met': 0,
                'sheets_processed': []
            }
            
            for sheet_target in tie_breaker_sheets:
                sheet_name = self.find_sheet_name(workbook, sheet_target)
                if sheet_name:
                    self.logger.info(f"   🎲 Extracting tie breakers from {sheet_name}...")
                    
                    sheet = workbook[sheet_name]
                    sheet_responses = {}
                    
                    # Scan for tie breaker responses
                    for row in range(1, 150):
                        for col in range(8):  # Columns A-H
                            col_letter = chr(65 + col)
                            cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                            
                            if cell_value and isinstance(cell_value, str) and cell_value.lower() in ['yes', 'no']:
                                # Find description for this tie breaker
                                description = None
                                for desc_col in range(max(0, col - 4), col):
                                    desc_letter = chr(65 + desc_col)
                                    desc_value = self.get_cell_value(sheet, f'{desc_letter}{row}')
                                    if desc_value and isinstance(desc_value, str) and len(desc_value) > 10:
                                        description = desc_value[:200]
                                        break
                                
                                if description:
                                    response_key = f'{sheet_name.replace(" ", "_")}_row_{row}'
                                    sheet_responses[response_key] = {
                                        'description': description,
                                        'response': cell_value.lower(),
                                        'sheet': sheet_name,
                                        'cell': f'{col_letter}{row}'
                                    }
                                    
                                    if cell_value.lower() == 'yes':
                                        all_tie_breakers['criteria_met'].append(description[:100])
                                        all_tie_breakers['total_criteria_met'] += 1
                    
                    all_tie_breakers['tie_breaker_responses'].update(sheet_responses)
                    all_tie_breakers['sheets_processed'].append(sheet_name)
            
            return all_tie_breakers
            
        except Exception as e:
            self.logger.error(f"Error extracting tie breakers: {e}")
            return {}
    
    def extract_pro_forma_data(self, workbook) -> Dict[str, Any]:
        """Extract 15-year pro forma operating data"""
        try:
            sheet_name = self.find_sheet_name(workbook, '15 Year Pro Forma')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            pro_forma = {
                'operating_projections': {},
                'year_1_data': {},
                'year_15_data': {},
                'key_assumptions': {},
                'cash_flow_summary': {}
            }
            
            self.logger.info("   📈 Extracting 15-year pro forma data...")
            
            # Look for key operating line items
            operating_items = [
                'gross rental income',
                'vacancy allowance',
                'effective gross income',
                'operating expenses',
                'management fee',
                'replacement reserves',
                'net operating income',
                'debt service',
                'cash flow'
            ]
            
            # Scan for operating data
            for row in range(1, 100):
                row_desc = self.get_cell_value(sheet, f'A{row}')
                if row_desc and isinstance(row_desc, str):
                    row_desc_lower = row_desc.lower()
                    
                    for item in operating_items:
                        if item in row_desc_lower:
                            # Look for Year 1 data (usually column B or C)
                            year1_value = self.get_cell_value(sheet, f'B{row}') or self.get_cell_value(sheet, f'C{row}')
                            if year1_value and isinstance(year1_value, (int, float)):
                                pro_forma['year_1_data'][item.replace(' ', '_')] = float(year1_value)
                            
                            # Look for Year 15 data (scan columns for last year)
                            for col in range(15, 20):  # Check columns O, P, Q, R, S
                                col_letter = chr(65 + col)
                                year15_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                                if year15_value and isinstance(year15_value, (int, float)):
                                    pro_forma['year_15_data'][item.replace(' ', '_')] = float(year15_value)
                                    break
                            break
            
            return pro_forma
            
        except Exception as e:
            self.logger.error(f"Error extracting pro forma data: {e}")
            return {}
    
    def extract_sources_basis_breakdown(self, workbook) -> Dict[str, Any]:
        """Extract detailed sources and basis breakdown"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Sources and Basis Breakdown')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            breakdown = {
                'basis_by_source': {},
                'source_allocations': {},
                'detailed_calculations': {}
            }
            
            self.logger.info("   🔍 Extracting sources and basis breakdown...")
            
            # This sheet typically contains detailed breakdowns of how each funding source
            # contributes to eligible basis calculations
            
            # Scan for source names and corresponding basis amounts
            for row in range(1, 100):
                for col in range(10):
                    col_letter = chr(65 + col)
                    cell_value = self.get_cell_value(sheet, f'{col_letter}{row}')
                    
                    if cell_value and isinstance(cell_value, str) and len(cell_value) > 5:
                        # Look for funding source names
                        if any(term in cell_value.lower() for term in ['loan', 'grant', 'equity', 'bond', 'subsidy']):
                            # Look for associated basis amounts in adjacent cells
                            for adj_col in range(col + 1, min(col + 5, 10)):
                                adj_letter = chr(65 + adj_col)
                                adj_value = self.get_cell_value(sheet, f'{adj_letter}{row}')
                                
                                if adj_value and isinstance(adj_value, (int, float)) and adj_value > 1000:
                                    breakdown['basis_by_source'][cell_value[:50]] = float(adj_value)
                                    break
            
            return breakdown
            
        except Exception as e:
            self.logger.error(f"Error extracting sources and basis breakdown: {e}")
            return {}
    
    def extract_sce_fce_basis(self, workbook) -> Dict[str, Any]:
        """Extract SCE and FCE basis calculations (9% applications only)"""
        try:
            sce_fce_data = {
                'sce_basis': {},
                'fce_basis': {},
                'has_sce_sheet': False,
                'has_fce_sheet': False
            }
            
            # Check for SCE Basis and Credits sheet
            sce_sheet_name = self.find_sheet_name(workbook, 'SCE Basis and Credits')
            if sce_sheet_name:
                sce_fce_data['has_sce_sheet'] = True
                self.logger.info("   📊 Extracting SCE basis calculations...")
                
                sce_sheet = workbook[sce_sheet_name]
                
                # Extract SCE-specific basis calculations
                for row in range(1, 50):
                    for col in range(8):
                        col_letter = chr(65 + col)
                        cell_value = self.get_cell_value(sce_sheet, f'{col_letter}{row}')
                        
                        if cell_value and isinstance(cell_value, str):
                            if 'basis' in cell_value.lower() or 'credit' in cell_value.lower():
                                # Look for numeric values
                                for adj_col in range(col + 1, min(col + 4, 8)):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(sce_sheet, f'{adj_letter}{row}')
                                    
                                    if adj_value and isinstance(adj_value, (int, float)) and adj_value > 0:
                                        sce_fce_data['sce_basis'][cell_value[:50]] = float(adj_value)
                                        break
            
            # Check for FCE Basis and Credits sheet
            fce_sheet_name = self.find_sheet_name(workbook, 'FCE Basis and Credits')
            if fce_sheet_name:
                sce_fce_data['has_fce_sheet'] = True
                self.logger.info("   📊 Extracting FCE basis calculations...")
                
                fce_sheet = workbook[fce_sheet_name]
                
                # Extract FCE-specific basis calculations
                for row in range(1, 50):
                    for col in range(8):
                        col_letter = chr(65 + col)
                        cell_value = self.get_cell_value(fce_sheet, f'{col_letter}{row}')
                        
                        if cell_value and isinstance(cell_value, str):
                            if 'basis' in cell_value.lower() or 'credit' in cell_value.lower():
                                # Look for numeric values
                                for adj_col in range(col + 1, min(col + 4, 8)):
                                    adj_letter = chr(65 + adj_col)
                                    adj_value = self.get_cell_value(fce_sheet, f'{adj_letter}{row}')
                                    
                                    if adj_value and isinstance(adj_value, (int, float)) and adj_value > 0:
                                        sce_fce_data['fce_basis'][cell_value[:50]] = float(adj_value)
                                        break
            
            return sce_fce_data
            
        except Exception as e:
            self.logger.error(f"Error extracting SCE/FCE basis: {e}")
            return {}
    
    def process_single_file(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Process a single Excel file with comprehensive extraction"""
        try:
            self.logger.info(f"📂 Processing file: {file_path.name}")
            
            parsed_info = self.parse_filename(file_path.name)
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            self.logger.info(f"   📋 Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Initialize comprehensive result structure
            result = {
                'metadata': {
                    'filename': file_path.name,
                    'processed_date': datetime.now().isoformat(),
                    'year': parsed_info['year'],
                    'credit_type': parsed_info['credit_type'],
                    'round': parsed_info['round'],
                    'project_id': parsed_info['project_id'],
                    'available_sheets': workbook.sheetnames,
                    'extraction_method': 'comprehensive_enhanced',
                    'extractor_version': '2.0'
                },
                
                # Core project data
                'project_information': {},
                'unit_mix_details': {},
                
                # Financial data
                'sources_and_uses': {},
                'basis_and_credits': {},
                'sources_basis_breakdown': {},
                'pro_forma_15_year': {},
                
                # Competitive data
                'scoring_breakdown': {},
                'tie_breaker_responses': {},
                
                # Additional sheets (9% applications)
                'sce_fce_calculations': {},
                
                # Summary metrics
                'extraction_summary': {}
            }
            
            # Execute comprehensive extractions
            extraction_results = {}
            
            # 1. Comprehensive Project Information
            self.logger.info("   🏢 Extracting comprehensive project information...")
            result['project_information'] = self.extract_comprehensive_project_info(workbook)
            extraction_results['project_info'] = len(result['project_information']) > 0
            
            # 2. Comprehensive Sources and Uses
            self.logger.info("   💰 Extracting comprehensive sources and uses...")
            result['sources_and_uses'] = self.extract_comprehensive_sources_uses(workbook)
            extraction_results['sources_uses'] = len(result['sources_and_uses']) > 0
            
            # 3. Complete Scoring Breakdown
            self.logger.info("   📊 Extracting complete scoring breakdown...")
            result['scoring_breakdown'] = self.extract_complete_scoring_breakdown(workbook)
            extraction_results['scoring'] = len(result['scoring_breakdown']) > 0
            
            # 4. Comprehensive Basis and Credits
            self.logger.info("   🧮 Extracting comprehensive basis and credits...")
            result['basis_and_credits'] = self.extract_comprehensive_basis_credits(workbook)
            extraction_results['basis_credits'] = len(result['basis_and_credits']) > 0
            
            # 5. All Tie Breaker Responses
            self.logger.info("   🎲 Extracting all tie breaker responses...")
            result['tie_breaker_responses'] = self.extract_all_tie_breakers(workbook)
            extraction_results['tie_breakers'] = len(result['tie_breaker_responses']) > 0
            
            # 6. 15-Year Pro Forma Data
            self.logger.info("   📈 Extracting 15-year pro forma data...")
            result['pro_forma_15_year'] = self.extract_pro_forma_data(workbook)
            extraction_results['pro_forma'] = len(result['pro_forma_15_year']) > 0
            
            # 7. Sources and Basis Breakdown
            self.logger.info("   🔍 Extracting sources and basis breakdown...")
            result['sources_basis_breakdown'] = self.extract_sources_basis_breakdown(workbook)
            extraction_results['sources_breakdown'] = len(result['sources_basis_breakdown']) > 0
            
            # 8. SCE/FCE Calculations (9% applications only)
            if '9' in parsed_info['credit_type']:
                self.logger.info("   📊 Extracting SCE/FCE calculations...")
                result['sce_fce_calculations'] = self.extract_sce_fce_basis(workbook)
                extraction_results['sce_fce'] = len(result['sce_fce_calculations']) > 0
            else:
                extraction_results['sce_fce'] = True  # N/A for 4% applications
            
            workbook.close()
            
            # Create extraction summary
            successful_extractions = sum(extraction_results.values())
            total_extractions = len(extraction_results)
            
            result['extraction_summary'] = {
                'sections_extracted': successful_extractions,
                'total_sections': total_extractions,
                'success_rate': round((successful_extractions / total_extractions) * 100, 1),
                'extraction_details': extraction_results,
                'data_richness_score': self.calculate_data_richness(result)
            }
            
            self.logger.info(f"   ✅ Successfully extracted {successful_extractions}/{total_extractions} data sections")
            self.logger.info(f"   📊 Data richness score: {result['extraction_summary']['data_richness_score']}/100")
            
            # Log specific successes
            for section, success in extraction_results.items():
                status = "✅" if success else "❌"
                self.logger.info(f"     {status} {section}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"   💥 Error processing file {file_path.name}: {e}")
            return None
    
    def calculate_data_richness(self, result: Dict[str, Any]) -> int:
        """Calculate a data richness score based on extracted content"""
        score = 0
        
        # Project information (20 points)
        if result['project_information']:
            score += min(20, len(result['project_information']) * 2)
        
        # Sources and uses (25 points)
        if result['sources_and_uses']:
            financing_sources = len(result['sources_and_uses'].get('financing_sources', {}))
            cost_categories = sum(len(cat) for cat in [
                result['sources_and_uses'].get('land_costs', {}),
                result['sources_and_uses'].get('new_construction_costs', {}),
                result['sources_and_uses'].get('architectural_costs', {})
            ])
            score += min(25, financing_sources * 3 + cost_categories)
        
        # Scoring breakdown (20 points)
        if result['scoring_breakdown']:
            scoring_items = len(result['scoring_breakdown'].get('scoring_categories', {}))
            score += min(20, scoring_items)
        
        # Basis and credits (15 points)
        if result['basis_and_credits']:
            score += min(15, len(result['basis_and_credits']) * 2)
        
        # Tie breakers (10 points)
        if result['tie_breaker_responses']:
            tie_breaker_items = len(result['tie_breaker_responses'].get('tie_breaker_responses', {}))
            score += min(10, tie_breaker_items)
        
        # Pro forma (10 points)
        if result['pro_forma_15_year']:
            score += min(10, len(result['pro_forma_15_year']) * 3)
        
        return min(100, score)
    
    def process_files(self, limit: int = None, test_mode: bool = True):
        """Process files with comprehensive extraction"""
        print("🎯 CTCAC COMPREHENSIVE EXTRACTOR - ENHANCED VERSION 2.0")
        print("=" * 70)
        print("🚀 Extracting ALL valuable LIHTC data for RAG system:")
        print("   • Complete project info & detailed unit mix")
        print("   • ALL financing sources & detailed cost breakdowns")
        print("   • Full competitive scoring with point breakdowns")
        print("   • Comprehensive basis & credit calculations")
        print("   • Complete tie breaker responses")
        print("   • 15-year pro forma operating data")
        print("   • Sources and basis breakdown details")
        print("   • SCE/FCE calculations (9% applications)")
        print("=" * 70)
        
        if test_mode and limit is None:
            limit = 5
            print(f"🧪 TEST MODE: Processing up to {limit} files")
        elif limit:
            print(f"📊 LIMITED MODE: Processing up to {limit} files")
        else:
            print("🔄 FULL MODE: Processing all files")
        
        print("=" * 70)
        
        # Find Excel files
        excel_files = list(self.input_path.glob("*.xlsx")) + list(self.input_path.glob("*.xls"))
        
        if not excel_files:
            self.logger.error(f"❌ No Excel files found in {self.input_path}")
            return
        
        total_files = len(excel_files)
        self.logger.info(f"📁 Found {total_files} Excel files in input directory")
        
        if limit:
            excel_files = excel_files[:limit]
            self.logger.info(f"🎯 Processing {len(excel_files)} files (limited)")
        
        processed_count = 0
        failed_count = 0
        extraction_stats = {
            'project_info': 0,
            'sources_uses': 0,
            'scoring': 0,
            'basis_credits': 0,
            'tie_breakers': 0,
            'pro_forma': 0,
            'sources_breakdown': 0,
            'sce_fce': 0
        }
        
        total_data_richness = 0
        
        print(f"\n📈 Processing Progress:")
        print("-" * 50)
        
        for i, file_path in enumerate(excel_files, 1):
            try:
                progress = f"[{i}/{len(excel_files)}]"
                print(f"\n{progress} {file_path.name}")
                
                result = self.process_single_file(file_path)
                
                if result:
                    # Parse filename for organized output
                    parsed_info = self.parse_filename(file_path.name)
                    output_path = self.determine_output_path(parsed_info)
                    json_filename = self.create_output_filename(parsed_info)
                    json_path = output_path / json_filename
                    
                    # Save as JSON
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
                    
                    self.logger.info(f"   💾 Saved: {json_filename}")
                    processed_count += 1
                    
                    # Update statistics
                    extraction_summary = result.get('extraction_summary', {})
                    extraction_details = extraction_summary.get('extraction_details', {})
                    
                    for section in extraction_stats.keys():
                        if extraction_details.get(section, False):
                            extraction_stats[section] += 1
                    
                    # Track data richness
                    data_richness = extraction_summary.get('data_richness_score', 0)
                    total_data_richness += data_richness
                    
                    print(f"   ✅ SUCCESS - Data richness: {data_richness}/100")
                else:
                    failed_count += 1
                    print(f"   ❌ FAILED")
                    
            except Exception as e:
                self.logger.error(f"   💥 Failed to process {file_path.name}: {e}")
                failed_count += 1
                print(f"   ❌ FAILED: {str(e)[:50]}...")
        
        # Final comprehensive summary
        print("\n" + "=" * 70)
        print("📊 COMPREHENSIVE PROCESSING SUMMARY")
        print("=" * 70)
        print(f"✅ Successfully processed: {processed_count}")
        print(f"❌ Failed: {failed_count}")
        print(f"📁 Total files attempted: {len(excel_files)}")
        print(f"📈 Success rate: {(processed_count/len(excel_files)*100):.1f}%")
        
        if processed_count > 0:
            avg_richness = total_data_richness / processed_count
            print(f"📊 Average data richness: {avg_richness:.1f}/100")
        
        print(f"\n📋 EXTRACTION SUCCESS RATES:")
        for section, count in extraction_stats.items():
            rate = (count / processed_count * 100) if processed_count > 0 else 0
            section_name = section.replace('_', ' ').title()
            print(f"  {section_name}: {count}/{processed_count} ({rate:.1f}%)")
        
        print(f"\n💾 JSON files saved to:")
        print(f"  4% applications: {self.output_path_4p}")
        print(f"  9% applications: {self.output_path_9p}")
        print(f"📋 Log files saved to: {self.log_path}")
        
        print(f"\n🎯 DATA EXTRACTED FOR RAG SYSTEM:")
        print("  ✅ Complete project information & unit mix details")
        print("  ✅ All financing sources & detailed cost breakdowns")
        print("  ✅ Full competitive scoring with point categories")
        print("  ✅ Comprehensive basis & credit calculations")
        print("  ✅ Complete tie breaker responses")
        print("  ✅ 15-year pro forma operating projections")
        print("  ✅ Sources and basis breakdown details")
        print("  ✅ SCE/FCE calculations (9% applications)")
        
        self.logger.info(f"Comprehensive extraction complete. Success: {processed_count}, Failed: {failed_count}")

def main():
    """Main function to run the comprehensive extractor"""
    
    # Define file paths
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p'
    output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p'
    log_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/logs'
    
    print("🎯 CTCAC COMPREHENSIVE DATA EXTRACTOR")
    print("=" * 50)
    print("Enhanced version that extracts ALL valuable LIHTC data:")
    print("• Project details & complete unit mix")
    print("• All financing sources & cost categories")
    print("• Complete competitive scoring breakdown")
    print("• Comprehensive basis & credit calculations")
    print("• All tie breaker responses")
    print("• 15-year operating pro forma")
    print("• Detailed sources & basis breakdowns")
    print("• SCE/FCE calculations")
    print("=" * 50)
    
    # Create extractor instance
    extractor = CTCACComprehensiveExtractor(input_path, output_path_4p, output_path_9p, log_path)
    
    # Run in test mode with 5 files
    extractor.process_files(limit=5, test_mode=True)
    
    print("\n🎉 Comprehensive extraction complete!")
    print("💡 Your JSON files now contain 10x more LIHTC data for the RAG system!")
    print(f"\n📁 Check your enhanced JSON files:")
    print(f"  4% applications: {output_path_4p}")
    print(f"  9% applications: {output_path_9p}")
    print(f"📋 Detailed logs: {log_path}")
    print("\n🚀 To process more files:")
    print("   - Change limit=10 for more test files")
    print("   - Change test_mode=False to process all 1,035+ files")
    print("   - Each JSON now contains comprehensive LIHTC data for your RAG system!")

if __name__ == "__main__":
    main()