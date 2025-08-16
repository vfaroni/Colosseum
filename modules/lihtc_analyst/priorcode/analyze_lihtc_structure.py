import os
import pandas as pd
import json
from openpyxl import load_workbook
import logging
import sys
from datetime import datetime

# Define all paths explicitly
SCRIPT_DIR = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/code/'
RAW_DATA_PATH = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
JSON_OUTPUT_4P = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p'
JSON_OUTPUT_9P = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p'
LOG_FILE = os.path.join(SCRIPT_DIR, 'lihtc_analysis.log')

# Set up logging to track our progress and any errors
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename=LOG_FILE
)

def analyze_excel_structure(file_path):
    """
    Analyze the structure of an Excel file containing LIHTC application data.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary with information about the Excel file structure
    """
    try:
        # Load the Excel workbook
        logging.info(f"Analyzing file: {file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Get basic information
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # Size in MB
        
        # Analyze each worksheet
        sheet_info = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Get the dimensions of data in the sheet
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Sample some column headers (assuming headers are in row 1)
            headers = []
            if max_row > 0 and max_col > 0:
                for col in range(1, min(max_col + 1, 11)):  # Sample up to 10 columns
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
            
            sheet_info.append({
                "sheet_name": sheet_name,
                "rows": max_row,
                "columns": max_col,
                "sample_headers": headers
            })
        
        # Create a summary of the file
        result = {
            "file_name": file_name,
            "file_path": file_path,
            "file_size_mb": round(file_size, 2),
            "total_sheets": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "detailed_sheet_info": sheet_info,
            "application_type": "Unknown"  # We'll try to determine this later
        }
        
        # Try to determine if it's a 4% or 9% application based on sheet names or content
        sheet_names_str = " ".join(wb.sheetnames).lower()
        if "9%" in sheet_names_str or "9 percent" in sheet_names_str:
            result["application_type"] = "9%"
        elif "4%" in sheet_names_str or "4 percent" in sheet_names_str:
            result["application_type"] = "4%"
        # If can't determine from sheet names, try to infer from file name
        if result["application_type"] == "Unknown":
            if "9p" in file_name.lower() or "9%" in file_name.lower():
                result["application_type"] = "9%"
            elif "4p" in file_name.lower() or "4%" in file_name.lower():
                result["application_type"] = "4%"
        
        logging.info(f"Successfully analyzed {file_name}")
        return result
    
    except Exception as e:
        logging.error(f"Error analyzing {file_path}: {str(e)}")
        return {
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "error": str(e)
        }

def ensure_directory_exists(directory_path):
    """
    Create a directory if it doesn't exist.
    
    Args:
        directory_path: Path to the directory
    """
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
        logging.info(f"Created directory: {directory_path}")
        print(f"Created directory: {directory_path}")

def save_analysis_to_json(analysis_results, output_file):
    """
    Save the analysis results to a JSON file.
    
    Args:
        analysis_results: List of analysis results
        output_file: Path to the output JSON file
    """
    # Ensure the output directory exists
    output_dir = os.path.dirname(output_file)
    ensure_directory_exists(output_dir)
    
    with open(output_file, 'w') as f:
        json.dump(analysis_results, f, indent=2)
    
    logging.info(f"Analysis results saved to {output_file}")
    print(f"Analysis results saved to {output_file}")

def main():
    """
    Main function to analyze LIHTC application structures.
    """
    print(f"Starting LIHTC Application Structure Analysis at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Log file: {LOG_FILE}")
    
    # Make sure script directory exists
    ensure_directory_exists(SCRIPT_DIR)
    
    # Check if raw data path exists
    if not os.path.exists(RAW_DATA_PATH):
        error_msg = f"ERROR: Raw data path does not exist: {RAW_DATA_PATH}"
        logging.error(error_msg)
        print(error_msg)
        return
    
    # Create output directories if they don't exist
    ensure_directory_exists(JSON_OUTPUT_4P)
    ensure_directory_exists(JSON_OUTPUT_9P)
    
    # Get all Excel files from the raw data directory
    all_files = [f for f in os.listdir(RAW_DATA_PATH) 
                if f.endswith('.xlsx') or f.endswith('.xls')]
    
    if not all_files:
        error_msg = f"ERROR: No Excel files found in {RAW_DATA_PATH}"
        logging.error(error_msg)
        print(error_msg)
        return
    
    print(f"Found {len(all_files)} Excel files in the raw data directory")
    
    # Filter files that appear to be 4% applications
    files_4p = [f for f in all_files if "4p" in f.lower() or "4%" in f.lower()]
    files_4p_paths = [os.path.join(RAW_DATA_PATH, f) for f in files_4p]
    
    # Filter files that appear to be 9% applications
    files_9p = [f for f in all_files if "9p" in f.lower() or "9%" in f.lower()]
    files_9p_paths = [os.path.join(RAW_DATA_PATH, f) for f in files_9p]
    
    # If no clear pattern is found, let's also check any remaining files
    remaining_files = [f for f in all_files if f not in files_4p and f not in files_9p]
    
    print(f"Found {len(files_4p)} potential 4% applications and {len(files_9p)} potential 9% applications.")
    print(f"There are {len(remaining_files)} additional Excel files that could not be classified by filename.")
    
    # Ask for confirmation before proceeding with analysis
    max_files = 10  # Default max files to analyze of each type
    
    # Analyze 4% applications
    print(f"\nAnalyzing up to {max_files} 4% applications...")
    results_4p = []
    for file_path in files_4p_paths[:max_files]:
        print(f"  Analyzing: {os.path.basename(file_path)}")
        result = analyze_excel_structure(file_path)
        results_4p.append(result)
    
    # Analyze 9% applications
    print(f"\nAnalyzing up to {max_files} 9% applications...")
    results_9p = []
    for file_path in files_9p_paths[:max_files]:
        print(f"  Analyzing: {os.path.basename(file_path)}")
        result = analyze_excel_structure(file_path)
        results_9p.append(result)
    
    # Save 4% application analysis
    output_file_4p = os.path.join(JSON_OUTPUT_4P, "application_structure_analysis.json")
    save_analysis_to_json({
        "application_type": "4%",
        "files_analyzed": len(results_4p),
        "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "results": results_4p
    }, output_file_4p)
    
    # Save 9% application analysis
    output_file_9p = os.path.join(JSON_OUTPUT_9P, "application_structure_analysis.json")
    save_analysis_to_json({
        "application_type": "9%",
        "files_analyzed": len(results_9p),
        "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "results": results_9p
    }, output_file_9p)
    
    print("\nAnalysis complete!")
    print(f"Analyzed {len(results_4p)} 4% applications and {len(results_9p)} 9% applications.")
    print(f"Results saved to:")
    print(f"  - {output_file_4p}")
    print(f"  - {output_file_9p}")

if __name__ == "__main__":
    main()