#!/usr/bin/env python3
"""
Create detailed analysis report for D'Marco's sites with improved geocoding
Uses PositionStack API for better precision
"""

import pandas as pd
import requests
import time
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

class DMarcoDetailedReporter:
    """Create comprehensive reports for D'Marco's site analysis"""
    
    def __init__(self):
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
        # PositionStack API key (from memory)
        self.positionstack_api_key = "YOUR_API_KEY_HERE"  # Would need to be provided
    
    def improve_geocoding_with_positionstack(self, df):
        """Re-geocode failed addresses using PositionStack"""
        improved_count = 0
        
        for idx, row in df.iterrows():
            if not row['Geocoding_Success'] or pd.isna(row['Latitude']):
                address = f"{row['Address']}, {row['City']}, {row['County']} County, Texas"
                
                try:
                    # PositionStack API call
                    url = "http://api.positionstack.com/v1/forward"
                    params = {
                        'access_key': self.positionstack_api_key,
                        'query': address,
                        'country': 'US',
                        'region': 'Texas',
                        'limit': 1
                    }
                    
                    if self.positionstack_api_key != "YOUR_API_KEY_HERE":
                        response = requests.get(url, params=params)
                        if response.status_code == 200:
                            data = response.json()
                            if data.get('data') and len(data['data']) > 0:
                                result = data['data'][0]
                                df.loc[idx, 'Latitude'] = result['latitude']
                                df.loc[idx, 'Longitude'] = result['longitude']
                                df.loc[idx, 'Geocoded_Address'] = result['label']
                                df.loc[idx, 'Geocoding_Success'] = True
                                improved_count += 1
                                self.logger.info(f"Improved geocoding for {row['MailingName']}")
                        
                        time.sleep(0.1)  # Rate limiting
                    
                except Exception as e:
                    self.logger.warning(f"PositionStack failed for {row['MailingName']}: {e}")
        
        return improved_count
    
    def create_detailed_report(self, analysis_file):
        """Create comprehensive Excel report with multiple analysis views"""
        
        # Load the analysis data
        df = pd.read_excel(analysis_file, sheet_name='All_DMarco_Sites')
        
        # Improve geocoding precision (if API key available)
        # improved = self.improve_geocoding_with_positionstack(df)
        # self.logger.info(f"Improved geocoding for {improved} additional sites")
        
        # Create comprehensive report
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"DMarco_Detailed_Analysis_Report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # 1. EXECUTIVE SUMMARY
            summary_data = {
                'Analysis Metric': [
                    'Total D\'Marco Sites Analyzed',
                    'Successfully Geocoded Sites',
                    'QCT/DDA Eligible Sites (30% Boost)',
                    'Non-QCT/DDA Sites',
                    'Excellent 4% Opportunities',
                    'Good 4% Opportunities', 
                    'Fair/Risk 4% Opportunities',
                    'Fatal Flaw 9% Sites',
                    'Viable 9% Sites',
                    'Average Site Size (Acres)',
                    'Total QCT/DDA Acreage'
                ],
                'Count/Value': [
                    len(df),
                    df['Geocoding_Success'].sum(),
                    len(df[df['Basis_Boost_Eligible'] == True]),
                    len(df[df['Basis_Boost_Eligible'] != True]),
                    len(df[df['Economic_Viability_4pct'] == 'Excellent']),
                    len(df[df['Economic_Viability_4pct'] == 'Good']),
                    len(df[df['Economic_Viability_4pct'].isin(['Fair', 'Fair - Competition Risk'])]),
                    len(df[df['One_Mile_Fatal_9pct'] == True]),
                    len(df[df['Economic_Viability_9pct'] == 'Viable']),
                    round(df['Acres'].mean(), 2),
                    round(df[df['Basis_Boost_Eligible'] == True]['Acres'].sum(), 2)
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive_Summary', index=False)
            
            # 2. QCT/DDA PRIORITY SITES (Main focus)
            qct_dda_sites = df[df['Basis_Boost_Eligible'] == True].copy()
            if len(qct_dda_sites) > 0:
                # Sort by viability and competition
                viability_order = {'Excellent': 1, 'Good': 2, 'Fair': 3, 'Fair - Competition Risk': 4}
                qct_dda_sites['sort_order'] = qct_dda_sites['Economic_Viability_4pct'].map(viability_order)
                qct_dda_sites = qct_dda_sites.sort_values(['sort_order', 'One_Mile_Competition_Count', 'Acres'], ascending=[True, True, False])
                
                # Select key columns for the priority view
                priority_cols = [
                    'MailingName', 'Address', 'City', 'County', 'Region', 'Acres',
                    'QCT_DDA_Combined', 'Economic_Viability_4pct', 'Economic_Viability_9pct',
                    'One_Mile_Competition_Count', 'Regional_Construction_Cost_Multiplier',
                    'Development_Recommendation', 'Key_Advantages', 'Risk_Factors'
                ]
                qct_dda_priority = qct_dda_sites[priority_cols].copy()
                qct_dda_priority.to_excel(writer, sheet_name='QCT_DDA_Priority_Sites', index=False)
            
            # 3. REGIONAL ANALYSIS
            regional_analysis = df.groupby('Region').agg({
                'MailingName': 'count',
                'Basis_Boost_Eligible': 'sum',
                'Acres': ['mean', 'sum'],
                'Economic_Viability_4pct': lambda x: (x == 'Excellent').sum()
            }).round(2)
            
            regional_analysis.columns = ['Total_Sites', 'QCT_DDA_Sites', 'Avg_Acres', 'Total_Acres', 'Excellent_4pct']
            regional_analysis['QCT_DDA_Percentage'] = (regional_analysis['QCT_DDA_Sites'] / regional_analysis['Total_Sites'] * 100).round(1)
            regional_analysis = regional_analysis.sort_values('QCT_DDA_Sites', ascending=False)
            regional_analysis.to_excel(writer, sheet_name='Regional_Analysis')
            
            # 4. COMPETITION ANALYSIS 
            competition_data = []
            for _, row in qct_dda_sites.iterrows():
                competition_data.append({
                    'Property': row['MailingName'],
                    'City': row['City'],
                    'County': row['County'],
                    'Region': row['Region'],
                    'One_Mile_Competition': row['One_Mile_Competition_Count'],
                    'Two_Mile_Competition': row['Two_Mile_Competition_Count'],
                    '4pct_Risk_Level': row['One_Mile_Risk_4pct'],
                    '9pct_Fatal_Flaw': row['One_Mile_Fatal_9pct'],
                    'Large_County_Rules': row['Large_County_Rules_Apply']
                })
            
            if competition_data:
                comp_df = pd.DataFrame(competition_data)
                comp_df = comp_df.sort_values('One_Mile_Competition')
                comp_df.to_excel(writer, sheet_name='Competition_Analysis', index=False)
            
            # 5. ECONOMIC FACTORS
            economic_data = []
            for _, row in qct_dda_sites.iterrows():
                economic_data.append({
                    'Property': row['MailingName'],
                    'City': row['City'],
                    'Region': row['Region'],
                    'Acres': row['Acres'],
                    'Construction_Cost_Multiplier': row['Regional_Construction_Cost_Multiplier'],
                    'Est_Construction_Cost_SF': row['Estimated_Construction_Cost_SF'],
                    'QCT_DDA_Type': row['QCT_DDA_Combined'],
                    '4pct_Viability': row['Economic_Viability_4pct'],
                    'Development_Notes': row['Development_Recommendation']
                })
            
            if economic_data:
                econ_df = pd.DataFrame(economic_data)
                econ_df = econ_df.sort_values(['Construction_Cost_Multiplier', 'Acres'], ascending=[True, False])
                econ_df.to_excel(writer, sheet_name='Economic_Analysis', index=False)
            
            # 6. NON-QCT/DDA SITES (For reference)
            non_qct_dda = df[df['Basis_Boost_Eligible'] != True].copy()
            if len(non_qct_dda) > 0:
                non_qct_cols = ['MailingName', 'Address', 'City', 'County', 'Region', 'Acres', 'Development_Recommendation']
                non_qct_dda[non_qct_cols].to_excel(writer, sheet_name='Non_QCT_DDA_Sites', index=False)
            
            # 7. GEOCODING ISSUES
            geocoding_issues = df[df['Geocoding_Success'] != True].copy()
            if len(geocoding_issues) > 0:
                issue_cols = ['MailingName', 'Address', 'City', 'County', 'Geocoded_Address', 'Basis_Boost_Eligible']
                geocoding_issues[issue_cols].to_excel(writer, sheet_name='Geocoding_Issues', index=False)
            
            # 8. ALL SITES MASTER LIST
            df.to_excel(writer, sheet_name='All_Sites_Complete', index=False)
        
        # Format the Excel file
        self._format_excel_report(output_file)
        
        self.logger.info(f"✅ Detailed report created: {output_file}")
        return output_file
    
    def _format_excel_report(self, filename):
        """Apply formatting to the Excel report"""
        try:
            wb = load_workbook(filename)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header formatting
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            self.logger.warning(f"Formatting error: {e}")
    
    def print_summary_report(self, analysis_file):
        """Print a text summary to console"""
        df = pd.read_excel(analysis_file, sheet_name='All_DMarco_Sites')
        qct_dda_sites = df[df['Basis_Boost_Eligible'] == True]
        
        print("="*80)
        print("D'MARCO SITES - DETAILED ANALYSIS SUMMARY")
        print("="*80)
        
        print(f"\n📊 OVERALL STATISTICS:")
        print(f"   Total Sites Analyzed: {len(df)}")
        print(f"   QCT/DDA Eligible: {len(qct_dda_sites)} ({len(qct_dda_sites)/len(df)*100:.1f}%)")
        print(f"   Non-QCT/DDA: {len(df) - len(qct_dda_sites)} ({(len(df) - len(qct_dda_sites))/len(df)*100:.1f}%)")
        
        if len(qct_dda_sites) > 0:
            print(f"\n🎯 QCT/DDA SITE BREAKDOWN:")
            print(f"   QCT Only: {len(qct_dda_sites[qct_dda_sites['QCT_DDA_Combined'] == 'QCT'])}")
            print(f"   DDA Only: {len(qct_dda_sites[qct_dda_sites['QCT_DDA_Combined'] == 'DDA'])}")
            print(f"   QCT+DDA: {len(qct_dda_sites[qct_dda_sites['QCT_DDA_Combined'] == 'QCT+DDA'])}")
            
            print(f"\n💰 4% DEAL VIABILITY:")
            viability_counts = qct_dda_sites['Economic_Viability_4pct'].value_counts()
            for viability, count in viability_counts.items():
                print(f"   {viability}: {count} sites")
            
            print(f"\n🏗️ TOP OPPORTUNITIES (Excellent 4% Sites):")
            excellent_sites = qct_dda_sites[qct_dda_sites['Economic_Viability_4pct'] == 'Excellent']
            for _, site in excellent_sites.head(10).iterrows():
                print(f"   • {site['MailingName']} - {site['City']}, {site['County']} ({site['Acres']} acres)")
            
            print(f"\n🗺️ REGIONAL DISTRIBUTION:")
            regional = qct_dda_sites.groupby('Region').size().sort_values(ascending=False)
            for region, count in regional.items():
                print(f"   {region}: {count} sites")
        
        print(f"\n⚠️ GEOCODING ACCURACY:")
        successful = df['Geocoding_Success'].sum()
        print(f"   Precise geocoding: {successful}/{len(df)} ({successful/len(df)*100:.1f}%)")
        print(f"   City-center fallback: {len(df) - successful} sites")
        
        print("\n" + "="*80)

def main():
    reporter = DMarcoDetailedReporter()
    
    # Use the most recent analysis file
    analysis_file = "DMarco_Sites_Analysis_20250618_233604.xlsx"
    
    # Create detailed report
    detailed_report = reporter.create_detailed_report(analysis_file)
    
    # Print summary
    reporter.print_summary_report(analysis_file)
    
    print(f"\n📁 DETAILED REPORT SAVED: {detailed_report}")
    print("\nKey sheets in the detailed report:")
    print("  • Executive_Summary: High-level metrics")
    print("  • QCT_DDA_Priority_Sites: Focus development targets")
    print("  • Regional_Analysis: Breakdown by TDHCA region")
    print("  • Competition_Analysis: LIHTC competition details")
    print("  • Economic_Analysis: Construction costs and viability")
    print("  • Geocoding_Issues: Sites needing address verification")

if __name__ == "__main__":
    main()