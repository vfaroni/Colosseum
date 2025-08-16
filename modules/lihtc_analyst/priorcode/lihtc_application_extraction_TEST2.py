import os
import re
import json
import logging
import traceback
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_application_extraction.log"),
        logging.StreamHandler()
    ]
)

class LIHTCApplicationExtractor:
    """
    Extract basic application info from LIHTC applications and save as JSON files.
    """
    
    def __init__(self, max_files_per_type=5):
        """
        Initialize the extractor with file paths.
        
        Args:
            max_files_per_type: Maximum number of files to process for each type (4p and 9p)
        """
        # Input path
        self.input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
        
        # Output paths - now with application subfolder
        self.output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p/application'
        self.output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p/application'
        
        # Maximum number of files to process per type
        self.max_files_per_type = max_files_per_type
        
        # Create output directories if they don't exist
        Path(self.output_path_4p).mkdir(parents=True, exist_ok=True)
        Path(self.output_path_9p).mkdir(parents=True, exist_ok=True)
        
        # Statistics
        self.files_processed = 0
        self.files_skipped = 0
        self.extraction_errors = 0
    
    def extract_application_tab(self, file_path, tab_name):
        """
        Focused extraction of application information with emphasis on unit counts
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Application tab from {file_name}")
        
        # Parse metadata from filename
        app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
        year_match = re.search(r'(\d{4})_\d+pct', file_name)
        round_match = re.search(r'_R(\d+)_', file_name)
        ctcac_match = re.search(r'_(\d+-\d+)', file_name)
        
        # Initialize result with metadata
        result = {
            "metadata": {
                "file_name": file_name,
                "application_type": app_type,
                "application_year": year_match.group(1) if year_match else "",
                "application_round": round_match.group(1) if round_match else "",
                "ctcac_number": ctcac_match.group(1) if ctcac_match else "",
                "extraction_date": datetime.now().isoformat()
            },
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # FIRST APPROACH: Look for tabs specifically focused on project data or units
            target_sheets = []
            unit_focused_tabs = [
                "Project", "Application", "Units", "Unit Mix", "Project Information", 
                "General", "Development", "Project Summary"
            ]
            
            # Add the specified tab name first if it exists
            if tab_name in wb.sheetnames:
                target_sheets.append(wb[tab_name])
            
            # Add any tabs that match our unit-focused list
            for sheet_name in wb.sheetnames:
                if any(unit_tab.lower() in sheet_name.lower() for unit_tab in unit_focused_tabs):
                    if sheet_name != tab_name:  # Don't add it twice
                        target_sheets.append(wb[sheet_name])
            
            # If no matching tabs found, check all tabs
            if not target_sheets:
                target_sheets = [wb[sheet_name] for sheet_name in wb.sheetnames]
            
            # Define focused search terms for finding unit counts - be very specific
            unit_terms = {
                "total_units": [
                    "total units", "total residential units", "total number of units", 
                    "total project units", "project total units"
                ],
                "low_income_units": [
                    "low income units", "low-income units", "affordable units", 
                    "tax credit units", "lihtc units", "restricted units", 
                    "affordable housing units"
                ],
                "manager_units": [
                    "manager units", "manager's units", "manager unit", "manager's unit",
                    "employee units", "staff units", "non-revenue units"
                ]
            }
            
            # Flag to track if we found any unit information
            found_units = False
            
            # Process each potential sheet until we find unit information
            for sheet in target_sheets:
                logging.info(f"Searching for unit information in sheet: {sheet.title}")
                
                # STRATEGY 1: Look for common unit count cells
                for row in range(1, min(300, sheet.max_row + 1)):
                    for col in range(1, min(30, sheet.max_column + 1)):
                        try:
                            cell_value = sheet.cell(row=row, column=col).value
                            
                            # Skip empty cells
                            if not cell_value:
                                continue
                            
                            # Convert to string for comparison
                            if not isinstance(cell_value, str):
                                cell_str = str(cell_value).lower()
                            else:
                                cell_str = cell_value.lower()
                            
                            # Strip any colons and spaces
                            cell_str = cell_str.strip().rstrip(':')
                            
                            # Check if this cell contains any of our unit terms
                            for unit_type, search_terms in unit_terms.items():
                                if any(term in cell_str for term in search_terms):
                                    logging.debug(f"Found potential {unit_type} label at {sheet.title} {row},{col}: {cell_value}")
                                    
                                    # Look for value to the right
                                    for offset in range(1, 5):
                                        if col + offset <= sheet.max_column:
                                            value_cell = sheet.cell(row=row, column=col+offset).value
                                            
                                            # If we found a numeric value
                                            if value_cell is not None:
                                                try:
                                                    # Try to convert to number
                                                    if isinstance(value_cell, str):
                                                        # Clean up numeric strings
                                                        num_str = re.sub(r'[^\d.]', '', value_cell)
                                                        if num_str:
                                                            unit_count = int(float(num_str))
                                                        else:
                                                            continue
                                                    else:
                                                        unit_count = int(float(value_cell))
                                                    
                                                    # Only accept reasonable unit counts
                                                    if 0 < unit_count <= 1000:
                                                        result["project"][unit_type] = unit_count
                                                        logging.info(f"Found {unit_type}: {unit_count}")
                                                        found_units = True
                                                        break
                                                except:
                                                    # Not a valid number, continue searching
                                                    pass
                        except Exception as e:
                            logging.debug(f"Error processing cell at {sheet.title} {row},{col}: {e}")
                            continue
                
                # STRATEGY 2: Scan for unit info in unit mix tables
                # Many applications have a unit mix table showing counts by bedroom type
                if not found_units:
                    # Look for table headers that suggest a unit mix
                    for row in range(1, min(200, sheet.max_row + 1)):
                        bedroom_col = None
                        count_col = None
                        
                        # First, look for table headers
                        for col in range(1, min(20, sheet.max_column + 1)):
                            cell_value = sheet.cell(row=row, column=col).value
                            if not cell_value:
                                continue
                                
                            if isinstance(cell_value, str):
                                cell_str = cell_value.lower().strip()
                                
                                # Look for column headers for a unit mix table
                                if any(x in cell_str for x in ["bedroom", "unit type", "br", "bedrooms"]):
                                    bedroom_col = col
                                elif any(x in cell_str for x in ["count", "number of units", "# units", "units", "quantity"]):
                                    count_col = col
                        
                        # If we found potential table headers
                        if bedroom_col is not None and count_col is not None:
                            logging.debug(f"Found potential unit mix table at row {row}")
                            
                            # Look in the rows below for unit counts
                            total_units = 0
                            low_income_units = 0
                            manager_units = 0
                            
                            # Scan rows below for unit counts
                            for data_row in range(row+1, min(row+20, sheet.max_row + 1)):
                                bed_type = sheet.cell(row=data_row, column=bedroom_col).value
                                count_val = sheet.cell(row=data_row, column=count_col).value
                                
                                # Check if this is a unit row
                                if bed_type and count_val:
                                    try:
                                        # Parse bedroom type
                                        bed_str = str(bed_type).lower().strip()
                                        
                                        # Parse count
                                        if isinstance(count_val, str):
                                            count_str = re.sub(r'[^\d.]', '', count_val)
                                            if count_str:
                                                count_num = int(float(count_str))
                                            else:
                                                continue
                                        else:
                                            count_num = int(float(count_val))
                                        
                                        # If it looks like a manager unit
                                        if "manager" in bed_str or "staff" in bed_str:
                                            manager_units += count_num
                                        # Otherwise it's probably a normal unit
                                        else:
                                            # Assume all units are low income unless marked otherwise
                                            if "market" not in bed_str:
                                                low_income_units += count_num
                                            total_units += count_num
                                    except:
                                        pass
                            
                            # If we found any units in the table, record them
                            if total_units > 0:
                                result["project"]["total_units"] = total_units
                                found_units = True
                                logging.info(f"Found total_units from unit mix table: {total_units}")
                                
                                if low_income_units > 0:
                                    result["project"]["low_income_units"] = low_income_units
                                    logging.info(f"Found low_income_units from unit mix table: {low_income_units}")
                                
                                if manager_units > 0:
                                    result["project"]["manager_units"] = manager_units
                                    logging.info(f"Found manager_units from unit mix table: {manager_units}")
                                
                                # If we found units, break out of the row loop
                                break
                
                # STRATEGY 3: Look for common project information patterns
                # This handles project name, address, city, etc.
                for row in range(1, min(100, sheet.max_row + 1)):
                    for col in range(1, min(15, sheet.max_column + 1)):
                        try:
                            cell_value = sheet.cell(row=row, column=col).value
                            
                            # Skip empty cells
                            if not cell_value:
                                continue
                            
                            # Must be a string to check for labels
                            if not isinstance(cell_value, str):
                                continue
                            
                            cell_str = cell_value.lower().strip().rstrip(':')
                            
                            # Check for project name
                            if (not result["project"]["name"] and
                                any(x in cell_str for x in ["project name", "development name", "property name"])):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    name_val = sheet.cell(row=row, column=col+offset).value
                                    if name_val and isinstance(name_val, str) and name_val.strip():
                                        result["project"]["name"] = name_val.strip()
                                        logging.info(f"Found project name: {name_val.strip()}")
                                        break
                            
                            # Check for address
                            elif (not result["project"]["address"] and
                                  any(x in cell_str for x in ["project address", "property address", "site address", "address"])):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    addr_val = sheet.cell(row=row, column=col+offset).value
                                    if addr_val and isinstance(addr_val, str) and addr_val.strip():
                                        result["project"]["address"] = addr_val.strip()
                                        logging.info(f"Found address: {addr_val.strip()}")
                                        break
                            
                            # Check for city
                            elif (not result["project"]["city"] and cell_str == "city"):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    city_val = sheet.cell(row=row, column=col+offset).value
                                    if city_val and isinstance(city_val, str) and city_val.strip():
                                        result["project"]["city"] = city_val.strip()
                                        logging.info(f"Found city: {city_val.strip()}")
                                        break
                            
                            # Check for county
                            elif (not result["project"]["county"] and cell_str == "county"):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    county_val = sheet.cell(row=row, column=col+offset).value
                                    if county_val and isinstance(county_val, str) and county_val.strip():
                                        result["project"]["county"] = county_val.strip()
                                        logging.info(f"Found county: {county_val.strip()}")
                                        break
                            
                            # Check for zip code
                            elif (not result["project"]["zip_code"] and 
                                  any(x in cell_str for x in ["zip code", "zip", "postal code"])):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    zip_val = sheet.cell(row=row, column=col+offset).value
                                    if zip_val:
                                        # Convert to string and clean
                                        zip_str = str(zip_val).strip()
                                        result["project"]["zip_code"] = zip_str
                                        logging.info(f"Found zip code: {zip_str}")
                                        break
                            
                            # Check for census tract
                            elif (not result["project"]["census_tract"] and
                                  any(x in cell_str for x in ["census tract", "tract", "census"])):
                                # Look to the right for the value
                                for offset in range(1, 4):
                                    tract_val = sheet.cell(row=row, column=col+offset).value
                                    if tract_val:
                                        # Convert to string and clean
                                        tract_str = str(tract_val).strip()
                                        result["project"]["census_tract"] = tract_str
                                        logging.info(f"Found census tract: {tract_str}")
                                        break
                        except Exception as e:
                            logging.debug(f"Error processing cell at {sheet.title} {row},{col}: {e}")
                            continue
                
                # If we found units for this sheet, no need to check other sheets
                if found_units:
                    break
            
            # If we still don't have a project name, try the filename
            if not result["project"]["name"]:
                # Try to extract from filename pattern
                ctcac_num = result["metadata"]["ctcac_number"]
                if ctcac_num:
                    result["project"]["name"] = f"Project {ctcac_num}"
            
            # Post-processing to fill in missing values
            
            # Validate unit counts - total should be at least low_income + manager
            low_income = result["project"]["low_income_units"]
            manager = result["project"]["manager_units"]
            total = result["project"]["total_units"]
            
            # If we have low_income and manager but no total, calculate it
            if low_income > 0 and manager > 0 and total == 0:
                result["project"]["total_units"] = low_income + manager
                logging.info(f"Calculated total_units: {low_income + manager}")
            
            # If we have total and low_income but no manager, estimate the manager count
            if total > 0 and low_income > 0 and manager == 0:
                estimated_manager = total - low_income
                if 0 < estimated_manager <= 3:  # Reasonable manager unit count
                    result["project"]["manager_units"] = estimated_manager
                    logging.info(f"Estimated manager_units: {estimated_manager}")
            
            # If we have total but no low_income, assume most units are low-income (typical for LIHTC)
            if total > 0 and low_income == 0 and manager == 0:
                # Assume 1-2 manager units based on project size
                estimated_manager = 1 if total < 50 else 2
                estimated_low_income = total - estimated_manager
                
                if estimated_low_income > 0:
                    result["project"]["low_income_units"] = estimated_low_income
                    result["project"]["manager_units"] = estimated_manager
                    logging.info(f"Estimated low_income_units: {estimated_low_income}")
                    logging.info(f"Estimated manager_units: {estimated_manager}")
            
            # Log the final results
            logging.info(f"Extracted project: {result['project']['name']}, Units: {result['project']['total_units']}")
            
            return result
        
        except Exception as e:
            logging.error(f"Error extracting Application tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def process_single_file(self, file_path):
        """
        Process a single LIHTC application file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            dict: Processed application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Processing file: {file_name}")
        
        try:
            # Determine application type
            app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
            
            if app_type == "unknown":
                logging.warning(f"Unknown application type for file: {file_name}, skipping")
                self.files_skipped += 1
                return None
            
            # Parse filename components for output naming
            year_match = re.search(r'(\d{4})_\d+pct', file_name)
            round_match = re.search(r'_R(\d+)_', file_name)
            ctcac_match = re.search(r'_(\d+-\d+)', file_name)
            
            year = year_match.group(1) if year_match else ""
            round_num = round_match.group(1) if round_match else ""
            ctcac_num = ctcac_match.group(1) if ctcac_match else ""
            
            # Extract application tab data
            application_data = self.extract_application_tab(file_path, "Application")
            
            # Use the original filename format for JSON output - maintain all the important identifiers
            # Format: YYYY_[4|9]pct_R#_##-###.json
            if year and round_num and ctcac_num:
                output_filename = f"{year}_{app_type}ct_R{round_num}_{ctcac_num}.json"
            else:
                # Fallback if we can't parse the filename components
                output_filename = file_name.replace(".xlsx", ".json")
            
            # Determine output directory
            output_dir = self.output_path_4p if app_type == "4p" else self.output_path_9p
            output_path = os.path.join(output_dir, output_filename)
            
            # Save to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(application_data, f, indent=2)
            
            logging.info(f"Saved extracted data to {output_path}")
            self.files_processed += 1
            
            return application_data
            
        except Exception as e:
            logging.error(f"Error processing file {file_name}: {e}")
            logging.error(traceback.format_exc())
            self.extraction_errors += 1
            return None
    
    def process_sample_files(self):
        """
        Process a sample of LIHTC application files in the input directory
        """
        logging.info(f"Starting processing of sample LIHTC application files in {self.input_path}")
        
        # Get list of all Excel files
        excel_files_4p = []
        excel_files_9p = []
        
        for root, dirs, files in os.walk(self.input_path):
            for file in files:
                if file.endswith('.xlsx'):
                    if '4pct' in file:
                        excel_files_4p.append(os.path.join(root, file))
                    elif '9pct' in file:
                        excel_files_9p.append(os.path.join(root, file))
        
        # Limit to maximum number of files per type
        excel_files_4p = excel_files_4p[:self.max_files_per_type]
        excel_files_9p = excel_files_9p[:self.max_files_per_type]
        
        logging.info(f"Selected {len(excel_files_4p)} 4% applications and {len(excel_files_9p)} 9% applications for processing")
        
        # Process 4% files
        print(f"\nProcessing {len(excel_files_4p)} 4% applications:")
        for i, file_path in enumerate(excel_files_4p):
            file_name = os.path.basename(file_path)
            print(f"  {i+1}. {file_name}")
            logging.info(f"Processing 4% file {i+1} of {len(excel_files_4p)}: {file_name}")
            self.process_single_file(file_path)
        
        # Process 9% files
        print(f"\nProcessing {len(excel_files_9p)} 9% applications:")
        for i, file_path in enumerate(excel_files_9p):
            file_name = os.path.basename(file_path)
            print(f"  {i+1}. {file_name}")
            logging.info(f"Processing 9% file {i+1} of {len(excel_files_9p)}: {file_name}")
            self.process_single_file(file_path)
        
        # Log summary
        logging.info(f"Completed processing sample files")
        logging.info(f"Successfully processed: {self.files_processed}")
        logging.info(f"Skipped: {self.files_skipped}")
        logging.info(f"Errors: {self.extraction_errors}")

def main():
    """Main entry point for the script."""
    # Ask the user how many files to process
    try:
        num_files = int(input("Enter the number of files to process for each type (4% and 9%): "))
        if num_files <= 0:
            print("Number must be positive. Using default of 5 files.")
            num_files = 5
    except ValueError:
        print("Invalid input. Using default of 5 files.")
        num_files = 5
    
    print(f"\nStarting LIHTC Application Tab Extractor...")
    print(f"Will process up to {num_files} files of each type (4% and 9%)")
    print(f"Looking for Excel files in folder path: /Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data")
    
    extractor = LIHTCApplicationExtractor(max_files_per_type=num_files)
    extractor.process_sample_files()
    
    print(f"\nFinished processing.")
    print(f"Successfully processed: {extractor.files_processed} files")
    print(f"Skipped: {extractor.files_skipped} files")
    print(f"Errors: {extractor.extraction_errors} files")
    print("\nSee lihtc_application_extraction.log for detailed information.")

if __name__ == "__main__":
    main()