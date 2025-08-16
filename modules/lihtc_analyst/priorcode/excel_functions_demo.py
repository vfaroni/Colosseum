#!/usr/bin/env python3
"""
Excel Functions Demo
Simple Python script to demonstrate Excel-like functions using pandas and openpyxl
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_sample_data():
    """Create sample data for demonstration"""
    data = {
        'Name': ['John', 'Jane', 'Bob', 'Alice', 'Charlie'],
        'Age': [25, 30, 35, 28, 32],
        'Salary': [50000, 60000, 70000, 55000, 65000],
        'Department': ['IT', 'HR', 'IT', 'Finance', 'IT']
    }
    return pd.DataFrame(data)

def excel_like_functions():
    """Demonstrate Excel-like functions"""
    df = create_sample_data()
    print("Sample Data:")
    print(df)
    print("\n" + "="*50 + "\n")
    
    # SUM equivalent
    total_salary = df['Salary'].sum()
    print(f"SUM of Salaries: ${total_salary:,}")
    
    # AVERAGE equivalent
    avg_salary = df['Salary'].mean()
    print(f"AVERAGE Salary: ${avg_salary:,.2f}")
    
    # COUNT equivalent
    count_employees = df['Name'].count()
    print(f"COUNT of Employees: {count_employees}")
    
    # MAX/MIN equivalent
    max_salary = df['Salary'].max()
    min_salary = df['Salary'].min()
    print(f"MAX Salary: ${max_salary:,}")
    print(f"MIN Salary: ${min_salary:,}")
    
    # COUNTIF equivalent (count IT employees)
    it_count = df[df['Department'] == 'IT'].shape[0]
    print(f"COUNTIF (IT Department): {it_count}")
    
    # SUMIF equivalent (sum IT salaries)
    it_salary_sum = df[df['Department'] == 'IT']['Salary'].sum()
    print(f"SUMIF (IT Department): ${it_salary_sum:,}")
    
    # VLOOKUP equivalent
    lookup_salary = df[df['Name'] == 'Jane']['Salary'].iloc[0]
    print(f"VLOOKUP (Jane's Salary): ${lookup_salary:,}")
    
    print("\n" + "="*50 + "\n")
    
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Data"
    
    # Write data to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Add formulas
    ws['F1'] = 'Total Salary'
    ws['F2'] = '=SUM(C2:C6)'
    ws['F3'] = 'Average Salary'
    ws['F4'] = '=AVERAGE(C2:C6)'
    
    # Save file
    filename = 'excel_functions_demo.xlsx'
    wb.save(filename)
    print(f"Excel file saved as: {filename}")
    
    return df

if __name__ == "__main__":
    excel_like_functions()