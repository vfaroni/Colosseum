#!/usr/bin/env python3
"""
D'Marco + CoStar Integrated Analysis System
Merges 65 D'Marco sites with 406 CoStar properties for comprehensive LIHTC analysis

Scoring Criteria:
1. QCT/DDA Status (30% basis boost) - Binary qualification
2. AMI Rent Levels by County/MSA - Revenue potential
3. FEMA Flood Risk - Construction cost impacts
4. TDHCA Regional Analysis - Cost multipliers
5. Competition Analysis - One Mile/Two Mile rules
6. Combined Scoring - Separate for 4% and 9% deals

Author: TDHCA Analysis System
Date: 2025-06-20
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import geopandas as gpd
from shapely.geometry import Point
from geopy.distance import geodesic
import requests
import json
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class IntegratedTDHCAAnalyzer:
    """Comprehensive analyzer combining D'Marco and CoStar properties"""
    
    def __init__(self):
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
        # Key data paths
        self.base_path = Path("/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects")
        self.code_path = self.base_path / "CTCAC_RAG/code"
        self.data_path = self.base_path / "Lvg_Bond_Execution/Data Sets"
        
        # Results storage
        self.merged_data = None
        self.analysis_results = None
        
    def load_dmarco_sites(self):
        """Load D'Marco's 65 sites from the final PositionStack analysis"""
        dmarco_file = self.code_path / "DMarco_Sites_Final_PositionStack_20250618_235606.xlsx"
        
        if not dmarco_file.exists():
            # Try alternative path
            dmarco_file = self.code_path / "analyze_dmarco_sites.py"
            self.logger.info("Loading D'Marco sites from analyzer output...")
            # Run the analyzer to get fresh data
            import subprocess
            subprocess.run(["python3", str(dmarco_file)], cwd=str(self.code_path))
            
        # Load the Excel file
        dmarco_df = pd.read_excel(dmarco_file, sheet_name='All_Sites_Final')
        self.logger.info(f"Loaded {len(dmarco_df)} D'Marco sites")
        
        # Standardize columns
        dmarco_df['Source'] = 'D\'Marco'
        return dmarco_df
        
    def load_costar_properties(self):
        """Load CoStar 406 properties from enhanced analysis"""
        costar_file = self.code_path / "FIXED_POSITIONSTACK_RESULTS/FIXED_Enhanced_CoStar_4pct_Bond_20250601_235030.xlsx"
        
        if not costar_file.exists():
            # Try alternative path
            costar_file = self.code_path / "CoStar_Land_Analysis_With_Counties_20250617_223737.xlsx"
            
        costar_df = pd.read_excel(costar_file)
        self.logger.info(f"Loaded {len(costar_df)} CoStar properties")
        
        # Standardize columns
        costar_df['Source'] = 'CoStar'
        return costar_df
        
    def merge_datasets(self, dmarco_df, costar_df):
        """Merge D'Marco and CoStar datasets with standardized fields"""
        # Standardize column names for both datasets
        column_mapping = {
            # D'Marco columns
            'MailingName': 'Property_Name',
            'Address': 'Street_Address',
            'City': 'City',
            'County': 'County',
            'Zip': 'Zip_Code',
            'Acres': 'Land_Acres',
            'Region': 'TDHCA_Region',
            'Latitude': 'Latitude',
            'Longitude': 'Longitude',
            'QCT_Status': 'Is_QCT',
            'DDA_Status': 'Is_DDA',
            'QCT_DDA_Eligible': 'QCT_DDA_Eligible',
            
            # CoStar columns (if different)
            'property_name': 'Property_Name',
            'address': 'Street_Address',
            'city': 'City',
            'county': 'County',
            'zip': 'Zip_Code',
            'acres': 'Land_Acres',
            'lat': 'Latitude',
            'lng': 'Longitude',
            'lon': 'Longitude',
        }
        
        # Rename columns
        dmarco_df = dmarco_df.rename(columns=column_mapping)
        costar_df = costar_df.rename(columns=column_mapping)
        
        # Select common columns
        common_cols = [
            'Property_Name', 'Street_Address', 'City', 'County', 'Zip_Code',
            'Land_Acres', 'Latitude', 'Longitude', 'Source', 'TDHCA_Region'
        ]
        
        # Add columns that might not exist
        for col in common_cols:
            if col not in dmarco_df.columns:
                dmarco_df[col] = None
            if col not in costar_df.columns:
                costar_df[col] = None
                
        # Merge datasets
        merged_df = pd.concat([
            dmarco_df[common_cols + [c for c in dmarco_df.columns if c not in common_cols]],
            costar_df[common_cols + [c for c in costar_df.columns if c not in common_cols]]
        ], ignore_index=True)
        
        self.logger.info(f"Merged dataset contains {len(merged_df)} total properties")
        return merged_df
        
    def load_qct_dda_data(self):
        """Load QCT/DDA shapefiles for analysis"""
        qct_path = self.data_path / "HUD DDA QCT/QUALIFIED_CENSUS_TRACTS_7341711606021821459.gpkg"
        dda_path = self.data_path / "HUD DDA QCT/Difficult_Development_Areas_-4200740390724245794.gpkg"
        
        self.qct_gdf = gpd.read_file(qct_path)
        self.dda_gdf = gpd.read_file(dda_path)
        
        # Convert to WGS84 for lat/lng comparison
        self.qct_gdf = self.qct_gdf.to_crs(epsg=4326)
        self.dda_gdf = self.dda_gdf.to_crs(epsg=4326)
        
        self.logger.info(f"Loaded {len(self.qct_gdf)} QCT and {len(self.dda_gdf)} DDA areas")
        
    def check_qct_dda_status(self, lat, lon):
        """Check if coordinates are in QCT or DDA"""
        point = Point(lon, lat)
        
        # Check QCT
        is_qct = any(self.qct_gdf.geometry.contains(point))
        
        # Check DDA
        is_dda = any(self.dda_gdf.geometry.contains(point))
        
        # Eligible if either QCT OR DDA (30% basis boost)
        eligible = is_qct or is_dda
        
        return {
            'Is_QCT': is_qct,
            'Is_DDA': is_dda,
            'QCT_DDA_Eligible': eligible,
            'Basis_Boost': 1.30 if eligible else 1.00
        }
        
    def load_ami_data(self):
        """Load HUD AMI rent data for Texas counties"""
        ami_file = self.data_path / "HUD AMI FMR/HUD2025_AMI_Rent_Data_Static.xlsx"
        
        # Read the Excel file
        ami_df = pd.read_excel(ami_file, sheet_name='Texas')
        
        # The columns are structured with AMI levels for each bedroom size
        # Looking for the pattern in column names
        self.ami_lookup = {}
        
        for _, row in ami_df.iterrows():
            county = str(row.iloc[1]).strip()  # County name is typically in column B
            
            # Extract all AMI rent levels
            # The last 20 columns contain rent data: Studio-4BR for 50%, 60%, 70%, 80% AMI
            rent_data = {}
            
            # Column mapping based on typical HUD format
            # Assuming columns follow pattern: Studio_50, 1BR_50, 2BR_50, 3BR_50, 4BR_50, Studio_60, etc.
            col_names = ami_df.columns.tolist()
            
            # Find the AMI 100% column (4-person household)
            ami_100_col = None
            for idx, col in enumerate(col_names):
                if '100%' in str(col) or 'AMI' in str(col) and '4' in str(col):
                    ami_100_col = idx
                    rent_data['AMI_100_4person'] = row.iloc[idx]
                    break
            
            # Get rent columns (last 20 columns)
            rent_cols_start = len(col_names) - 20
            
            # 50% AMI rents (columns 1-5 from the last 20)
            rent_data['Studio_50pct'] = row.iloc[rent_cols_start]
            rent_data['1BR_50pct'] = row.iloc[rent_cols_start + 1]
            rent_data['2BR_50pct'] = row.iloc[rent_cols_start + 2]
            rent_data['3BR_50pct'] = row.iloc[rent_cols_start + 3]
            rent_data['4BR_50pct'] = row.iloc[rent_cols_start + 4]
            
            # 60% AMI rents (columns 6-10 from the last 20)
            rent_data['Studio_60pct'] = row.iloc[rent_cols_start + 5]
            rent_data['1BR_60pct'] = row.iloc[rent_cols_start + 6]
            rent_data['2BR_60pct'] = row.iloc[rent_cols_start + 7]
            rent_data['3BR_60pct'] = row.iloc[rent_cols_start + 8]
            rent_data['4BR_60pct'] = row.iloc[rent_cols_start + 9]
            
            # 70% AMI rents (columns 11-15 from the last 20)
            rent_data['Studio_70pct'] = row.iloc[rent_cols_start + 10]
            rent_data['1BR_70pct'] = row.iloc[rent_cols_start + 11]
            rent_data['2BR_70pct'] = row.iloc[rent_cols_start + 12]
            rent_data['3BR_70pct'] = row.iloc[rent_cols_start + 13]
            rent_data['4BR_70pct'] = row.iloc[rent_cols_start + 14]
            
            # 80% AMI rents (columns 16-20 from the last 20)
            rent_data['Studio_80pct'] = row.iloc[rent_cols_start + 15]
            rent_data['1BR_80pct'] = row.iloc[rent_cols_start + 16]
            rent_data['2BR_80pct'] = row.iloc[rent_cols_start + 17]
            rent_data['3BR_80pct'] = row.iloc[rent_cols_start + 18]
            rent_data['4BR_80pct'] = row.iloc[rent_cols_start + 19]
            
            # Add MSA name if available
            rent_data['MSA_Name'] = row.iloc[2] if len(row) > 2 else 'Non-Metro'
            
            self.ami_lookup[county] = rent_data
            
        self.logger.info(f"Loaded AMI data for {len(self.ami_lookup)} Texas counties with all bedroom sizes and AMI levels")
        
    def get_fema_flood_risk(self, lat, lon):
        """Query FEMA flood zone for given coordinates"""
        try:
            # FEMA ArcGIS REST API endpoint
            url = "https://hazards.fema.gov/gis/nfhl/rest/services/public/NFHL/MapServer/28/query"
            
            params = {
                'geometry': f'{lon},{lat}',
                'geometryType': 'esriGeometryPoint',
                'inSR': '4326',
                'spatialRel': 'esriSpatialRelIntersects',
                'outFields': 'FLD_ZONE,ZONE_SUBTY',
                'returnGeometry': 'false',
                'f': 'json'
            }
            
            response = requests.get(url, params=params, timeout=10)
            data = response.json()
            
            if 'features' in data and len(data['features']) > 0:
                zone = data['features'][0]['attributes'].get('FLD_ZONE', 'X')
                subtype = data['features'][0]['attributes'].get('ZONE_SUBTY', '')
                
                # Determine construction cost impact
                if zone in ['VE', 'V']:
                    cost_impact = 1.30  # 30% increase
                elif zone in ['AE', 'A']:
                    cost_impact = 1.20  # 20% increase
                elif zone == 'X':
                    cost_impact = 1.00  # No increase
                else:
                    cost_impact = 1.10  # 10% increase for other zones
                    
                return {
                    'FEMA_Zone': zone,
                    'FEMA_Subtype': subtype,
                    'Flood_Cost_Impact': cost_impact
                }
            else:
                return {
                    'FEMA_Zone': 'X',
                    'FEMA_Subtype': '',
                    'Flood_Cost_Impact': 1.00
                }
                
        except Exception as e:
            self.logger.warning(f"FEMA API error: {e}")
            return {
                'FEMA_Zone': 'Unknown',
                'FEMA_Subtype': '',
                'Flood_Cost_Impact': 1.05
            }
            
    def load_tdhca_projects(self):
        """Load TDHCA project list for competition analysis"""
        project_file = self.data_path / "State Specific/TX/Project_List/TX_TDHCA_Project_List_05252025.xlsx"
        
        self.tdhca_projects = pd.read_excel(project_file)
        
        # Filter for relevant years (2021-2024 for competition)
        self.tdhca_projects['Year'] = pd.to_numeric(self.tdhca_projects['Year'], errors='coerce')
        self.tdhca_projects = self.tdhca_projects[
            self.tdhca_projects['Year'].between(2021, 2024)
        ]
        
        self.logger.info(f"Loaded {len(self.tdhca_projects)} TDHCA projects for competition analysis")
        
    def check_competition(self, lat, lon, county):
        """Check One Mile Three Year and Two Mile Same Year rules"""
        results = {
            'One_Mile_Count': 0,
            'One_Mile_Projects': [],
            'Two_Mile_Count': 0,
            'Two_Mile_Projects': [],
            'Competition_Risk_4pct': 'Low',
            'Competition_Risk_9pct': 'Low'
        }
        
        site_point = (lat, lon)
        
        # Large counties for Two Mile rule (9% only)
        large_counties = ['Harris', 'Dallas', 'Tarrant', 'Bexar', 'Travis', 'Collin', 'Denton']
        is_large_county = county in large_counties
        
        for _, project in self.tdhca_projects.iterrows():
            if pd.isna(project['Latitude']) or pd.isna(project['Longitude']):
                continue
                
            project_point = (project['Latitude'], project['Longitude'])
            distance = geodesic(site_point, project_point).miles
            
            # One Mile Three Year Rule
            if distance <= 1.0 and project['Year'] >= 2022:
                results['One_Mile_Count'] += 1
                results['One_Mile_Projects'].append({
                    'Name': project['Project_Name'],
                    'Year': project['Year'],
                    'Distance': round(distance, 2)
                })
                
            # Two Mile Same Year Rule (9% in large counties)
            if is_large_county and distance <= 2.0 and project['Year'] == 2024:
                results['Two_Mile_Count'] += 1
                results['Two_Mile_Projects'].append({
                    'Name': project['Project_Name'],
                    'Year': project['Year'],
                    'Distance': round(distance, 2)
                })
                
        # Determine risk levels
        if results['One_Mile_Count'] > 0:
            results['Competition_Risk_4pct'] = 'Medium'  # Risk indicator for 4%
            results['Competition_Risk_9pct'] = 'Fatal'   # Fatal flaw for 9%
            
        if results['Two_Mile_Count'] > 0:
            results['Competition_Risk_9pct'] = 'Fatal'   # Fatal flaw for 9% in large counties
            
        return results
        
    def calculate_scores(self, row):
        """Calculate comprehensive scores for 4% and 9% deals"""
        scores = {}
        
        # Base construction cost per SF
        base_cost = 150
        
        # Regional cost multiplier
        region_multipliers = {
            'Region 1': 0.90, 'Region 2': 0.95, 'Region 3': 1.15,
            'Region 4': 0.98, 'Region 5': 1.00, 'Region 6': 1.18,
            'Region 7': 1.20, 'Region 8': 1.05, 'Region 9': 1.10,
            'Region 10': 1.08, 'Region 11': 1.12, 'Region 12': 0.92,
            'Region 13': 0.95
        }
        
        region = row.get('TDHCA_Region', 'Region 5')
        regional_multiplier = region_multipliers.get(region, 1.00)
        
        # Total construction cost impact
        flood_impact = row.get('Flood_Cost_Impact', 1.00)
        total_cost = base_cost * regional_multiplier * flood_impact
        
        # AMI rent revenue potential (2BR 60% as baseline)
        ami_rent = row.get('AMI_2BR_60pct', 1200)
        
        # Economic score (rent to cost ratio)
        economic_score = (ami_rent * 12) / (total_cost * 850)  # 850 SF for 2BR
        economic_score = min(economic_score * 100, 100)  # Scale to 0-100
        
        # QCT/DDA bonus
        qct_dda_score = 30 if row.get('QCT_DDA_Eligible', False) else 0
        
        # Competition penalty
        competition_penalty_4pct = 0
        competition_penalty_9pct = 0
        
        if row.get('Competition_Risk_4pct') == 'Medium':
            competition_penalty_4pct = -10
        if row.get('Competition_Risk_9pct') == 'Fatal':
            competition_penalty_9pct = -100  # Fatal flaw
            
        # Land efficiency (smaller is better for urban, larger for rural)
        acres = row.get('Land_Acres', 10)
        if region in ['Region 3', 'Region 6', 'Region 7']:  # Urban regions
            land_score = max(0, 20 - acres * 2)
        else:  # Rural regions
            land_score = min(acres * 0.5, 20)
            
        # 4% Score (focus on economics and feasibility)
        scores['Score_4pct'] = max(0, 
            economic_score * 0.5 +  # 50% weight
            qct_dda_score * 0.3 +   # 30% weight
            land_score * 0.2 +      # 20% weight
            competition_penalty_4pct
        )
        
        # 9% Score (focus on competitive points)
        scores['Score_9pct'] = max(0,
            economic_score * 0.3 +  # 30% weight
            qct_dda_score * 0.4 +   # 40% weight (more important for 9%)
            land_score * 0.3 +      # 30% weight
            competition_penalty_9pct
        )
        
        # Overall recommendation
        if scores['Score_9pct'] < 0:
            scores['Recommendation'] = 'Fatal Flaw - Avoid'
        elif scores['Score_4pct'] >= 70:
            scores['Recommendation'] = 'Strong 4% Candidate'
        elif scores['Score_9pct'] >= 70 and competition_penalty_9pct == 0:
            scores['Recommendation'] = 'Strong 9% Candidate'
        elif scores['Score_4pct'] >= 50:
            scores['Recommendation'] = 'Moderate 4% Potential'
        else:
            scores['Recommendation'] = 'Limited Potential'
            
        scores['Construction_Cost_PSF'] = round(total_cost, 2)
        scores['Annual_Revenue_Potential'] = round(ami_rent * 12 * 100, 0)  # 100 units
        
        return scores
        
    def analyze_all_sites(self):
        """Run comprehensive analysis on all sites"""
        # Load all data sources
        self.load_qct_dda_data()
        self.load_ami_data()
        self.load_tdhca_projects()
        
        # Process each site
        results = []
        total_sites = len(self.merged_data)
        
        for idx, row in self.merged_data.iterrows():
            if idx % 50 == 0:
                self.logger.info(f"Processing site {idx+1}/{total_sites}")
                
            site_results = row.to_dict()
            
            # Skip if no coordinates
            if pd.isna(row.get('Latitude')) or pd.isna(row.get('Longitude')):
                self.logger.warning(f"Skipping {row.get('Property_Name')} - no coordinates")
                continue
                
            lat = row['Latitude']
            lon = row['Longitude']
            
            # QCT/DDA analysis
            qct_dda = self.check_qct_dda_status(lat, lon)
            site_results.update(qct_dda)
            
            # AMI rent lookup
            county = row.get('County', '')
            if county in self.ami_lookup:
                site_results.update(self.ami_lookup[county])
            else:
                site_results.update({
                    'AMI_2BR_60pct': 1200,
                    'AMI_2BR_50pct': 1000,
                    'AMI_2BR_30pct': 600,
                    'MSA_Name': 'Non-Metro'
                })
                
            # FEMA flood risk
            fema_data = self.get_fema_flood_risk(lat, lon)
            site_results.update(fema_data)
            
            # Competition analysis
            competition = self.check_competition(lat, lon, county)
            site_results.update(competition)
            
            # Calculate scores
            scores = self.calculate_scores(site_results)
            site_results.update(scores)
            
            results.append(site_results)
            
        self.analysis_results = pd.DataFrame(results)
        self.logger.info(f"Analysis complete for {len(self.analysis_results)} sites")
        
    def generate_report(self):
        """Generate comprehensive Excel report for D'Marco"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.code_path / f"DMarco_CoStar_Integrated_Analysis_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Executive Summary
            summary_data = {
                'Metric': [
                    'Total Properties Analyzed',
                    'D\'Marco Sites',
                    'CoStar Properties',
                    'QCT/DDA Eligible (30% Boost)',
                    'Strong 4% Candidates (Score 70+)',
                    'Strong 9% Candidates (Score 70+)',
                    'Fatal Flaw Sites',
                    'Average Construction Cost/SF',
                    'Average AMI 2BR Rent'
                ],
                'Value': [
                    len(self.analysis_results),
                    len(self.analysis_results[self.analysis_results['Source'] == 'D\'Marco']),
                    len(self.analysis_results[self.analysis_results['Source'] == 'CoStar']),
                    len(self.analysis_results[self.analysis_results['QCT_DDA_Eligible'] == True]),
                    len(self.analysis_results[self.analysis_results['Score_4pct'] >= 70]),
                    len(self.analysis_results[self.analysis_results['Score_9pct'] >= 70]),
                    len(self.analysis_results[self.analysis_results['Competition_Risk_9pct'] == 'Fatal']),
                    f"${self.analysis_results['Construction_Cost_PSF'].mean():.2f}",
                    f"${self.analysis_results['AMI_2BR_60pct'].mean():.0f}"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
            
            # Top 4% Opportunities
            top_4pct = self.analysis_results.nlargest(50, 'Score_4pct')
            top_4pct.to_excel(writer, sheet_name='Top_4pct_Sites', index=False)
            
            # Top 9% Opportunities (excluding fatal flaws)
            viable_9pct = self.analysis_results[
                self.analysis_results['Competition_Risk_9pct'] != 'Fatal'
            ]
            top_9pct = viable_9pct.nlargest(50, 'Score_9pct')
            top_9pct.to_excel(writer, sheet_name='Top_9pct_Sites', index=False)
            
            # QCT/DDA Sites Only
            qct_dda_sites = self.analysis_results[
                self.analysis_results['QCT_DDA_Eligible'] == True
            ].sort_values('Score_4pct', ascending=False)
            qct_dda_sites.to_excel(writer, sheet_name='QCT_DDA_Sites', index=False)
            
            # All Sites
            self.analysis_results.to_excel(writer, sheet_name='All_Sites', index=False)
            
            # Regional Analysis
            regional_summary = self.analysis_results.groupby('TDHCA_Region').agg({
                'Property_Name': 'count',
                'QCT_DDA_Eligible': 'sum',
                'Score_4pct': 'mean',
                'Score_9pct': 'mean',
                'Construction_Cost_PSF': 'mean',
                'AMI_2BR_60pct': 'mean'
            }).round(2)
            regional_summary.columns = [
                'Total_Sites', 'QCT_DDA_Count', 'Avg_4pct_Score',
                'Avg_9pct_Score', 'Avg_Construction_Cost', 'Avg_AMI_Rent'
            ]
            regional_summary.to_excel(writer, sheet_name='Regional_Analysis')
            
            # Format the workbook
            workbook = writer.book
            
            # Format Executive Summary
            ws = workbook['Executive_Summary']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
                if 'Strong' in str(row[0].value) or 'QCT/DDA' in str(row[0].value):
                    row[0].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif 'Fatal' in str(row[0].value):
                    row[0].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    
        self.logger.info(f"Report generated: {output_file}")
        return output_file
        
    def run_analysis(self):
        """Execute the complete analysis pipeline"""
        self.logger.info("Starting D'Marco + CoStar integrated analysis...")
        
        # Load datasets
        dmarco_df = self.load_dmarco_sites()
        costar_df = self.load_costar_properties()
        
        # Merge datasets
        self.merged_data = self.merge_datasets(dmarco_df, costar_df)
        
        # Run comprehensive analysis
        self.analyze_all_sites()
        
        # Generate report
        report_file = self.generate_report()
        
        # Print summary
        print("\n" + "="*80)
        print("ANALYSIS COMPLETE - D'MARCO + COSTAR INTEGRATED SYSTEM")
        print("="*80)
        print(f"\nTotal Properties Analyzed: {len(self.analysis_results)}")
        print(f"- D'Marco Sites: {len(self.analysis_results[self.analysis_results['Source'] == 'D\'Marco'])}")
        print(f"- CoStar Properties: {len(self.analysis_results[self.analysis_results['Source'] == 'CoStar'])}")
        print(f"\nQCT/DDA Eligible (30% Boost): {len(self.analysis_results[self.analysis_results['QCT_DDA_Eligible'] == True])}")
        print(f"\nTop Recommendations:")
        print(f"- Strong 4% Candidates: {len(self.analysis_results[self.analysis_results['Score_4pct'] >= 70])}")
        print(f"- Strong 9% Candidates: {len(self.analysis_results[self.analysis_results['Score_9pct'] >= 70])}")
        print(f"- Fatal Flaw Sites: {len(self.analysis_results[self.analysis_results['Competition_Risk_9pct'] == 'Fatal'])}")
        print(f"\nReport saved to: {report_file}")
        print("\nSCORING CRITERIA USED:")
        print("4% Deals: Economic viability (50%), QCT/DDA bonus (30%), Land efficiency (20%)")
        print("9% Deals: QCT/DDA bonus (40%), Economic viability (30%), Land efficiency (30%)")
        print("Competition: One Mile Rule (Fatal for 9%, Risk for 4%), Two Mile Rule (Fatal for 9% in large counties)")
        print("="*80)
        
        return report_file


if __name__ == "__main__":
    analyzer = IntegratedTDHCAAnalyzer()
    analyzer.run_analysis()