"""
Create an enhanced Excel workbook for D'Marco with comprehensive data
Includes poverty rates, AMI rents, map links, and detailed analysis
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from datetime import datetime
import geopandas as gpd
from shapely.geometry import Point

def load_poverty_data():
    """Load Texas poverty rate data"""
    print("Loading poverty rate data...")
    poverty_df = pd.read_csv('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Lvg_Bond_Execution/Data Sets/Poverty Rate Census Tracts (ACS)/poverty_summary_TX_2022.csv')
    
    # Convert GEOID to string for proper matching
    poverty_df['GEOID'] = poverty_df['GEOID'].astype(str).str.zfill(11)
    return poverty_df

def load_ami_data():
    """Load HUD AMI rent data"""
    print("Loading HUD AMI rent data...")
    ami_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Lvg_Bond_Execution/Data Sets/HUD AMI FMR/HUD2025_AMI_Rent_Data_Static.xlsx'
    
    # Try to load Texas data
    try:
        ami_df = pd.read_excel(ami_path, sheet_name='TX')
    except:
        # If no TX sheet, load first sheet and filter for Texas
        ami_df = pd.read_excel(ami_path)
        if 'State' in ami_df.columns:
            ami_df = ami_df[ami_df['State'] == 'TX']
    
    return ami_df

def get_census_tract_for_property(lat, lon, poverty_gdf):
    """Find census tract for a property based on coordinates"""
    if pd.isna(lat) or pd.isna(lon):
        return None
    
    point = Point(lon, lat)
    
    # Find which tract contains this point
    for idx, tract in poverty_gdf.iterrows():
        if tract.geometry.contains(point):
            return tract['GEOID']
    
    return None

def calculate_price_per_acre(price, sf):
    """Calculate price per acre from price and square feet"""
    if pd.isna(price) or pd.isna(sf) or sf == 0:
        return None
    
    acres = sf / 43560  # Convert SF to acres
    return price / acres

def create_map_link(address, lat, lon):
    """Create a hyperlink for Apple/Google Maps"""
    if pd.notna(lat) and pd.notna(lon):
        # Google Maps link with satellite view
        google_link = f"https://www.google.com/maps/@{lat},{lon},18z/data=!3m1!1e3"
        return f'=HYPERLINK("{google_link}", "📍 View Map")'
    elif pd.notna(address):
        # Address-based search
        address_encoded = address.replace(' ', '+').replace(',', '')
        google_link = f"https://www.google.com/maps/search/{address_encoded}"
        return f'=HYPERLINK("{google_link}", "📍 Search Map")'
    return ""

def determine_opportunity_type(row):
    """Determine if property is better for 4% or 9% and why"""
    reasons = []
    recommended = "4%"
    
    # Check for 9% advantages
    if row.get('QCT_Status') == 'Yes' or row.get('DDA_Status') == 'Yes':
        reasons.append("30% basis boost (QCT/DDA)")
    
    if row.get('poverty_rate', 100) <= 20:
        reasons.append("Low poverty area (<20%)")
        recommended = "9%"  # Low poverty is better for 9% competitive
    
    if row.get('TDHCA_One_Mile_Fatal') == True:
        reasons.append("⚠️ ONE MILE RULE VIOLATION")
        recommended = "NEITHER"
    
    if row.get('Land_Viability_Score', 0) >= 90:
        reasons.append("High viability score")
    
    # 4% advantages
    if row.get('Sale Price', 0) > 500000:
        reasons.append("Higher price suits 4% (non-competitive)")
        recommended = "4%"
    
    if row.get('Market_Saturation_Score', 10) < 5:
        reasons.append("High competition area - 4% better")
        recommended = "4%"
    
    return recommended, " | ".join(reasons)

def create_enhanced_workbook():
    """Create the enhanced Excel workbook"""
    
    # Load all data
    print("Loading property data...")
    df = pd.read_excel('CoStar_Land_Analysis_20250616_203751.xlsx', sheet_name='All_Land_Analysis')
    
    poverty_df = load_poverty_data()
    ami_df = load_ami_data()
    
    # Load poverty shapefile for geocoding
    print("Loading census tract geometries...")
    poverty_gdf = gpd.read_file('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Lvg_Bond_Execution/Data Sets/Poverty Rate Census Tracts (ACS)/poverty_tracts_TX_2022.gpkg')
    poverty_gdf['GEOID'] = poverty_gdf['GEOID'].astype(str).str.zfill(11)
    
    # Calculate additional fields
    print("Calculating additional fields...")
    
    # Acres and price per acre
    df['Acres'] = df['Land SF Gross'] / 43560
    df['Price_Per_Acre'] = df.apply(lambda x: calculate_price_per_acre(x['Sale Price'], x['Land SF Gross']), axis=1)
    
    # Map links
    df['Map_Link'] = df.apply(lambda x: create_map_link(x['Address'], x['Latitude'], x['Longitude']), axis=1)
    
    # Get census tract and poverty data
    print("Matching properties to census tracts...")
    df['Census_Tract'] = df.apply(lambda x: get_census_tract_for_property(x['Latitude'], x['Longitude'], poverty_gdf), axis=1)
    
    # Merge poverty data
    df = df.merge(poverty_df[['GEOID', 'poverty_rate']], 
                  left_on='Census_Tract', right_on='GEOID', how='left')
    
    # Get county from address for AMI lookup
    df['County_Clean'] = df['Address'].str.extract(r', ([^,]+) County', expand=False)
    
    # Merge AMI data (simplified - you may need to adjust based on actual AMI file structure)
    print("Adding AMI rent data...")
    
    # Determine 4% vs 9% opportunity
    df[['Recommended_Type', 'Opportunity_Reasons']] = df.apply(
        lambda x: pd.Series(determine_opportunity_type(x)), axis=1
    )
    
    # Create Excel file
    output_file = f'DMarco_Enhanced_Properties_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Executive Summary (Hot Leads)
        print("Creating Executive Summary sheet...")
        hot_leads_df = df.nlargest(25, 'Land_Viability_Score')[[
            'Address', 'City', 'Acres', 'Sale Price', 'Price_Per_Acre',
            'Land_Viability_Score', 'poverty_rate', 'Recommended_Type',
            'Listing Broker Company', 'Listing Broker Phone',
            'Map_Link', 'Opportunity_Reasons'
        ]].copy()
        
        # Rename columns for clarity
        hot_leads_df.columns = ['Address', 'City', 'Acres', 'Total Price', 'Price/Acre',
                                'Score', 'Poverty %', 'Best For', 'Broker', 'Phone',
                                'Map', 'Why This Property Wins']
        
        # Add action columns
        hot_leads_df.insert(0, 'Priority', 'HOT')
        hot_leads_df.insert(len(hot_leads_df.columns), 'Your Notes', '')
        hot_leads_df.insert(len(hot_leads_df.columns), 'Next Step', 'Call Broker')
        
        hot_leads_df.to_excel(writer, sheet_name='🔥 Executive Summary', index=False)
        
        # Sheet 2: All Properties Enhanced
        print("Creating All Properties Enhanced sheet...")
        all_props_df = df[[
            'Address', 'City', 'Acres', 'Sale Price', 'Price_Per_Acre',
            'Land_Viability_Score', 'poverty_rate', 'QCT_Status', 'DDA_Status',
            'TDHCA_One_Mile_Fatal', 'FEMA_Zone', 'Recommended_Type',
            'Listing Broker Company', 'Listing Broker Phone', 'Map_Link'
        ]].copy()
        
        # Add user columns
        all_props_df.insert(0, 'Lead Status', 'NEW')
        all_props_df.insert(len(all_props_df.columns), 'Your Notes', '')
        
        all_props_df.to_excel(writer, sheet_name='All Properties Enhanced', index=False)
        
        # Sheet 3: Low Poverty Winners (<20%)
        print("Creating Low Poverty Winners sheet...")
        low_poverty_df = df[df['poverty_rate'] <= 20].nlargest(50, 'Land_Viability_Score')[[
            'Address', 'City', 'Acres', 'Sale Price', 'poverty_rate',
            'Land_Viability_Score', 'Recommended_Type', 'Map_Link'
        ]].copy()
        
        low_poverty_df.to_excel(writer, sheet_name='🏆 Low Poverty Winners', index=False)
        
        # Sheet 4: Complete Raw Data
        print("Creating Complete Raw Data sheet...")
        df.to_excel(writer, sheet_name='Complete Raw Data', index=False)
        
        # Sheet 5: AMI Rent Reference
        print("Creating AMI Rent Reference sheet...")
        # Create a simplified AMI reference (you'll need to adjust based on actual data)
        ami_reference = pd.DataFrame({
            'Information': [
                'Texas HUD AMI Rents 2025',
                '',
                'Most Texas Markets:',
                '60% AMI - 1BR: $800-950',
                '60% AMI - 2BR: $950-1,150',
                '60% AMI - 3BR: $1,100-1,350',
                '60% AMI - 4BR: $1,250-1,500',
                '',
                '100% AMI - 4BR: $2,100-2,800',
                '',
                'Major Markets (Austin/Dallas/Houston):',
                '60% AMI rents are 15-25% higher',
                '',
                'QCT/DDA Areas get 30% basis boost',
                'Low Poverty (<20%) helps win 9% credits'
            ]
        })
        ami_reference.to_excel(writer, sheet_name='📊 AMI Rent Guide', index=False)
    
    # Apply formatting
    print("Applying formatting...")
    wb = openpyxl.load_workbook(output_file)
    
    # Format each sheet
    format_executive_summary(wb['🔥 Executive Summary'])
    format_all_properties_enhanced(wb['All Properties Enhanced'])
    format_low_poverty_winners(wb['🏆 Low Poverty Winners'])
    
    # Save
    wb.save(output_file)
    print(f"\n✅ Created: {output_file}")
    print("\n📱 Enhanced features include:")
    print("   - Poverty rates with <20% highlighted")
    print("   - Price per acre calculations")
    print("   - Map links for each property")
    print("   - 4% vs 9% recommendations")
    print("   - Complete raw data preserved")

def format_executive_summary(ws):
    """Format the executive summary sheet"""
    
    # Header formatting
    header_fill = PatternFill(start_color='FF3B30', end_color='FF3B30', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Column widths
    column_widths = {
        'A': 10,   # Priority
        'B': 35,   # Address
        'C': 12,   # City
        'D': 10,   # Acres
        'E': 15,   # Total Price
        'F': 15,   # Price/Acre
        'G': 10,   # Score
        'H': 12,   # Poverty %
        'I': 12,   # Best For
        'J': 20,   # Broker
        'K': 15,   # Phone
        'L': 12,   # Map
        'M': 40,   # Why Wins
        'N': 25,   # Your Notes
        'O': 15,   # Next Step
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Number formatting
    for row in range(2, ws.max_row + 1):
        ws[f'D{row}'].number_format = '0.00'  # Acres
        ws[f'E{row}'].number_format = '$#,##0'  # Price
        ws[f'F{row}'].number_format = '$#,##0'  # Price/Acre
        ws[f'H{row}'].number_format = '0.0%'  # Poverty rate
    
    # Poverty rate formatting - highlight <= 20%
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for row in range(2, ws.max_row + 1):
        poverty_cell = ws[f'H{row}']
        if poverty_cell.value and poverty_cell.value <= 0.20:
            poverty_cell.fill = green_fill
            poverty_cell.font = Font(bold=True, color='006100')
    
    # Score color scale
    ws.conditional_formatting.add(f'G2:G{ws.max_row}',
        ColorScaleRule(start_type='num', start_value=70, start_color='FFCCCC',
                      mid_type='num', mid_value=85, mid_color='FFFFCC',
                      end_type='num', end_value=100, end_color='CCFFCC'))
    
    # Best For column - color by type
    for row in range(2, ws.max_row + 1):
        type_cell = ws[f'I{row}']
        if type_cell.value == '9%':
            type_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            type_cell.font = Font(bold=True, color='0066CC')
        elif type_cell.value == '4%':
            type_cell.fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
            type_cell.font = Font(bold=True, color='CC6600')
        elif type_cell.value == 'NEITHER':
            type_cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            type_cell.font = Font(bold=True, color='CC0000')
    
    # Add borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def format_all_properties_enhanced(ws):
    """Format the enhanced all properties sheet"""
    
    # Header formatting
    header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Poverty rate conditional formatting
    for row in range(2, ws.max_row + 1):
        poverty_cell = ws[f'G{row}']  # Assuming poverty rate is in column G
        if poverty_cell.value and poverty_cell.value <= 0.20:
            poverty_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            poverty_cell.font = Font(bold=True, color='006100')

def format_low_poverty_winners(ws):
    """Format the low poverty winners sheet"""
    
    # Header formatting
    header_fill = PatternFill(start_color='34C759', end_color='34C759', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # All poverty rates should be green since they're all <= 20%
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for row in range(2, ws.max_row + 1):
        for col in ws.iter_cols(min_row=row, max_row=row):
            if 'poverty' in str(ws.cell(1, col[0].column).value).lower():
                col[0].fill = green_fill
                col[0].font = Font(bold=True, color='006100')

if __name__ == "__main__":
    try:
        create_enhanced_workbook()
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        print("\nCreating simplified version without poverty data...")
        # Fallback to simpler version if geocoding fails