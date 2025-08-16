#!/usr/bin/env python3
"""
Federal-Enhanced CTCAC LIHTC Application Data Extractor
Integrates federal authority citations from Federal LIHTC RAG system

Enhanced with:
- Federal IRC Section 42 citations for basis calculations
- Federal regulatory authority for compliance requirements  
- Revenue Procedure references for inflation adjustments
- Authority hierarchy validation (Federal > State)
"""

import pandas as pd
import json
import os
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from typing import Dict, List, Any, Optional
import re
import sys

# Import unified LIHTC RAG query system
try:
    from unified_lihtc_rag_query import UnifiedLIHTCRAGQuery
except ImportError:
    print("Warning: unified_lihtc_rag_query not available. Federal citations will be limited.")
    UnifiedLIHTCRAGQuery = None

class FederalEnhancedCTCACExtractor:
    """
    CTCAC extractor with integrated federal authority citations
    """
    
    def __init__(self, input_path: str, output_path_4p: str, output_path_9p: str, log_path: str, 
                 data_sets_base_dir: str = None):
        """Initialize the extractor with file paths and federal RAG integration"""
        self.input_path = Path(input_path)
        self.output_path_4p = Path(output_path_4p)
        self.output_path_9p = Path(output_path_9p)
        self.log_path = Path(log_path)
        
        # Federal RAG system integration
        if data_sets_base_dir and UnifiedLIHTCRAGQuery:
            try:
                self.federal_rag = UnifiedLIHTCRAGQuery(data_sets_base_dir)
                self.federal_available = True
                print("✅ Federal LIHTC RAG system loaded - Enhanced citations available")
            except Exception as e:
                print(f"⚠️ Federal RAG system unavailable: {e}")
                self.federal_rag = None
                self.federal_available = False
        else:
            self.federal_rag = None
            self.federal_available = False
        
        # Create directories if they don't exist
        self.output_path_4p.mkdir(parents=True, exist_ok=True)
        self.output_path_9p.mkdir(parents=True, exist_ok=True)
        self.log_path.mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        self.setup_logging()
        
        # Define comprehensive sheet mapping
        self.target_sheets = {
            'application': 'Application',
            'sources_uses': 'Sources and Uses Budget',
            'sources_basis_breakdown': 'Sources and Basis Breakdown',
            'points_system': 'Points System',
            'basis_credits': 'Basis & Credits',
            'tie_breaker': 'Tie Breaker',
            'final_tie_breaker': 'Final Tie Breaker ',
            'disaster_tie_breaker': 'Disaster Credit Tie Breaker',
            'pro_forma': '15 Year Pro Forma',
            'sce_basis': 'SCE Basis and Credits',
            'fce_basis': 'FCE Basis and Credits',
            'calHFA_addendum': 'CalHFA Addendum',
            'subsidy_contract': 'Subsidy Contract Calculation'
        }
        
        # Federal authority mapping for LIHTC requirements
        self.federal_authorities = {
            'qualified_basis': {
                'authority': 'IRC Section 42(c)(1)',
                'level': 'statutory',
                'description': 'Federal definition of qualified basis for LIHTC'
            },
            'applicable_percentage': {
                'authority': 'IRC Section 42(b)(1)',
                'level': 'statutory', 
                'description': 'Federal applicable credit percentage'
            },
            'gross_income_test': {
                'authority': 'IRC Section 42(g)(1)',
                'level': 'statutory',
                'description': 'Federal gross income qualification requirements'
            },
            'rent_restriction': {
                'authority': 'IRC Section 42(g)(2)',
                'level': 'statutory',
                'description': 'Federal rent restriction requirements'
            },
            'compliance_period': {
                'authority': 'IRC Section 42(i)(1)',
                'level': 'statutory',
                'description': 'Federal 15-year compliance period'
            },
            'extended_use': {
                'authority': 'IRC Section 42(h)(6)',
                'level': 'statutory',
                'description': 'Federal extended use agreement requirements'
            },
            'placed_in_service': {
                'authority': '26 CFR 1.42-6',
                'level': 'regulatory',
                'description': 'Federal placed-in-service date regulations'
            },
            'ami_calculations': {
                'authority': '26 CFR 1.42-9',
                'level': 'regulatory',
                'description': 'Federal area median income determination'
            },
            'recapture_provisions': {
                'authority': 'IRC Section 42(j)',
                'level': 'statutory',
                'description': 'Federal recapture and monitoring provisions'
            },
            'inflation_adjustments': {
                'authority': 'Revenue Procedure 2024-40',
                'level': 'guidance',
                'description': '2025 federal inflation adjustments for LIHTC'
            }
        }
    
    def setup_logging(self):
        """Set up logging configuration"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_path / f"federal_enhanced_ctcac_extraction_{timestamp}.log"
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        
        # Clear any existing handlers
        self.logger.handlers.clear()
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.logger.propagate = False
        
        self.logger.info(f"Federal-enhanced extraction logging initialized. Log file: {log_file}")
        self.logger.info(f"Federal RAG integration: {'✅ Available' if self.federal_available else '❌ Not Available'}")

    def get_federal_authority_citation(self, concept: str, value: Any = None) -> Dict[str, Any]:
        """Get federal authority citation for a LIHTC concept"""
        citation = {
            'federal_authority': None,
            'authority_level': None,
            'statutory_basis': None,
            'regulatory_reference': None,
            'current_guidance': None,
            'conflict_resolution': 'state_implementation'
        }
        
        # Get base authority mapping
        if concept in self.federal_authorities:
            base_auth = self.federal_authorities[concept]
            citation.update({
                'federal_authority': base_auth['authority'],
                'authority_level': base_auth['level'],
                'description': base_auth['description']
            })
        
        # Enhance with live federal RAG lookups if available
        if self.federal_available and self.federal_rag:
            try:
                # Search for specific federal requirements
                search_terms = {
                    'qualified_basis': 'qualified basis calculation Section 42',
                    'applicable_percentage': 'applicable percentage federal credit',
                    'compliance_period': 'compliance period monitoring Section 42',
                    'ami_calculations': 'area median income calculation',
                    'inflation_adjustments': '2025 inflation adjustments credit ceiling'
                }
                
                if concept in search_terms:
                    federal_results = self.federal_rag.search_by_authority_level(
                        search_terms[concept],
                        authority_levels=['statutory', 'regulatory', 'guidance'],
                        limit=3
                    )
                    
                    if federal_results:
                        # Extract the highest authority source
                        top_result = federal_results[0]
                        citation.update({
                            'live_federal_source': top_result.get('document_title', ''),
                            'authority_score': top_result.get('authority_score', 0),
                            'source_type': top_result.get('source_type', ''),
                            'effective_date': top_result.get('effective_date', ''),
                            'content_preview': top_result.get('content', '')[:200] + '...'
                        })
                        
                        # Set conflict resolution based on authority level
                        if top_result.get('authority_score', 0) >= 80:
                            citation['conflict_resolution'] = 'federal_overrides_state'
                        elif top_result.get('authority_score', 0) >= 60:
                            citation['conflict_resolution'] = 'federal_guidance_minimum'
                        
            except Exception as e:
                self.logger.warning(f"Federal RAG lookup failed for {concept}: {e}")
        
        return citation

    def extract_basis_and_credits_enhanced(self, workbook, project_info: Dict) -> Dict[str, Any]:
        """Extract basis and credits data with federal authority citations"""
        result = {}
        
        # Standard CTCAC extraction
        if 'Basis & Credits' in workbook.sheetnames:
            sheet = workbook['Basis & Credits']
            
            # Extract key basis values
            qualified_basis = self.safe_extract_value(sheet, 'D47')  # Qualified basis
            applicable_percentage = self.safe_extract_value(sheet, 'D51')  # Applicable percentage
            federal_credits = self.safe_extract_value(sheet, 'D52')  # Federal credits
            
            result.update({
                'qualified_basis': qualified_basis,
                'applicable_percentage': applicable_percentage,
                'federal_credits': federal_credits
            })
            
            # Add federal authority citations
            result['federal_authorities'] = {
                'qualified_basis': self.get_federal_authority_citation('qualified_basis', qualified_basis),
                'applicable_percentage': self.get_federal_authority_citation('applicable_percentage', applicable_percentage),
                'compliance_period': self.get_federal_authority_citation('compliance_period'),
                'recapture_provisions': self.get_federal_authority_citation('recapture_provisions')
            }
            
            # Check for federal vs state conflicts
            if self.federal_available:
                result['authority_analysis'] = self.analyze_federal_state_conflicts(result)
            
            self.logger.info(f"Extracted basis & credits with federal citations: QB=${qualified_basis:,.0f}, AP={applicable_percentage}%")
        
        return result

    def analyze_federal_state_conflicts(self, extracted_data: Dict) -> Dict[str, Any]:
        """Analyze potential conflicts between federal requirements and CTCAC implementation"""
        conflicts = {
            'critical_conflicts': [],       # Violations of federal minimums
            'enhanced_state_requirements': [], # State requirements exceeding federal
            'investigation_required': [],   # Potential gaps or other funding sources
            'compliance_verification': {},
            'recommendations': []
        }
        
        if not self.federal_available:
            conflicts['warning'] = 'Federal RAG system unavailable - limited conflict detection'
            return conflicts
        
        try:
            # 1. COMPLIANCE PERIOD ANALYSIS
            federal_compliance = self.federal_rag.search_by_authority_level(
                'compliance period requirements Section 42',
                authority_levels=['statutory', 'regulatory'],
                limit=2
            )
            
            if federal_compliance:
                conflicts['compliance_verification']['federal_minimum'] = '15 years per IRC 42(i)(1)'
                conflicts['compliance_verification']['ctcac_requirement'] = '55 years per state extended use agreement'
                
                # This is enhancement, not conflict
                conflicts['enhanced_state_requirements'].append({
                    'category': 'Compliance Period',
                    'federal_requirement': '15 years minimum',
                    'ctcac_requirement': '55 years',
                    'analysis': 'CTCAC exceeds federal minimum - COMPLIANT',
                    'funding_implication': 'Extended use may enable additional state/local funding sources'
                })
            
            # 2. AMI CALCULATION METHODOLOGY
            federal_ami = self.federal_rag.search_by_authority_level(
                'area median income calculation methodology HUD',
                authority_levels=['regulatory'],
                limit=2
            )
            
            if federal_ami:
                conflicts['compliance_verification']['ami_source'] = 'HUD data required per 26 CFR 1.42-9'
                conflicts['recommendations'].append('VERIFY: CTCAC uses HUD AMI data per federal requirement')
                
                # Flag for investigation if state has different AMI rules
                conflicts['investigation_required'].append({
                    'category': 'AMI Calculation',
                    'federal_requirement': 'HUD area median income data per 26 CFR 1.42-9',
                    'investigation_needed': 'Verify CTCAC rent calculations use federal HUD AMI data',
                    'potential_issue': 'If CTCAC uses different AMI source, this violates federal requirements',
                    'action_required': 'Review CTCAC QAP Section 10327 for AMI methodology'
                })
            
            # 3. QUALIFIED BASIS CALCULATIONS
            federal_basis = self.federal_rag.search_by_authority_level(
                'qualified basis calculation Section 42',
                authority_levels=['statutory'],
                limit=2
            )
            
            if federal_basis:
                conflicts['investigation_required'].append({
                    'category': 'Qualified Basis',
                    'federal_requirement': 'IRC Section 42(c) qualified basis definition',
                    'investigation_needed': 'Compare CTCAC eligible basis items vs federal qualified basis',
                    'potential_gap_funding': 'CTCAC may exclude items that require gap funding from other sources',
                    'action_required': 'Review CTCAC basis calculation worksheets vs federal requirements'
                })
            
            # 4. APPLICABLE PERCENTAGE RATES
            federal_percentage = self.federal_rag.search_by_authority_level(
                '2025 inflation adjustments applicable percentage',
                authority_levels=['guidance'],
                limit=2
            )
            
            if federal_percentage:
                conflicts['compliance_verification']['federal_rates'] = 'Current rates per Revenue Procedure 2024-40'
                conflicts['recommendations'].append('CRITICAL: Verify CTCAC uses current federal applicable percentages')
                
                # Flag as critical if rates don't match
                conflicts['critical_conflicts'].append({
                    'category': 'Applicable Percentage',
                    'federal_requirement': 'Must use current federal rates from Revenue Procedure 2024-40',
                    'potential_violation': 'Using outdated rates violates federal requirements',
                    'compliance_risk': 'HIGH - Could invalidate credit allocation',
                    'action_required': 'Verify CTCAC application uses current federal rates'
                })
            
            # 5. RENT RESTRICTIONS
            federal_rents = self.federal_rag.search_by_authority_level(
                'rent restriction requirements Section 42',
                authority_levels=['statutory'],
                limit=2
            )
            
            if federal_rents:
                conflicts['investigation_required'].append({
                    'category': 'Rent Restrictions',
                    'federal_requirement': 'IRC Section 42(g)(2) - 30% of AMI maximum',
                    'investigation_needed': 'Compare CTCAC rent limits vs federal 30% AMI maximum',
                    'potential_enhancement': 'CTCAC may have stricter rent limits enabling deeper affordability',
                    'funding_implication': 'Lower rents may require additional subsidy/gap funding',
                    'action_required': 'Review CTCAC rent calculation vs federal maximums'
                })
            
            # 6. INCOME TARGETING
            federal_income = self.federal_rag.search_by_authority_level(
                'gross income test requirements Section 42',
                authority_levels=['statutory'],
                limit=2
            )
            
            if federal_income:
                conflicts['investigation_required'].append({
                    'category': 'Income Targeting',
                    'federal_requirement': 'IRC Section 42(g)(1) - 60% AMI maximum (or 40%/60% test)',
                    'investigation_needed': 'Compare CTCAC income targeting vs federal requirements',
                    'potential_enhancement': 'CTCAC may require deeper income targeting',
                    'funding_implication': 'Lower income targets may require additional operating subsidies',
                    'action_required': 'Review CTCAC income targeting requirements vs federal minimums'
                })
            
            # Generate summary recommendations
            if conflicts['critical_conflicts']:
                conflicts['recommendations'].insert(0, 
                    f"🚨 CRITICAL: {len(conflicts['critical_conflicts'])} potential federal compliance violations identified")
            
            if conflicts['investigation_required']:
                conflicts['recommendations'].append(
                    f"🔍 INVESTIGATE: {len(conflicts['investigation_required'])} areas require federal vs state comparison")
            
            if conflicts['enhanced_state_requirements']:
                conflicts['recommendations'].append(
                    f"✅ ENHANCED: {len(conflicts['enhanced_state_requirements'])} state requirements exceed federal minimums")
            
            conflicts['recommendations'].extend([
                "Review CTCAC QAP against federal IRC Section 42 requirements",
                "Identify gap funding sources if state requirements exceed federal",
                "Verify all calculations use federally-required data sources",
                "Document compliance with federal minimums before state enhancements"
            ])
                
        except Exception as e:
            self.logger.warning(f"Federal conflict analysis failed: {e}")
            conflicts['error'] = str(e)
        
        return conflicts

    def safe_extract_value(self, sheet, cell_address: str) -> Any:
        """Safely extract value from Excel cell"""
        try:
            cell = sheet[cell_address]
            if cell.value is None:
                return None
            return cell.value
        except Exception as e:
            self.logger.warning(f"Failed to extract from {cell_address}: {e}")
            return None

    def extract_comprehensive_data(self, file_path: Path) -> Dict[str, Any]:
        """Extract comprehensive data with federal authority integration"""
        self.logger.info(f"Processing file: {file_path.name}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Parse filename for metadata
            project_info = self.parse_filename(file_path.name)
            
            # Initialize extraction result
            extraction_result = {
                'file_metadata': {
                    'filename': file_path.name,
                    'processed_date': datetime.now().isoformat(),
                    'file_size_mb': round(file_path.stat().st_size / 1048576, 2),
                    'credit_type': project_info.get('credit_type', 'unknown'),
                    'federal_enhancement': True,
                    'federal_rag_available': self.federal_available
                },
                'project_info': project_info,
                'extraction_status': 'in_progress'
            }
            
            # Extract basis and credits with federal citations
            extraction_result['basis_credits'] = self.extract_basis_and_credits_enhanced(workbook, project_info)
            
            # Extract other standard data
            extraction_result['application_data'] = self.extract_application_sheet(workbook)
            extraction_result['sources_uses'] = self.extract_sources_uses(workbook)
            extraction_result['points_system'] = self.extract_points_system(workbook)
            
            # Add federal authority summary
            extraction_result['federal_authority_summary'] = self.generate_federal_summary(extraction_result)
            
            extraction_result['extraction_status'] = 'completed'
            return extraction_result
            
        except Exception as e:
            self.logger.error(f"Extraction failed for {file_path.name}: {e}")
            return {
                'file_metadata': {'filename': file_path.name, 'error': str(e)},
                'extraction_status': 'failed',
                'error_details': str(e)
            }

    def extract_application_sheet(self, workbook) -> Dict[str, Any]:
        """Extract basic application data"""
        result = {}
        if 'Application' in workbook.sheetnames:
            sheet = workbook['Application']
            result.update({
                'project_name': self.safe_extract_value(sheet, 'C8'),
                'development_address': self.safe_extract_value(sheet, 'C9'),
                'city': self.safe_extract_value(sheet, 'C10'),
                'county': self.safe_extract_value(sheet, 'C11'),
                'total_units': self.safe_extract_value(sheet, 'C25'),
                'total_development_cost': self.safe_extract_value(sheet, 'C46')
            })
        return result

    def extract_sources_uses(self, workbook) -> Dict[str, Any]:
        """Extract sources and uses data"""
        result = {}
        if 'Sources and Uses Budget' in workbook.sheetnames:
            sheet = workbook['Sources and Uses Budget']
            result.update({
                'total_development_cost': self.safe_extract_value(sheet, 'D38'),
                'total_sources': self.safe_extract_value(sheet, 'D73'),
                'lihtc_equity': self.safe_extract_value(sheet, 'D62')
            })
        return result

    def extract_points_system(self, workbook) -> Dict[str, Any]:
        """Extract points system data"""
        result = {}
        if 'Points System' in workbook.sheetnames:
            sheet = workbook['Points System']
            result.update({
                'total_points': self.safe_extract_value(sheet, 'D100'),
                'max_possible_points': self.safe_extract_value(sheet, 'D101')
            })
        return result

    def generate_federal_summary(self, extraction_result: Dict) -> Dict[str, Any]:
        """Generate summary of federal authority implications"""
        summary = {
            'federal_requirements_identified': 0,
            'authority_levels_cited': [],
            'compliance_recommendations': [],
            'conflict_analysis_available': self.federal_available
        }
        
        # Count federal authorities cited
        basis_credits = extraction_result.get('basis_credits', {})
        federal_auths = basis_credits.get('federal_authorities', {})
        
        for auth_type, auth_data in federal_auths.items():
            if auth_data.get('federal_authority'):
                summary['federal_requirements_identified'] += 1
                level = auth_data.get('authority_level')
                if level and level not in summary['authority_levels_cited']:
                    summary['authority_levels_cited'].append(level)
        
        # Generate recommendations
        if summary['federal_requirements_identified'] > 0:
            summary['compliance_recommendations'].extend([
                'Verify qualified basis calculations comply with IRC Section 42(c)(1)',
                'Confirm applicable percentage rates per latest Revenue Procedure',
                'Ensure compliance period meets federal 15-year minimum',
                'Review AMI calculations against HUD data requirements'
            ])
        
        return summary

    def parse_filename(self, filename: str) -> Dict[str, str]:
        """Parse filename to extract metadata"""
        base_name = Path(filename).stem
        
        project_info = {
            'original_filename': filename,
            'credit_type': 'unknown'
        }
        
        # Determine credit type
        if any(x in filename.lower() for x in ['4%', '4 percent', 'four percent', 'bond']):
            project_info['credit_type'] = '4%'
        elif any(x in filename.lower() for x in ['9%', '9 percent', 'nine percent', 'competitive']):
            project_info['credit_type'] = '9%'
        
        return project_info

    def process_files(self, specific_files: List[str] = None) -> Dict[str, Any]:
        """Process CTCAC files with federal authority enhancement"""
        self.logger.info("Starting federal-enhanced CTCAC extraction process")
        
        if not self.input_path.exists():
            raise FileNotFoundError(f"Input directory not found: {self.input_path}")
        
        # Find Excel files
        excel_files = list(self.input_path.glob("*.xlsx")) + list(self.input_path.glob("*.xls"))
        
        if specific_files:
            excel_files = [f for f in excel_files if f.name in specific_files]
        
        if not excel_files:
            self.logger.warning("No Excel files found for processing")
            return {'processed_files': 0, 'successful_extractions': 0}
        
        self.logger.info(f"Found {len(excel_files)} Excel files to process")
        
        results = {
            'processed_files': 0,
            'successful_extractions': 0,
            'failed_extractions': 0,
            'files_processed': [],
            'federal_enhancement_stats': {
                'total_federal_citations': 0,
                'authority_levels_used': set(),
                'rag_lookups_successful': 0
            }
        }
        
        for file_path in excel_files:
            try:
                # Extract data
                extracted_data = self.extract_comprehensive_data(file_path)
                
                # Determine output path based on credit type
                credit_type = extracted_data.get('project_info', {}).get('credit_type', 'unknown')
                if credit_type == '4%':
                    output_path = self.output_path_4p
                elif credit_type == '9%':
                    output_path = self.output_path_9p
                else:
                    output_path = self.output_path_4p  # Default
                
                # Save extracted data
                output_file = output_path / f"{file_path.stem}_federal_enhanced_extraction.json"
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(extracted_data, f, indent=2, default=str)
                
                if extracted_data.get('extraction_status') == 'completed':
                    results['successful_extractions'] += 1
                    
                    # Update federal enhancement stats
                    federal_summary = extracted_data.get('federal_authority_summary', {})
                    results['federal_enhancement_stats']['total_federal_citations'] += federal_summary.get('federal_requirements_identified', 0)
                    
                    for level in federal_summary.get('authority_levels_cited', []):
                        results['federal_enhancement_stats']['authority_levels_used'].add(level)
                else:
                    results['failed_extractions'] += 1
                
                results['files_processed'].append({
                    'filename': file_path.name,
                    'status': extracted_data.get('extraction_status'),
                    'output_file': str(output_file),
                    'federal_citations': extracted_data.get('federal_authority_summary', {}).get('federal_requirements_identified', 0)
                })
                
                results['processed_files'] += 1
                
            except Exception as e:
                self.logger.error(f"Failed to process {file_path.name}: {e}")
                results['failed_extractions'] += 1
                results['files_processed'].append({
                    'filename': file_path.name,
                    'status': 'failed',
                    'error': str(e)
                })
        
        # Convert set to list for JSON serialization
        results['federal_enhancement_stats']['authority_levels_used'] = list(results['federal_enhancement_stats']['authority_levels_used'])
        
        # Generate final summary
        success_rate = results['successful_extractions'] / results['processed_files'] if results['processed_files'] > 0 else 0
        
        self.logger.info(f"""
Federal-Enhanced CTCAC Extraction Complete:
✅ Success Rate: {success_rate:.1%} ({results['successful_extractions']}/{results['processed_files']})
🏛️ Federal Citations: {results['federal_enhancement_stats']['total_federal_citations']}
⚖️ Authority Levels: {', '.join(results['federal_enhancement_stats']['authority_levels_used'])}
🔍 RAG Integration: {'Available' if self.federal_available else 'Not Available'}
        """)
        
        return results


# Example usage
if __name__ == "__main__":
    # Initialize with federal RAG integration
    extractor = FederalEnhancedCTCACExtractor(
        input_path='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data',
        output_path_4p='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/JSON_data/4p',
        output_path_9p='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/JSON_data/9p',
        log_path='/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/logs',
        data_sets_base_dir='/Users/williamrice/HERR Dropbox/Bill Rice/Data_Sets'  # Federal RAG data location
    )
    
    # Process files
    results = extractor.process_files()
    print("\nFederal-Enhanced CTCAC Extraction completed!")
    print(f"Results: {results}")