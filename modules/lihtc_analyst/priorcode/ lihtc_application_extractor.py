import os
import re
import json
import logging
import traceback
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_application_extraction.log"),
        logging.StreamHandler()
    ]
)

class LIHTCApplicationExtractor:
    """
    Extract basic application info from LIHTC applications and save as JSON files.
    """
    
    def __init__(self):
        """Initialize the extractor with file paths."""
        # Input path
        self.input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
        
        # Output paths
        self.output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p'
        self.output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p'
        
        # Create output directories if they don't exist
        Path(self.output_path_4p).mkdir(parents=True, exist_ok=True)
        Path(self.output_path_9p).mkdir(parents=True, exist_ok=True)
        
        # Statistics
        self.files_processed = 0
        self.files_skipped = 0
        self.extraction_errors = 0
    
    def extract_application_tab(self, file_path, tab_name):
        """
        Enhanced extraction of application information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Application tab from {file_name}")
        
        # Parse metadata from filename
        app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
        year_match = re.search(r'(\d{4})_\d+pct', file_name)
        round_match = re.search(r'_R(\d+)_', file_name)
        ctcac_match = re.search(r'_(\d+-\d+)', file_name)
        
        # Initialize result with metadata
        result = {
            "metadata": {
                "file_name": file_name,
                "application_type": app_type,
                "application_year": year_match.group(1) if year_match else "",
                "application_round": round_match.group(1) if round_match else "",
                "ctcac_number": ctcac_match.group(1) if ctcac_match else "",
                "extraction_date": datetime.now().isoformat()
            },
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Find relevant tab - check multiple possible tab names
            sheet = None
            potential_tab_names = [
                "Application", "Project", "General", "Project Information", 
                "Project Data", "General Information", "Development", "Basic Information"
            ]
            
            # First try the exact tab name provided
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
            else:
                # If not found, try other potential tab names
                for potential_name in potential_tab_names:
                    for sheet_name in wb.sheetnames:
                        if potential_name.lower() in sheet_name.lower():
                            sheet = wb[sheet_name]
                            logging.info(f"Found Application data in tab: {sheet_name}")
                            break
                    if sheet:
                        break
            
            # If still not found, try the first sheet as a last resort
            if not sheet and len(wb.sheetnames) > 0:
                sheet = wb[wb.sheetnames[0]]
                logging.info(f"Using first sheet ({wb.sheetnames[0]}) for Application data")
            
            if not sheet:
                logging.warning(f"No suitable sheet found for Application data in {file_name}")
                return result
            
            # Enhanced patterns for field matching
            field_patterns = {
                "project.name": [
                    r"(?i)project\s*name",
                    r"(?i)development\s*name",
                    r"(?i)property\s*name",
                    r"(?i)apartment\s*name",
                    r"(?i)name\s*of\s*project"
                ],
                "project.address": [
                    r"(?i)project\s*address",
                    r"(?i)property\s*address",
                    r"(?i)development\s*address",
                    r"(?i)site\s*address",
                    r"(?i)address"
                ],
                "project.city": [
                    r"(?i)city",
                    r"(?i)city/town",
                    r"(?i)municipality"
                ],
                "project.county": [
                    r"(?i)county",
                    r"(?i)parish"
                ],
                "project.zip_code": [
                    r"(?i)zip\s*code",
                    r"(?i)zip",
                    r"(?i)postal\s*code",
                    r"(?i)zip\s*\+\s*4"
                ],
                "project.census_tract": [
                    r"(?i)census\s*tract",
                    r"(?i)tract",
                    r"(?i)census\s*id"
                ],
                "project.total_units": [
                    r"(?i)total\s*units",
                    r"(?i)total\s*number\s*of\s*units",
                    r"(?i)total\s*residential\s*units",
                    r"(?i)total\s*dwelling\s*units",
                    r"(?i)number\s*of\s*units"
                ],
                "project.low_income_units": [
                    r"(?i)low[\s-]*income\s*units",
                    r"(?i)affordable\s*units",
                    r"(?i)tax\s*credit\s*units",
                    r"(?i)lihtc\s*units",
                    r"(?i)income\s*restricted\s*units",
                    r"(?i)restricted\s*units"
                ],
                "project.manager_units": [
                    r"(?i)manager['s]*\s*units?",
                    r"(?i)employee\s*units?",
                    r"(?i)staff\s*units?",
                    r"(?i)on[-\s]site\s*manager['s]*\s*units?"
                ]
            }
            
            # Helper function to check if a value is worth storing
            def is_valid_value(field_path, value):
                # Don't override existing values unless empty
                parts = field_path.split('.')
                target = result
                for part in parts[:-1]:
                    target = target[part]
                
                current_value = target[parts[-1]]
                if current_value and current_value != 0 and field_path != "project.total_units":
                    return False
                    
                if value is None:
                    return False
                    
                if isinstance(value, str) and value.strip() == "":
                    return False
                    
                # For numeric fields, check if it makes sense
                if parts[-1] in ['total_units', 'low_income_units', 'manager_units']:
                    try:
                        # Try to convert to a number
                        if isinstance(value, str):
                            # Clean up the string - remove non-numeric characters
                            num_str = re.sub(r'[^\d.]', '', value)
                            if not num_str:
                                return False
                            value = float(num_str)
                        else:
                            value = float(value)
                        
                        # Check if it's a reasonable unit count (between 1 and 1000)
                        if value <= 0 or value > 1000:
                            return False
                    except:
                        return False
                
                return True
            
            # Process the fields in the worksheet - scan more rows and columns
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(30, sheet.max_column + 1)):
                    try:
                        cell_value = sheet.cell(row=row, column=col).value
                        
                        # Skip empty cells
                        if not cell_value:
                            continue
                        
                        # Convert to string for regex matching
                        if not isinstance(cell_value, str):
                            continue
                        
                        # Check against all field patterns
                        for field_path, patterns in field_patterns.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a match, now look for the data value
                                logging.debug(f"Found potential match for {field_path} at row {row}, col {col}")
                                
                                # Check for value in adjacent cells - more exhaustive search
                                data_found = False
                                data_value = None
                                
                                # Check to the right - several cells
                                for r_offset in range(1, 5):
                                    if col + r_offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+r_offset).value
                                        if data_value and is_valid_value(field_path, data_value):
                                            data_found = True
                                            break
                                
                                # If not found to the right, check below
                                if not data_found:
                                    for d_offset in range(1, 4):
                                        if row + d_offset <= sheet.max_row:
                                            data_value = sheet.cell(row=row+d_offset, column=col).value
                                            if data_value and is_valid_value(field_path, data_value):
                                                data_found = True
                                                break
                                
                                # If not found in adjacent cells, check the same cell (might be a merged cell with label and value)
                                if not data_found:
                                    # Some applications put the value in the same cell as the label
                                    # Check if the label is followed by a colon with a value
                                    match = re.search(r'(?i)[:]\s*(.+)', cell_value)
                                    if match:
                                        data_value = match.group(1).strip()
                                        if data_value and is_valid_value(field_path, data_value):
                                            data_found = True
                                
                                # Update the result if data was found
                                if data_found:
                                    # Parse field path to update nested dict
                                    parts = field_path.split('.')
                                    target = result
                                    for part in parts[:-1]:
                                        target = target[part]
                                    
                                    # Process the value based on field type
                                    if parts[-1] in ['total_units', 'low_income_units', 'manager_units']:
                                        # Clean and convert to integer for unit counts
                                        try:
                                            if isinstance(data_value, str):
                                                # Remove any non-numeric characters
                                                num_str = re.sub(r'[^\d.]', '', data_value)
                                                if num_str:
                                                    data_value = int(float(num_str))
                                            else:
                                                data_value = int(float(data_value))
                                        except:
                                            # If conversion fails, skip
                                            continue
                                    elif isinstance(data_value, str):
                                        # Clean up string values
                                        data_value = data_value.strip()
                                    
                                    # Set the value
                                    target[parts[-1]] = data_value
                                    logging.debug(f"Set {field_path} to {data_value}")
                    except Exception as e:
                        logging.debug(f"Error processing cell at row {row}, col {col}: {e}")
                        continue
            
            # Special processing for project name - try to extract from filename if not found
            if not result["project"]["name"]:
                # Look for project name in the filename (often included)
                name_match = re.search(r'(?i)_([^_]+)_\d+pct', file_name)
                if name_match:
                    result["project"]["name"] = name_match.group(1).replace('_', ' ')
            
            # Validation check - total units should be at least the sum of low_income and manager units
            low_income = result["project"]["low_income_units"]
            manager = result["project"]["manager_units"]
            total = result["project"]["total_units"]
            
            if low_income > 0 and manager > 0 and total == 0:
                result["project"]["total_units"] = low_income + manager
            
            # If we have total and low_income but no manager, estimate manager
            if total > 0 and low_income > 0 and manager == 0:
                # Typical LIHTC projects have 1-2 manager units
                estimated_manager = total - low_income
                if 0 < estimated_manager <= 3:  # Reasonable manager unit count
                    result["project"]["manager_units"] = estimated_manager
            
            # After all extraction, log what we found
            logging.info(f"Extracted project: {result['project']['name']}, Units: {result['project']['total_units']}")
            
            return result
        
        except Exception as e:
            logging.error(f"Error extracting Application tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def process_single_file(self, file_path):
        """
        Process a single LIHTC application file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            dict: Processed application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Processing file: {file_name}")
        
        try:
            # Determine application type
            app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
            
            if app_type == "unknown":
                logging.warning(f"Unknown application type for file: {file_name}, skipping")
                self.files_skipped += 1
                return None
            
            # Extract application tab data
            application_data = self.extract_application_tab(file_path, "Application")
            
            # Determine output filename
            ctcac_num = application_data["metadata"]["ctcac_number"]
            if not ctcac_num:
                # Generate a filename based on project name and application type
                project_name = application_data["project"]["name"] or "unknown_project"
                # Clean project name for filename
                project_name = re.sub(r'[^\w\-]', '_', project_name)
                output_filename = f"{project_name}_{app_type}.json"
            else:
                output_filename = f"{ctcac_num}_{app_type}.json"
            
            # Determine output directory
            output_dir = self.output_path_4p if app_type == "4p" else self.output_path_9p
            output_path = os.path.join(output_dir, output_filename)
            
            # Save to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(application_data, f, indent=2)
            
            logging.info(f"Saved extracted data to {output_path}")
            self.files_processed += 1
            
            return application_data
            
        except Exception as e:
            logging.error(f"Error processing file {file_name}: {e}")
            logging.error(traceback.format_exc())
            self.extraction_errors += 1
            return None
    
    def process_all_files(self):
        """
        Process all LIHTC application files in the input directory
        """
        logging.info(f"Starting processing of all LIHTC application files in {self.input_path}")
        
        # Get list of all Excel files
        excel_files = []
        for root, dirs, files in os.walk(self.input_path):
            for file in files:
                if file.endswith('.xlsx') and ('4pct' in file or '9pct' in file):
                    excel_files.append(os.path.join(root, file))
        
        logging.info(f"Found {len(excel_files)} LIHTC application files to process")
        
        # Process each file
        for i, file_path in enumerate(excel_files):
            logging.info(f"Processing file {i+1} of {len(excel_files)}: {os.path.basename(file_path)}")
            self.process_single_file(file_path)
        
        # Log summary
        logging.info(f"Completed processing {len(excel_files)} files")
        logging.info(f"Successfully processed: {self.files_processed}")
        logging.info(f"Skipped: {self.files_skipped}")
        logging.info(f"Errors: {self.extraction_errors}")

def main():
    """Main entry point for the script."""
    print("Starting LIHTC Application Tab Extractor...")
    print("This script will extract basic project information from LIHTC applications")
    print(f"Looking for Excel files in folder path: /Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data")
    
    extractor = LIHTCApplicationExtractor()
    extractor.process_all_files()
    
    print(f"\nFinished processing.")
    print(f"Successfully processed: {extractor.files_processed} files")
    print(f"Skipped: {extractor.files_skipped} files")
    print(f"Errors: {extractor.extraction_errors} files")
    print("\nSee lihtc_application_extraction.log for detailed information.")

if __name__ == "__main__":
    main()