# excel_merged_cells_test.py

import pandas as pd
from openpyxl import load_workbook
import logging
from pathlib import Path
import os
import re

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("excel_merged_cells_test.log"),
        logging.StreamHandler()
    ]
)

def get_cell_value(sheet, cell_ref):
    """
    Get cell value, handling merged cells properly.
    
    Args:
        sheet: The worksheet
        cell_ref: Cell reference (e.g., 'B10')
        
    Returns:
        The cell value
    """
    try:
        # Convert cell reference to coordinates
        col, row = re.findall(r'([A-Z]+)(\d+)', cell_ref)[0]
        col_idx = 0
        for c in col:
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        row_idx = int(row)
        
        # Check if cell is part of merged range
        for merged_range in sheet.merged_cells.ranges:
            if (col_idx >= merged_range.min_col and col_idx <= merged_range.max_col and
                row_idx >= merged_range.min_row and row_idx <= merged_range.max_row):
                # Get the top-left cell of the merged range
                top_left = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                logging.info(f"Cell {cell_ref} is part of merged range {merged_range}, using {top_left} instead")
                return sheet[top_left].value
        
        # If not part of merged range, return the value directly
        return sheet[cell_ref].value
    except Exception as e:
        logging.error(f"Error getting value for {cell_ref}: {e}")
        return None

def get_column_letter(col_idx):
    """Convert column index to letter (1 -> A, 2 -> B, etc.)"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def find_nearby_data(sheet, cell_ref, radius=3):
    """
    Look for data in cells near the specified cell.
    
    Args:
        sheet: The worksheet
        cell_ref: Starting cell reference (e.g., 'B10')
        radius: How many cells to search in each direction
        
    Returns:
        Dictionary with found values and their locations
    """
    try:
        # Parse cell reference
        col, row = re.findall(r'([A-Z]+)(\d+)', cell_ref)[0]
        col_idx = 0
        for c in col:
            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
        row_idx = int(row)
        
        # Search nearby cells
        results = {}
        for r in range(row_idx - radius, row_idx + radius + 1):
            for c in range(col_idx - radius, col_idx + radius + 1):
                if r <= 0 or c <= 0:  # Skip invalid coordinates
                    continue
                
                # Skip the original cell
                if r == row_idx and c == col_idx:
                    continue
                
                nearby_ref = f"{get_column_letter(c)}{r}"
                try:
                    value = sheet[nearby_ref].value
                    if value is not None and str(value).strip():
                        results[nearby_ref] = value
                except:
                    pass
        
        return results
    except Exception as e:
        logging.error(f"Error searching near {cell_ref}: {e}")
        return {}

def test_excel_reading_with_merged_cells(file_path):
    """Test reading key data from an Excel file, handling merged cells."""
    logging.info(f"Testing Excel reading with merged cells for: {os.path.basename(file_path)}")
    
    # List of test locations to check
    test_cells = [
        {"sheet": "Application", "cell": "B10", "description": "CTCAC Number"},
        {"sheet": "Application", "cell": "B14", "description": "Project Name"},
        {"sheet": "Application", "cell": "B15", "description": "Project Address"},
        {"sheet": "Application", "cell": "D18", "description": "Total Units"},
        {"sheet": "Sources and Uses Budget", "cell": "F134", "description": "Total Project Cost"},
        {"sheet": "Sources and Uses Budget", "cell": "F58", "description": "Construction Costs"},
        {"sheet": "Basis & Credits", "cell": "D98", "description": "Total Eligible Basis"},
        {"sheet": "Basis & Credits", "cell": "G135", "description": "Federal Annual Credit"}
    ]
    
    # First check if file exists
    if not os.path.exists(file_path):
        logging.error(f"File not found: {file_path}")
        return
    
    try:
        # Load workbook with values instead of formulas
        wb = load_workbook(file_path, data_only=True)
        
        # List all sheets in the workbook
        logging.info(f"Sheets in workbook: {wb.sheetnames}")
        
        # Check each test cell
        results = []
        for test in test_cells:
            sheet_name = test["sheet"]
            cell_ref = test["cell"]
            description = test["description"]
            
            if sheet_name not in wb.sheetnames:
                results.append({
                    "description": description,
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "status": "Sheet not found",
                    "value": None,
                    "nearby": {}
                })
                continue
            
            try:
                sheet = wb[sheet_name]
                
                # Get value handling merged cells
                value = get_cell_value(sheet, cell_ref)
                
                # If no value found, look in nearby cells
                nearby = {}
                if value is None or (isinstance(value, str) and not value.strip()):
                    nearby = find_nearby_data(sheet, cell_ref)
                
                results.append({
                    "description": description,
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "status": "Success" if value is not None else "Empty cell",
                    "value": value,
                    "nearby": nearby
                })
            except Exception as e:
                results.append({
                    "description": description,
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "status": f"Error: {str(e)}",
                    "value": None,
                    "nearby": {}
                })
        
        # Print results
        logging.info("Cell Reading Results:")
        logging.info("-" * 80)
        for result in results:
            logging.info(f"{result['description']} ({result['sheet']}!{result['cell']}): {result['status']}")
            if result['value'] is not None:
                logging.info(f"  Value: {result['value']}")
            
            if result['nearby']:
                logging.info("  Nearby data found:")
                for cell, value in result['nearby'].items():
                    logging.info(f"    {cell}: {value}")
        logging.info("-" * 80)
        
        # Check for merged cells in key sheets
        for sheet_name in ["Application", "Sources and Uses Budget", "Basis & Credits"]:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                merged_ranges = list(sheet.merged_cells.ranges)
                logging.info(f"Sheet '{sheet_name}' has {len(merged_ranges)} merged cell ranges")
                if merged_ranges:
                    logging.info(f"First 5 merged ranges: {merged_ranges[:5]}")
        
    except Exception as e:
        logging.error(f"Error opening workbook: {e}")

if __name__ == "__main__":
    # Provide one LIHTC application file for testing
    file_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/2023_4pct_R1_23-477.xlsx'
    
    # You can replace with any specific file you want to test
    test_excel_reading_with_merged_cells(file_path)