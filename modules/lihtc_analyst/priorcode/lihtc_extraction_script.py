#!/usr/bin/env python3
# LIHTC Application Data Extraction Script
# This script extracts data from LIHTC application Excel files (4% and 9%)
# and converts them to structured JSON files for RAG system usage.

import os
import re
import json
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_extraction.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class LIHTCApplicationExtractor:
    """
    Extracts data from LIHTC application Excel files and converts to JSON
    """
    
    def __init__(self, input_path, output_path_4p, output_path_9p, max_files=None):
        """
        Initialize the extractor with input and output paths
        
        Args:
            input_path (str): Path to directory containing the application Excel files
            output_path_4p (str): Path to directory for 4% JSON output
            output_path_9p (str): Path to directory for 9% JSON output
            max_files (int, optional): Maximum number of files to process for testing
        """
        self.input_path = Path(input_path)
        self.output_path_4p = Path(output_path_4p)
        self.output_path_9p = Path(output_path_9p)
        self.max_files = max_files
        
        # Create output directories if they don't exist
        self.output_path_4p.mkdir(parents=True, exist_ok=True)
        self.output_path_9p.mkdir(parents=True, exist_ok=True)
        
        # Dictionary to hold section mappings
        self.section_mapping = {}
        
    def extract_application_type_info(self, filename):
        """
        Extract application type, year, round, and CTCAC number from filename
        
        Args:
            filename (str): The filename to parse
            
        Returns:
            tuple: (year, app_type, round, ctcac_number)
        """
        # Example expected format: 2024_4pct_R1_24408.xlsx or 2024_9pct_R1_24001.xlsx
        pattern = r'(\d{4})_([\d\w]+)_R(\d+)_(\d+)'
        match = re.match(pattern, filename)
        
        if match:
            year, app_type, round_num, ctcac_num = match.groups()
            return year, app_type, round_num, ctcac_num
        else:
            # If pattern doesn't match, try to extract what we can
            parts = filename.split('_')
            year = parts[0] if len(parts) > 0 and parts[0].isdigit() else "Unknown"
            app_type = parts[1] if len(parts) > 1 else "Unknown"
            round_num = parts[2].replace('R', '') if len(parts) > 2 and 'R' in parts[2] else "Unknown"
            ctcac_num = parts[3].split('.')[0] if len(parts) > 3 else "Unknown"
            
            logger.warning(f"Filename {filename} did not match expected pattern. Extracted: {year}, {app_type}, {round_num}, {ctcac_num}")
            return year, app_type, round_num, ctcac_num
    
    def is_yellow_cell(self, cell):
        """
        Check if a cell has yellow fill (indicating an input cell)
        
        Args:
            cell: The openpyxl cell object
            
        Returns:
            bool: True if cell has yellow fill, False otherwise
        """
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            # Different yellow shades that might be used
            yellow_colors = ['FFFFFF00', 'FFFF00', 'FFFFC000']
            return cell.fill.start_color.rgb in yellow_colors
        return False
    
    def get_column_name(self, idx):
        """
        Convert column index to column letter (e.g., 1 -> A, 27 -> AA)
        
        Args:
            idx (int): The column index
            
        Returns:
            str: The column letter
        """
        return get_column_letter(idx)
    
    def clean_value(self, value):
        """
        Clean cell values for consistent output
        
        Args:
            value: The cell value
            
        Returns:
            str, int, float, or None: The cleaned value
        """
        if value is None or pd.isna(value) or value == '':
            return ""
        
        # Handle different value types
        if isinstance(value, (int, float)):
            # Keep integers as integers, format floats consistently
            if value == int(value):
                return int(value)
            return value
        
        # Convert to string and strip whitespace
        value = str(value).strip()
        
        # Handle placeholder values
        placeholders = ["(select)", "(select one)", "(specify here)", "N/A"]
        if value in placeholders:
            return value
        
        return value
    
    def find_section_header(self, row_idx, col_idx, worksheet):
        """
        Find the section header for a given cell
        
        Args:
            row_idx (int): The row index of the cell
            col_idx (int): The column index of the cell
            worksheet: The openpyxl worksheet
            
        Returns:
            str: The section header
        """
        # Look for headers up to 20 rows above the current row
        for i in range(row_idx, max(1, row_idx - 20), -1):
            header_cell = worksheet.cell(row=i, column=1)
            value = header_cell.value
            
            if value and isinstance(value, str) and "SECTION" in value:
                return value
        
        # If no specific section header found, return a default
        return "General Information"
    
    def extract_yellow_cells(self, worksheet):
        """
        Extract all yellow (input) cells from a worksheet
        
        Args:
            worksheet: The openpyxl worksheet
            
        Returns:
            dict: Dictionary of yellow cell data with cell references as keys
        """
        yellow_cells = {}
        
        # First, scan for section headers to build a mapping
        self.section_mapping = {}
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str) and ("SECTION" in cell.value or cell.value.startswith("III") or cell.value.startswith("II")):
                for col_idx in range(1, worksheet.max_column + 1):
                    self.section_mapping[(row_idx, col_idx)] = cell.value
        
        # Now extract yellow cells
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if self.is_yellow_cell(cell):
                    col_letter = self.get_column_name(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"
                    
                    # Get header from the same row, or from a nearby row if not available
                    header_cell = worksheet.cell(row=row_idx, column=1)
                    header = header_cell.value if header_cell.value else self.find_section_header(row_idx, col_idx, worksheet)
                    
                    # Find section by checking the section mapping
                    section = None
                    for i in range(row_idx, 0, -1):
                        if (i, 1) in self.section_mapping:
                            section = self.section_mapping[(i, 1)]
                            break
                    
                    if not section:
                        section = "General Information"
                    
                    yellow_cells[cell_ref] = {
                        "value": self.clean_value(cell.value),
                        "header": header if header else "",
                        "section": section
                    }
        
        return yellow_cells
    
    def extract_total_cells(self, worksheet):
        """
        Extract important total cells from the worksheet
        
        Args:
            worksheet: The openpyxl worksheet
            
        Returns:
            dict: Dictionary of total cell data with descriptive keys
        """
        total_cells = {}
        
        # Common total cells and their descriptions
        total_cell_patterns = [
            (r"Total # Units: \((.*?)\)", "Total # Units"),
            (r"Total: \((.*?)\)", "Total"),
            (r"Total # Low Income Units: \((.*?)\)", "Total # Low Income Units"),
            (r"Total # Market Rate Units: \((.*?)\)", "Total # Market Rate Units"),
            (r"Total Annual Potential Gross Income: \((.*?)\)", "Total Annual Potential Gross Income"),
            (r"Total Annual Residential Operating Expenses: \((.*?)\)", "Total Annual Residential Operating Expenses"),
            (r"Total Projected Annual Rental Subsidy: \((.*?)\)", "Total Projected Annual Rental Subsidy")
        ]
        
        # Look for cells containing total patterns
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    for pattern, description in total_cell_patterns:
                        match = re.search(pattern, cell.value)
                        if match:
                            # Extract the cell reference from the pattern match
                            cell_ref = match.group(1)
                            if ":" in cell_ref:
                                cell_ref = cell_ref.split(":")[0]
                            
                            # Extract column and row from the reference
                            col_match = re.match(r'([A-Z]+)', cell_ref)
                            row_match = re.search(r'(\d+)', cell_ref)
                            
                            if col_match and row_match:
                                col = col_match.group(1)
                                row = int(row_match.group(1))
                                
                                # Get the value from the referenced cell
                                value_cell = worksheet[f"{col}{row}"]
                                
                                # Find section by checking the section mapping
                                section = None
                                for i in range(row_idx, 0, -1):
                                    if (i, 1) in self.section_mapping:
                                        section = self.section_mapping[(i, 1)]
                                        break
                                
                                if not section:
                                    section = "General Information"
                                
                                # Store the cell data with the original description and cell reference
                                key = f"{cell.value}"
                                total_cells[key] = {
                                    "value": self.clean_value(value_cell.value),
                                    "section": section
                                }
        
        # Also look for specific cells that are commonly used as totals
        key_total_cells = {
            # Basic project information
            "G730": "Total # Units",
            "T730": "Total Monthly Rents",
            "O756": "Total # Manager Units",
            "AC756": "Total Manager Monthly Rents",
            "O776": "Total # Market Rate Units",
            "AC776": "Total Market Rate Monthly Rents",
            "Z788": "Total Projected Annual Rental Subsidy",
            "Z796": "Total Miscellaneous Income",
            "Z797": "Total Annual Potential Gross Income",
            
            # Operating expenses
            "W826": "Total Administrative",
            "W834": "Total Utilities",
            "W841": "Total Payroll / Payroll Taxes",
            "W851": "Total Maintenance",
            "W858": "Total Other Expenses",
            "AC862": "Total Annual Residential Operating Expenses",
            "AC863": "Total Number of Units in the Project",
            "AC865": "Total 3-Month Operating Reserve",
            "AC867": "Total Annual Services Amenities Budget",
            "AC868": "Total Annual Reserve for Replacement",
            "AC870": "Total Annual Real Estate Taxes",
            "Z880": "Total Annual Commercial/Non-Residential Net Income"
        }
        
        for cell_ref, description in key_total_cells.items():
            try:
                if cell_ref in worksheet:
                    value_cell = worksheet[cell_ref]
                    
                    # Find the section for this cell
                    row = int(re.search(r'(\d+)', cell_ref).group(1))
                    section = None
                    for i in range(row, 0, -1):
                        if (i, 1) in self.section_mapping:
                            section = self.section_mapping[(i, 1)]
                            break
                    
                    if not section:
                        section = "General Information"
                    
                    key = f"{description}: ({cell_ref})"
                    total_cells[key] = {
                        "value": self.clean_value(value_cell.value),
                        "section": section
                    }
            except Exception as e:
                logger.warning(f"Error extracting key total cell {cell_ref}: {e}")
        
        return total_cells
    
    def extract_threshold_basis_limits(self, worksheet):
        """
        Extract threshold basis limit data from the worksheet
        
        Args:
            worksheet: The openpyxl worksheet
            
        Returns:
            dict: Dictionary of threshold basis limit data
        """
        threshold_basis_limits = {}
        
        # Look for cells in the threshold basis limit section
        basis_limit_cells = {
            # SRO/Studio limits
            "R961": "SRO/STUDIO Basis Limit",
            # 1 Bedroom limits
            "R962": "1 Bedroom Basis Limit",
            # 2 Bedroom limits
            "R963": "2 Bedroom Basis Limit",
            "AB963": "2 Bedroom Units",
            "AJ963": "2 Bedroom Total",
            # 3 Bedroom limits
            "R964": "3 Bedroom Basis Limit",
            "AB964": "3 Bedroom Units",
            "AJ964": "3 Bedroom Total",
            # 4+ Bedroom limits
            "R965": "4+ Bedroom Basis Limit",
            # Project totals
            "AB966": "Total Units",
            "AJ967": "Total Basis Limit",
            # Adjustment factors
            "BA989": "Adjustment Factor 1",
            "BA990": "Adjustment Factor 2",
            "AJ991": "Adjustment Amount",
            "BA992": "Adjustment Factor 3",
            "AJ1020": "Total Adjusted Threshold Basis Limit",
            "X1024": "Total Eligible Basis",
            "X1025": "Percentage of Adjusted Threshold Basis Limit",
            # High cost test values
            "C1036": "High Cost Test Value 1",
            "C1042": "High Cost Test Value 2",
            "C1048": "High Cost Test Value 3"
        }
        
        for cell_ref, description in basis_limit_cells.items():
            try:
                if cell_ref in worksheet:
                    value_cell = worksheet[cell_ref]
                    threshold_basis_limits[cell_ref] = {
                        "description": description,
                        "value": self.clean_value(value_cell.value)
                    }
            except Exception as e:
                logger.warning(f"Error extracting threshold basis limit cell {cell_ref}: {e}")
        
        return threshold_basis_limits
    
    def extract_section_headers(self, worksheet):
        """
        Extract main section headers from the worksheet
        
        Args:
            worksheet: The openpyxl worksheet
            
        Returns:
            dict: Dictionary of section headers with cell references as keys
        """
        section_headers = {}
        
        # Look for main section headers in column A
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str) and (
                cell.value.startswith("II.") or 
                cell.value.startswith("III.") or
                cell.value.startswith("CALIFORNIA") or
                "SECTION" in cell.value
            ):
                cell_ref = f"A{row_idx}"
                section_headers[cell_ref] = cell.value
        
        return section_headers
    
    def extract_project_info(self, worksheet):
        """
        Extract basic project information from the worksheet
        
        Args:
            worksheet: The openpyxl worksheet
            
        Returns:
            dict: Dictionary of project information
        """
        project_info = {
            "name": "",
            "address": "",
            "city": "",
            "county": "",
            "zip_code": "",
            "census_tract": "",
            "total_units": 0,
            "low_income_units": 0,
            "manager_units": 0
        }
        
        # Look for specific cells containing project information
        project_info_cells = {
            "Project Name": ["A17", "A18", "B17", "B18", "C17", "C18"],
            "Site Address": ["F354", "G354", "H354", "I354", "J354"],
            "City": ["F357", "G357", "H357"],
            "County": ["R357", "S357", "T357"],
            "Zip Code": ["F358", "G358", "H358"],
            "Census Tract": ["R358", "S358", "T358"]
        }
        
        # Try to find values in potential cell locations
        for field, cell_refs in project_info_cells.items():
            for cell_ref in cell_refs:
                try:
                    if cell_ref in worksheet and worksheet[cell_ref].value:
                        value = self.clean_value(worksheet[cell_ref].value)
                        field_key = field.lower().replace(" ", "_")
                        if field_key in project_info:
                            project_info[field_key] = value
                            break
                except Exception as e:
                    logger.debug(f"Error checking cell {cell_ref} for {field}: {e}")
        
        # Try to extract unit counts from common locations
        unit_count_cells = {
            "total_units": ["G730", "AC863", "AB966"],
            "low_income_units": ["G687", "AG733"],
            "manager_units": ["O756", "T837"]
        }
        
        for field, cell_refs in unit_count_cells.items():
            for cell_ref in cell_refs:
                try:
                    if cell_ref in worksheet and worksheet[cell_ref].value:
                        value = self.clean_value(worksheet[cell_ref].value)
                        if isinstance(value, (int, float)):
                            project_info[field] = value
                            break
                except Exception as e:
                    logger.debug(f"Error checking cell {cell_ref} for {field}: {e}")
        
        return project_info
    
    def process_file(self, file_path):
        """
        Process a single application Excel file
        
        Args:
            file_path (Path): Path to the Excel file
            
        Returns:
            dict: Structured data extracted from the file
            str: The output file path where the JSON will be saved
        """
        logger.info(f"Processing file: {file_path}")
        filename = file_path.name
        
        # Extract application info from filename
        year, app_type, round_num, ctcac_num = self.extract_application_type_info(filename)
        
        # Determine output directory based on application type
        if "4pct" in app_type.lower() or "4p" in app_type.lower():
            output_dir = self.output_path_4p
        else:
            output_dir = self.output_path_9p
        
        # Generate output filename
        output_filename = f"{filename.split('.')[0]}.json"
        output_path = output_dir / output_filename
        
        # Load the workbook
        try:
            workbook = load_workbook(file_path, data_only=True)
            main_sheet = workbook.active  # Usually the first sheet contains the application
            
            # Extract data
            yellow_cells = self.extract_yellow_cells(main_sheet)
            total_cells = self.extract_total_cells(main_sheet)
            threshold_basis_limits = self.extract_threshold_basis_limits(main_sheet)
            section_headers = self.extract_section_headers(main_sheet)
            project_info = self.extract_project_info(main_sheet)
            
            # Create structured data
            data = {
                "metadata": {
                    "file_name": filename,
                    "extraction_date": datetime.now().isoformat(),
                    "application_year": year,
                    "application_type": app_type,
                    "application_round": round_num,
                    "ctcac_number": ctcac_num
                },
                "yellow_input_cells": yellow_cells,
                "total_cells": total_cells,
                "section_headers": section_headers,
                "threshold_basis_limits": threshold_basis_limits,
                "project": project_info
            }
            
            return data, output_path
            
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}", exc_info=True)
            return None, None
    
    def extract_data(self):
        """
        Extract data from all application files in the input directory
        """
        logger.info(f"Starting extraction from directory: {self.input_path}")
        
        # Get all Excel files in the input directory
        excel_files = [f for f in self.input_path.glob("*.xlsx") if not f.name.startswith("~$")]
        
        if self.max_files:
            excel_files = excel_files[:self.max_files]
        
        logger.info(f"Found {len(excel_files)} Excel files to process")
        
        processed_count = 0
        for file_path in excel_files:
            data, output_path = self.process_file(file_path)
            
            if data and output_path:
                # Save the JSON data
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2)
                
                processed_count += 1
                logger.info(f"Saved output to {output_path}")
        
        logger.info(f"Extraction complete. Processed {processed_count} files.")
        return processed_count

def main():
    """
    Main function to run the extraction
    """
    # Define paths
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p/application'
    output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p/application'
    
    # Number of files to process for testing (set to None for all files)
    max_files = 5  # Process only first 5 files for testing
    
    # Create extractor instance
    extractor = LIHTCApplicationExtractor(input_path, output_path_4p, output_path_9p, max_files)
    
    # Run extraction
    extractor.extract_data()

if __name__ == "__main__":
    main()
