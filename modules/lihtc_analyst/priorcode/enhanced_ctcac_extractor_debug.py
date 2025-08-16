#!/usr/bin/env python3
"""
Enhanced CTCAC LIHTC Application Data Extractor - Debug Version
Extracts comprehensive data from CTCAC applications for RAG system
"""

import pandas as pd
import json
import os
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from typing import Dict, List, Any, Optional
import re
import sys

def main():
    """Main function with debugging output"""
    
    print("🎯 CTCAC COMPREHENSIVE DATA EXTRACTOR - DEBUG VERSION")
    print("=" * 60)
    
    # Define file paths
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p'
    output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p'
    log_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/logs'
    
    print(f"📁 Input path: {input_path}")
    print(f"📁 4% output path: {output_path_4p}")
    print(f"📁 9% output path: {output_path_9p}")
    print(f"📁 Log path: {log_path}")
    print("=" * 60)
    
    # Check if input directory exists
    input_dir = Path(input_path)
    print(f"\n🔍 Checking input directory...")
    print(f"Directory exists: {input_dir.exists()}")
    
    if not input_dir.exists():
        print(f"❌ ERROR: Input directory does not exist!")
        print(f"Please check the path: {input_path}")
        return
    
    # Check for Excel files
    print(f"\n🔍 Looking for Excel files...")
    excel_files = list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls"))
    print(f"Found {len(excel_files)} Excel files")
    
    if len(excel_files) == 0:
        print("❌ No Excel files found in the input directory!")
        print("Let's check what files are in the directory:")
        
        # List all files in the directory
        all_files = list(input_dir.iterdir())
        print(f"Total files/folders in directory: {len(all_files)}")
        
        # Show first 10 files
        print("\nFirst 10 items in directory:")
        for i, file_path in enumerate(all_files[:10]):
            print(f"  {i+1}. {file_path.name} ({'file' if file_path.is_file() else 'folder'})")
        
        if len(all_files) > 10:
            print(f"  ... and {len(all_files) - 10} more items")
        
        return
    
    # Show some Excel files found
    print(f"\nFirst few Excel files found:")
    for i, file_path in enumerate(excel_files[:5]):
        print(f"  {i+1}. {file_path.name}")
    
    if len(excel_files) > 5:
        print(f"  ... and {len(excel_files) - 5} more files")
    
    print(f"\n🚀 Ready to create extractor and process files...")
    
    try:
        # Create extractor instance
        print("Creating extractor instance...")
        extractor = CTCACComprehensiveExtractor(input_path, output_path_4p, output_path_9p, log_path)
        print("✅ Extractor created successfully!")
        
        # Run in test mode with 3 files
        print("\n🧪 Starting test mode processing...")
        extractor.process_files(limit=3, test_mode=True)
        
        print("\n🎉 Processing complete!")
        
    except Exception as e:
        print(f"❌ ERROR creating or running extractor: {e}")
        import traceback
        traceback.print_exc()

class CTCACComprehensiveExtractor:
    """
    Comprehensive extractor that captures ALL valuable data from CTCAC applications
    """
    
    def __init__(self, input_path: str, output_path_4p: str, output_path_9p: str, log_path: str):
        """Initialize the extractor with file paths"""
        print(f"📝 Initializing extractor...")
        
        self.input_path = Path(input_path)
        self.output_path_4p = Path(output_path_4p)
        self.output_path_9p = Path(output_path_9p)
        self.log_path = Path(log_path)
        
        print(f"   Input path: {self.input_path}")
        print(f"   4% output: {self.output_path_4p}")
        print(f"   9% output: {self.output_path_9p}")
        print(f"   Log path: {self.log_path}")
        
        # Create directories if they don't exist
        print(f"📁 Creating output directories...")
        self.output_path_4p.mkdir(parents=True, exist_ok=True)
        self.output_path_9p.mkdir(parents=True, exist_ok=True)
        self.log_path.mkdir(parents=True, exist_ok=True)
        print(f"✅ Directories created/verified")
        
        # Set up logging
        print(f"📋 Setting up logging...")
        self.setup_logging()
        print(f"✅ Logging configured")
        
        # Define comprehensive sheet mapping
        self.target_sheets = {
            'application': 'Application',
            'sources_uses': 'Sources and Uses Budget',
            'sources_basis_breakdown': 'Sources and Basis Breakdown',
            'points_system': 'Points System',
            'basis_credits': 'Basis & Credits',
            'tie_breaker': 'Tie Breaker',
            'final_tie_breaker': 'Final Tie Breaker ',
            'disaster_tie_breaker': 'Disaster Credit Tie Breaker',
            'pro_forma': '15 Year Pro Forma',
            'sce_basis': 'SCE Basis and Credits',
            'fce_basis': 'FCE Basis and Credits',
            'calHFA_addendum': 'CalHFA Addendum',
            'subsidy_contract': 'Subsidy Contract Calculation'
        }
        
        print(f"✅ Extractor initialization complete!")
    
    def setup_logging(self):
        """Set up logging configuration"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_path / f"ctcac_comprehensive_extraction_{timestamp}.log"
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        
        # Clear any existing handlers
        self.logger.handlers.clear()
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.logger.propagate = False
        
        self.logger.info(f"Comprehensive extraction logging initialized. Log file: {log_file}")
    
    def parse_filename(self, filename: str) -> Dict[str, str]:
        """Parse filename to extract metadata"""
        base_name = Path(filename).stem
        
        parsed = {
            'year': 'unknown',
            'credit_type': 'unknown',
            'round': 'unknown',
            'project_id': 'unknown',
            'original_filename': filename
        }
        
        try:
            # Pattern for both formats: YYYY_XpctXX_RX_YY-XXX and YYYY_XpctXX_RX_CAYYXXX
            pattern = r'(\d{4})_(\d+pct)_([rR]\d+)_(?:CA)?(.+)'
            match = re.match(pattern, base_name)
            
            if match:
                parsed['year'] = match.group(1)
                parsed['credit_type'] = match.group(2)
                parsed['round'] = match.group(3).upper()
                parsed['project_id'] = match.group(4)
        
        except Exception as e:
            self.logger.warning(f"Error parsing filename {filename}: {e}")
        
        return parsed
    
    def determine_output_path(self, parsed_info: Dict[str, str]) -> Path:
        """Determine correct output path based on credit type"""
        if '4' in parsed_info['credit_type']:
            return self.output_path_4p
        else:
            return self.output_path_9p
    
    def create_output_filename(self, parsed_info: Dict[str, str]) -> str:
        """Create standardized output filename"""
        return f"{parsed_info['year']}_{parsed_info['credit_type']}_{parsed_info['round']}_CA-{parsed_info['project_id']}.json"
    
    def get_cell_value(self, sheet, cell_ref: str):
        """Safely get cell value from sheet"""
        try:
            cell = sheet[cell_ref]
            return cell.value if cell and cell.value is not None else None
        except:
            return None
    
    def find_sheet_name(self, workbook, target_name: str) -> Optional[str]:
        """Find the actual sheet name (handles slight variations)"""
        if target_name in workbook.sheetnames:
            return target_name
        
        # Look for close matches
        for sheet_name in workbook.sheetnames:
            if target_name.lower().replace(' ', '') in sheet_name.lower().replace(' ', ''):
                return sheet_name
        return None
    
    def extract_basic_project_info(self, workbook) -> Dict[str, Any]:
        """Extract basic project information for testing"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Application')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            project_info = {
                'ctcac_applicant': self.get_cell_value(sheet, 'H16'),
                'project_name': self.get_cell_value(sheet, 'H18'),
                'project_name_alt': self.get_cell_value(sheet, 'H17'),
            }
            
            # Clean up None values
            project_info = {k: v for k, v in project_info.items() if v is not None}
            
            return project_info
            
        except Exception as e:
            self.logger.error(f"Error extracting basic project info: {e}")
            return {}
    
    def process_single_file(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Process a single Excel file - simplified for testing"""
        try:
            self.logger.info(f"📂 Processing file: {file_path.name}")
            
            parsed_info = self.parse_filename(file_path.name)
            
            # Load workbook
            self.logger.info("   Loading workbook...")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            self.logger.info(f"   📋 Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Initialize result structure
            result = {
                'metadata': {
                    'filename': file_path.name,
                    'processed_date': datetime.now().isoformat(),
                    'year': parsed_info['year'],
                    'credit_type': parsed_info['credit_type'],
                    'round': parsed_info['round'],
                    'project_id': parsed_info['project_id'],
                    'available_sheets': workbook.sheetnames,
                    'extraction_method': 'basic_test',
                    'extractor_version': '2.1_debug'
                },
                'project_information': {},
                'extraction_summary': {}
            }
            
            # Extract basic project information
            self.logger.info("   🏢 Extracting basic project information...")
            result['project_information'] = self.extract_basic_project_info(workbook)
            
            workbook.close()
            
            # Create extraction summary
            result['extraction_summary'] = {
                'sections_extracted': 1 if result['project_information'] else 0,
                'total_sections': 1,
                'success_rate': 100.0 if result['project_information'] else 0.0,
                'test_mode': True
            }
            
            self.logger.info(f"   ✅ Basic extraction complete")
            
            return result
            
        except Exception as e:
            self.logger.error(f"   💥 Error processing file {file_path.name}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def process_files(self, limit: int = None, test_mode: bool = True):
        """Process files - simplified for testing"""
        print("🎯 CTCAC COMPREHENSIVE EXTRACTOR - TEST MODE")
        print("=" * 50)
        
        if test_mode and limit is None:
            limit = 3
            print(f"🧪 TEST MODE: Processing up to {limit} files")
        elif limit:
            print(f"📊 LIMITED MODE: Processing up to {limit} files")
        else:
            print("🔄 FULL MODE: Processing all files")
        
        print("=" * 50)
        
        # Find Excel files
        print("🔍 Finding Excel files...")
        excel_files = list(self.input_path.glob("*.xlsx")) + list(self.input_path.glob("*.xls"))
        
        if not excel_files:
            self.logger.error(f"❌ No Excel files found in {self.input_path}")
            print(f"❌ No Excel files found in {self.input_path}")
            return
        
        total_files = len(excel_files)
        self.logger.info(f"📁 Found {total_files} Excel files in input directory")
        print(f"📁 Found {total_files} Excel files")
        
        if limit:
            excel_files = excel_files[:limit]
            self.logger.info(f"🎯 Processing {len(excel_files)} files (limited)")
            print(f"🎯 Processing {len(excel_files)} files (limited)")
        
        processed_count = 0
        failed_count = 0
        
        print(f"\n📈 Processing Progress:")
        print("-" * 30)
        
        for i, file_path in enumerate(excel_files, 1):
            try:
                progress = f"[{i}/{len(excel_files)}]"
                print(f"\n{progress} {file_path.name}")
                
                result = self.process_single_file(file_path)
                
                if result:
                    # Parse filename for organized output
                    parsed_info = self.parse_filename(file_path.name)
                    output_path = self.determine_output_path(parsed_info)
                    json_filename = self.create_output_filename(parsed_info)
                    json_path = output_path / json_filename
                    
                    # Save as JSON
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(result, f, indent=2, ensure_ascii=False)
                    
                    self.logger.info(f"   💾 Saved: {json_filename}")
                    processed_count += 1
                    
                    print(f"   ✅ SUCCESS - Saved to: {json_filename}")
                else:
                    failed_count += 1
                    print(f"   ❌ FAILED")
                    
            except Exception as e:
                self.logger.error(f"   💥 Failed to process {file_path.name}: {e}")
                failed_count += 1
                print(f"   ❌ FAILED: {str(e)[:50]}...")
        
        # Final summary
        print("\n" + "=" * 50)
        print("📊 PROCESSING SUMMARY")
        print("=" * 50)
        print(f"✅ Successfully processed: {processed_count}")
        print(f"❌ Failed: {failed_count}")
        print(f"📁 Total files attempted: {len(excel_files)}")
        
        if len(excel_files) > 0:
            success_rate = (processed_count/len(excel_files)*100)
            print(f"📈 Success rate: {success_rate:.1f}%")
        
        print(f"\n💾 JSON files saved to:")
        print(f"  4% applications: {self.output_path_4p}")
        print(f"  9% applications: {self.output_path_9p}")
        print(f"📋 Log files saved to: {self.log_path}")
        
        self.logger.info(f"Test extraction complete. Success: {processed_count}, Failed: {failed_count}")

if __name__ == "__main__":
    main()