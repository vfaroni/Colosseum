#!/usr/bin/env python3
"""
BOTN-HFC Integration Tool
Integrates CoStar data and AMI analysis into BOTN underwriting template
Specifically designed for Texas HFC distressed property analysis
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BOTNHFCIntegrator:
    def __init__(self):
        # File paths
        self.botn_template_path = "/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Workforce_Housing/BOTN_Proformas/Workforce BOTN 05.12.25.xlsx"
        self.hfc_analysis_path = "/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Workforce_Housing/TX WFH/HFC_Project/analysis_outputs"
        self.output_path = "/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Workforce_Housing/TX WFH/HFC_Project/analysis_outputs"
        
        # Create output directory
        Path(self.output_path).mkdir(parents=True, exist_ok=True)
        
        self.hfc_properties = None
        self.botn_template = None
        
    def load_hfc_property_data(self):
        """Load HFC property analysis results"""
        try:
            # Find the most recent HFC analysis file
            analysis_files = list(Path(self.hfc_analysis_path).glob("HFC_Property_Analysis_*.xlsx"))
            if not analysis_files:
                logger.error("No HFC analysis files found")
                return False
            
            latest_file = max(analysis_files, key=lambda x: x.stat().st_mtime)
            self.hfc_properties = pd.read_excel(latest_file, sheet_name='Property_Analysis')
            logger.info(f"Loaded HFC data from: {latest_file.name}")
            logger.info(f"Properties loaded: {len(self.hfc_properties)}")
            
            return True
        except Exception as e:
            logger.error(f"Error loading HFC property data: {e}")
            return False
    
    def load_botn_template(self):
        """Load BOTN template structure"""
        try:
            self.botn_template = openpyxl.load_workbook(self.botn_template_path)
            logger.info("BOTN template loaded successfully")
            return True
        except Exception as e:
            logger.error(f"Error loading BOTN template: {e}")
            return False
    
    def create_property_botn_analysis(self, property_data, costar_data=None):
        """Create BOTN analysis for a single HFC property"""
        try:
            # Create a copy of the template
            wb = openpyxl.load_workbook(self.botn_template_path)
            inputs_sheet = wb['Inputs']
            
            # Update basic property information
            property_updates = {
                'B2': property_data.get('property_name', ''),  # Property Name
                'B4': property_data.get('address', ''),        # Address
                'B6': 'TX',                                     # State (change from CA)
                'B7': f"{property_data.get('county', '')} County",  # County
                'B8': 'LIHTC',                                  # Keep LIHTC designation
                'B10': 'Manual',                                # Rent Type
            }
            
            # Apply basic updates
            for cell_ref, value in property_updates.items():
                inputs_sheet[cell_ref] = value
            
            # Add HFC-specific data if available
            if costar_data:
                self.integrate_costar_data(inputs_sheet, costar_data)
            
            # Add AMI analysis
            self.add_ami_analysis(inputs_sheet, property_data)
            
            # Add HFC risk analysis
            self.add_hfc_risk_analysis(inputs_sheet, property_data)
            
            # Save property-specific file
            property_name = property_data.get('property_name', 'Unknown').replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"{self.output_path}/BOTN_{property_name}_{timestamp}.xlsx"
            wb.save(output_file)
            
            logger.info(f"Created BOTN analysis for {property_data.get('property_name')}: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Error creating BOTN analysis for {property_data.get('property_name')}: {e}")
            return None
    
    def integrate_costar_data(self, inputs_sheet, costar_data):
        """Integrate CoStar financial data into BOTN inputs"""
        try:
            # Financial data integration
            financial_updates = {}
            
            # Current rents (if available)
            if 'avg_rent_2br' in costar_data:
                financial_updates['B25'] = costar_data['avg_rent_2br']  # 2BR rent
            
            # Occupancy rate
            if 'occupancy_rate' in costar_data:
                financial_updates['B30'] = costar_data['occupancy_rate']  # Occupancy %
            
            # Property tax information
            if 'annual_property_tax' in costar_data:
                financial_updates['B35'] = costar_data['annual_property_tax']  # Current tax
            
            # NOI information
            if 'net_operating_income' in costar_data:
                financial_updates['B40'] = costar_data['net_operating_income']  # T12 NOI
            
            # Apply updates
            for cell_ref, value in financial_updates.items():
                inputs_sheet[cell_ref] = value
                
        except Exception as e:
            logger.error(f"Error integrating CoStar data: {e}")
    
    def add_ami_analysis(self, inputs_sheet, property_data):
        """Add AMI rent analysis to BOTN template"""
        try:
            # AMI rent data (from our analysis)
            ami_updates = {}
            
            if property_data.get('50_ami_2br'):
                ami_updates['D25'] = property_data['50_ami_2br']  # 50% AMI 2BR
            
            if property_data.get('60_ami_2br'):
                ami_updates['D26'] = property_data['60_ami_2br']  # 60% AMI 2BR
            
            if property_data.get('80_ami_2br'):
                ami_updates['D27'] = property_data['80_ami_2br']  # 80% AMI 2BR
            
            # Calculate rent gaps
            current_rent = self.parse_rent_midpoint(property_data.get('current_rent_range', ''))
            if current_rent and property_data.get('60_ami_2br'):
                gap_60_ami = current_rent - property_data['60_ami_2br']
                ami_updates['D28'] = gap_60_ami  # Gap to 60% AMI
            
            # Apply AMI updates
            for cell_ref, value in ami_updates.items():
                inputs_sheet[cell_ref] = value
                
        except Exception as e:
            logger.error(f"Error adding AMI analysis: {e}")
    
    def add_hfc_risk_analysis(self, inputs_sheet, property_data):
        """Add HFC-specific risk analysis"""
        try:
            # HFC risk data
            hfc_updates = {
                'F25': property_data.get('hb21_risk_level', ''),      # HB21 Risk Level
                'F26': property_data.get('conversion_priority', ''),  # Priority ranking
                'F27': 'January 1, 2027',                            # HB21 deadline
            }
            
            # LIHTC eligibility flags
            if property_data.get('lihtc_low_poverty_bonus'):
                hfc_updates['F28'] = 'Eligible'  # Low poverty bonus
            else:
                hfc_updates['F28'] = 'Not Eligible'
            
            # Poverty rate
            if property_data.get('poverty_rate'):
                hfc_updates['F29'] = f"{property_data['poverty_rate']:.1f}%"
            
            # Apply HFC updates
            for cell_ref, value in hfc_updates.items():
                inputs_sheet[cell_ref] = value
                
        except Exception as e:
            logger.error(f"Error adding HFC risk analysis: {e}")
    
    def parse_rent_midpoint(self, rent_range_str):
        """Parse rent range string to get midpoint"""
        try:
            if not rent_range_str or rent_range_str == 'TBD':
                return None
            
            # Remove $ and , and split on -
            cleaned = str(rent_range_str).replace('$', '').replace(',', '')
            if '-' in cleaned:
                parts = cleaned.split('-')
                if len(parts) == 2:
                    low = float(parts[0])
                    high = float(parts[1])
                    return (low + high) / 2
            return None
        except:
            return None
    
    def create_portfolio_summary(self):
        """Create portfolio-level summary analysis"""
        try:
            # Create summary workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "HFC Portfolio Summary"
            
            # Headers
            headers = [
                'Property Name', 'County', 'Units', 'HB21 Risk', 'Priority',
                'Current Rent Range', '60% AMI Rent', 'Rent Gap', 'Poverty Rate',
                'LIHTC Eligible', 'Conversion Feasibility'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Property data
            for row, (_, property_data) in enumerate(self.hfc_properties.iterrows(), 2):
                current_rent = self.parse_rent_midpoint(property_data.get('current_rent_range', ''))
                ami_60_rent = property_data.get('60_ami_2br', 0)
                rent_gap = current_rent - ami_60_rent if current_rent and ami_60_rent else 0
                
                # Determine conversion feasibility
                if rent_gap <= 0:
                    feasibility = "Highly Feasible"
                elif rent_gap <= 200:
                    feasibility = "Feasible with Support"
                elif rent_gap <= 400:
                    feasibility = "Challenging"
                else:
                    feasibility = "Deep Affordability Required"
                
                row_data = [
                    property_data.get('property_name', ''),
                    property_data.get('county', ''),
                    property_data.get('units', ''),
                    property_data.get('hb21_risk_level', ''),
                    property_data.get('conversion_priority', ''),
                    property_data.get('current_rent_range', ''),
                    ami_60_rent,
                    rent_gap,
                    property_data.get('poverty_rate', ''),
                    'Yes' if property_data.get('lihtc_low_poverty_bonus') else 'No',
                    feasibility
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Save portfolio summary
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            summary_file = f"{self.output_path}/HFC_Portfolio_BOTN_Summary_{timestamp}.xlsx"
            wb.save(summary_file)
            
            logger.info(f"Portfolio summary created: {summary_file}")
            return summary_file
            
        except Exception as e:
            logger.error(f"Error creating portfolio summary: {e}")
            return None
    
    def run_full_integration(self, costar_data_dict=None):
        """Run complete BOTN-HFC integration"""
        if not self.load_hfc_property_data():
            logger.error("Failed to load HFC property data")
            return
        
        if not self.load_botn_template():
            logger.error("Failed to load BOTN template")
            return
        
        logger.info("Starting BOTN-HFC integration...")
        
        # Create individual property analyses
        created_files = []
        
        for _, property_data in self.hfc_properties.iterrows():
            property_name = property_data.get('property_name', '')
            
            # Get CoStar data for this property if provided
            costar_data = None
            if costar_data_dict and property_name in costar_data_dict:
                costar_data = costar_data_dict[property_name]
            
            # Create BOTN analysis
            output_file = self.create_property_botn_analysis(property_data, costar_data)
            if output_file:
                created_files.append(output_file)
        
        # Create portfolio summary
        summary_file = self.create_portfolio_summary()
        if summary_file:
            created_files.append(summary_file)
        
        logger.info(f"Integration complete. Created {len(created_files)} files:")
        for file in created_files:
            logger.info(f"  - {Path(file).name}")
        
        return created_files

def main():
    """Main execution function"""
    integrator = BOTNHFCIntegrator()
    
    # Example CoStar data structure (to be populated with actual data)
    example_costar_data = {
        '2828 Royal Oaks': {
            'avg_rent_2br': 1500,
            'occupancy_rate': 0.92,
            'annual_property_tax': 850000,
            'net_operating_income': 4200000
        },
        'Shannon Creek': {
            'avg_rent_2br': 1200,
            'occupancy_rate': 0.95,
            'annual_property_tax': 0,  # Currently HFC exempt
            'net_operating_income': 5800000
        }
    }
    
    # Run integration
    results = integrator.run_full_integration(example_costar_data)
    
    if results:
        print(f"\nBOTN-HFC Integration completed successfully!")
        print(f"Created {len(results)} analysis files")
        print(f"Files saved to: {integrator.output_path}")
    else:
        print("Integration failed. Check logs for details.")

if __name__ == "__main__":
    main()