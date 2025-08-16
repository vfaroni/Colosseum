"""
Create a user-friendly Excel workbook for D'Marco
Simple, color-coded, and ready to use
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

def create_dmarco_workbook():
    """Create a simplified, user-friendly Excel for D'Marco"""
    
    # Load the analysis data
    print("Loading property data...")
    df = pd.read_excel('CoStar_Land_Analysis_20250616_203751.xlsx', sheet_name='All_Land_Analysis')
    
    # Create a new workbook
    output_file = f'DMarco_Properties_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Hot Leads (Top 25 properties)
        print("Creating Hot Leads sheet...")
        hot_leads_df = df.nlargest(25, 'Land_Viability_Score')[
            ['Address', 'City', 'Sale Price', 'Land_Viability_Score', 
             'Listing Broker Company', 'Listing Broker Phone',
             'TDHCA_One_Mile_Fatal', 'FEMA_Insurance_Risk']
        ].copy()
        
        # Add action columns
        hot_leads_df.insert(0, 'Priority', 'HOT')
        hot_leads_df.insert(1, 'Contact Today?', 'YES')
        hot_leads_df.insert(len(hot_leads_df.columns), 'Your Notes', '')
        hot_leads_df.insert(len(hot_leads_df.columns), 'Last Contact', '')
        hot_leads_df.insert(len(hot_leads_df.columns), 'Next Step', 'Call Broker')
        
        hot_leads_df.to_excel(writer, sheet_name='🔥 HOT LEADS - CALL TODAY', index=False)
        
        # Sheet 2: All Properties (simplified)
        print("Creating All Properties sheet...")
        all_props_df = df[[
            'Address', 'City', 'Sale Price', 'Land_Viability_Score',
            'Listing Broker Company', 'Listing Broker Phone',
            'QCT_Status', 'DDA_Status', 'TDHCA_One_Mile_Fatal', 
            'FEMA_Zone', 'Size (SF)'
        ]].copy()
        
        # Add user columns
        all_props_df.insert(0, 'Lead Status', 'NEW')
        all_props_df.insert(len(all_props_df.columns), 'Your Notes', '')
        all_props_df.insert(len(all_props_df.columns), 'Last Contact', '')
        
        all_props_df.to_excel(writer, sheet_name='All Properties', index=False)
        
        # Sheet 3: Simple Instructions
        print("Creating Instructions sheet...")
        instructions = pd.DataFrame({
            'D\'Marco\'s Quick Guide': [
                '🎯 YOUR DAILY WORKFLOW:',
                '',
                '1. START HERE: Open "HOT LEADS" tab',
                '2. CALL: Work through the list top to bottom',
                '3. UPDATE: Add notes after each call',
                '4. TRACK: Change Lead Status (HOT/WARM/COLD)',
                '',
                '📞 WHAT TO SAY TO BROKERS:',
                '',
                '"Hi, I\'m calling about [ADDRESS]. Is it still available?"',
                '"Why is the seller selling?" (Listen for motivation!)',
                '"Are they flexible on price?"',
                '"How quickly can they close?"',
                '',
                '🟢 GOOD SIGNS (Mark as HOT):',
                '- Seller needs quick sale',
                '- Broker responds immediately',
                '- Price negotiable',
                '- No other offers yet',
                '',
                '🔴 RED FLAGS (Mark as COLD):',
                '- Already under contract',
                '- Seller not motivated',
                '- Environmental issues',
                '- Broker unresponsive',
                '',
                '💡 QUICK TIPS:',
                '- Best call times: 10am-12pm, 2pm-4pm',
                '- Always text if no answer',
                '- Green cells = GOOD',
                '- Red cells = BAD',
                '- Yellow cells = CAUTION',
                '',
                '📊 UNDERSTANDING THE SCORES:',
                '- Score 90-100 = Excellent opportunity',
                '- Score 80-89 = Good opportunity',
                '- Score 70-79 = Has challenges',
                '- Below 70 = Probably skip',
                '',
                'Remember: Every NO gets you closer to a YES!'
            ]
        })
        instructions.to_excel(writer, sheet_name='📚 START HERE - Instructions', index=False)
        
    # Now apply formatting
    print("Applying formatting and colors...")
    wb = openpyxl.load_workbook(output_file)
    
    # Format Hot Leads sheet
    ws_hot = wb['🔥 HOT LEADS - CALL TODAY']
    format_hot_leads_sheet(ws_hot)
    
    # Format All Properties sheet
    ws_all = wb['All Properties']
    format_all_properties_sheet(ws_all)
    
    # Format Instructions
    ws_inst = wb['📚 START HERE - Instructions']
    format_instructions_sheet(ws_inst)
    
    # Save
    wb.save(output_file)
    print(f"\n✅ Created: {output_file}")
    print("\n📱 This file is ready for D'Marco to use immediately!")
    print("   - No software needed")
    print("   - Works on phone/tablet/computer")
    print("   - Color-coded for easy scanning")
    print("   - Simple action items included")

def format_hot_leads_sheet(ws):
    """Format the hot leads sheet with colors and styling"""
    
    # Header formatting
    header_fill = PatternFill(start_color='FF3B30', end_color='FF3B30', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    column_widths = {
        'A': 12,  # Priority
        'B': 15,  # Contact Today
        'C': 40,  # Address
        'D': 15,  # City
        'E': 15,  # Price
        'F': 12,  # Score
        'G': 25,  # Broker Company
        'H': 20,  # Broker Phone
        'I': 15,  # Fatal Flaw
        'J': 15,  # Flood Risk
        'K': 30,  # Your Notes
        'L': 15,  # Last Contact
        'M': 20,  # Next Step
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Priority column (RED)
    red_fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
    red_font = Font(bold=True, color='FF3B30')
    
    for row in range(2, ws.max_row + 1):
        ws[f'A{row}'].fill = red_fill
        ws[f'A{row}'].font = red_font
        ws[f'B{row}'].fill = red_fill
        ws[f'B{row}'].font = red_font
    
    # Score column - color scale (red to green)
    ws.conditional_formatting.add(f'F2:F{ws.max_row}',
        ColorScaleRule(start_type='num', start_value=70, start_color='FFCCCC',
                      mid_type='num', mid_value=85, mid_color='FFFFCC',
                      end_type='num', end_value=100, end_color='CCFFCC'))
    
    # Fatal flaw column - highlight if TRUE
    red_text = Font(color='FF0000', bold=True)
    ws.conditional_formatting.add(f'I2:I{ws.max_row}',
        CellIsRule(operator='equal', formula=['TRUE'], font=red_text))
    
    # Price formatting
    for row in range(2, ws.max_row + 1):
        ws[f'E{row}'].number_format = '$#,##0'
    
    # Phone number formatting - make it blue and underlined for easy clicking
    phone_font = Font(color='007AFF', underline='single')
    for row in range(2, ws.max_row + 1):
        ws[f'H{row}'].font = phone_font
    
    # Add borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def format_all_properties_sheet(ws):
    """Format the all properties sheet"""
    
    # Header formatting
    header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 15  # Lead Status
    ws.column_dimensions['B'].width = 40  # Address
    ws.column_dimensions['C'].width = 15  # City
    ws.column_dimensions['D'].width = 15  # Price
    ws.column_dimensions['E'].width = 12  # Score
    
    # Conditional formatting for Lead Status column
    # HOT = Red
    ws.conditional_formatting.add('A2:A1000',
        CellIsRule(operator='equal', formula=['"HOT"'],
                  fill=PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid'),
                  font=Font(color='FF3B30', bold=True)))
    
    # WARM = Orange
    ws.conditional_formatting.add('A2:A1000',
        CellIsRule(operator='equal', formula=['"WARM"'],
                  fill=PatternFill(start_color='FFF3E5', end_color='FFF3E5', fill_type='solid'),
                  font=Font(color='FF9500', bold=True)))
    
    # COLD = Blue
    ws.conditional_formatting.add('A2:A1000',
        CellIsRule(operator='equal', formula=['"COLD"'],
                  fill=PatternFill(start_color='E5F2FF', end_color='E5F2FF', fill_type='solid'),
                  font=Font(color='007AFF', bold=True)))
    
    # Score column - color scale
    ws.conditional_formatting.add(f'E2:E{ws.max_row}',
        ColorScaleRule(start_type='num', start_value=0, start_color='FFCCCC',
                      mid_type='num', mid_value=80, mid_color='FFFFCC',
                      end_type='num', end_value=100, end_color='CCFFCC'))
    
    # QCT/DDA columns - green if YES
    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    
    for col in ['G', 'H']:  # QCT and DDA columns
        ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    
    # Fatal flaw column - red if TRUE
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    ws.conditional_formatting.add(f'I2:I{ws.max_row}',
        CellIsRule(operator='equal', formula=['TRUE'], fill=red_fill))
    
    # Price formatting
    for row in range(2, ws.max_row + 1):
        ws[f'D{row}'].number_format = '$#,##0'

def format_instructions_sheet(ws):
    """Format the instructions sheet"""
    
    # Set column width
    ws.column_dimensions['A'].width = 80
    
    # Format different sections
    for row in range(1, ws.max_row + 1):
        cell = ws[f'A{row}']
        value = str(cell.value)
        
        # Headers
        if '🎯' in value or '📞' in value or '🟢' in value or '🔴' in value or '💡' in value or '📊' in value:
            cell.font = Font(bold=True, size=14, color='007AFF')
            cell.fill = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
        
        # Important lines
        elif value.startswith('"'):
            cell.font = Font(italic=True, color='FF3B30')
            cell.fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
        
        # Tips
        elif 'Score' in value and '-' in value:
            cell.font = Font(bold=True)
        
        # Final message
        elif 'Remember:' in value:
            cell.font = Font(bold=True, size=12, color='34C759')
            cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
        
        # Alignment
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add some row height for readability
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

if __name__ == "__main__":
    create_dmarco_workbook()