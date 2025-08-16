#!/usr/bin/env python3
"""
Analyze BOTN Workforce Housing Underwriting Template
This script reads and analyzes the Excel underwriting template structure
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime
from pathlib import Path

class BOTNUnderwritingAnalyzer:
    def __init__(self, template_path):
        self.template_path = template_path
        self.workbook = None
        self.analysis_results = {
            'template_info': {},
            'sheet_analysis': {},
            'inputs_structure': {},
            'output_structure': {},
            'data_connections': {}
        }
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            print(f"Loading workbook: {self.template_path}")
            self.workbook = load_workbook(self.template_path, read_only=True, data_only=True)
            
            # Get basic info
            self.analysis_results['template_info'] = {
                'file_name': Path(self.template_path).name,
                'sheet_count': len(self.workbook.sheetnames),
                'sheet_names': self.workbook.sheetnames,
                'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            print(f"Found {len(self.workbook.sheetnames)} sheets: {', '.join(self.workbook.sheetnames)}")
            return True
            
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def analyze_inputs_sheet(self):
        """Analyze the Inputs sheet structure"""
        try:
            sheet = self.workbook['Inputs']
            print("\n=== ANALYZING INPUTS SHEET ===")
            
            # Find key input sections by scanning first column
            input_sections = {}
            current_section = None
            
            for row in range(1, min(sheet.max_row + 1, 200)):  # Scan first 200 rows
                cell_a = sheet.cell(row=row, column=1).value
                cell_b = sheet.cell(row=row, column=2).value
                
                if cell_a:
                    # Look for section headers (typically in bold or different formatting)
                    if isinstance(cell_a, str) and not cell_b:
                        # Likely a section header
                        current_section = cell_a.strip()
                        input_sections[current_section] = {'row_start': row, 'items': []}
                    elif current_section and cell_b is not None:
                        # Data row within a section
                        input_sections[current_section]['items'].append({
                            'label': str(cell_a),
                            'value': cell_b,
                            'row': row
                        })
            
            # Store analysis
            self.analysis_results['inputs_structure'] = {
                'sections': list(input_sections.keys()),
                'section_details': {}
            }
            
            # Analyze each section
            for section, data in input_sections.items():
                if data['items']:
                    print(f"\nSection: {section}")
                    print(f"  Items: {len(data['items'])}")
                    # Show first few items
                    for item in data['items'][:3]:
                        print(f"    - {item['label']}: {item['value']}")
                    
                    self.analysis_results['inputs_structure']['section_details'][section] = {
                        'item_count': len(data['items']),
                        'sample_items': data['items'][:5]
                    }
            
        except Exception as e:
            print(f"Error analyzing Inputs sheet: {e}")
    
    def analyze_output_sheet(self):
        """Analyze the Output sheet structure"""
        try:
            sheet = self.workbook['Output']
            print("\n=== ANALYZING OUTPUT SHEET ===")
            
            # Find key output sections
            output_sections = {}
            
            # Scan for common financial model sections
            keywords = ['revenue', 'income', 'expense', 'noi', 'cash flow', 'debt', 'equity', 
                       'return', 'irr', 'npv', 'yield', 'cap rate', 'dscr', 'ltv']
            
            for row in range(1, min(sheet.max_row + 1, 200)):
                cell_a = sheet.cell(row=row, column=1).value
                if cell_a and isinstance(cell_a, str):
                    cell_lower = cell_a.lower()
                    for keyword in keywords:
                        if keyword in cell_lower:
                            # Found a relevant section
                            section_name = cell_a.strip()
                            output_sections[section_name] = {'row': row, 'values': []}
                            
                            # Get values from the row
                            for col in range(2, min(sheet.max_column + 1, 20)):
                                val = sheet.cell(row=row, column=col).value
                                if val is not None:
                                    output_sections[section_name]['values'].append(val)
                            break
            
            self.analysis_results['output_structure'] = {
                'sections_found': list(output_sections.keys()),
                'section_count': len(output_sections)
            }
            
            print(f"\nFound {len(output_sections)} key output sections:")
            for section in list(output_sections.keys())[:10]:
                print(f"  - {section}")
                
        except Exception as e:
            print(f"Error analyzing Output sheet: {e}")
    
    def analyze_data_sheets(self):
        """Analyze the data sheets (FMRs, SAFMRs, etc.)"""
        try:
            print("\n=== ANALYZING DATA SHEETS ===")
            
            data_sheets = ['FY25_FMRs', 'SAFMRs', 'Section8-FY25', 'Counties']
            
            for sheet_name in data_sheets:
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                    
                    # Get dimensions
                    rows = sheet.max_row
                    cols = sheet.max_column
                    
                    # Get headers (first row)
                    headers = []
                    for col in range(1, min(cols + 1, 20)):
                        header = sheet.cell(row=1, column=col).value
                        if header:
                            headers.append(str(header))
                    
                    print(f"\n{sheet_name}:")
                    print(f"  Dimensions: {rows} rows x {cols} columns")
                    print(f"  Headers: {', '.join(headers[:10])}")
                    
                    self.analysis_results['sheet_analysis'][sheet_name] = {
                        'rows': rows,
                        'columns': cols,
                        'headers': headers
                    }
                    
        except Exception as e:
            print(f"Error analyzing data sheets: {e}")
    
    def identify_formulas_and_connections(self):
        """Identify key formulas and data connections"""
        try:
            print("\n=== IDENTIFYING KEY FORMULAS ===")
            
            # In read_only mode, we can't access formulas directly
            # But we can identify likely calculation patterns
            
            output_sheet = self.workbook['Output']
            
            # Look for calculation indicators
            calc_patterns = []
            
            for row in range(1, min(output_sheet.max_row + 1, 100)):
                row_values = []
                for col in range(1, min(output_sheet.max_column + 1, 15)):
                    val = output_sheet.cell(row=row, column=col).value
                    row_values.append(val)
                
                # Check if row has numeric progression (likely calculations)
                numeric_count = sum(1 for v in row_values[1:] if isinstance(v, (int, float)))
                if numeric_count >= 3:
                    label = row_values[0] if row_values[0] else f"Row {row}"
                    calc_patterns.append({
                        'row': row,
                        'label': str(label),
                        'numeric_values': numeric_count
                    })
            
            self.analysis_results['data_connections']['calculation_rows'] = len(calc_patterns)
            self.analysis_results['data_connections']['sample_calculations'] = calc_patterns[:10]
            
            print(f"Found {len(calc_patterns)} rows with likely calculations")
            
        except Exception as e:
            print(f"Error identifying formulas: {e}")
    
    def create_template_summary(self):
        """Create a summary of the template structure"""
        try:
            print("\n=== TEMPLATE SUMMARY ===")
            
            summary = {
                'purpose': 'BOTN Workforce Housing Underwriting Model',
                'key_features': [],
                'data_sources': [],
                'output_metrics': []
            }
            
            # Identify key features based on sheet analysis
            if 'FY25_FMRs' in self.workbook.sheetnames:
                summary['data_sources'].append('HUD Fair Market Rents (FY2025)')
            if 'SAFMRs' in self.workbook.sheetnames:
                summary['data_sources'].append('Small Area Fair Market Rents')
            if 'Section8-FY25' in self.workbook.sheetnames:
                summary['data_sources'].append('Section 8 Payment Standards')
            
            # Key features from inputs
            if self.analysis_results['inputs_structure'].get('sections'):
                summary['key_features'] = self.analysis_results['inputs_structure']['sections'][:10]
            
            # Output metrics
            if self.analysis_results['output_structure'].get('sections_found'):
                summary['output_metrics'] = self.analysis_results['output_structure']['sections_found'][:10]
            
            self.analysis_results['summary'] = summary
            
            print(f"\nTemplate Purpose: {summary['purpose']}")
            print(f"Data Sources: {', '.join(summary['data_sources'])}")
            print(f"Key Input Sections: {len(summary['key_features'])}")
            print(f"Output Metrics: {len(summary['output_metrics'])}")
            
        except Exception as e:
            print(f"Error creating summary: {e}")
    
    def save_analysis(self):
        """Save analysis results to JSON"""
        try:
            output_path = Path(self.template_path).parent / f"BOTN_Template_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            with open(output_path, 'w') as f:
                json.dump(self.analysis_results, f, indent=2, default=str)
            
            print(f"\nAnalysis saved to: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"Error saving analysis: {e}")
            return None

def main():
    """Main analysis function"""
    template_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Workforce_Housing/BOTN_Proformas/Workforce BOTN 05.12.25.xlsx'
    
    analyzer = BOTNUnderwritingAnalyzer(template_path)
    
    if analyzer.load_workbook():
        analyzer.analyze_inputs_sheet()
        analyzer.analyze_output_sheet()
        analyzer.analyze_data_sheets()
        analyzer.identify_formulas_and_connections()
        analyzer.create_template_summary()
        analyzer.save_analysis()
        
        print("\n=== ANALYSIS COMPLETE ===")
        print(f"Template: {analyzer.analysis_results['template_info']['file_name']}")
        print(f"Sheets Analyzed: {len(analyzer.analysis_results['sheet_analysis'])}")

if __name__ == "__main__":
    main()