# extract_lihtc_tabs.py

import os
import json
import pandas as pd
from openpyxl import load_workbook
import logging
from pathlib import Path
import re
from datetime import datetime
import traceback

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_tab_extraction.log"),
        logging.StreamHandler()
    ]
)

class LIHTCTabExtractor:
    """
    Extract data from specific tabs in LIHTC applications
    """
    
    def __init__(self, input_path, output_path):
        """
        Initialize the tab extractor
        
        Args:
            input_path: Path to LIHTC Excel files
            output_path: Base path for output JSON files
        """
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        
        # Create main output directories
        for app_type in ['4p', '9p']:
            os.makedirs(os.path.join(self.output_path, app_type), exist_ok=True)
        
        # Stats tracking
        self.stats = {
            "start_time": datetime.now().isoformat(),
            "files_processed": 0,
            "tabs_extracted": {},
            "errors": []
        }
    
    def get_application_files(self, app_type, limit=None):
        """
        Get list of application files to process
        
        Args:
            app_type: Application type (4p or 9p)
            limit: Maximum number of files to return
            
        Returns:
            list: List of file paths
        """
        pattern = '4pct' if app_type == '4p' else '9pct'
        files = []
        
        for file in os.listdir(self.input_path):
            if file.endswith('.xlsx') and pattern in file:
                files.append(os.path.join(self.input_path, file))
                if limit and len(files) >= limit:
                    break
        
        logging.info(f"Found {len(files)} {app_type} application files")
        return files
    
    def extract_application_tab(self, file_path, tab_name):
        """
        Extract basic application information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Application tab from {file_name}")
        
        # Parse metadata from filename
        app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
        year_match = re.search(r'(\d{4})_\d+pct', file_name)
        round_match = re.search(r'_R(\d+)_', file_name)
        ctcac_match = re.search(r'_(\d+-\d+)', file_name)
        
        # Initialize result with metadata
        result = {
            "metadata": {
                "file_name": file_name,
                "application_type": app_type,
                "application_year": year_match.group(1) if year_match else "",
                "application_round": round_match.group(1) if round_match else "",
                "ctcac_number": ctcac_match.group(1) if ctcac_match else "",
                "extraction_date": datetime.now().isoformat()
            },
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Try to extract key information
            # We'll search for label patterns and adjacent data
            
            # Dictionary of label patterns and their corresponding fields in the result
            labels = {
                r"(?i)project\s*name\s*:": "project.name",
                r"(?i)project\s*address\s*:": "project.address",
                r"(?i)city\s*:": "project.city",
                r"(?i)county\s*:": "project.county",
                r"(?i)zip\s*code\s*:": "project.zip_code",
                r"(?i)census\s*tract\s*:": "project.census_tract",
                r"(?i)total\s*units?\s*:": "project.total_units",
                r"(?i)low\s*income\s*units?\s*:": "project.low_income_units",
                r"(?i)manager\s*units?\s*:": "project.manager_units"
            }
            
            # Search for labels and extract data
            for row in range(1, min(100, sheet.max_row + 1)):  # Limit to first 100 rows
                for col in range(1, min(20, sheet.max_column + 1)):  # Limit to first 20 columns
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against label patterns
                        for pattern, field_path in labels.items():
                            if re.search(pattern, cell_value):
                                # Found a label, look for adjacent data
                                # Check cell to the right first
                                data_value = sheet.cell(row=row, column=col+1).value
                                
                                # If no value to the right, check below
                                if not data_value:
                                    data_value = sheet.cell(row=row+1, column=col).value
                                
                                # Update result if data found
                                if data_value:
                                    # Parse field path to update nested dict
                                    parts = field_path.split('.')
                                    target = result
                                    for part in parts[:-1]:
                                        target = target[part]
                                    
                                    # Convert numeric fields
                                    if parts[-1] in ['total_units', 'low_income_units', 'manager_units']:
                                        try:
                                            data_value = int(float(data_value))
                                        except:
                                            pass
                                    
                                    target[parts[-1]] = data_value
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Application tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_sources_uses_tab(self, file_path, tab_name):
        """
        Extract Sources and Uses Budget information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted sources and uses data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Sources and Uses tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "costs": {
                "total_project_cost": 0,
                "acquisition_costs": {
                    "land": 0,
                    "improvements": 0,
                    "total": 0
                },
                "construction_costs": {
                    "new_construction": 0,
                    "rehabilitation": 0,
                    "total": 0
                },
                "developer_fee": {
                    "total": 0,
                    "deferred": 0
                }
            },
            "sources": {
                "permanent": [],
                "construction": []
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Common labels to search for in Sources and Uses
            key_fields = {
                "total_project_cost": [
                    r"(?i)total\s*project\s*costs?",
                    r"(?i)total\s*development\s*costs?"
                ],
                "land": [
                    r"(?i)land\s*costs?",
                    r"(?i)land\s*acquisition"
                ],
                "improvements": [
                    r"(?i)existing\s*improvements",
                    r"(?i)building\s*acquisition"
                ],
                "new_construction": [
                    r"(?i)new\s*construction",
                    r"(?i)structures"
                ],
                "rehabilitation": [
                    r"(?i)rehabilitation",
                    r"(?i)rehab"
                ],
                "developer_fee": [
                    r"(?i)developer\s*fee",
                    r"(?i)developer\s*costs?"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a field label, look for value to the right
                                # Search a few columns to the right
                                for offset in range(1, 5):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value
                                            if field == "total_project_cost":
                                                result["costs"]["total_project_cost"] = data_value
                                            elif field == "land":
                                                result["costs"]["acquisition_costs"]["land"] = data_value
                                            elif field == "improvements":
                                                result["costs"]["acquisition_costs"]["improvements"] = data_value
                                            elif field == "new_construction":
                                                result["costs"]["construction_costs"]["new_construction"] = data_value
                                            elif field == "rehabilitation":
                                                result["costs"]["construction_costs"]["rehabilitation"] = data_value
                                            elif field == "developer_fee":
                                                result["costs"]["developer_fee"]["total"] = data_value
                                            break
            
            # Calculate acquisition total
            land = result["costs"]["acquisition_costs"]["land"]
            improvements = result["costs"]["acquisition_costs"]["improvements"]
            result["costs"]["acquisition_costs"]["total"] = land + improvements
            
            # Calculate construction total
            new_construction = result["costs"]["construction_costs"]["new_construction"]
            rehabilitation = result["costs"]["construction_costs"]["rehabilitation"]
            result["costs"]["construction_costs"]["total"] = new_construction + rehabilitation
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Sources and Uses tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_basis_credits_tab(self, file_path, tab_name):
        """
        Extract Basis & Credits information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted basis and credits data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Basis & Credits tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "basis": {
                "total_eligible_basis": 0,
                "threshold_basis_limit": 0
            },
            "credits": {
                "federal_annual": 0,
                "federal_total": 0,
                "state_annual": 0,
                "state_total": 0,
                "credit_price": 0,
                "equity_amount": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Common labels to search for in Basis & Credits
            key_fields = {
                "total_eligible_basis": [
                    r"(?i)total\s*eligible\s*basis",
                    r"(?i)eligible\s*basis\s*total"
                ],
                "threshold_basis_limit": [
                    r"(?i)threshold\s*basis\s*limit",
                    r"(?i)adjusted\s*threshold\s*basis\s*limit"
                ],
                "federal_annual": [
                    r"(?i)annual\s*federal\s*credit",
                    r"(?i)federal\s*credit\s*annual"
                ],
                "state_annual": [
                    r"(?i)annual\s*state\s*credit",
                    r"(?i)state\s*credit\s*annual"
                ],
                "credit_price": [
                    r"(?i)credit\s*price",
                    r"(?i)price\s*per\s*credit"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a field label, look for value to the right
                                # Search a few columns to the right
                                for offset in range(1, 5):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value
                                            if field == "total_eligible_basis":
                                                result["basis"]["total_eligible_basis"] = data_value
                                            elif field == "threshold_basis_limit":
                                                result["basis"]["threshold_basis_limit"] = data_value
                                            elif field == "federal_annual":
                                                result["credits"]["federal_annual"] = data_value
                                            elif field == "state_annual":
                                                result["credits"]["state_annual"] = data_value
                                            elif field == "credit_price":
                                                result["credits"]["credit_price"] = data_value
                                            break
            
            # Calculate derived fields
            if result["credits"]["federal_annual"]:
                result["credits"]["federal_total"] = result["credits"]["federal_annual"] * 10
            
            if result["credits"]["state_annual"]:
                result["credits"]["state_total"] = result["credits"]["state_annual"]
            
            if result["credits"]["federal_annual"] and result["credits"]["credit_price"]:
                result["credits"]["equity_amount"] = result["credits"]["federal_annual"] * 10 * result["credits"]["credit_price"]
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Basis & Credits tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_points_system_tab(self, file_path, tab_name):
        """
        Extract Points System information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted points data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Points System tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "scoring": {
                "total_score": 0,
                "general_partner_experience": 0,
                "management_company_experience": 0,
                "housing_needs": 0,
                "site_amenities": 0,
                "service_amenities": 0,
                "sustainable_building_methods": 0,
                "lowest_income": 0,
                "readiness_to_proceed": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Common labels to search for in Points System
            key_fields = {
                "total_score": [
                    r"(?i)total\s*score",
                    r"(?i)total\s*points"
                ],
                "general_partner_experience": [
                    r"(?i)general\s*partner\s*experience",
                    r"(?i)gp\s*experience"
                ],
                "management_company_experience": [
                    r"(?i)management\s*company\s*experience",
                    r"(?i)management\s*experience"
                ],
                "housing_needs": [
                    r"(?i)housing\s*needs",
                    r"(?i)housing\s*type"
                ],
                "site_amenities": [
                    r"(?i)site\s*amenities"
                ],
                "service_amenities": [
                    r"(?i)service\s*amenities"
                ],
                "sustainable_building_methods": [
                    r"(?i)sustainable\s*building",
                    r"(?i)energy\s*efficiency"
                ],
                "lowest_income": [
                    r"(?i)lowest\s*income",
                    r"(?i)serving\s*lowest\s*income"
                ],
                "readiness_to_proceed": [
                    r"(?i)readiness\s*to\s*proceed"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a field label, look for value to the right
                                # Search a few columns to the right
                                for offset in range(1, 8):  # Search more columns for points
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value
                                            result["scoring"][field] = data_value
                                            break
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Points System tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_tie_breaker_tab(self, file_path, tab_name):
        """
        Extract Tie Breaker information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted tie breaker data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Tie Breaker tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "tie_breaker": {
                "score": 0,
                "leveraged_soft_resources": 0,
                "ratio_requested_to_total": 0,
                "resource_area_bonus": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Common labels to search for in Tie Breaker
            key_fields = {
                "score": [
                    r"(?i)tie\s*breaker\s*score",
                    r"(?i)final\s*tie\s*breaker"
                ],
                "leveraged_soft_resources": [
                    r"(?i)leveraged\s*soft\s*resources",
                    r"(?i)soft\s*resources\s*ratio"
                ],
                "ratio_requested_to_total": [
                    r"(?i)ratio\s*of\s*requested\s*unadjusted\s*eligible\s*basis",
                    r"(?i)1\s*minus\s*ratio\s*of\s*requested"
                ],
                "resource_area_bonus": [
                    r"(?i)resource\s*area\s*bonus",
                    r"(?i)highest\s*resource\s*area"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a field label, look for value to the right
                                # Search a few columns to the right
                                for offset in range(1, 8):  # Search more columns
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value
                                            result["tie_breaker"][field] = data_value
                                            break
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Tie Breaker tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_pro_forma_tab(self, file_path, tab_name):
        """
        Extract 15 Year Pro Forma information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted pro forma data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Pro Forma tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "pro_forma": {
                "year_1": {
                    "gross_income": 0,
                    "vacancy": 0,
                    "effective_gross_income": 0,
                    "operating_expenses": 0,
                    "net_operating_income": 0,
                    "debt_service": 0,
                    "cash_flow": 0,
                    "debt_service_coverage": 0
                }
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Check if tab exists
            if tab_name not in wb.sheetnames:
                logging.warning(f"Tab '{tab_name}' not found in {file_name}")
                return result
            
            sheet = wb[tab_name]
            
            # Common labels to search for in Pro Forma
            key_fields = {
                "gross_income": [
                    r"(?i)gross\s*income",
                    r"(?i)gross\s*potential\s*income"
                ],
                "vacancy": [
                    r"(?i)vacancy",
                    r"(?i)vacancy\s*loss"
                ],
                "effective_gross_income": [
                    r"(?i)effective\s*gross\s*income",
                    r"(?i)total\s*income"
                ],
                "operating_expenses": [
                    r"(?i)operating\s*expenses",
                    r"(?i)total\s*expenses"
                ],
                "net_operating_income": [
                    r"(?i)net\s*operating\s*income",
                    r"(?i)NOI"
                ],
                "debt_service": [
                    r"(?i)debt\s*service",
                    r"(?i)total\s*debt\s*service"
                ],
                "cash_flow": [
                    r"(?i)cash\s*flow",
                    r"(?i)cash\s*flow\s*after\s*debt\s*service"
                ],
                "debt_service_coverage": [
                    r"(?i)debt\s*service\s*coverage",
                    r"(?i)debt\s*coverage\s*ratio"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(15, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value) for pattern in patterns):
                                # Found a field label, look for value to the right
                                # Search a few columns to the right - since this is year 1, it's usually the first data column
                                for offset in range(1, 5):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value for year 1
                                            result["pro_forma"]["year_1"][field] = data_value
                                            break
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Pro Forma tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def save_tab_data(self, data, app_type, file_name, tab_name):
        """
        Save tab data to JSON file
        
        Args:
            data: Extracted data
            app_type: Application type (4p or 9p)
            file_name: Original file name
            tab_name: Tab name
            
        Returns:
            str: Path to saved file
        """
        # Create output directory structure
        output_dir = os.path.join(self.output_path, app_type, tab_name.lower().replace(' ', '_'))
        os.makedirs(output_dir, exist_ok=True)
        
        # Create output filename
        output_file = os.path.join(output_dir, f"{file_name.replace('.xlsx', '')}.json")
        
        # Save data
        with open(output_file, 'w') as f:
            json.dump(data, f, indent=2)
        
        logging.info(f"Saved {tab_name} data to {output_file}")
        return output_file
    
    def process_files(self, app_type, limit=None):
        """
        Process a batch of files
        
        Args:
            app_type: Application type (4p or 9p)
            limit: Maximum number of files to process
        """
        files = self.get_application_files(app_type, limit)
        
        # Define tabs to extract and their corresponding extractors
        tabs_to_extract = {
            "Application": self.extract_application_tab,
            "Sources and Uses Budget": self.extract_sources_uses_tab,
            "Basis & Credits": self.extract_basis_credits_tab,
            "Points System": self.extract_points_system_tab,
            "Tie Breaker": self.extract_tie_breaker_tab,
            "15 Year Pro Forma": self.extract_pro_forma_tab
        }
        
        # Initialize tab stats
        for tab in tabs_to_extract.keys():
            self.stats["tabs_extracted"][tab] = 0
        
        # Process each file
        for file_path in files:
            file_name = os.path.basename(file_path)
            logging.info(f"Processing {file_name}")
            
            try:
                # Extract each tab
                for tab_name, extractor in tabs_to_extract.items():
                    try:
                        data = extractor(file_path, tab_name)
                        self.save_tab_data(data, app_type, file_name, tab_name)
                        self.stats["tabs_extracted"][tab_name] += 1
                    except Exception as e:
                        logging.error(f"Error extracting {tab_name} from {file_name}: {e}")
                        self.stats["errors"].append(f"Error extracting {tab_name} from {file_name}: {str(e)}")
                
                self.stats["files_processed"] += 1
                logging.info(f"Completed processing {file_name}")
                
            except Exception as e:
                logging.error(f"Error processing {file_name}: {e}")
                logging.error(traceback.format_exc())
                self.stats["errors"].append(f"Error processing {file_name}: {str(e)}")
    
    def generate_summary(self):
        """
        Generate summary of extraction process
        
        Returns:
            dict: Summary statistics
        """
        self.stats["end_time"] = datetime.now().isoformat()
        start_time = datetime.fromisoformat(self.stats["start_time"])
        end_time = datetime.fromisoformat(self.stats["end_time"])
        duration = end_time - start_time
        
        self.stats["duration_seconds"] = duration.total_seconds()
        self.stats["duration_formatted"] = str(duration)
        
        # Log summary
        logging.info(f"Extraction completed. Processed {self.stats['files_processed']} files in {self.stats['duration_formatted']}")
        for tab, count in self.stats["tabs_extracted"].items():
            logging.info(f"Extracted {count} {tab} tabs")
        
        if self.stats["errors"]:
            logging.warning(f"Encountered {len(self.stats['errors'])} errors")
        
        # Save summary to file
        summary_file = os.path.join(self.output_path, "extraction_summary.json")
        with open(summary_file, 'w') as f:
            json.dump(self.stats, f, indent=2)
        
        logging.info(f"Saved summary to {summary_file}")
        return self.stats

# Main execution function
def main():
    # Define paths
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    json_base_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    
    # Create extractor
    extractor = LIHTCTabExtractor(input_path, json_base_path)
    
    # Process just 3 files from each application type for testing
    logging.info("Starting extraction of 4% applications (test sample)...")
    extractor.process_files('4p', limit=3)  # Process only 3 files for testing
    
    # Process 9% applications
    logging.info("Starting extraction of 9% applications (test sample)...")
    extractor.process_files('9p', limit=3)  # Process only 3 files for testing
    
    # Generate summary
    summary = extractor.generate_summary()
    print(f"Extraction complete! Processed {summary['files_processed']} files.")
    
if __name__ == "__main__":
    main()