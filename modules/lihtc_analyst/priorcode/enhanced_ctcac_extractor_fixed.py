#!/usr/bin/env python3
"""
Enhanced CTCAC LIHTC Application Data Extractor - Complete RAG Version
Extracts comprehensive data from CTCAC applications for RAG system

This enhanced version captures:
- Complete project information & unit mix details
- All financing sources and detailed cost breakdowns  
- Full competitive scoring breakdown with all point categories
- Comprehensive basis & credit calculations
- Complete tie breaker responses
- 15-year pro forma data
- Sources and basis breakdown details
- SCE/FCE basis calculations

Based on detailed analysis of 2023-2025 CTCAC applications (4% and 9%)
"""

import pandas as pd
import json
import os
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
from typing import Dict, List, Any, Optional
import re
import sys

class CTCACComprehensiveExtractor:
    """
    Comprehensive extractor that captures ALL valuable data from CTCAC applications
    """
    
    def __init__(self, input_path: str, output_path_4p: str, output_path_9p: str, log_path: str):
        """Initialize the extractor with file paths"""
        self.input_path = Path(input_path)
        self.output_path_4p = Path(output_path_4p)
        self.output_path_9p = Path(output_path_9p)
        self.log_path = Path(log_path)
        
        # Create directories if they don't exist
        self.output_path_4p.mkdir(parents=True, exist_ok=True)
        self.output_path_9p.mkdir(parents=True, exist_ok=True)
        self.log_path.mkdir(parents=True, exist_ok=True)
        
        # Set up logging
        self.setup_logging()
        
        # Define comprehensive sheet mapping
        self.target_sheets = {
            'application': 'Application',
            'sources_uses': 'Sources and Uses Budget',
            'sources_basis_breakdown': 'Sources and Basis Breakdown',
            'points_system': 'Points System',
            'basis_credits': 'Basis & Credits',
            'tie_breaker': 'Tie Breaker',
            'final_tie_breaker': 'Final Tie Breaker ',
            'disaster_tie_breaker': 'Disaster Credit Tie Breaker',
            'pro_forma': '15 Year Pro Forma',
            'sce_basis': 'SCE Basis and Credits',
            'fce_basis': 'FCE Basis and Credits',
            'calHFA_addendum': 'CalHFA Addendum',
            'subsidy_contract': 'Subsidy Contract Calculation'
        }
    
    def setup_logging(self):
        """Set up logging configuration"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = self.log_path / f"ctcac_comprehensive_extraction_{timestamp}.log"
        
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        self.logger.propagate = False
        
        self.logger.info(f"Comprehensive extraction logging initialized. Log file: {log_file}")
    
    def parse_filename(self, filename: str) -> Dict[str, str]:
        """Parse filename to extract metadata"""
        base_name = Path(filename).stem
        
        parsed = {
            'year': 'unknown',
            'credit_type': 'unknown',
            'round': 'unknown',
            'project_id': 'unknown',
            'original_filename': filename
        }
        
        try:
            # Pattern for both formats: YYYY_XpctXX_RX_YY-XXX and YYYY_XpctXX_RX_CAYYXXX
            pattern = r'(\d{4})_(\d+pct)_([rR]\d+)_(?:CA)?(.+)'
            match = re.match(pattern, base_name)
            
            if match:
                parsed['year'] = match.group(1)
                parsed['credit_type'] = match.group(2)
                parsed['round'] = match.group(3).upper()
                parsed['project_id'] = match.group(4)
        
        except Exception as e:
            self.logger.warning(f"Error parsing filename {filename}: {e}")
        
        return parsed
    
    def determine_output_path(self, parsed_info: Dict[str, str]) -> Path:
        """Determine correct output path based on credit type"""
        if '4' in parsed_info['credit_type']:
            return self.output_path_4p
        else:
            return self.output_path_9p
    
    def create_output_filename(self, parsed_info: Dict[str, str]) -> str:
        """Create standardized output filename"""
        return f"{parsed_info['year']}_{parsed_info['credit_type']}_{parsed_info['round']}_CA-{parsed_info['project_id']}.json"
    
    def get_cell_value(self, sheet, cell_ref: str):
        """Safely get cell value from sheet, handling merged cells"""
        try:
            cell = sheet[cell_ref]
            
            # If cell is part of a merged range, get the value from the top-left cell
            if hasattr(sheet, 'merged_cells'):
                for merged_range in sheet.merged_cells.ranges:
                    if cell_ref in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                        return top_left_cell.value if top_left_cell and top_left_cell.value is not None else None
            
            return cell.value if cell and cell.value is not None else None
        except:
            return None
    
    def scan_merged_cell_area(self, sheet, start_row: int, end_row: int, start_col: int = 0, end_col: int = 15):
        """Scan an area and return all non-None values, handling merged cells properly"""
        values = {}
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                col_letter = chr(65 + col)
                cell_ref = f'{col_letter}{row}'
                value = self.get_cell_value(sheet, cell_ref)
                
                if value is not None:
                    values[cell_ref] = value
        
        return values
    
    def find_sheet_name(self, workbook, target_name: str) -> Optional[str]:
        """Find the actual sheet name (handles slight variations)"""
        if target_name in workbook.sheetnames:
            return target_name
        
        # Look for close matches
        for sheet_name in workbook.sheetnames:
            if target_name.lower().replace(' ', '') in sheet_name.lower().replace(' ', ''):
                return sheet_name
        return None
    
    def extract_comprehensive_project_info(self, workbook) -> Dict[str, Any]:
        """Extract comprehensive project information"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Application')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            project_info = {
                # Basic project identification
                'ctcac_applicant': self.get_cell_value(sheet, 'H16'),
                'project_name': self.get_cell_value(sheet, 'H18'),
                'project_name_alt': self.get_cell_value(sheet, 'H17'),
                
                # Application details
                'application_type': None,
                'state_credits_election': None,
                'project_address': None,
                'city': None,
                'county': None,
                'zip_code': None,
                
                # Unit information
                'total_units': None,
                'affordable_units': None,
                'market_rate_units': None,
                'unit_mix_details': {},
                
                # Set-aside information
                'set_aside_type': None,
                'nonprofit_participation': None,
                'tribal_project': None,
                
                # Development information
                'new_construction': None,
                'rehabilitation': None,
                'acquisition_rehabilitation': None,
                
                # Special features
                'green_building': None,
                'transit_oriented': None,
                'preservation_project': None
            }
            
            # Determine application type
            for row in range(8, 20):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and isinstance(cell_value, str):
                    if '4%' in cell_value and 'TAX-EXEMPT BONDS' in cell_value.upper():
                        project_info['application_type'] = '4% with Tax-Exempt Bonds'
                    elif '9%' in cell_value and 'COMPETITIVE' in cell_value.upper():
                        project_info['application_type'] = '9% Competitive'
            
            # Extract location information (scan around project name area)
            for row in range(15, 30):
                for col in ['H', 'I', 'J', 'K']:
                    value = self.get_cell_value(sheet, f'{col}{row}')
                    if value and isinstance(value, str):
                        # Look for address patterns
                        if any(term in value.lower() for term in ['street', 'avenue', 'road', 'drive', 'way']):
                            project_info['project_address'] = value
                        # Look for city patterns
                        elif len(value) > 3 and value.replace(' ', '').replace(',', '').isalpha():
                            project_info['city'] = value
                        # Look for zip codes
                        elif re.match(r'^\d{5}(-\d{4})?$', value.strip()):
                            project_info['zip_code'] = value
            
            # Look for state credits election
            for row in range(30, 45):
                cell_value = self.get_cell_value(sheet, f'A{row}')
                if cell_value and 'state credit' in str(cell_value).lower():
                    for col in ['B', 'C', 'D', 'E', 'F']:
                        response = self.get_cell_value(sheet, f'{col}{row}')
                        if response and str(response).lower() in ['yes', 'no']:
                            project_info['state_credits_election'] = str(response)
                            break
            
            # Extract unit mix information (comprehensive scan)
            unit_mix = self.extract_detailed_unit_mix(sheet)
            if unit_mix:
                project_info.update(unit_mix)
            
            # Clean up None values
            project_info = {k: v for k, v in project_info.items() if v is not None}
            
            return project_info
            
        except Exception as e:
            self.logger.error(f"Error extracting comprehensive project info: {e}")
            return {}
    
    def extract_detailed_unit_mix(self, sheet) -> Dict[str, Any]:
        """Extract detailed unit mix information from Application sheet - enhanced for merged cells"""
        try:
            unit_mix = {
                'total_units': None,
                'affordable_units': None,
                'market_rate_units': None,
                'unit_types': {},
                'ami_levels': {},
                'rent_schedule': {},
                'unit_mix_table': []
            }
            
            # Enhanced search for unit mix data with merged cell handling
            unit_keywords = [
                'bedroom', 'studio', '1 br', '2 br', '3 br', '4 br', '5 br',
                'unit type', 'unit mix', 'ami', 'rent', 'income',
                '30%', '40%', '50%', '60%', '70%', '80%', '100%', '120%'
            ]
            
            self.logger.info("   🏠 Enhanced unit mix extraction (merged cell aware)...")
            
            # Scan for unit mix tables in typical locations
            for search_start in [100, 150, 200, 250, 300, 350]:
                search_end = search_start + 100
                
                # Look for table headers that indicate unit mix
                table_found = False
                table_start_row = None
                
                for row in range(search_start, search_end):
                    row_values = self.scan_merged_cell_area(sheet, row, row, 0, 20)  # Columns A-T
                    
                    # Check if this row contains unit mix headers
                    header_indicators = 0
                    for cell_ref, value in row_values.items():
                        if isinstance(value, str):
                            value_lower = value.lower()
                            if any(keyword in value_lower for keyword in unit_keywords):
                                header_indicators += 1
                    
                    # If we found multiple unit mix indicators in one row, this might be a table header
                    if header_indicators >= 2:
                        table_found = True
                        table_start_row = row
                        self.logger.info(f"     Found potential unit mix table at row {row}")
                        break
                
                if table_found and table_start_row:
                    # Extract data from the table
                    self.extract_unit_mix_table(sheet, table_start_row, unit_mix)
                    
                    # If we found substantial data, break
                    if unit_mix['total_units'] or len(unit_mix['unit_types']) > 0:
                        break
            
            # Additional search for total units
            if not unit_mix['total_units']:
                for row in range(50, 400):
                    row_values = self.scan_merged_cell_area(sheet, row, row, 0, 15)
                    
                    for cell_ref, value in row_values.items():
                        if isinstance(value, str) and 'total' in value.lower() and 'unit' in value.lower():
                            # Look for numeric value in same row
                            for val_ref, val in row_values.items():
                                if isinstance(val, (int, float)) and 10 <= val <= 1000:  # Reasonable unit count
                                    unit_mix['total_units'] = int(val)
                                    self.logger.info(f"     Found total units: {val} at {val_ref}")
                                    break
                            if unit_mix['total_units']:
                                break
                    if unit_mix['total_units']:
                        break
            
            return unit_mix
            
        except Exception as e:
            self.logger.error(f"Error extracting detailed unit mix: {e}")
            return {}
    
    def extract_unit_mix_table(self, sheet, start_row: int, unit_mix: Dict):
        """Extract unit mix data from a table starting at start_row"""
        try:
            # Analyze the next 20 rows for unit mix data
            for row_offset in range(0, 20):
                current_row = start_row + row_offset
                row_values = self.scan_merged_cell_area(sheet, current_row, current_row, 0, 20)
                
                # Look for unit type information
                unit_type = None
                unit_count = None
                ami_level = None
                rent_amount = None
                
                for cell_ref, value in row_values.items():
                    if isinstance(value, str):
                        value_lower = value.lower()
                        
                        # Identify unit types
                        if any(term in value_lower for term in ['studio', '1 br', '2 br', '3 br', '4 br', 'bedroom']):
                            unit_type = value.strip()
                        
                        # Identify AMI levels
                        ami_match = re.search(r'(\d+)%.*ami', value_lower)
                        if ami_match:
                            ami_level = f"{ami_match.group(1)}%_AMI"
                    
                    elif isinstance(value, (int, float)):
                        # Could be unit count, rent amount, or other numeric data
                        if 1 <= value <= 500:  # Reasonable unit count
                            unit_count = int(value)
                        elif 500 <= value <= 5000:  # Reasonable rent amount
                            rent_amount = float(value)
                
                # If we found meaningful unit data, store it
                if unit_type and unit_count:
                    unit_mix['unit_types'][unit_type] = unit_count
                    self.logger.info(f"     Found unit type: {unit_type} = {unit_count} units")
                
                if ami_level and unit_count:
                    unit_mix['ami_levels'][ami_level] = unit_count
                    self.logger.info(f"     Found AMI level: {ami_level} = {unit_count} units")
                
                if unit_type and rent_amount:
                    unit_mix['rent_schedule'][unit_type] = rent_amount
                    self.logger.info(f"     Found rent: {unit_type} = ${rent_amount}")
        
        except Exception as e:
            self.logger.error(f"Error extracting unit mix table: {e}")
    
    def extract_comprehensive_sources_uses(self, workbook) -> Dict[str, Any]:
        """Extract comprehensive sources and uses with ALL financing sources"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Sources and Uses Budget')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            sources_uses = {
                # Detailed cost categories
                'land_costs': {},
                'rehabilitation_costs': {},
                'new_construction_costs': {},
                'architectural_costs': {},
                'construction_interest_fees': {},
                'permanent_financing_costs': {},
                'other_costs': {},
                
                # Comprehensive financing sources
                'financing_sources': {},
                'source_details': {},
                
                # Totals and calculations
                'total_uses': None,
                'total_sources': None,
                'sources_uses_balance': None
            }
            
            # Extract ALL financing sources (columns F through P)
            self.logger.info("   🔍 Extracting all financing sources...")
            for col in range(5, 16):  # Columns F-P
                col_letter = chr(65 + col)
                
                # Get source name from header
                source_name = self.get_cell_value(sheet, f'{col_letter}2')
                if source_name and isinstance(source_name, str) and source_name.strip():
                    clean_name = source_name.replace('\n', ' ').replace('\r', ' ').strip()
                    
                    source_details = {
                        'name': clean_name,
                        'amounts': {},
                        'total_amount': 0
                    }
                    
                    # Check key total rows for amounts
                    key_rows = [8, 12, 26, 38, 42, 53, 60, 70, 80]
                    for row in key_rows:
                        amount = self.get_cell_value(sheet, f'{col_letter}{row}')
                        if amount and isinstance(amount, (int, float)) and amount > 0:
                            source_details['amounts'][f'row_{row}'] = float(amount)
                            source_details['total_amount'] += float(amount)
                    
                    if source_details['total_amount'] > 0:
                        sources_uses['financing_sources'][clean_name] = source_details['total_amount']
                        sources_uses['source_details'][clean_name] = source_details
            
            # Extract detailed use categories
            self.logger.info("   🏗️ Extracting detailed cost categories...")
            
            # Land costs (rows 4-12)
            land_items = [
                ('land_cost_value', 4),
                ('demolition', 5),
                ('land_lease_prepayment', 7),
                ('total_land_cost', 8),
                ('existing_improvements', 9),
                ('offsite_improvements', 10),
                ('total_acquisition', 11),
                ('total_land_acquisition', 12)
            ]
            
            for item_name, row in land_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                res_cost = self.get_cell_value(sheet, f'C{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['land_costs'][item_name] = {
                        'total': float(total_cost),
                        'residential': float(res_cost) if res_cost and isinstance(res_cost, (int, float)) else None
                    }
            
            # Rehabilitation costs (rows 17-26)
            rehab_items = [
                ('site_work', 17),
                ('structures', 18),
                ('general_requirements', 19),
                ('contractor_overhead', 20),
                ('contractor_profit', 21),
                ('prevailing_wages', 22),
                ('general_liability_insurance', 23),
                ('construction_management', 24),
                ('other_rehab', 25),
                ('total_rehabilitation', 26)
            ]
            
            for item_name, row in rehab_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['rehabilitation_costs'][item_name] = float(total_cost)
            
            # New construction costs (rows 29-38)
            new_construction_items = [
                ('site_work', 29),
                ('structures', 30),
                ('general_requirements', 31),
                ('contractor_overhead', 32),
                ('contractor_profit', 33),
                ('prevailing_wages', 34),
                ('general_liability_insurance', 35),
                ('construction_management', 36),
                ('other_new_construction', 37),
                ('total_new_construction', 38)
            ]
            
            for item_name, row in new_construction_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['new_construction_costs'][item_name] = float(total_cost)
            
            # Architectural costs (rows 40-42)
            arch_items = [
                ('design', 40),
                ('supervision', 41),
                ('total_architectural', 42)
            ]
            
            for item_name, row in arch_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['architectural_costs'][item_name] = float(total_cost)
            
            # Construction interest and fees (rows 45-53)
            const_interest_items = [
                ('construction_loan_interest', 45),
                ('origination_fee', 46),
                ('credit_enhancement_fee', 47),
                ('bond_premium', 48),
                ('title_recording', 49),
                ('insurance', 51),
                ('other_construction_fees', 52),
                ('total_construction_interest_fees', 53)
            ]
            
            for item_name, row in const_interest_items:
                total_cost = self.get_cell_value(sheet, f'B{row}')
                if total_cost and isinstance(total_cost, (int, float)) and total_cost > 0:
                    sources_uses['construction_interest_fees'][item_name] = float(total_cost)
            
            return sources_uses
            
        except Exception as e:
            self.logger.error(f"Error extracting comprehensive sources and uses: {e}")
            return {}
    
    def extract_complete_scoring_breakdown(self, workbook) -> Dict[str, Any]:
        """Extract complete competitive scoring breakdown - enhanced for merged cells"""
        try:
            sheet_name = self.find_sheet_name(workbook, 'Points System')
            if not sheet_name:
                return {}
            
            sheet = workbook[sheet_name]
            
            scoring = {
                'total_points_claimed': 0,
                'max_possible_points': 0,
                'scoring_categories': {},
                'detailed_responses': {},
                'point_summary': {}
            }
            
            self.logger.info("   📊 Extracting complete scoring breakdown (merged cell aware)...")
            
            # Enhanced scan for Points System sheet with merged cell handling
            for row in range(1, 300):  # Extended range
                
                # Get values across the row, handling merged cells
                row_values = self.scan_merged_cell_area(sheet, row, row, 0, 10)  # Columns A-K
                
                # Look for scoring patterns
                description = None
                response = None
                points_value = None
                max_points = None
                
                # Find description (usually in column A, but could be merged)
                for cell_ref, value in row_values.items():
                    if isinstance(value, str) and len(value) > 10:
                        # Could be a description
                        if any(keyword in value.lower() for keyword in [
                            'point', 'experience', 'development', 'site', 'amenity', 
                            'green', 'solar', 'financing', 'leveraging', 'cost', 
                            'service', 'housing', 'preservation', 'management',
                            'general partner', 'nonprofit', 'special needs'
                        ]):
                            description = value
                            break
                
                # Find Yes/No responses (usually in column B or C)
                for cell_ref, value in row_values.items():
                    if isinstance(value, str) and value.lower() in ['yes', 'no']:
                        response = value.lower()
                        break
                
                # Find point values (look for numbers 1-50 that could be points)
                potential_points = []
                for cell_ref, value in row_values.items():
                    if isinstance(value, (int, float)) and 0 < value <= 50:
                        potential_points.append(value)
                
                # If we found potential points, use them
                if potential_points:
                    if len(potential_points) == 1:
                        points_value = potential_points[0]
                        max_points = potential_points[0]
                    else:
                        # Multiple point values - might be max and claimed
                        max_points = max(potential_points)
                        points_value = max(potential_points) if response == 'yes' else 0
                
                # If we have a meaningful scoring item, record it
                if description and (response or points_value):
                    category_key = f'row_{row}'
                    
                    # Determine points awarded
                    if response == 'yes' and points_value:
                        points_awarded = float(points_value)
                    elif response == 'no':
                        points_awarded = 0
                    elif points_value and not response:
                        points_awarded = float(points_value)  # Assume claimed if value present
                    else:
                        points_awarded = 0
                    
                    # Determine max possible points
                    if max_points:
                        max_possible = float(max_points)
                    elif points_value:
                        max_possible = float(points_value)
                    else:
                        max_possible = 0
                    
                    scoring['scoring_categories'][category_key] = {
                        'description': description[:200],
                        'response': response,
                        'points_awarded': points_awarded,
                        'max_points': max_possible,
                        'claimed': points_awarded > 0,
                        'row_data': {k: v for k, v in row_values.items() if v is not None}  # Debug info
                    }
                    
                    scoring['total_points_claimed'] += points_awarded
                    scoring['max_possible_points'] += max_possible
                    
                    # Create summary by category type
                    category_type = self.categorize_scoring_item(description)
                    if category_type not in scoring['point_summary']:
                        scoring['point_summary'][category_type] = {
                            'points_claimed': 0,
                            'max_points': 0,
                            'items': 0
                        }
                    
                    scoring['point_summary'][category_type]['points_claimed'] += points_awarded
                    scoring['point_summary'][category_type]['max_points'] += max_possible
                    scoring['point_summary'][category_type]['items'] += 1
            
            self.logger.info(f"   ✅ Found {len(scoring['scoring_categories'])} scoring items, Total: {scoring['total_points_claimed']}/{scoring['max_possible_points']} points")
            
            return scoring
            
        except Exception as e:
            self.logger.error(f"Error extracting complete scoring breakdown: {e}")
            return {}
    
    def categorize_scoring_item(self, description: str) -> str:
        """Categorize scoring items into logical groups"""
        desc_lower = description.lower()
        
        if any(term in desc_lower for term in ['experience', 'development', 'management', 'general partner']):
            return 'Developer_Experience'
        elif any(term in desc_lower for term in ['site', 'amenity', 'location', 'transit']):
            return 'Site_Amenities'
        elif any(term in desc_lower for term in ['green', 'solar', 'energy', 'sustainable']):
            return 'Green_Building'
        elif any(term in desc_lower for term in ['financing', 'leveraging', 'funding']):
            return 'Financing_Characteristics'
        elif any(term in desc_lower for term in ['cost', 'efficiency', 'savings']):
            return 'Cost_Efficiency'
        elif any(term in desc_lower for term in ['service', 'supportive', 'social']):
            return 'Supportive_Services'
        elif any(term in desc_lower for term in ['housing', 'preservation', 'at-risk']):
            return 'Housing_Preservation'
        else: