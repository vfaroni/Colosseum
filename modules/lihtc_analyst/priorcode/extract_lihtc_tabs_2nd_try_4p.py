# extract_lihtc_tabs.py

import os
import json
import pandas as pd
from openpyxl import load_workbook
import logging
from pathlib import Path
import re
from datetime import datetime
import traceback

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("lihtc_tab_extraction.log"),
        logging.StreamHandler()
    ]
)

class LIHTCTabExtractor:
    """
    Extract data from specific tabs in LIHTC applications
    """
    
    def __init__(self, input_path, output_path_4p, output_path_9p):
        """
        Initialize the tab extractor
        
        Args:
            input_path: Path to LIHTC Excel files
            output_path_4p: Path for 4% JSON output
            output_path_9p: Path for 9% JSON output
        """
        self.input_path = Path(input_path)
        self.output_paths = {
            '4p': Path(output_path_4p),
            '9p': Path(output_path_9p)
        }
        
        # Create main output directories
        for path in self.output_paths.values():
            path.mkdir(parents=True, exist_ok=True)
        
        # Stats tracking
        self.stats = {
            "start_time": datetime.now().isoformat(),
            "files_processed": 0,
            "tabs_extracted": {},
            "errors": []
        }
    
    def get_application_files(self, app_type, limit=None):
        """
        Get list of application files to process
        
        Args:
            app_type: Application type (4p or 9p)
            limit: Maximum number of files to return
            
        Returns:
            list: List of file paths
        """
        pattern = '4pct' if app_type == '4p' else '9pct'
        files = []
        
        for file in os.listdir(self.input_path):
            if file.endswith('.xlsx') and pattern in file:
                files.append(os.path.join(self.input_path, file))
                if limit and len(files) >= limit:
                    break
        
        logging.info(f"Found {len(files)} {app_type} application files")
        return files
    
    def extract_application_tab(self, file_path, tab_name):
        """
        Extract basic application information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted application data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Application tab from {file_name}")
        
        # Parse metadata from filename
        app_type = "4p" if "4pct" in file_name else "9p" if "9pct" in file_name else "unknown"
        year_match = re.search(r'(\d{4})_\d+pct', file_name)
        round_match = re.search(r'_R(\d+)_', file_name)
        ctcac_match = re.search(r'_(\d+-\d+)', file_name)
        
        # Initialize result with metadata
        result = {
            "metadata": {
                "file_name": file_name,
                "application_type": app_type,
                "application_year": year_match.group(1) if year_match else "",
                "application_round": round_match.group(1) if round_match else "",
                "ctcac_number": ctcac_match.group(1) if ctcac_match else "",
                "extraction_date": datetime.now().isoformat()
            },
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Search all sheets for Application data, since the tab name might vary
            found_tab = False
            for sheet_name in wb.sheetnames:
                if "Application" in sheet_name or "Project" in sheet_name:
                    sheet = wb[sheet_name]
                    found_tab = True
                    break
            
            # If we didn't find a specific tab, try the one passed in
            if not found_tab:
                if tab_name in wb.sheetnames:
                    sheet = wb[tab_name]
                    found_tab = True
                else:
                    # Try the first sheet as a last resort
                    sheet = wb.worksheets[0]
            
            # Dictionary of label patterns and their corresponding fields in the result
            labels = {
                r"(?i)project\s*name": "project.name",
                r"(?i)development\s*name": "project.name",
                r"(?i)property\s*name": "project.name",
                r"(?i)project\s*address": "project.address",
                r"(?i)property\s*address": "project.address",
                r"(?i)address": "project.address",
                r"(?i)city": "project.city",
                r"(?i)county": "project.county",
                r"(?i)zip\s*code": "project.zip_code",
                r"(?i)census\s*tract": "project.census_tract",
                r"(?i)total\s*units?\s*": "project.total_units",
                r"(?i)total\s*number\s*of\s*units": "project.total_units",
                r"(?i)low\s*income\s*units": "project.low_income_units",
                r"(?i)affordable\s*units": "project.low_income_units",
                r"(?i)lihtc\s*units": "project.low_income_units",
                r"(?i)manager\s*units": "project.manager_units",
                r"(?i)manager'?s?\s*units?": "project.manager_units"
            }
            
            # Search for labels and extract data - scan more rows
            for row in range(1, min(200, sheet.max_row + 1)):
                for col in range(1, min(30, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against label patterns
                        for pattern, field_path in labels.items():
                            if re.search(pattern, cell_value, re.IGNORECASE):
                                # Found a label, search nearby cells for data
                                data_found = False
                                
                                # Check horizontally - several cells to the right
                                for offset in range(1, 5):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value is not None and data_value != "":
                                            data_found = True
                                            break
                                
                                # If not found, check vertically - cell below
                                if not data_found:
                                    data_value = sheet.cell(row=row+1, column=col).value
                                    if data_value is not None and data_value != "":
                                        data_found = True
                                
                                # Update result if data found
                                if data_found:
                                    # Parse field path to update nested dict
                                    parts = field_path.split('.')
                                    target = result
                                    for part in parts[:-1]:
                                        target = target[part]
                                    
                                    # Convert numeric fields
                                    if parts[-1] in ['total_units', 'low_income_units', 'manager_units']:
                                        try:
                                            if isinstance(data_value, str):
                                                # Remove any non-numeric characters
                                                clean_value = re.sub(r'[^\d.]', '', data_value)
                                                if clean_value:
                                                    data_value = int(float(clean_value))
                                            else:
                                                data_value = int(float(data_value))
                                        except:
                                            pass
                                    
                                    # Only update if we found a value and the field is currently empty
                                    if data_value and (target[parts[-1]] == "" or target[parts[-1]] == 0):
                                        target[parts[-1]] = data_value
            
            # Special check for units - if we still don't have values, search for units in a table
            if result["project"]["total_units"] == 0:
                for row in range(1, min(200, sheet.max_row + 1)):
                    for col in range(1, min(30, sheet.max_column + 1)):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            if "total units" in cell_value.lower() or "residential units" in cell_value.lower():
                                # Search for totals in this row and below
                                for r in range(row, min(row + 10, sheet.max_row + 1)):
                                    for c in range(1, min(30, sheet.max_column + 1)):
                                        value = sheet.cell(row=r, column=c).value
                                        if isinstance(value, (int, float)) and value > 0 and value < 1000:
                                            # Likely a unit count
                                            if result["project"]["total_units"] == 0:
                                                result["project"]["total_units"] = int(value)
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Application tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_sources_uses_tab(self, file_path, tab_name):
        """
        Extract Sources and Uses Budget information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted sources and uses data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Sources and Uses tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "costs": {
                "total_project_cost": 0,
                "acquisition_costs": {
                    "land": 0,
                    "improvements": 0,
                    "total": 0
                },
                "construction_costs": {
                    "new_construction": 0,
                    "rehabilitation": 0,
                    "total": 0
                },
                "developer_fee": {
                    "total": 0,
                    "deferred": 0
                }
            },
            "sources": {
                "permanent": [],
                "construction": []
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Look for relevant tab names if the exact one isn't found
            found_tab = False
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                found_tab = True
            else:
                for sheet_name in wb.sheetnames:
                    if "Source" in sheet_name or "Budget" in sheet_name or "Cost" in sheet_name:
                        sheet = wb[sheet_name]
                        found_tab = True
                        break
                
                if not found_tab:
                    logging.warning(f"No Sources and Uses tab found in {file_name}")
                    return result
            
            # Common labels to search for in Sources and Uses - expanded patterns
            key_fields = {
                "total_project_cost": [
                    r"(?i)total\s*project\s*costs?",
                    r"(?i)total\s*development\s*costs?",
                    r"(?i)total\s*costs?",
                    r"(?i)sum\s*of\s*costs"
                ],
                "land": [
                    r"(?i)land\s*costs?",
                    r"(?i)land\s*acquisition",
                    r"(?i)land\s*value",
                    r"(?i)cost\s*of\s*land"
                ],
                "improvements": [
                    r"(?i)existing\s*improvements",
                    r"(?i)building\s*acquisition",
                    r"(?i)existing\s*structures",
                    r"(?i)improvements\s*acquisition"
                ],
                "new_construction": [
                    r"(?i)new\s*construction",
                    r"(?i)structures",
                    r"(?i)building\s*construction"
                ],
                "rehabilitation": [
                    r"(?i)rehabilitation",
                    r"(?i)rehab",
                    r"(?i)renovation"
                ],
                "developer_fee": [
                    r"(?i)developer\s*fee",
                    r"(?i)developer\s*costs?",
                    r"(?i)development\s*fee"
                ],
                "deferred_developer_fee": [
                    r"(?i)deferred\s*developer\s*fee",
                    r"(?i)deferred\s*fee"
                ]
            }
            
            # Scan for key fields - search more rows
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in patterns):
                                # Found a field label, look for value to the right and in cells below
                                found_value = False
                                
                                # Search to the right first - try more columns
                                for offset in range(1, 8):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)) and data_value > 0:
                                            # Found a numeric value
                                            found_value = True
                                            if field == "total_project_cost":
                                                result["costs"]["total_project_cost"] = data_value
                                            elif field == "land":
                                                result["costs"]["acquisition_costs"]["land"] = data_value
                                            elif field == "improvements":
                                                result["costs"]["acquisition_costs"]["improvements"] = data_value
                                            elif field == "new_construction":
                                                result["costs"]["construction_costs"]["new_construction"] = data_value
                                            elif field == "rehabilitation":
                                                result["costs"]["construction_costs"]["rehabilitation"] = data_value
                                            elif field == "developer_fee":
                                                result["costs"]["developer_fee"]["total"] = data_value
                                            elif field == "deferred_developer_fee":
                                                result["costs"]["developer_fee"]["deferred"] = data_value
                                            break
                                
                                # If not found, try below - for a few rows
                                if not found_value:
                                    for offset in range(1, 4):
                                        if row + offset <= sheet.max_row:
                                            data_value = sheet.cell(row=row+offset, column=col).value
                                            if data_value and isinstance(data_value, (int, float)) and data_value > 0:
                                                # Found a numeric value
                                                if field == "total_project_cost":
                                                    result["costs"]["total_project_cost"] = data_value
                                                elif field == "land":
                                                    result["costs"]["acquisition_costs"]["land"] = data_value
                                                elif field == "improvements":
                                                    result["costs"]["acquisition_costs"]["improvements"] = data_value
                                                elif field == "new_construction":
                                                    result["costs"]["construction_costs"]["new_construction"] = data_value
                                                elif field == "rehabilitation":
                                                    result["costs"]["construction_costs"]["rehabilitation"] = data_value
                                                elif field == "developer_fee":
                                                    result["costs"]["developer_fee"]["total"] = data_value
                                                elif field == "deferred_developer_fee":
                                                    result["costs"]["developer_fee"]["deferred"] = data_value
                                                break
            
            # Try to extract sources information - look for common loan patterns
            source_patterns = [
                r"(?i)permanent\s*loan",
                r"(?i)tax\s*credit\s*equity",
                r"(?i)(state|county|city|local)\s*funding",
                r"(?i)soft\s*loan",
                r"(?i)deferred\s*fee"
            ]
            
            # Scan for sources
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check if this might be a source
                        if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in source_patterns):
                            # Look for amount to the right
                            for offset in range(1, 5):
                                if col + offset <= sheet.max_column:
                                    amount = sheet.cell(row=row, column=col+offset).value
                                    if amount and isinstance(amount, (int, float)) and amount > 0:
                                        # Found a source with amount
                                        source = {
                                            "name": cell_value.strip(),
                                            "amount": amount
                                        }
                                        
                                        # Determine if permanent or construction
                                        if "construction" in cell_value.lower() or "bridge" in cell_value.lower():
                                            result["sources"]["construction"].append(source)
                                        else:
                                            result["sources"]["permanent"].append(source)
                                        break
            
            # Calculate acquisition total
            land = result["costs"]["acquisition_costs"]["land"]
            improvements = result["costs"]["acquisition_costs"]["improvements"]
            result["costs"]["acquisition_costs"]["total"] = land + improvements
            
            # Calculate construction total
            new_construction = result["costs"]["construction_costs"]["new_construction"]
            rehabilitation = result["costs"]["construction_costs"]["rehabilitation"]
            result["costs"]["construction_costs"]["total"] = new_construction + rehabilitation
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Sources and Uses tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_basis_credits_tab(self, file_path, tab_name):
        """
        Extract Basis & Credits information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted basis and credits data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Basis & Credits tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "basis": {
                "total_eligible_basis": 0,
                "threshold_basis_limit": 0
            },
            "credits": {
                "federal_annual": 0,
                "federal_total": 0,
                "state_annual": 0,
                "state_total": 0,
                "credit_price": 0,
                "equity_amount": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Look for relevant tab names if the exact one isn't found
            found_tab = False
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                found_tab = True
            else:
                for sheet_name in wb.sheetnames:
                    if "Basis" in sheet_name or "Credit" in sheet_name:
                        sheet = wb[sheet_name]
                        found_tab = True
                        break
                
                if not found_tab:
                    logging.warning(f"No Basis & Credits tab found in {file_name}")
                    return result
            
            # Common labels to search for in Basis & Credits - more variations
            key_fields = {
                "total_eligible_basis": [
                    r"(?i)total\s*eligible\s*basis",
                    r"(?i)eligible\s*basis\s*total",
                    r"(?i)total\s*basis",
                    r"(?i)sum\s*of\s*eligible\s*basis"
                ],
                "threshold_basis_limit": [
                    r"(?i)threshold\s*basis\s*limit",
                    r"(?i)adjusted\s*threshold\s*basis\s*limit",
                    r"(?i)basis\s*limit",
                    r"(?i)maximum\s*basis"
                ],
                "federal_annual": [
                    r"(?i)annual\s*federal\s*credit",
                    r"(?i)federal\s*credit\s*annual",
                    r"(?i)federal\s*tax\s*credit\s*annual",
                    r"(?i)annual\s*federal\s*tax\s*credit"
                ],
                "state_annual": [
                    r"(?i)annual\s*state\s*credit",
                    r"(?i)state\s*credit\s*annual",
                    r"(?i)state\s*tax\s*credit\s*annual",
                    r"(?i)annual\s*state\s*tax\s*credit"
                ],
                "credit_price": [
                    r"(?i)credit\s*price",
                    r"(?i)price\s*per\s*credit",
                    r"(?i)tax\s*credit\s*factor",
                    r"(?i)syndication\s*rate"
                ]
            }
            
            # Scan for key fields - more rows and columns
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(25, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in patterns):
                                # Found a field label, look for value around it
                                found_value = False
                                
                                # Search to the right first
                                for offset in range(1, 8):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value and isinstance(data_value, (int, float)):
                                            # Found a numeric value
                                            found_value = True
                                            if field == "total_eligible_basis":
                                                result["basis"]["total_eligible_basis"] = data_value
                                            elif field == "threshold_basis_limit":
                                                result["basis"]["threshold_basis_limit"] = data_value
                                            elif field == "federal_annual":
                                                result["credits"]["federal_annual"] = data_value
                                            elif field == "state_annual":
                                                result["credits"]["state_annual"] = data_value
                                            elif field == "credit_price":
                                                result["credits"]["credit_price"] = data_value
                                            break
                                
                                # If not found, try below
                                if not found_value:
                                    for offset in range(1, 4):
                                        if row + offset <= sheet.max_row:
                                            data_value = sheet.cell(row=row+offset, column=col).value
                                            if data_value and isinstance(data_value, (int, float)):
                                                if field == "total_eligible_basis":
                                                    result["basis"]["total_eligible_basis"] = data_value
                                                elif field == "threshold_basis_limit":
                                                    result["basis"]["threshold_basis_limit"] = data_value
                                                elif field == "federal_annual":
                                                    result["credits"]["federal_annual"] = data_value
                                                elif field == "state_annual":
                                                    result["credits"]["state_annual"] = data_value
                                                elif field == "credit_price":
                                                    result["credits"]["credit_price"] = data_value
                                                break
            
            # Calculate derived fields
            if result["credits"]["federal_annual"]:
                result["credits"]["federal_total"] = result["credits"]["federal_annual"] * 10
            
            if result["credits"]["state_annual"]:
                result["credits"]["state_total"] = result["credits"]["state_annual"]
            
            if result["credits"]["federal_annual"] and result["credits"]["credit_price"]:
                result["credits"]["equity_amount"] = result["credits"]["federal_annual"] * 10 * result["credits"]["credit_price"]
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Basis & Credits tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_points_system_tab(self, file_path, tab_name):
        """
        Extract Points System information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted points data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Points System tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "scoring": {
                "total_score": 0,
                "general_partner_experience": 0,
                "management_company_experience": 0,
                "housing_needs": 0,
                "site_amenities": 0,
                "service_amenities": 0,
                "sustainable_building_methods": 0,
                "lowest_income": 0,
                "readiness_to_proceed": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Look for relevant tab names if the exact one isn't found
            found_tab = False
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                found_tab = True
            else:
                for sheet_name in wb.sheetnames:
                    if "Point" in sheet_name or "Score" in sheet_name:
                        sheet = wb[sheet_name]
                        found_tab = True
                        break
                
                if not found_tab:
                    logging.warning(f"No Points System tab found in {file_name}")
                    return result
            
            # Common labels to search for in Points System - more variations
            key_fields = {
                "total_score": [
                    r"(?i)total\s*score",
                    r"(?i)total\s*points",
                    r"(?i)sum\s*of\s*points",
                    r"(?i)overall\s*score"
                ],
                "general_partner_experience": [
                    r"(?i)general\s*partner\s*experience",
                    r"(?i)gp\s*experience",
                    r"(?i)developer\s*experience"
                ],
                "management_company_experience": [
                    r"(?i)management\s*company\s*experience",
                    r"(?i)management\s*experience",
                    r"(?i)property\s*management\s*experience"
                ],
                "housing_needs": [
                    r"(?i)housing\s*needs",
                    r"(?i)housing\s*type",
                    r"(?i)housing\s*category"
                ],
                "site_amenities": [
                    r"(?i)site\s*amenities",
                    r"(?i)location\s*amenities"
                ],
                "service_amenities": [
                    r"(?i)service\s*amenities",
                    r"(?i)resident\s*services"
                ],
                "sustainable_building_methods": [
                    r"(?i)sustainable\s*building",
                    r"(?i)energy\s*efficiency",
                    r"(?i)green\s*building"
                ],
                "lowest_income": [
                    r"(?i)lowest\s*income",
                    r"(?i)serving\s*lowest\s*income",
                    r"(?i)deeper\s*targeting"
                ],
                "readiness_to_proceed": [
                    r"(?i)readiness\s*to\s*proceed",
                    r"(?i)ready\s*to\s*proceed"
                ]
            }
            
            # Scan for key fields - more rows
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(25, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in patterns):
                                # Found a field label, search for points value
                                found_value = False
                                
                                # Search more columns to the right for potential points
                                for offset in range(1, 12):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value is not None and isinstance(data_value, (int, float)):
                                            # Found a potential points value
                                            # Points should typically be between 0 an# Check if it's a reasonable value for points
                                            # LIHTC points typically range from 0-100, with individual categories usually 0-20
                                            if 0 <= data_value <= 100:
                                                found_value = True
                                                result["scoring"][field] = data_value
                                                break
                                
                                # If not found horizontally, try looking below
                                if not found_value:
                                    for offset in range(1, 5):
                                        if row + offset <= sheet.max_row:
                                            data_value = sheet.cell(row=row+offset, column=col).value
                                            if data_value is not None and isinstance(data_value, (int, float)):
                                                if 0 <= data_value <= 100:
                                                    result["scoring"][field] = data_value
                                                    break
            
            # If we didn't find a total score, try to calculate it from components
            if result["scoring"]["total_score"] == 0:
                total = sum(value for key, value in result["scoring"].items() if key != "total_score")
                if total > 0:
                    result["scoring"]["total_score"] = total
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Points System tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_tie_breaker_tab(self, file_path, tab_name):
        """
        Extract Tie Breaker information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted tie breaker data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Tie Breaker tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "tie_breaker": {
                "score": 0,
                "leveraged_soft_resources": 0,
                "ratio_requested_to_total": 0,
                "resource_area_bonus": 0
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Look for relevant tab names if the exact one isn't found
            found_tab = False
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                found_tab = True
            else:
                for sheet_name in wb.sheetnames:
                    if "Tie" in sheet_name or "Breaker" in sheet_name or "Tiebreaker" in sheet_name:
                        sheet = wb[sheet_name]
                        found_tab = True
                        break
                
                # Check Points System tab as a fallback
                if not found_tab:
                    for sheet_name in wb.sheetnames:
                        if "Point" in sheet_name or "Score" in sheet_name:
                            sheet = wb[sheet_name]
                            found_tab = True
                            break
                
                if not found_tab:
                    logging.warning(f"No Tie Breaker tab found in {file_name}")
                    return result
            
            # Common labels to search for in Tie Breaker - more variations
            key_fields = {
                "score": [
                    r"(?i)tie\s*breaker\s*score",
                    r"(?i)final\s*tie\s*breaker",
                    r"(?i)tiebreaker\s*score",
                    r"(?i)overall\s*tiebreaker"
                ],
                "leveraged_soft_resources": [
                    r"(?i)leveraged\s*soft\s*resources",
                    r"(?i)soft\s*resources\s*ratio",
                    r"(?i)public\s*funds",
                    r"(?i)leveraged\s*funds"
                ],
                "ratio_requested_to_total": [
                    r"(?i)ratio\s*of\s*requested\s*unadjusted\s*eligible\s*basis",
                    r"(?i)1\s*minus\s*ratio\s*of\s*requested",
                    r"(?i)adjusted\s*ratio",
                    r"(?i)basis\s*ratio"
                ],
                "resource_area_bonus": [
                    r"(?i)resource\s*area\s*bonus",
                    r"(?i)highest\s*resource\s*area",
                    r"(?i)resource\s*area\s*score",
                    r"(?i)location\s*bonus"
                ]
            }
            
            # Scan for key fields - more thorough search
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(25, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in patterns):
                                # Found a field label, search for value
                                found_value = False
                                
                                # Search more columns to the right
                                for offset in range(1, 12):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value is not None and isinstance(data_value, (int, float)):
                                            found_value = True
                                            result["tie_breaker"][field] = data_value
                                            break
                                
                                # If not found, look below
                                if not found_value:
                                    for offset in range(1, 5):
                                        if row + offset <= sheet.max_row:
                                            data_value = sheet.cell(row=row+offset, column=col).value
                                            if data_value is not None and isinstance(data_value, (int, float)):
                                                result["tie_breaker"][field] = data_value
                                                break
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Tie Breaker tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def extract_pro_forma_tab(self, file_path, tab_name):
        """
        Extract 15 Year Pro Forma information
        
        Args:
            file_path: Path to Excel file
            tab_name: Name of the tab to extract
            
        Returns:
            dict: Extracted pro forma data
        """
        file_name = os.path.basename(file_path)
        logging.info(f"Extracting Pro Forma tab from {file_name}")
        
        # Initialize result
        result = {
            "metadata": {
                "file_name": file_name,
                "tab": tab_name,
                "extraction_date": datetime.now().isoformat()
            },
            "pro_forma": {
                "year_1": {
                    "gross_income": 0,
                    "vacancy": 0,
                    "effective_gross_income": 0,
                    "operating_expenses": 0,
                    "net_operating_income": 0,
                    "debt_service": 0,
                    "cash_flow": 0,
                    "debt_service_coverage": 0
                }
            }
        }
        
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            
            # Look for relevant tab names if the exact one isn't found
            found_tab = False
            if tab_name in wb.sheetnames:
                sheet = wb[tab_name]
                found_tab = True
            else:
                for sheet_name in wb.sheetnames:
                    if "Pro" in sheet_name or "Forma" in sheet_name or "Cash" in sheet_name or "Flow" in sheet_name:
                        sheet = wb[sheet_name]
                        found_tab = True
                        break
                
                if not found_tab:
                    logging.warning(f"No Pro Forma tab found in {file_name}")
                    return result
            
            # Common labels to search for in Pro Forma - more variations
            key_fields = {
                "gross_income": [
                    r"(?i)gross\s*income",
                    r"(?i)gross\s*potential\s*income",
                    r"(?i)potential\s*gross\s*income",
                    r"(?i)total\s*potential\s*income"
                ],
                "vacancy": [
                    r"(?i)vacancy",
                    r"(?i)vacancy\s*loss",
                    r"(?i)vacancy\s*rate",
                    r"(?i)vacancy\s*factor"
                ],
                "effective_gross_income": [
                    r"(?i)effective\s*gross\s*income",
                    r"(?i)total\s*income",
                    r"(?i)net\s*rental\s*income",
                    r"(?i)adjusted\s*gross\s*income"
                ],
                "operating_expenses": [
                    r"(?i)operating\s*expenses",
                    r"(?i)total\s*expenses",
                    r"(?i)total\s*operating\s*expenses",
                    r"(?i)annual\s*expenses"
                ],
                "net_operating_income": [
                    r"(?i)net\s*operating\s*income",
                    r"(?i)NOI",
                    r"(?i)income\s*before\s*debt\s*service"
                ],
                "debt_service": [
                    r"(?i)debt\s*service",
                    r"(?i)total\s*debt\s*service",
                    r"(?i)loan\s*payment",
                    r"(?i)mortgage\s*payment"
                ],
                "cash_flow": [
                    r"(?i)cash\s*flow",
                    r"(?i)cash\s*flow\s*after\s*debt\s*service",
                    r"(?i)net\s*cash\s*flow",
                    r"(?i)residual\s*cash\s*flow"
                ],
                "debt_service_coverage": [
                    r"(?i)debt\s*service\s*coverage",
                    r"(?i)debt\s*coverage\s*ratio",
                    r"(?i)DCR",
                    r"(?i)DSCR"
                ]
            }
            
            # Scan for key fields
            for row in range(1, min(300, sheet.max_row + 1)):
                for col in range(1, min(25, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value and isinstance(cell_value, str):
                        # Check against field patterns
                        for field, patterns in key_fields.items():
                            if any(re.search(pattern, cell_value, re.IGNORECASE) for pattern in patterns):
                                # Found a field label, look for Year 1 value to the right
                                # Pro forma tables typically have years in columns, so scan horizontally
                                
                                # First, try to identify which column is Year 1
                                year1_col = None
                                for y_col in range(col, min(col + 15, sheet.max_column + 1)):
                                    y_header = sheet.cell(row=row-1, column=y_col).value
                                    if y_header is not None:
                                        # Look for variations of Year 1
                                        if (isinstance(y_header, str) and ("Year 1" in y_header or "Year-1" in y_header or 
                                             "Year One" in y_header or "1st Year" in y_header)) or y_header == 1:
                                            year1_col = y_col
                                            break
                                
                                # If we identified Year 1 column, use it
                                if year1_col:
                                    data_value = sheet.cell(row=row, column=year1_col).value
                                    if data_value is not None and isinstance(data_value, (int, float)):
                                        result["pro_forma"]["year_1"][field] = data_value
                                        continue
                                
                                # If we didn't find Year 1 specifically, try the first numeric column
                                # This assumes the first year is the first data column
                                for offset in range(1, 10):
                                    if col + offset <= sheet.max_column:
                                        data_value = sheet.cell(row=row, column=col+offset).value
                                        if data_value is not None and isinstance(data_value, (int, float)):
                                            result["pro_forma"]["year_1"][field] = data_value
                                            break
            
            return result
            
        except Exception as e:
            logging.error(f"Error extracting Pro Forma tab from {file_name}: {e}")
            logging.error(traceback.format_exc())
            return result
    
    def save_tab_data(self, data, app_type, file_name, tab_name):
        """
        Save tab data to JSON file
        
        Args:
            data: Extracted data
            app_type: Application type (4p or 9p)
            file_name: Original file name
            tab_name: Tab name
            
        Returns:
            str: Path to saved file
        """
        # Create output directory structure - using the correct paths
        output_dir = os.path.join(self.output_paths[app_type], tab_name.lower().replace(' ', '_'))
        os.makedirs(output_dir, exist_ok=True)
        
        # Create output filename
        output_file = os.path.join(output_dir, f"{file_name.replace('.xlsx', '')}.json")
        
        # Save data
        with open(output_file, 'w') as f:
            json.dump(data, f, indent=2)
        
        logging.info(f"Saved {tab_name} data to {output_file}")
        return output_file
    
    def process_files(self, app_type, limit=None):
        """
        Process a batch of files
        
        Args:
            app_type: Application type (4p or 9p)
            limit: Maximum number of files to process
        """
        files = self.get_application_files(app_type, limit)
        
        # Define tabs to extract and their corresponding extractors
        tabs_to_extract = {
            "Application": self.extract_application_tab,
            "Sources and Uses Budget": self.extract_sources_uses_tab,
            "Basis & Credits": self.extract_basis_credits_tab,
            "Points System": self.extract_points_system_tab,
            "Tie Breaker": self.extract_tie_breaker_tab,
            "15 Year Pro Forma": self.extract_pro_forma_tab
        }
        
        # Initialize tab stats
        for tab in tabs_to_extract.keys():
            self.stats["tabs_extracted"][tab] = 0
        
        # Process each file
        for file_path in files:
            file_name = os.path.basename(file_path)
            logging.info(f"Processing {file_name}")
            
            try:
                # Extract each tab
                for tab_name, extractor in tabs_to_extract.items():
                    try:
                        data = extractor(file_path, tab_name)
                        self.save_tab_data(data, app_type, file_name, tab_name)
                        self.stats["tabs_extracted"][tab_name] += 1
                    except Exception as e:
                        logging.error(f"Error extracting {tab_name} from {file_name}: {e}")
                        self.stats["errors"].append(f"Error extracting {tab_name} from {file_name}: {str(e)}")
                
                self.stats["files_processed"] += 1
                logging.info(f"Completed processing {file_name}")
                
            except Exception as e:
                logging.error(f"Error processing {file_name}: {e}")
                logging.error(traceback.format_exc())
                self.stats["errors"].append(f"Error processing {file_name}: {str(e)}")
    
    def generate_summary(self):
        """
        Generate summary of extraction process
        
        Returns:
            dict: Summary statistics
        """
        self.stats["end_time"] = datetime.now().isoformat()
        start_time = datetime.fromisoformat(self.stats["start_time"])
        end_time = datetime.fromisoformat(self.stats["end_time"])
        duration = end_time - start_time
        
        self.stats["duration_seconds"] = duration.total_seconds()
        self.stats["duration_formatted"] = str(duration)
        
        # Log summary
        logging.info(f"Extraction completed. Processed {self.stats['files_processed']} files in {self.stats['duration_formatted']}")
        for tab, count in self.stats["tabs_extracted"].items():
            logging.info(f"Extracted {count} {tab} tabs")
        
        if self.stats["errors"]:
            logging.warning(f"Encountered {len(self.stats['errors'])} errors")
        
        # Save summary to file - to input directory root to make it easy to find
        summary_file = os.path.join(self.input_path, "extraction_summary.json")
        with open(summary_file, 'w') as f:
            json.dump(self.stats, f, indent=2)
        
        logging.info(f"Saved summary to {summary_file}")
        return self.stats

# Main execution function
def main():
    # Define paths using the exact paths specified
    input_path = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data'
    output_path_4p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p'
    output_path_9p = '/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p'
    
    # Create extractor with the correct paths
    extractor = LIHTCTabExtractor(input_path, output_path_4p, output_path_9p)
    
    # Process just 3 files from each application type for testing
    logging.info("Starting extraction of 4% applications (test sample)...")
    extractor.process_files('4p', limit=3)  # Process only 3 files for testing
    
    # Process 9% applications
    logging.info("Starting extraction of 9% applications (test sample)...")
    extractor.process_files('9p', limit=3)  # Process only 3 files for testing
    
    # Generate summary
    summary = extractor.generate_summary()
    print(f"Extraction complete! Processed {summary['files_processed']} files.")

if __name__ == "__main__":
    main()