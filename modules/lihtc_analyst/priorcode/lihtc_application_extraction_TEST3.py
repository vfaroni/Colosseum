import os
import sys
import json
import pandas as pd
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from datetime import datetime
import logging
import re
from pathlib import Path
import colorama
from colorama import Fore, Style

# Initialize colorama for colored terminal output
colorama.init()

# Set up logging
log_dir = Path('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/logs')
log_dir.mkdir(parents=True, exist_ok=True)
log_file = log_dir / f"extraction_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Define paths
raw_data_path = Path('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data')
output_path_4p = Path('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/4p')
output_path_9p = Path('/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CTCAC_RAG/raw_data/JSON_data/9p')

# Create output directories if they don't exist
output_path_4p.mkdir(parents=True, exist_ok=True)
output_path_9p.mkdir(parents=True, exist_ok=True)

def is_yellow_cell(cell):
    """
    Check if a cell has yellow fill color.
    Yellow in Excel can be represented in different ways.
    """
    # Common representations of yellow in Excel
    yellow_indices = ['FFFF00', 'FFFF00FF', 'FFFFFF00', 6]
    
    # Get the fill color index/RGB/etc
    fill_color = None
    if hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'index'):
        fill_color = cell.fill.start_color.index
    elif hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'rgb'):
        fill_color = cell.fill.start_color.rgb
    elif hasattr(cell.fill, 'fgColor') and hasattr(cell.fill.fgColor, 'rgb'):
        fill_color = cell.fill.fgColor.rgb
    
    # Check if the cell's fill color matches any of the yellow representations
    if fill_color and (fill_color in yellow_indices or str(fill_color) in yellow_indices):
        return True
    
    # Check for specific fill_type that indicates a solid fill
    if hasattr(cell.fill, 'fill_type') and cell.fill.fill_type == 'solid':
        # Some versions might use different attributes
        if hasattr(cell.fill, 'fgColor') and hasattr(cell.fill.fgColor, 'rgb'):
            if cell.fill.fgColor.rgb and cell.fill.fgColor.rgb.startswith('FFFF'):
                return True
    
    return False

def get_cell_value(sheet, cell_reference, default=""):
    """
    Safely get a cell value from a sheet.
    """
    try:
        value = sheet[cell_reference].value
        if value is None:
            return default
        return str(value).strip()
    except Exception:
        return default

def find_section_header(sheet, row):
    """
    Find the section header for a given row by looking up to find
    a row with a header (typically in column A with ** markers).
    """
    for i in range(row, 0, -1):
        cell_value = get_cell_value(sheet, f"A{i}")
        # Look for section headers which often have ** or are in ALL CAPS
        if "**" in cell_value or (cell_value.isupper() and len(cell_value) > 5):
            return cell_value
    return ""

def find_row_header(sheet, row, col):
    """
    Find the row header for a cell, which is typically in column A 
    of the same row.
    """
    # Check column A in the same row
    row_header = get_cell_value(sheet, f"A{row}")
    
    # If no header in column A, check neighboring columns
    if not row_header:
        for c in range(1, col):
            candidate = get_cell_value(sheet, f"{get_column_letter(c)}{row}")
            if candidate and len(candidate) > 3:  # Reasonable length for a header
                row_header = candidate
                break
    
    return row_header

def find_header_for_cell(sheet, row, col):
    """
    Find an appropriate header for a cell by looking at section headers
    and row headers.
    """
    section_header = find_section_header(sheet, row)
    row_header = find_row_header(sheet, row, col)
    
    # Look for a merged cell that might contain a header
    for merged_range in sheet.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1):
            # If it's a header merged cell and not the cell we're looking at
            if merged_range.min_col == 1 and col > merged_range.max_col:
                merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                if merged_value:
                    row_header = str(merged_value).strip()
    
    # Combine headers if both exist
    if section_header and row_header:
        return f"{section_header} - {row_header}"
    elif section_header:
        return section_header
    elif row_header:
        return row_header
    else:
        # If no headers found, look at cells to the left as potential labels
        for c in range(col-1, 0, -1):
            label = get_cell_value(sheet, f"{get_column_letter(c)}{row}")
            if label and len(label) > 2:  # Reasonable length for a label
                return label
        
        # If still no header, use the cell reference
        return f"Cell {get_column_letter(col)}{row}"

def parse_filename(file_name):
    """
    Parse CTCAC application information from the filename.
    Examples:
    - 2025_4pct_R1_25494.xlsx
    - 2024_9pct_R1_24001.xlsx
    """
    metadata = {
        "file_name": file_name,
        "extraction_date": datetime.now().isoformat()
    }
    
    # Pattern to match: YYYY_9pct_RZ_NNNNN or YYYY_4pct_RZ_NNNNN
    pattern = r"(\d{4})_(\d[p|P](?:ct)?)_R(\d+)_(\d+)"
    match = re.search(pattern, file_name)
    
    if match:
        metadata["application_year"] = match.group(1)
        
        # Handle app type (4p or 9p)
        app_type = match.group(2).lower()
        if app_type.startswith("4"):
            metadata["application_type"] = "4p"
        elif app_type.startswith("9"):
            metadata["application_type"] = "9p"
        
        metadata["application_round"] = match.group(3)
        metadata["ctcac_number"] = match.group(4)
    else:
        # Try alternative patterns
        year_match = re.search(r"(\d{4})", file_name)
        if year_match:
            metadata["application_year"] = year_match.group(1)
        
        if "4pct" in file_name.lower() or "4p" in file_name.lower():
            metadata["application_type"] = "4p"
        elif "9pct" in file_name.lower() or "9p" in file_name.lower():
            metadata["application_type"] = "9p"
        
        round_match = re.search(r"R(\d+)", file_name)
        if round_match:
            metadata["application_round"] = round_match.group(1)
        
        # Try to find CTCAC number at the end
        number_match = re.search(r"_(\d+)(?:\.xlsx)?$", file_name)
        if number_match:
            metadata["ctcac_number"] = number_match.group(1)
    
    return metadata

def extract_application_data(file_path):
    """
    Extract data from the Application tab of a LIHTC application.
    """
    logger.info(f"Processing file: {file_path}")
    print(f"{Fore.CYAN}Processing: {os.path.basename(file_path)}{Style.RESET_ALL}")
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Check if 'Application' tab exists
        if 'Application' not in workbook.sheetnames:
            logger.error(f"No 'Application' tab found in {file_path}")
            print(f"{Fore.RED}Error: No 'Application' tab found{Style.RESET_ALL}")
            return None
        
        sheet = workbook['Application']
        
        # Initialize data structure
        metadata = parse_filename(os.path.basename(file_path))
        data = {
            "metadata": metadata,
            "yellow_input_cells": {},
            "total_cells": {},
            "section_headers": {},
            "threshold_basis_limits": {},
            "project": {
                "name": "",
                "address": "",
                "city": "",
                "county": "",
                "zip_code": "",
                "census_tract": "",
                "total_units": 0,
                "low_income_units": 0,
                "manager_units": 0
            }
        }
        
        # Track the current section
        current_section = ""
        
        # Get max rows and columns to scan
        max_row = min(2000, sheet.max_row)  # Limit to 2000 rows to prevent excessive scanning
        max_col = min(100, sheet.max_column)  # Limit to 100 columns
        
        print(f"{Fore.YELLOW}Scanning {max_row} rows and {max_col} columns...{Style.RESET_ALL}")
        
        # Extract project information first (usually at the top)
        for row in range(1, 100):  # Check first 100 rows for project info
            a_cell_value = get_cell_value(sheet, f"A{row}")
            if "PROJECT" in a_cell_value.upper() and "INFORMATION" in a_cell_value.upper():
                # Found project information section
                project_section_start = row
                # Look within the next 20 rows for key project info
                for r in range(project_section_start, project_section_start + 20):
                    for c in range(1, 15):  # First 15 columns should cover it
                        cell = sheet.cell(row=r, column=c)
                        if is_yellow_cell(cell) and cell.value:
                            # Try to determine what field this is
                            row_header = find_row_header(sheet, r, c)
                            if row_header:
                                row_header_upper = row_header.upper()
                                if "PROJECT NAME" in row_header_upper or "NAME OF PROJECT" in row_header_upper:
                                    data["project"]["name"] = str(cell.value)
                                elif "ADDRESS" in row_header_upper and "EMAIL" not in row_header_upper:
                                    data["project"]["address"] = str(cell.value)
                                elif "CITY" in row_header_upper:
                                    data["project"]["city"] = str(cell.value)
                                elif "COUNTY" in row_header_upper:
                                    data["project"]["county"] = str(cell.value)
                                elif "ZIP" in row_header_upper:
                                    data["project"]["zip_code"] = str(cell.value)
                                elif "CENSUS" in row_header_upper and "TRACT" in row_header_upper:
                                    data["project"]["census_tract"] = str(cell.value)
                                elif ("TOTAL" in row_header_upper and "UNITS" in row_header_upper) or "# OF UNITS" in row_header_upper:
                                    try:
                                        data["project"]["total_units"] = int(float(cell.value)) if cell.value else 0
                                    except (ValueError, TypeError):
                                        pass
                                elif "LOW INCOME" in row_header_upper and "UNITS" in row_header_upper:
                                    try:
                                        data["project"]["low_income_units"] = int(float(cell.value)) if cell.value else 0
                                    except (ValueError, TypeError):
                                        pass
                                elif "MANAGER" in row_header_upper and "UNITS" in row_header_upper:
                                    try:
                                        data["project"]["manager_units"] = int(float(cell.value)) if cell.value else 0
                                    except (ValueError, TypeError):
                                        pass
                break
        
        # Process the whole sheet to gather all data
        yellow_cell_count = 0
        total_cell_count = 0
        threshold_basis_count = 0
        
        for row in range(1, max_row + 1):
            # Check for section headers in column A
            cell_a = sheet.cell(row=row, column=1)
            cell_a_value = cell_a.value
            
            if cell_a_value and isinstance(cell_a_value, str):
                # If it looks like a section header (** or all caps or similar formatting)
                if "**" in cell_a_value or (cell_a_value.isupper() and len(cell_a_value) > 5):
                    current_section = cell_a_value.strip()
                    data["section_headers"][f"A{row}"] = current_section
                    
                    # Check if this is the threshold basis limit section
                    if "THRESHOLD BASIS LIMIT" in current_section.upper():
                        logger.info(f"Found threshold basis limit section at row {row}")
                        print(f"{Fore.GREEN}Found threshold basis limit section at row {row}{Style.RESET_ALL}")
                        
                        # Special handling for threshold basis limit section
                        threshold_section_row = row
                        
                        # Scan the next 100 rows for basis adjustment data
                        for basis_row in range(threshold_section_row, threshold_section_row + 100):
                            if basis_row <= max_row:
                                # Look in A column for descriptions
                                basis_desc = get_cell_value(sheet, f"A{basis_row}")
                                
                                # Look for adjusted basis amounts in columns
                                for basis_col in range(2, max_col + 1):
                                    basis_cell = sheet.cell(row=basis_row, column=basis_col)
                                    if basis_cell.value and basis_cell.value != 0:
                                        # Try to convert to a number to see if it's a value
                                        try:
                                            float(basis_cell.value)
                                            # It's a numeric value, likely an adjustment
                                            cell_ref = f"{get_column_letter(basis_col)}{basis_row}"
                                            data["threshold_basis_limits"][cell_ref] = {
                                                "description": basis_desc,
                                                "value": str(basis_cell.value)
                                            }
                                            threshold_basis_count += 1
                                        except (ValueError, TypeError):
                                            # Not a numeric value, skip
                                            pass
            
            # Process all cells in this row
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"
                
                # 1. Check for yellow input cells
                if is_yellow_cell(cell):
                    header = find_header_for_cell(sheet, row, col)
                    data["yellow_input_cells"][cell_address] = {
                        "value": str(cell.value) if cell.value is not None else "",
                        "header": header,
                        "section": current_section
                    }
                    yellow_cell_count += 1
                
                # 2. Check for cells containing "Total"
                elif cell.value and isinstance(cell.value, str) and "Total" in cell.value:
                    # Look to the right for total values (up to 5 cells)
                    for offset in range(1, 6):
                        total_col = col + offset
                        if total_col <= max_col:
                            total_cell = sheet.cell(row=row, column=total_col)
                            if total_cell.value is not None:
                                # Try to see if it's a number
                                try:
                                    float(total_cell.value)
                                    is_numeric = True
                                except (ValueError, TypeError):
                                    is_numeric = False
                                
                                if is_numeric:
                                    total_key = f"{cell.value.strip()} ({get_column_letter(total_col)}{row})"
                                    data["total_cells"][total_key] = {
                                        "value": str(total_cell.value),
                                        "section": current_section
                                    }
                                    total_cell_count += 1
                                    break
        
        # Specifically look for the threshold basis section if not already found
        if threshold_basis_count == 0:
            # Try to find "THRESHOLD BASIS LIMIT" section
            for row in range(1, max_row + 1):
                cell_value = get_cell_value(sheet, f"A{row}")
                if "THRESHOLD BASIS LIMIT" in cell_value.upper():
                    print(f"{Fore.GREEN}Found threshold basis limit section at row {row} (second pass){Style.RESET_ALL}")
                    
                    # Scan for the adjusted basis amounts
                    for basis_row in range(row, min(row + 100, max_row + 1)):
                        a_cell = sheet.cell(row=basis_row, column=1)
                        a_value = a_cell.value
                        
                        if a_value and isinstance(a_value, str) and ("adjust" in a_value.lower() or "increas" in a_value.lower() or "basis" in a_value.lower()):
                            # This row likely has basis adjustments
                            description = a_value
                            
                            # Look across the row for values
                            for col in range(2, max_col + 1):
                                value_cell = sheet.cell(row=basis_row, column=col)
                                if value_cell.value and value_cell.value != 0:
                                    try:
                                        float(value_cell.value)  # Test if it's a number
                                        cell_ref = f"{get_column_letter(col)}{basis_row}"
                                        data["threshold_basis_limits"][cell_ref] = {
                                            "description": description,
                                            "value": str(value_cell.value)
                                        }
                                        threshold_basis_count += 1
                                    except (ValueError, TypeError):
                                        pass
        
        # If project name is still empty, try to find it in the yellow input cells
        if not data["project"]["name"]:
            for cell_ref, cell_info in data["yellow_input_cells"].items():
                if "PROJECT NAME" in cell_info["header"].upper() or "NAME OF PROJECT" in cell_info["header"].upper():
                    data["project"]["name"] = cell_info["value"]
                    break
        
        # If we still don't have a project name, use the CTCAC number
        if not data["project"]["name"] and "ctcac_number" in metadata:
            data["project"]["name"] = f"Project {metadata['ctcac_number']}"
        
        print(f"{Fore.GREEN}Extracted {yellow_cell_count} yellow cells, {total_cell_count} totals, and {threshold_basis_count} threshold basis items{Style.RESET_ALL}")
        logger.info(f"Extracted {yellow_cell_count} yellow cells, {total_cell_count} totals, and {threshold_basis_count} threshold basis items")
        
        return data
    
    except Exception as e:
        logger.error(f"Error processing {file_path}: {str(e)}", exc_info=True)
        print(f"{Fore.RED}Error: {str(e)}{Style.RESET_ALL}")
        return None

def main():
    """Main function to process files."""
    print(f"{Fore.CYAN}CTCAC Application Data Extraction Tool{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'=' * 50}{Style.RESET_ALL}")
    
    # Get list of Excel files
    excel_files = [f for f in os.listdir(raw_data_path) if f.endswith('.xlsx')]
    
    # Ask user how many files to process
    print(f"\nFound {len(excel_files)} Excel files in {raw_data_path}")
    user_input = input("How many files do you want to process? (Enter a number or 'ALL'): ")
    
    if user_input.upper() == 'ALL':
        files_to_process = excel_files
    else:
        try:
            num_files = int(user_input)
            files_to_process = excel_files[:num_files]
        except ValueError:
            logger.error("Invalid input. Please enter a number or 'ALL'")
            print(f"{Fore.RED}Invalid input. Please enter a number or 'ALL'{Style.RESET_ALL}")
            return
    
    logger.info(f"Will process {len(files_to_process)} files")
    print(f"\n{Fore.CYAN}Processing {len(files_to_process)} files...{Style.RESET_ALL}\n")
    
    # Process each file
    successful_files = 0
    failed_files = 0
    
    for i, file_name in enumerate(files_to_process):
        file_path = raw_data_path / file_name
        
        # Determine if it's a 4% or 9% application based on filename
        app_type = None
        if '4pct' in file_name.lower() or '4p' in file_name.lower():
            app_type = '4p'
            output_path = output_path_4p
        elif '9pct' in file_name.lower() or '9p' in file_name.lower():
            app_type = '9p'
            output_path = output_path_9p
        else:
            logger.warning(f"Can't determine application type for {file_name}, skipping")
            print(f"{Fore.YELLOW}Can't determine application type for {file_name}, skipping{Style.RESET_ALL}")
            failed_files += 1
            continue
        
        print(f"\n{Fore.CYAN}[{i+1}/{len(files_to_process)}] Processing: {file_name} (Type: {app_type}){Style.RESET_ALL}")
        
        # Extract data
        data = extract_application_data(file_path)
        
        if data:
            # Create output JSON filename
            json_filename = f"{os.path.splitext(file_name)[0]}.json"
            json_path = output_path / json_filename
            
            # Write to JSON file
            with open(json_path, 'w') as json_file:
                json.dump(data, json_file, indent=2)
            
            print(f"{Fore.GREEN}Saved data to {json_path}{Style.RESET_ALL}")
            logger.info(f"Saved data to {json_path}")
            successful_files += 1
        else:
            logger.error(f"Failed to extract data from {file_name}")
            print(f"{Fore.RED}Failed to extract data from {file_name}{Style.RESET_ALL}")
            failed_files += 1
    
    print(f"\n{Fore.CYAN}{'=' * 50}{Style.RESET_ALL}")
    print(f"{Fore.GREEN}Processing complete! Successfully processed {successful_files} files.{Style.RESET_ALL}")
    if failed_files > 0:
        print(f"{Fore.YELLOW}Failed to process {failed_files} files. See log for details.{Style.RESET_ALL}")
    
    print(f"\n{Fore.CYAN}Log file: {log_file}{Style.RESET_ALL}")
    
    logger.info("Processing complete!")

if __name__ == "__main__":
    main()