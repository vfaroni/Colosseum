#!/usr/bin/env python3
"""
TDHCA QCT/DDA Focused Land Analyzer with Regional Tracking
Per user preference: Focus only on QCT/DDA eligible sites (30% basis boost)
Incorporates TDHCA's 13 regions for better opportunity tracking

Author: TDHCA Analysis System
Date: 2025-06-19
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from geopy.distance import geodesic
import geopandas as gpd
from shapely.geometry import Point
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

class TDHCAQCTFocusedAnalyzer:
    """QCT/DDA focused analyzer with TDHCA regional tracking"""
    
    def __init__(self):
        # Set up logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
        # Data paths
        self.base_path = "/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/Lvg_Bond_Execution/Data Sets"
        self.tdhca_file = f"{self.base_path}/State Specific/TX/Project_List/TX_TDHCA_Project_List_05252025.xlsx"
        self.ami_file = f"{self.base_path}/HUD AMI FMR/HUD2025_AMI_Rent_Data_Static.xlsx"
        self.qct_file = f"{self.base_path}/HUD DDA QCT/QUALIFIED_CENSUS_TRACTS_7341711606021821459.gpkg"
        self.dda_file = f"{self.base_path}/HUD DDA QCT/Difficult_Development_Areas_-4200740390724245794.gpkg"
        self.schools_file = f"{self.base_path}/TX_Public_Schools/Schools_2024_to_2025.csv"
        self.regions_file = "/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/TDHCA_RAG/TDHCA_Regions/TDHCA_Regions.xlsx"
        
        # Load data
        self.tdhca_data = None
        self.ami_data = None
        self.qct_gdf = None
        self.dda_gdf = None
        self.schools_data = None
        self.regions_data = None
        
        self.load_all_data()
        
        # Large counties for 9% new construction family deals
        self.large_counties_9pct = {
            'Harris', 'Dallas', 'Tarrant', 'Bexar', 'Travis', 'Collin', 'Denton'
        }
        
        # FEMA flood zone construction cost impacts
        self.flood_cost_impacts = {
            'VE': 1.30, 'V': 1.30,   # +30% construction
            'AE': 1.20, 'A': 1.20,   # +20% construction
            'AH': 1.12, 'AO': 1.12,  # +12% construction
            'X': 1.05, 'D': 1.05     # +5% construction
        }
    
    def load_all_data(self):
        """Load all required datasets"""
        self.logger.info("Loading all datasets...")
        
        # Load TDHCA regions
        try:
            if Path(self.regions_file).exists():
                df_regions = pd.read_excel(self.regions_file, header=0)
                # Clean the data
                df_regions.columns = ['Empty', 'County', 'Region'] + list(df_regions.columns[3:])
                self.regions_data = df_regions[['County', 'Region']].dropna()
                # Remove header row if it exists
                self.regions_data = self.regions_data[self.regions_data['County'] != 'County']
                # Create lookup dictionary
                self.county_to_region = dict(zip(
                    self.regions_data['County'].str.strip(),
                    self.regions_data['Region'].str.strip()
                ))
                self.logger.info(f"✅ Loaded {len(self.regions_data)} counties with TDHCA regions")
        except Exception as e:
            self.logger.error(f"Error loading regions data: {e}")
        
        # Load TDHCA projects
        try:
            if Path(self.tdhca_file).exists():
                self.tdhca_data = pd.read_excel(self.tdhca_file, sheet_name='PropInventory')
                self.logger.info(f"✅ Loaded {len(self.tdhca_data)} TDHCA projects")
                
                # Clean coordinates
                self.tdhca_data['Latitude11'] = pd.to_numeric(self.tdhca_data['Latitude11'], errors='coerce')
                self.tdhca_data['Longitude11'] = pd.to_numeric(self.tdhca_data['Longitude11'], errors='coerce')
                self.tdhca_data = self.tdhca_data.dropna(subset=['Latitude11', 'Longitude11'])
                
                if 'Year' in self.tdhca_data.columns:
                    self.tdhca_data['Year'] = pd.to_numeric(self.tdhca_data['Year'], errors='coerce')
        except Exception as e:
            self.logger.error(f"Error loading TDHCA data: {e}")
        
        # Load AMI data
        try:
            if Path(self.ami_file).exists():
                self.ami_data = pd.read_excel(self.ami_file, sheet_name=0)
                self.ami_data = self.ami_data[self.ami_data['stusps'] == 'TX']
                self.logger.info(f"✅ Loaded AMI data for {len(self.ami_data)} Texas counties")
        except Exception as e:
            self.logger.error(f"Error loading AMI data: {e}")
        
        # Load QCT/DDA data
        try:
            if Path(self.qct_file).exists():
                self.qct_gdf = gpd.read_file(self.qct_file)
                self.qct_gdf = self.qct_gdf[self.qct_gdf.geometry.is_valid]
                # Convert to WGS84 for lat/lng comparison
                if self.qct_gdf.crs != 'EPSG:4326':
                    self.qct_gdf = self.qct_gdf.to_crs('EPSG:4326')
                self.logger.info(f"✅ Loaded {len(self.qct_gdf)} QCT areas (converted to WGS84)")
                
            if Path(self.dda_file).exists():
                self.dda_gdf = gpd.read_file(self.dda_file)
                self.dda_gdf = self.dda_gdf[self.dda_gdf.geometry.is_valid]
                # Convert to WGS84 for lat/lng comparison
                if self.dda_gdf.crs != 'EPSG:4326':
                    self.dda_gdf = self.dda_gdf.to_crs('EPSG:4326')
                self.logger.info(f"✅ Loaded {len(self.dda_gdf)} DDA areas (converted to WGS84)")
        except Exception as e:
            self.logger.error(f"Error loading QCT/DDA data: {e}")
    
    def get_tdhca_region(self, county):
        """Get TDHCA region number for a county"""
        if not county or not self.county_to_region:
            return "Unknown"
        
        # Clean county name for matching
        county_clean = county.strip().title()
        
        # Try exact match first
        if county_clean in self.county_to_region:
            return self.county_to_region[county_clean]
        
        # Try without 'County' suffix
        if ' County' in county_clean:
            county_no_suffix = county_clean.replace(' County', '')
            if county_no_suffix in self.county_to_region:
                return self.county_to_region[county_no_suffix]
        
        # Try adding 'County' suffix
        county_with_suffix = f"{county_clean} County"
        if county_with_suffix in self.county_to_region:
            return self.county_to_region[county_with_suffix]
        
        return "Unknown"
    
    def check_qct_dda_status(self, lat, lng):
        """Check if location is in QCT or DDA"""
        if pd.isna(lat) or pd.isna(lng):
            return {
                'in_qct': False,
                'in_dda': False,
                'basis_boost_eligible': False,
                'status': 'Unknown - No coordinates'
            }
        
        if self.qct_gdf is None or self.dda_gdf is None:
            return {
                'in_qct': False,
                'in_dda': False,
                'basis_boost_eligible': False,
                'status': 'Error - QCT/DDA data not loaded'
            }
        
        point = Point(lng, lat)
        
        # Check QCT
        in_qct = False
        try:
            for idx, tract in self.qct_gdf.iterrows():
                if tract.geometry and tract.geometry.contains(point):
                    in_qct = True
                    break
        except Exception as e:
            self.logger.error(f"Error checking QCT: {e}")
        
        # Check DDA
        in_dda = False
        try:
            for idx, area in self.dda_gdf.iterrows():
                if area.geometry and area.geometry.contains(point):
                    in_dda = True
                    break
        except Exception as e:
            self.logger.error(f"Error checking DDA: {e}")
        
        # Determine status
        if in_qct and in_dda:
            status = "QCT+DDA"
        elif in_qct:
            status = "QCT"
        elif in_dda:
            status = "DDA"
        else:
            status = "Neither"
        
        return {
            'in_qct': in_qct,
            'in_dda': in_dda,
            'basis_boost_eligible': in_qct or in_dda,
            'status': status
        }
    
    def analyze_competition_by_type(self, lat, lng, county, deal_type='4%'):
        """Analyze competition with correct rules for 4% vs 9%"""
        competition = {
            'one_mile_projects': [],
            'one_mile_recent_count': 0,
            'one_mile_fatal_flaw': False,  # 9% only
            'one_mile_risk_level': 'Low',  # 4% only
            'two_mile_violation': False,   # 9% family new construction only
            'market_saturation_1mi': 0,
            'market_saturation_2mi': 0
        }
        
        if self.tdhca_data is None:
            return competition
        
        try:
            site_coords = (lat, lng)
            current_year = datetime.now().year
            
            for idx, project in self.tdhca_data.iterrows():
                try:
                    project_coords = (project['Latitude11'], project['Longitude11'])
                    distance_miles = geodesic(site_coords, project_coords).miles
                    project_year = int(project['Year'])
                    years_ago = current_year - project_year
                    
                    # Count all projects within 1 mile for market saturation
                    if distance_miles <= 1.0:
                        competition['market_saturation_1mi'] += 1
                    elif distance_miles <= 2.0:
                        competition['market_saturation_2mi'] += 1
                    
                    # One Mile Three Year Rule
                    if distance_miles <= 1.0 and years_ago <= 3:
                        competition['one_mile_recent_count'] += 1
                        competition['one_mile_projects'].append({
                            'name': project.get('Development Name', 'Unknown'),
                            'year': project_year,
                            'distance': round(distance_miles, 2)
                        })
                        
                        if deal_type == '9%':
                            competition['one_mile_fatal_flaw'] = True
                        else:  # 4%
                            if competition['one_mile_recent_count'] >= 2:
                                competition['one_mile_risk_level'] = 'High'
                            else:
                                competition['one_mile_risk_level'] = 'Medium'
                    
                    # Two Mile Same Year (9% family new construction in large counties)
                    if (deal_type == '9%' and 
                        county in self.large_counties_9pct and
                        distance_miles <= 2.0 and 
                        project_year == current_year):
                        competition['two_mile_violation'] = True
                
                except Exception:
                    continue
        
        except Exception as e:
            self.logger.error(f"Competition analysis error: {e}")
        
        return competition
    
    def process_qct_dda_focused_analysis(self, input_file):
        """
        Process land data focusing ONLY on QCT/DDA eligible sites
        
        Args:
            input_file: Excel file with land data
        """
        self.logger.info(f"Processing QCT/DDA focused analysis for: {input_file}")
        
        # Load input data
        df = pd.read_excel(input_file)
        initial_count = len(df)
        self.logger.info(f"Loaded {initial_count} properties")
        
        # Initialize results columns
        analysis_columns = [
            'TDHCA_Region',
            'QCT_Status', 
            'DDA_Status',
            'QCT_DDA_Combined',
            'Basis_Boost_Eligible',
            'Deal_Type_Analyzed',
            'One_Mile_Risk_4pct',
            'One_Mile_Fatal_9pct',
            'Competition_1mi_Count',
            'Competition_2mi_Count',
            'Flood_Cost_Impact',
            'Overall_Viability_4pct',
            'Overall_Viability_9pct',
            'Development_Notes'
        ]
        
        for col in analysis_columns:
            df[col] = None
        
        qct_dda_eligible_rows = []
        
        # Analyze each property
        for idx, row in df.iterrows():
            lat = row.get('Latitude', row.get('lat', row.get('Lat')))
            lng = row.get('Longitude', row.get('lng', row.get('Lon', row.get('Long'))))
            county = row.get('County', row.get('county', row.get('county_name', '')))
            
            if pd.isna(lat) or pd.isna(lng):
                df.loc[idx, 'Development_Notes'] = 'Missing coordinates'
                continue
            
            # Check if data already has QCT/DDA status
            if 'Federal_Basis_Boost' in df.columns and pd.notna(row.get('Federal_Basis_Boost')):
                # Use existing QCT/DDA status
                is_qct_dda = row['Federal_Basis_Boost']
                if 'QCT_Status' in df.columns:
                    qct_status = row.get('QCT_Status', False)
                else:
                    qct_status = is_qct_dda  # Assume if basis boost, then QCT or DDA
                
                if 'DDA_Status' in df.columns:
                    dda_status = row.get('DDA_Status', False)
                else:
                    dda_status = False  # Can't determine separately
                
                # Determine combined status
                if qct_status and dda_status:
                    status = "QCT+DDA"
                elif qct_status:
                    status = "QCT"
                elif dda_status:
                    status = "DDA"
                elif is_qct_dda:
                    status = "QCT or DDA"  # We know it's eligible but not which
                else:
                    status = "Neither"
                
                qct_dda = {
                    'in_qct': qct_status,
                    'in_dda': dda_status,
                    'basis_boost_eligible': is_qct_dda,
                    'status': status
                }
            else:
                # Check QCT/DDA status using shapefiles
                qct_dda = self.check_qct_dda_status(lat, lng)
            
            # ONLY PROCESS QCT/DDA ELIGIBLE SITES
            if not qct_dda['basis_boost_eligible']:
                df.loc[idx, 'QCT_DDA_Combined'] = qct_dda['status']
                df.loc[idx, 'Basis_Boost_Eligible'] = False
                df.loc[idx, 'Development_Notes'] = 'EXCLUDED - No 30% basis boost'
                continue
            
            # This is a QCT/DDA site - full analysis
            qct_dda_eligible_rows.append(idx)
            
            # Get TDHCA region
            region = self.get_tdhca_region(county)
            df.loc[idx, 'TDHCA_Region'] = region
            
            # QCT/DDA details
            df.loc[idx, 'QCT_Status'] = qct_dda['in_qct']
            df.loc[idx, 'DDA_Status'] = qct_dda['in_dda']
            df.loc[idx, 'QCT_DDA_Combined'] = qct_dda['status']
            df.loc[idx, 'Basis_Boost_Eligible'] = True
            
            # Analyze for both 4% and 9%
            comp_4pct = self.analyze_competition_by_type(lat, lng, county, '4%')
            comp_9pct = self.analyze_competition_by_type(lat, lng, county, '9%')
            
            # Competition results
            df.loc[idx, 'One_Mile_Risk_4pct'] = comp_4pct['one_mile_risk_level']
            df.loc[idx, 'One_Mile_Fatal_9pct'] = comp_9pct['one_mile_fatal_flaw']
            df.loc[idx, 'Competition_1mi_Count'] = comp_4pct['market_saturation_1mi']
            df.loc[idx, 'Competition_2mi_Count'] = comp_4pct['market_saturation_2mi']
            
            # Flood cost impact
            flood_zone = row.get('Flood Zone', row.get('flood_zone', 'X'))
            df.loc[idx, 'Flood_Cost_Impact'] = self.flood_cost_impacts.get(flood_zone, 1.05)
            
            # Overall viability assessment
            # 4% viability
            if comp_4pct['one_mile_risk_level'] == 'Low':
                df.loc[idx, 'Overall_Viability_4pct'] = 'Excellent'
            elif comp_4pct['one_mile_risk_level'] == 'Medium':
                df.loc[idx, 'Overall_Viability_4pct'] = 'Good'
            else:
                df.loc[idx, 'Overall_Viability_4pct'] = 'Fair - Competition Risk'
            
            # 9% viability
            if comp_9pct['one_mile_fatal_flaw']:
                df.loc[idx, 'Overall_Viability_9pct'] = 'Not Viable - Fatal Flaw'
            elif comp_9pct['two_mile_violation']:
                df.loc[idx, 'Overall_Viability_9pct'] = 'Challenging - Two Mile Rule'
            else:
                df.loc[idx, 'Overall_Viability_9pct'] = 'Viable - Full Scoring Needed'
            
            # Development notes
            notes = []
            if region != "Unknown":
                notes.append(f"TDHCA {region}")
            notes.append(f"30% basis boost ({qct_dda['status']})")
            if comp_4pct['one_mile_recent_count'] > 0:
                notes.append(f"{comp_4pct['one_mile_recent_count']} recent LIHTC within 1mi")
            
            df.loc[idx, 'Development_Notes'] = '; '.join(notes)
        
        # Create output with only QCT/DDA sites
        df_qct_dda = df.loc[qct_dda_eligible_rows].copy()
        
        # Sort by region and 4% viability
        viability_order = {'Excellent': 1, 'Good': 2, 'Fair - Competition Risk': 3}
        df_qct_dda['viability_sort'] = df_qct_dda['Overall_Viability_4pct'].map(
            lambda x: viability_order.get(x, 4)
        )
        df_qct_dda = df_qct_dda.sort_values(['TDHCA_Region', 'viability_sort'])
        df_qct_dda = df_qct_dda.drop('viability_sort', axis=1)
        
        # Save results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"QCT_DDA_Focused_Analysis_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Total Properties Analyzed',
                    'QCT/DDA Eligible Sites',
                    'Non-QCT/DDA Sites (Excluded)',
                    'Excellent 4% Opportunities',
                    'Good 4% Opportunities',
                    'Fair 4% Opportunities',
                    'Fatal Flaw 9% Sites'
                ],
                'Count': [
                    initial_count,
                    len(df_qct_dda),
                    initial_count - len(df_qct_dda),
                    len(df_qct_dda[df_qct_dda['Overall_Viability_4pct'] == 'Excellent']) if len(df_qct_dda) > 0 else 0,
                    len(df_qct_dda[df_qct_dda['Overall_Viability_4pct'] == 'Good']) if len(df_qct_dda) > 0 else 0,
                    len(df_qct_dda[df_qct_dda['Overall_Viability_4pct'] == 'Fair - Competition Risk']) if len(df_qct_dda) > 0 else 0,
                    len(df_qct_dda[df_qct_dda['One_Mile_Fatal_9pct'] == True]) if len(df_qct_dda) > 0 else 0
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # QCT/DDA sites only (if any found)
            if len(df_qct_dda) > 0:
                df_qct_dda.to_excel(writer, sheet_name='QCT_DDA_Sites', index=False)
                
                # Regional summary
                regional_summary = df_qct_dda.groupby('TDHCA_Region').agg({
                    'TDHCA_Region': 'count',
                    'Overall_Viability_4pct': lambda x: (x == 'Excellent').sum()
                }).rename(columns={'TDHCA_Region': 'Total_Sites', 'Overall_Viability_4pct': 'Excellent_4pct'})
                regional_summary.to_excel(writer, sheet_name='Regional_Summary')
                
                # Excellent 4% opportunities
                excellent_4pct = df_qct_dda[df_qct_dda['Overall_Viability_4pct'] == 'Excellent']
                if len(excellent_4pct) > 0:
                    excellent_4pct.to_excel(writer, sheet_name='Best_4pct_Sites', index=False)
            else:
                # Add a note if no QCT/DDA sites found
                no_sites_df = pd.DataFrame({'Note': ['No QCT/DDA eligible sites found in the input data']})
                no_sites_df.to_excel(writer, sheet_name='No_QCT_DDA_Sites', index=False)
            
            # Format summary sheet
            if 'Summary' in writer.sheets:
                ws = writer.sheets['Summary']
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
        
        self.logger.info(f"\n📊 QCT/DDA Focused Analysis Complete:")
        self.logger.info(f"   Total properties: {initial_count}")
        self.logger.info(f"   QCT/DDA eligible: {len(df_qct_dda)} ({len(df_qct_dda)/initial_count*100:.1f}%)")
        self.logger.info(f"   Excluded (no basis boost): {initial_count - len(df_qct_dda)}")
        self.logger.info(f"\n💾 Results saved to: {output_file}")
        
        return df_qct_dda, output_file
    
    def create_dmarco_import_template(self):
        """Create Excel template for D'Marco's deals"""
        template_data = {
            'Property_Name': ['Example Property'],
            'Address': ['123 Main St'],
            'City': ['Austin'],
            'County': ['Travis'],
            'Latitude': [30.2672],
            'Longitude': [-97.7431],
            'Acreage': [2.5],
            'Price': [1500000],
            'Price_Per_Acre': [600000],
            'Flood_Zone': ['X'],
            'Source': ['Broker - ABC Realty'],
            'Broker_Name': ['John Smith'],
            'Contact': ['(512) 555-1234'],
            'Notes': ['Near transit, good schools']
        }
        
        df = pd.DataFrame(template_data)
        
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"DMarco_QCT_Import_Template_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Import_Data', index=False)
            
            # Add instructions
            instructions = {
                'Instructions': [
                    'Enter one property per row',
                    'Latitude/Longitude are required for QCT/DDA analysis',
                    'County name must match TDHCA regions exactly',
                    'We will ONLY analyze properties in QCT/DDA areas',
                    'Focus is on 4% deals but we analyze both 4% and 9% viability'
                ]
            }
            pd.DataFrame(instructions).to_excel(writer, sheet_name='Instructions', index=False)
        
        self.logger.info(f"✅ Created import template: {filename}")
        return filename

def main():
    """Example usage"""
    analyzer = TDHCAQCTFocusedAnalyzer()
    
    # Create import template
    template = analyzer.create_dmarco_import_template()
    print(f"\nCreated D'Marco import template: {template}")
    
    # Example: Analyze existing CoStar data
    # Assuming we have the file with counties already added
    costar_file = "CoStar_Land_Analysis_With_Counties_20250618_114519.xlsx"
    if Path(costar_file).exists():
        df_results, output_file = analyzer.process_qct_dda_focused_analysis(costar_file)
        print(f"\nAnalysis complete. QCT/DDA sites saved to: {output_file}")

if __name__ == "__main__":
    main()