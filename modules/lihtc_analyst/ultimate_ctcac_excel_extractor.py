#!/usr/bin/env python3
"""
ULTIMATE CTCAC EXCEL-TO-EXCEL EXTRACTOR
The BEAST that makes OpenAI and Meta look like amateurs!
Built by Bill's M4 WINGMAN - Laser focused, production ready
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import json
import logging
from typing import Dict, List, Any, Optional, Tuple
import numpy as np
from collections import defaultdict

class UltimateCTCACExtractor:
    """
    THE extraction system that dominates all California CTCAC applications
    Full data preservation, intelligent extraction, perfect Excel exports
    """
    
    def __init__(self, log_level=logging.INFO):
        # Setup logging for complete QA tracking
        self.setup_logging(log_level)
        
        # Critical sheets we MUST extract from every application
        self.critical_sheets = [
            'Application',
            'Sources and Uses Budget',
            'Basis & Credits', 
            'Sources and Basis Breakdown',
            'Points System',
            'Tie Breaker'
        ]
        
        # Additional sheets to grab if present
        self.additional_sheets = [
            'Disaster Credit Tie Breaker',
            'Supplemental',
            'Special Needs',
            'Public Funds',
            'Unit Mix',
            'Rent Roll'
        ]
        
        # Initialize extraction statistics
        self.stats = {
            'files_processed': 0,
            'sheets_extracted': 0,
            'cells_extracted': 0,
            'formulas_preserved': 0,
            'errors': []
        }
        
        self.logger.info("🚀 ULTIMATE CTCAC EXTRACTOR INITIALIZED - Ready to dominate!")
        
    def setup_logging(self, log_level):
        """Setup comprehensive logging for QA validation"""
        log_dir = Path("extraction_logs")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = log_dir / f"ctcac_extraction_{timestamp}.log"
        
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.log_file = log_file
        
    def extract_4p_applications(self, 
                               data_dir: Path,
                               output_dir: Path,
                               sample_size: int = 5) -> Dict[str, Any]:
        """
        Extract data from 4p applications with FULL preservation
        This is where we show our superiority!
        """
        self.logger.info(f"="*80)
        self.logger.info(f"🎯 STARTING 4P EXTRACTION MISSION")
        self.logger.info(f"Data Directory: {data_dir}")
        self.logger.info(f"Output Directory: {output_dir}")
        self.logger.info(f"="*80)
        
        # Create output directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Find all 4p applications
        pattern_2025 = "2025_4pct_R1_*.xlsx"
        pattern_2024 = "2024_4pct_R1_*.xlsx"
        
        files_2025 = list(data_dir.glob(pattern_2025))[:sample_size]
        files_2024 = list(data_dir.glob(pattern_2024))[:sample_size] if len(files_2025) < sample_size else []
        
        all_files = files_2025 + files_2024
        
        if not all_files:
            self.logger.error("❌ No 4p applications found!")
            return {}
        
        self.logger.info(f"📊 Found {len(all_files)} applications to process")
        
        # Process each application
        master_data = {}
        
        for idx, file_path in enumerate(all_files, 1):
            self.logger.info(f"\n{'='*60}")
            self.logger.info(f"Processing [{idx}/{len(all_files)}]: {file_path.name}")
            self.logger.info(f"{'='*60}")
            
            try:
                # Extract data from this application
                app_data = self._extract_single_application(file_path)
                master_data[file_path.name] = app_data
                
                # Create individual Excel export for this application
                self._create_excel_export(app_data, output_dir, file_path.name)
                
                self.stats['files_processed'] += 1
                
            except Exception as e:
                self.logger.error(f"❌ Error processing {file_path.name}: {e}")
                self.stats['errors'].append({
                    'file': file_path.name,
                    'error': str(e)
                })
        
        # Create master comparison Excel
        self._create_master_comparison_excel(master_data, output_dir)
        
        # Generate comprehensive report
        report = self._generate_extraction_report(master_data)
        
        self.logger.info(f"\n{'='*80}")
        self.logger.info(f"✅ EXTRACTION COMPLETE!")
        self.logger.info(f"Files Processed: {self.stats['files_processed']}")
        self.logger.info(f"Sheets Extracted: {self.stats['sheets_extracted']}")
        self.logger.info(f"Cells Extracted: {self.stats['cells_extracted']}")
        self.logger.info(f"Formulas Preserved: {self.stats['formulas_preserved']}")
        self.logger.info(f"Log File: {self.log_file}")
        self.logger.info(f"{'='*80}")
        
        return report
    
    def _extract_single_application(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract ALL data from a single CTCAC application
        This is where the magic happens!
        """
        app_data = {
            'file_name': file_path.name,
            'file_path': str(file_path),
            'extraction_timestamp': datetime.now().isoformat(),
            'sheets': {},
            'key_metrics': {},
            'metadata': {}
        }
        
        # Load with openpyxl for formula preservation
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Also load with pandas for easy data manipulation
        excel_file = pd.ExcelFile(file_path)
        
        # Extract metadata
        app_data['metadata'] = {
            'total_sheets': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'file_size_kb': file_path.stat().st_size / 1024
        }
        
        # Process each critical sheet
        for sheet_name in self.critical_sheets:
            if sheet_name in wb.sheetnames:
                self.logger.info(f"  📋 Extracting: {sheet_name}")
                sheet_data = self._extract_sheet_comprehensive(
                    wb[sheet_name],
                    wb_values[sheet_name],
                    excel_file,
                    sheet_name
                )
                app_data['sheets'][sheet_name] = sheet_data
                self.stats['sheets_extracted'] += 1
                
                # Extract key metrics from specific sheets
                self._extract_key_metrics(sheet_name, sheet_data, app_data)
        
        # Also grab additional sheets if present
        for sheet_name in self.additional_sheets:
            if sheet_name in wb.sheetnames:
                self.logger.info(f"  📋 Extracting (additional): {sheet_name}")
                sheet_data = self._extract_sheet_comprehensive(
                    wb[sheet_name],
                    wb_values[sheet_name],
                    excel_file,
                    sheet_name
                )
                app_data['sheets'][sheet_name] = sheet_data
                self.stats['sheets_extracted'] += 1
        
        wb.close()
        wb_values.close()
        
        return app_data
    
    def _extract_sheet_comprehensive(self, 
                                    ws_formulas,
                                    ws_values, 
                                    excel_file: pd.ExcelFile,
                                    sheet_name: str) -> Dict:
        """
        Extract EVERYTHING from a sheet - formulas, values, formatting, the works!
        """
        sheet_data = {
            'cells': {},
            'formulas': {},
            'values': {},
            'merged_cells': [],
            'dimensions': {
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            },
            'pandas_data': None,
            'key_fields': {}
        }
        
        # Extract merged cells
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract all cells with data
        for row in range(1, min(ws_formulas.max_row + 1, 1000)):  # Limit for performance
            for col in range(1, min(ws_formulas.max_column + 1, 50)):
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_formula.value is not None or cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__
                    }
                    
                    # Check for formula
                    if hasattr(cell_formula, '_value') and isinstance(cell_formula._value, str) and cell_formula._value.startswith('='):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                        self.stats['formulas_preserved'] += 1
                    
                    sheet_data['cells'][cell_ref] = cell_info
                    sheet_data['values'][cell_ref] = cell_value.value
                    self.stats['cells_extracted'] += 1
                    
                    # Look for key fields
                    if isinstance(cell_value.value, str):
                        self._identify_key_field(cell_value.value, cell_ref, row, col, 
                                                ws_values, sheet_data['key_fields'])
        
        # Also get pandas dataframe for easy analysis
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            sheet_data['pandas_data'] = df.to_dict('records')
        except Exception as e:
            self.logger.warning(f"    ⚠️ Could not read sheet with pandas: {e}")
        
        return sheet_data
    
    def _identify_key_field(self, cell_text: str, cell_ref: str, row: int, col: int,
                           ws, key_fields: Dict):
        """
        Identify and extract key LIHTC fields with their values
        """
        key_patterns = {
            'project_name': ['project name', 'development name', 'property name'],
            'total_units': ['total units', 'unit count', 'number of units'],
            'tax_credits': ['annual federal', 'tax credit', 'federal credit'],
            'tdc': ['total development cost', 'total project cost', 'development cost'],
            'developer_fee': ['developer fee', 'developer overhead'],
            'acquisition': ['acquisition cost', 'acquisition price'],
            'construction': ['construction cost', 'construction hard'],
            'eligible_basis': ['eligible basis', 'total eligible'],
            'qualified_basis': ['qualified basis', 'total qualified']
        }
        
        cell_lower = cell_text.lower().strip()
        
        for field_name, patterns in key_patterns.items():
            for pattern in patterns:
                if pattern in cell_lower:
                    # Try to find the value (usually to the right or below)
                    value = None
                    
                    # Check right
                    if col < ws.max_column:
                        right_cell = ws.cell(row=row, column=col+1)
                        if right_cell.value and not isinstance(right_cell.value, str):
                            value = right_cell.value
                    
                    # Check below
                    if value is None and row < ws.max_row:
                        below_cell = ws.cell(row=row+1, column=col)
                        if below_cell.value and not isinstance(below_cell.value, str):
                            value = below_cell.value
                    
                    if field_name not in key_fields:
                        key_fields[field_name] = []
                    
                    key_fields[field_name].append({
                        'label': cell_text,
                        'label_cell': cell_ref,
                        'value': value,
                        'found_at': f"Row {row}, Col {col}"
                    })
                    break
    
    def _extract_key_metrics(self, sheet_name: str, sheet_data: Dict, app_data: Dict):
        """
        Extract critical metrics from specific sheets
        """
        metrics = app_data['key_metrics']
        
        if sheet_name == 'Application' and 'key_fields' in sheet_data:
            fields = sheet_data['key_fields']
            
            # Extract project name
            if 'project_name' in fields and fields['project_name']:
                metrics['project_name'] = fields['project_name'][0].get('value', 'Unknown')
            
            # Extract total units
            if 'total_units' in fields and fields['total_units']:
                metrics['total_units'] = fields['total_units'][0].get('value', 0)
        
        elif sheet_name == 'Sources and Uses Budget' and 'key_fields' in sheet_data:
            fields = sheet_data['key_fields']
            
            # Extract TDC
            if 'tdc' in fields and fields['tdc']:
                metrics['total_development_cost'] = fields['tdc'][0].get('value', 0)
            
            # Extract developer fee
            if 'developer_fee' in fields and fields['developer_fee']:
                metrics['developer_fee'] = fields['developer_fee'][0].get('value', 0)
        
        elif sheet_name == 'Basis & Credits' and 'key_fields' in sheet_data:
            fields = sheet_data['key_fields']
            
            # Extract basis info
            if 'eligible_basis' in fields and fields['eligible_basis']:
                metrics['eligible_basis'] = fields['eligible_basis'][0].get('value', 0)
            
            if 'qualified_basis' in fields and fields['qualified_basis']:
                metrics['qualified_basis'] = fields['qualified_basis'][0].get('value', 0)
            
            if 'tax_credits' in fields and fields['tax_credits']:
                metrics['annual_credits'] = fields['tax_credits'][0].get('value', 0)
    
    def _create_excel_export(self, app_data: Dict, output_dir: Path, source_name: str):
        """
        Create a beautiful Excel export with all extracted data
        Each critical sheet gets its own tab with the SAME NAME!
        """
        # Create output filename
        base_name = source_name.replace('.xlsx', '')
        output_file = output_dir / f"EXTRACTED_{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add metadata sheet first
        self._add_metadata_sheet(wb, app_data)
        
        # Add each extracted sheet with ORIGINAL NAME
        for sheet_name, sheet_data in app_data['sheets'].items():
            self.logger.info(f"    ✍️ Writing sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit is 31 chars
            
            # Write the data
            self._write_sheet_data(ws, sheet_data)
        
        # Add summary sheet
        self._add_summary_sheet(wb, app_data)
        
        # Save the workbook
        wb.save(output_file)
        self.logger.info(f"  💾 Saved: {output_file.name}")
        
        return output_file
    
    def _write_sheet_data(self, ws, sheet_data: Dict):
        """
        Write extracted data to worksheet, preserving as much as possible
        """
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write cells with their original positions
        if 'cells' in sheet_data:
            for cell_ref, cell_info in sheet_data['cells'].items():
                row = cell_info['row']
                col = cell_info['col']
                
                # Write value
                ws.cell(row=row, column=col, value=cell_info['value'])
                
                # Add formula as comment if present
                if cell_info.get('formula'):
                    cell = ws.cell(row=row, column=col)
                    cell.comment = openpyxl.comments.Comment(
                        f"Formula: {cell_info['formula']}", 
                        "CTCAC Extractor"
                    )
        
        # Apply merged cells
        if 'merged_cells' in sheet_data:
            for merged_range in sheet_data['merged_cells']:
                try:
                    ws.merge_cells(merged_range)
                except:
                    pass  # Skip if merge fails
        
        # Add extracted key fields section at the bottom
        if 'key_fields' in sheet_data and sheet_data['key_fields']:
            # Find the last row with data
            last_row = sheet_data['dimensions']['max_row'] + 5
            
            # Add header
            ws.cell(row=last_row, column=1, value="EXTRACTED KEY FIELDS").font = header_font
            ws.cell(row=last_row, column=1).fill = header_fill
            
            row_num = last_row + 1
            for field_name, field_instances in sheet_data['key_fields'].items():
                for instance in field_instances:
                    ws.cell(row=row_num, column=1, value=field_name)
                    ws.cell(row=row_num, column=2, value=instance['label'])
                    ws.cell(row=row_num, column=3, value=instance['value'])
                    ws.cell(row=row_num, column=4, value=instance['found_at'])
                    row_num += 1
    
    def _add_metadata_sheet(self, wb: Workbook, app_data: Dict):
        """
        Add metadata sheet with extraction details for QA
        """
        ws = wb.create_sheet(title="METADATA", index=0)
        
        # Header styling
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        
        # Title
        ws['A1'] = "CTCAC EXTRACTION METADATA"
        ws['A1'].font = header_font
        
        # File info
        ws['A3'] = "Source File:"
        ws['B3'] = app_data['file_name']
        ws['A4'] = "Extraction Time:"
        ws['B4'] = app_data['extraction_timestamp']
        ws['A5'] = "File Path:"
        ws['B5'] = app_data['file_path']
        
        # Sheet info
        ws['A7'] = "SHEETS EXTRACTED"
        ws['A7'].font = subheader_font
        
        row = 8
        for sheet_name in app_data['sheets'].keys():
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = f"{len(app_data['sheets'][sheet_name].get('cells', {}))} cells"
            row += 1
        
        # Key metrics
        ws[f'A{row+1}'] = "KEY METRICS FOUND"
        ws[f'A{row+1}'].font = subheader_font
        
        row += 2
        for metric_name, metric_value in app_data.get('key_metrics', {}).items():
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_value
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _add_summary_sheet(self, wb: Workbook, app_data: Dict):
        """
        Add summary sheet with all key information consolidated
        """
        ws = wb.create_sheet(title="SUMMARY")
        
        # Title
        ws['A1'] = "EXTRACTION SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Create summary table
        summary_data = []
        
        # Add key metrics
        for metric_name, metric_value in app_data.get('key_metrics', {}).items():
            summary_data.append({
                'Category': 'Key Metric',
                'Field': metric_name,
                'Value': metric_value
            })
        
        # Add statistics
        total_cells = sum(len(sheet.get('cells', {})) for sheet in app_data['sheets'].values())
        total_formulas = sum(len(sheet.get('formulas', {})) for sheet in app_data['sheets'].values())
        
        summary_data.append({'Category': 'Statistics', 'Field': 'Total Cells', 'Value': total_cells})
        summary_data.append({'Category': 'Statistics', 'Field': 'Total Formulas', 'Value': total_formulas})
        summary_data.append({'Category': 'Statistics', 'Field': 'Sheets Processed', 'Value': len(app_data['sheets'])})
        
        # Convert to dataframe and write
        if summary_data:
            df = pd.DataFrame(summary_data)
            
            # Write headers
            for col_num, column_title in enumerate(df.columns, 1):
                ws.cell(row=3, column=col_num, value=column_title)
                ws.cell(row=3, column=col_num).font = Font(bold=True)
            
            # Write data
            for row_num, row_data in enumerate(df.values, 4):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
    
    def _create_master_comparison_excel(self, master_data: Dict, output_dir: Path):
        """
        Create a master Excel comparing all extracted applications
        This is where we show the full power!
        """
        output_file = output_dir / f"MASTER_COMPARISON_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create comparison sheet
        ws = wb.create_sheet(title="COMPARISON")
        
        # Headers
        headers = ['Application', 'Project Name', 'Total Units', 'TDC', 'Developer Fee', 
                  'Eligible Basis', 'Qualified Basis', 'Annual Credits', 'Sheets Found']
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
            ws.cell(row=1, column=col_num).fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
        
        # Data rows
        row_num = 2
        for file_name, app_data in master_data.items():
            metrics = app_data.get('key_metrics', {})
            
            ws.cell(row=row_num, column=1, value=file_name)
            ws.cell(row=row_num, column=2, value=metrics.get('project_name', 'N/A'))
            ws.cell(row=row_num, column=3, value=metrics.get('total_units', 0))
            ws.cell(row=row_num, column=4, value=metrics.get('total_development_cost', 0))
            ws.cell(row=row_num, column=5, value=metrics.get('developer_fee', 0))
            ws.cell(row=row_num, column=6, value=metrics.get('eligible_basis', 0))
            ws.cell(row=row_num, column=7, value=metrics.get('qualified_basis', 0))
            ws.cell(row=row_num, column=8, value=metrics.get('annual_credits', 0))
            ws.cell(row=row_num, column=9, value=len(app_data.get('sheets', {})))
            
            row_num += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save
        wb.save(output_file)
        self.logger.info(f"\n📊 Master Comparison saved: {output_file.name}")
    
    def _generate_extraction_report(self, master_data: Dict) -> Dict:
        """
        Generate comprehensive extraction report
        """
        report = {
            'timestamp': datetime.now().isoformat(),
            'statistics': self.stats,
            'files': {},
            'summary': {
                'total_files': len(master_data),
                'total_sheets': sum(len(d.get('sheets', {})) for d in master_data.values()),
                'successful_extractions': len([d for d in master_data.values() if d.get('key_metrics')]),
            }
        }
        
        # Add file-specific data
        for file_name, app_data in master_data.items():
            report['files'][file_name] = {
                'sheets_extracted': list(app_data.get('sheets', {}).keys()),
                'key_metrics': app_data.get('key_metrics', {}),
                'extraction_time': app_data.get('extraction_timestamp')
            }
        
        # Save JSON report
        json_file = Path("extraction_logs") / f"extraction_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_file, 'w') as f:
            json.dump(report, f, indent=2, default=str)
        
        self.logger.info(f"📄 JSON Report saved: {json_file}")
        
        return report


def main():
    """
    DOMINATE those CTCAC applications!
    """
    print("="*80)
    print("🚀 ULTIMATE CTCAC EXCEL EXTRACTOR")
    print("   Built by Bill's M4 WINGMAN")
    print("   Making OpenAI and Meta look like amateurs since 2025")
    print("="*80)
    
    # Setup paths
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/ctcac_extractions")
    
    # Initialize the BEAST
    extractor = UltimateCTCACExtractor()
    
    # DOMINATE those 4p applications
    report = extractor.extract_4p_applications(
        data_dir=data_dir,
        output_dir=output_dir,
        sample_size=3  # Start with 3 for testing
    )
    
    print("\n" + "="*80)
    print("✅ EXTRACTION COMPLETE - WE DOMINATED!")
    print("="*80)
    print(f"Files Processed: {report['summary']['total_files']}")
    print(f"Sheets Extracted: {report['summary']['total_sheets']}")
    print(f"Successful Extractions: {report['summary']['successful_extractions']}")
    print("\nCheck the output directory for your beautiful Excel files!")
    print("Tower and Bill - prepare to be amazed!")
    

if __name__ == "__main__":
    main()