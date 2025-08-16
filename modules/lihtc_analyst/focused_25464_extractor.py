#!/usr/bin/env python3
"""
FOCUSED 25-464 CTCAC EXTRACTION
Enhanced extraction specifically for the application we analyzed with PDFs
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import logging

class Focused25464Extractor:
    """Enhanced extractor focused on the 25-464 application"""
    
    def __init__(self):
        self.setup_logging()
        
        # Enhanced patterns based on our PDF analysis
        self.critical_patterns = {
            'Application': {
                'project_name': ['Project Name', 'project name', 'Development Name'],
                'address': ['Site Address', 'Project Address', '1300 Jefferson'],
                'city': ['City', 'Napa'],
                'county': ['County'],
                'total_units': ['Total Units', 'Unit Count', '84'],
                'applicant': ['CTCAC APPLICANT', 'Applicant Name', 'RAHD'],
                'contact_person': ['Contact Person', 'David Beacham'],
                'phone': ['Phone', '760-579-2093'],
                'email': ['Email', 'Dave@rahdgroup.com'],
                'census_tract': ['Census Tract', '2002.02'],
                'annual_credits': ['annual Federal Credits', 'Annual Federal Tax Credit', '$1,417,201']
            },
            
            'Sources and Uses Budget': {
                'tax_exempt_bond': ['Tax-Exempt', 'Berkadia', '$16,500,000'],
                'tax_credit_equity': ['LIHTC Investor', 'Tax Credit Equity', '$5,979,512'],
                'taxable_loan': ['Taxable', '$4,750,000'],
                'developer_fee': ['Developer Fee', '$5,000,000'],
                'total_development_cost': ['Total Development Cost', 'TDC', '$34,129,512'],
                'cost_per_unit': ['$406,304', 'Cost Per Unit']
            },
            
            'Basis & Credits': {
                'eligible_basis': ['Eligible Basis', '$32,985,512'],
                'qualified_basis': ['Qualified Basis'],
                'applicable_percentage': ['Applicable Percentage', '4%'],
                'annual_credits': ['Annual Credits', '$1,417,201']
            }
        }
        
        self.logger.info("🎯 FOCUSED 25-464 EXTRACTOR Ready!")
    
    def setup_logging(self):
        """Setup logging"""
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def extract_comprehensive(self, file_path: Path) -> dict:
        """Comprehensive extraction with full data preservation"""
        self.logger.info(f"🚀 COMPREHENSIVE EXTRACTION: {file_path.name}")
        
        result = {
            'file_info': {
                'name': file_path.name,
                'path': str(file_path),
                'timestamp': datetime.now().isoformat()
            },
            'sheets_data': {},
            'intelligent_findings': {},
            'raw_cell_data': {},
            'summary': {}
        }
        
        # Load Excel file
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Process each sheet comprehensively
        for sheet_name in wb_formulas.sheetnames:
            self.logger.info(f"  📋 Processing: {sheet_name}")
            
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]
            
            sheet_result = self._extract_sheet_comprehensive(
                ws_formulas, ws_values, sheet_name
            )
            
            result['sheets_data'][sheet_name] = sheet_result
            
            # Apply intelligent pattern matching for critical sheets
            if sheet_name in self.critical_patterns:
                findings = self._apply_intelligent_extraction(
                    sheet_result, sheet_name
                )
                result['intelligent_findings'][sheet_name] = findings
        
        wb_formulas.close()
        wb_values.close()
        
        # Generate summary
        result['summary'] = self._generate_summary(result)
        
        return result
    
    def _extract_sheet_comprehensive(self, ws_formulas, ws_values, sheet_name: str) -> dict:
        """Extract everything from a sheet"""
        sheet_data = {
            'metadata': {
                'sheet_name': sheet_name,
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            },
            'cells': {},
            'formulas': {},
            'merged_cells': []
        }
        
        # Get merged cells
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract all meaningful cells (expanded range)
        for row in range(1, min(ws_formulas.max_row + 1, 500)):
            for col in range(1, min(ws_formulas.max_column + 1, 50)):
                
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__
                    }
                    
                    # Check for formula
                    if (hasattr(cell_formula, '_value') and 
                        isinstance(cell_formula._value, str) and 
                        cell_formula._value.startswith('=')):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                    
                    sheet_data['cells'][cell_ref] = cell_info
        
        return sheet_data
    
    def _apply_intelligent_extraction(self, sheet_data: dict, sheet_name: str) -> dict:
        """Apply intelligent pattern matching"""
        findings = {}
        patterns = self.critical_patterns[sheet_name]
        
        for field_name, search_terms in patterns.items():
            found_instances = []
            
            # Search through all cells for patterns
            for cell_ref, cell_info in sheet_data['cells'].items():
                if isinstance(cell_info['value'], str):
                    cell_text = str(cell_info['value']).lower()
                    
                    for term in search_terms:
                        if str(term).lower() in cell_text:
                            # Try to find associated value
                            associated_value = self._find_associated_value(
                                cell_info, sheet_data
                            )
                            
                            found_instances.append({
                                'label': cell_info['value'],
                                'location': cell_ref,
                                'row': cell_info['row'],
                                'col': cell_info['col'],
                                'associated_value': associated_value,
                                'search_term': term
                            })
                            break
            
            if found_instances:
                findings[field_name] = found_instances
        
        return findings
    
    def _find_associated_value(self, label_cell: dict, sheet_data: dict):
        """Find value associated with a label"""
        row = label_cell['row']
        col = label_cell['col']
        
        # Check common locations for values
        check_positions = [
            (row, col + 1),      # Right
            (row, col + 2),      # Two right
            (row + 1, col),      # Below
            (row - 1, col),      # Above
        ]
        
        for check_row, check_col in check_positions:
            check_ref = f"{openpyxl.utils.get_column_letter(check_col)}{check_row}"
            if check_ref in sheet_data['cells']:
                value = sheet_data['cells'][check_ref]['value']
                if value is not None and not isinstance(value, str):
                    return value
                elif isinstance(value, str) and value.strip() and len(value.strip()) < 100:
                    # Short string values might be what we want
                    return value.strip()
        
        return None
    
    def _generate_summary(self, result: dict) -> dict:
        """Generate comprehensive summary"""
        summary = {
            'total_sheets': len(result['sheets_data']),
            'sheets_processed': list(result['sheets_data'].keys()),
            'total_cells_extracted': 0,
            'total_formulas': 0,
            'intelligent_findings_count': 0,
            'key_project_info': {}
        }
        
        # Count totals
        for sheet_data in result['sheets_data'].values():
            summary['total_cells_extracted'] += len(sheet_data['cells'])
            summary['total_formulas'] += len(sheet_data['formulas'])
        
        # Count intelligent findings
        for findings in result['intelligent_findings'].values():
            summary['intelligent_findings_count'] += len(findings)
        
        # Extract key project info from Application sheet
        if 'Application' in result['intelligent_findings']:
            app_findings = result['intelligent_findings']['Application']
            for field, instances in app_findings.items():
                if instances:
                    # Take the first (usually best) instance
                    best_instance = instances[0]
                    summary['key_project_info'][field] = {
                        'label': best_instance['label'],
                        'value': best_instance['associated_value'],
                        'location': best_instance['location']
                    }
        
        return summary
    
    def create_comprehensive_excel(self, result: dict, output_path: Path):
        """Create comprehensive Excel output"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Executive Summary sheet
        self._create_executive_summary(wb, result)
        
        # Create Intelligent Findings sheet
        self._create_intelligent_findings_sheet(wb, result)
        
        # Create detailed sheets for each source sheet
        for sheet_name, sheet_data in result['sheets_data'].items():
            if sheet_name in ['Application', 'Sources and Uses Budget', 'Basis & Credits', 'Points System']:
                self._create_detailed_sheet(wb, sheet_name, sheet_data, 
                                          result['intelligent_findings'].get(sheet_name, {}))
        
        # Save the workbook
        wb.save(output_path)
        self.logger.info(f"💾 Comprehensive Excel saved: {output_path.name}")
        
        return output_path
    
    def _create_executive_summary(self, wb: Workbook, result: dict):
        """Create executive summary sheet"""
        ws = wb.create_sheet(title="EXECUTIVE SUMMARY", index=0)
        
        # Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        bold_font = Font(bold=True)
        
        # Title
        ws['A1'] = "COMPREHENSIVE CTCAC EXTRACTION - 25-464"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Extraction Statistics
        ws[f'A{row}'] = "EXTRACTION STATISTICS"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        summary = result['summary']
        stats = [
            ("Total Sheets Processed", summary['total_sheets']),
            ("Total Cells Extracted", summary['total_cells_extracted']),
            ("Total Formulas Preserved", summary['total_formulas']),
            ("Intelligent Findings", summary['intelligent_findings_count'])
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Key Project Information
        ws[f'A{row}'] = "KEY PROJECT INFORMATION"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        key_info = summary.get('key_project_info', {})
        for field, data in key_info.items():
            if isinstance(data, dict) and data.get('value'):
                ws[f'A{row}'] = field.replace('_', ' ').title()
                ws[f'B{row}'] = str(data['value'])
                ws[f'C{row}'] = data.get('location', '')
                row += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_intelligent_findings_sheet(self, wb: Workbook, result: dict):
        """Create intelligent findings sheet"""
        ws = wb.create_sheet(title="INTELLIGENT FINDINGS")
        
        # Headers
        headers = ['Sheet', 'Field', 'Label Found', 'Associated Value', 'Location', 'Search Term']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 2
        
        # Add all findings
        for sheet_name, findings in result['intelligent_findings'].items():
            for field_name, instances in findings.items():
                for instance in instances:
                    ws.cell(row=row, column=1, value=sheet_name)
                    ws.cell(row=row, column=2, value=field_name)
                    ws.cell(row=row, column=3, value=str(instance['label']))
                    ws.cell(row=row, column=4, value=str(instance['associated_value']) if instance['associated_value'] else 'N/A')
                    ws.cell(row=row, column=5, value=instance['location'])
                    ws.cell(row=row, column=6, value=instance['search_term'])
                    row += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_sheet(self, wb: Workbook, sheet_name: str, sheet_data: dict, findings: dict):
        """Create detailed sheet for specific source sheet"""
        ws = wb.create_sheet(title=f"DETAIL_{sheet_name[:20]}")
        
        # Sheet title
        ws['A1'] = f"DETAILED EXTRACTION: {sheet_name}"
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        
        # Intelligent findings for this sheet
        if findings:
            ws[f'A{row}'] = "INTELLIGENT FINDINGS"
            ws[f'A{row}'].font = Font(bold=True, color="0066CC")
            row += 1
            
            # Headers
            headers = ['Field', 'Label', 'Value', 'Location']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1
            
            # Data
            for field_name, instances in findings.items():
                for instance in instances:
                    ws.cell(row=row, column=1, value=field_name)
                    ws.cell(row=row, column=2, value=str(instance['label']))
                    ws.cell(row=row, column=3, value=str(instance['associated_value']) if instance['associated_value'] else 'N/A')
                    ws.cell(row=row, column=4, value=instance['location'])
                    row += 1
            
            row += 3
        
        # Raw data sample (first 50 cells with data)
        ws[f'A{row}'] = "RAW DATA SAMPLE"
        ws[f'A{row}'].font = Font(bold=True, color="006600")
        row += 1
        
        headers = ['Location', 'Value', 'Data Type', 'Formula']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        # Show first 50 cells with data
        count = 0
        for cell_ref, cell_info in sheet_data['cells'].items():
            if count >= 50:
                break
            
            ws.cell(row=row, column=1, value=cell_ref)
            ws.cell(row=row, column=2, value=str(cell_info['value'])[:100])  # Truncate long values
            ws.cell(row=row, column=3, value=cell_info['data_type'])
            ws.cell(row=row, column=4, value=cell_info.get('formula', ''))
            row += 1
            count += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Run focused extraction on 25-464"""
    print("="*80)
    print("🎯 FOCUSED 25-464 COMPREHENSIVE EXTRACTION")
    print("   Enhanced extraction for PDF-analyzed application")
    print("="*80)
    
    # Paths
    input_file = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data/2025_4pct_R1_25-464.xlsx")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/ctcac_extractions_v2")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    if not input_file.exists():
        print(f"❌ Input file not found: {input_file}")
        return
    
    # Initialize extractor
    extractor = Focused25464Extractor()
    
    # Run comprehensive extraction
    result = extractor.extract_comprehensive(input_file)
    
    # Create comprehensive Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f"COMPREHENSIVE_25464_{timestamp}.xlsx"
    extractor.create_comprehensive_excel(result, output_file)
    
    # Print results
    print("\n" + "="*80)
    print("✅ COMPREHENSIVE EXTRACTION COMPLETE!")
    print("="*80)
    
    summary = result['summary']
    print(f"📊 Sheets Processed: {summary['total_sheets']}")
    print(f"📋 Cells Extracted: {summary['total_cells_extracted']}")
    print(f"🧮 Formulas Preserved: {summary['total_formulas']}")
    print(f"🎯 Intelligent Findings: {summary['intelligent_findings_count']}")
    
    # Show key project info
    print("\n🏠 KEY PROJECT INFO:")
    for field, data in summary.get('key_project_info', {}).items():
        if isinstance(data, dict) and data.get('value'):
            print(f"   {field.title()}: {data['value']}")
    
    print(f"\n💾 Comprehensive Excel: {output_file.name}")
    print("\nThis file contains full extraction with intelligent findings!")

if __name__ == "__main__":
    main()