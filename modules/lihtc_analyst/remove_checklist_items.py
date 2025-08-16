#!/usr/bin/env python3
"""
Remove Checklist Items sheet and analyze 15 Year Pro Forma differences
"""

import openpyxl
from pathlib import Path

def analyze_proforma_difference():
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    
    # Load the raw files to see actual sheet dimensions
    file_2024 = data_dir / "2024_4pct_R1_24-409.xlsx"
    file_2025 = data_dir / "2025_4pct_R1_25-464.xlsx"
    
    print("🔍 ANALYZING 15 YEAR PRO FORMA DIFFERENCES")
    print("=" * 50)
    
    wb_2024 = openpyxl.load_workbook(file_2024, read_only=True)
    wb_2025 = openpyxl.load_workbook(file_2025, read_only=True)
    
    # Find 15 Year Pro Forma sheets
    proforma_2024 = None
    proforma_2025 = None
    
    for sheet in wb_2024.worksheets:
        if "15 Year Pro Forma" in sheet.title:
            proforma_2024 = sheet
            break
    
    for sheet in wb_2025.worksheets:
        if "15 Year Pro Forma" in sheet.title:
            proforma_2025 = sheet
            break
    
    if proforma_2024 and proforma_2025:
        print(f"📊 15 YEAR PRO FORMA COMPARISON:")
        print(f"2024: {proforma_2024.max_row} rows x {proforma_2024.max_column} cols = {proforma_2024.max_row * proforma_2024.max_column:,} cells")
        print(f"2025: {proforma_2025.max_row} rows x {proforma_2025.max_column} cols = {proforma_2025.max_row * proforma_2025.max_column:,} cells")
        
        diff_cells = (proforma_2024.max_row * proforma_2024.max_column) - (proforma_2025.max_row * proforma_2025.max_column)
        print(f"Difference: {diff_cells:+,} cells")
        
        if abs(diff_cells) > 100:
            print(f"⚠️  Pro Forma should be similar size - investigating...")
            
            # Check if 2024 has corrupted range
            if proforma_2024.max_row > 300 or proforma_2024.max_column > 50:
                print(f"🔴 2024 Pro Forma has corrupted range: {proforma_2024.max_row}x{proforma_2024.max_column}")
            
            if proforma_2025.max_row > 300 or proforma_2025.max_column > 50:
                print(f"🔴 2025 Pro Forma has corrupted range: {proforma_2025.max_row}x{proforma_2025.max_column}")
    
    # Check for Checklist Items
    checklist_2024 = None
    checklist_2025 = None
    
    for sheet in wb_2024.worksheets:
        if "Checklist Items" in sheet.title:
            checklist_2024 = sheet
            break
    
    for sheet in wb_2025.worksheets:
        if "Checklist Items" in sheet.title:
            checklist_2025 = sheet
            break
    
    print(f"\n📋 CHECKLIST ITEMS ANALYSIS:")
    print(f"2024 has Checklist Items: {'Yes' if checklist_2024 else 'No'}")
    print(f"2025 has Checklist Items: {'Yes' if checklist_2025 else 'No'}")
    
    if checklist_2024:
        cells = checklist_2024.max_row * checklist_2024.max_column
        print(f"2024 Checklist Items: {checklist_2024.max_row} rows x {checklist_2024.max_column} cols = {cells:,} cells")
        print(f"✅ Recommendation: Remove Checklist Items sheet (saves {cells:,} cells)")
    
    wb_2024.close()
    wb_2025.close()
    
    return checklist_2024 is not None

def create_enhanced_processor_v19():
    """Create V1.9 processor that removes Checklist Items sheet"""
    
    print("\n🚀 Creating Enhanced Processor V1.9 - Remove Checklist Items")
    
    # Read the V1.8 processor
    v18_path = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/enhanced_ctcac_processor_v18.py")
    
    if not v18_path.exists():
        print("❌ V1.8 processor not found")
        return
    
    with open(v18_path, 'r') as f:
        v18_content = f.read()
    
    # Add checklist removal logic
    checklist_removal_code = '''
    def _remove_checklist_items_sheet(self, wb):
        """Remove Checklist Items sheet if it exists"""
        sheets_to_remove = []
        for sheet in wb.worksheets:
            if "Checklist Items" in sheet.title or "checklist" in sheet.title.lower():
                sheets_to_remove.append(sheet.title)
        
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
                print(f"✅ Removed sheet: {sheet_name}")
        
        return len(sheets_to_remove) > 0
    '''
    
    # Insert the removal method before the process_single_file method
    insertion_point = v18_content.find("def process_single_file(self, file_path: Path) -> Dict:")
    if insertion_point != -1:
        v19_content = v18_content[:insertion_point] + checklist_removal_code + "\n    " + v18_content[insertion_point:]
        
        # Add checklist removal call in process_single_file
        wb_load_pattern = "wb = openpyxl.load_workbook(file_path, read_only=False)"
        replacement = """wb = openpyxl.load_workbook(file_path, read_only=False)
            
            # Remove Checklist Items sheet if present
            self._remove_checklist_items_sheet(wb)"""
        
        v19_content = v19_content.replace(wb_load_pattern, replacement)
        
        # Update version number and description
        v19_content = v19_content.replace("Enhanced CTCAC Processor V1.8", "Enhanced CTCAC Processor V1.9")
        v19_content = v19_content.replace("V1.8 - SMART RANGE DETECTION", "V1.9 - REMOVE CHECKLIST ITEMS")
        
        # Write V1.9
        v19_path = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/enhanced_ctcac_processor_v19.py")
        with open(v19_path, 'w') as f:
            f.write(v19_content)
        
        print(f"✅ Created V1.9 processor: {v19_path}")
        return True
    
    return False

def main():
    has_checklist = analyze_proforma_difference()
    
    if has_checklist:
        print(f"\n🔧 CREATING ENHANCED V1.9 TO REMOVE CHECKLIST ITEMS")
        success = create_enhanced_processor_v19()
        if success:
            print(f"✅ V1.9 processor ready - removes unnecessary Checklist Items sheets")

if __name__ == "__main__":
    main()