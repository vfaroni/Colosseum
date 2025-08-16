#!/usr/bin/env python3
"""
BATCH COMPREHENSIVE EXTRACTOR
Run the enhanced extraction on multiple applications for comparison
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
import logging
from typing import Dict, List

class BatchComprehensiveExtractor:
    """Enhanced batch extractor for multiple CTCAC applications"""
    
    def __init__(self):
        self.setup_logging()
        
        # Enhanced patterns based on our analysis
        self.critical_patterns = {
            'Application': {
                'project_name': ['Project Name', 'project name', 'Development Name'],
                'address': ['Site Address', 'Project Address'],
                'city': ['City'],
                'county': ['County'],
                'total_units': ['Total Units', 'Unit Count'],
                'applicant': ['CTCAC APPLICANT', 'Applicant Name'],
                'contact_person': ['Contact Person'],
                'phone': ['Phone'],
                'email': ['Email'],
                'census_tract': ['Census Tract'],
                'annual_credits': ['annual Federal Credits', 'Annual Federal Tax Credit']
            },
            
            'Sources and Uses Budget': {
                'tax_exempt_bond': ['Tax-Exempt', 'Bond'],
                'tax_credit_equity': ['LIHTC Investor', 'Tax Credit Equity'],
                'taxable_loan': ['Taxable'],
                'developer_fee': ['Developer Fee'],
                'total_development_cost': ['Total Development Cost', 'TDC'],
                'cost_per_unit': ['Cost Per Unit']
            },
            
            'Basis & Credits': {
                'eligible_basis': ['Eligible Basis'],
                'qualified_basis': ['Qualified Basis'],
                'applicable_percentage': ['Applicable Percentage'],
                'annual_credits': ['Annual Credits']
            },
            
            'Points System': {
                'tie_breaker_points': ['Tie Breaker'],
                'readiness_points': ['Readiness'],
                'sustainability_points': ['Sustainability', 'Green Building']
            }
        }
        
        self.logger.info("🚀 BATCH COMPREHENSIVE EXTRACTOR Ready!")
    
    def setup_logging(self):
        """Setup logging"""
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def extract_batch_applications(self, data_dir: Path, sample_size: int = 3) -> Dict:
        """Extract multiple applications comprehensively"""
        self.logger.info(f"🎯 BATCH EXTRACTION: Processing {sample_size} applications")
        
        # Find applications (same logic as original)
        pattern_2025 = "2025_4pct_R1_*.xlsx"
        pattern_2024 = "2024_4pct_R1_*.xlsx"
        
        files_2025 = list(data_dir.glob(pattern_2025))[:sample_size]
        files_2024 = list(data_dir.glob(pattern_2024))[:sample_size] if len(files_2025) < sample_size else []
        
        all_files = files_2025 + files_2024
        all_files = all_files[:sample_size]  # Ensure we only get the requested number
        
        if not all_files:
            self.logger.error("❌ No applications found!")
            return {}
        
        self.logger.info(f"📋 Found {len(all_files)} applications:")
        for f in all_files:
            self.logger.info(f"   - {f.name}")
        
        # Extract each application
        batch_results = {
            'timestamp': datetime.now().isoformat(),
            'applications': {},
            'batch_summary': {
                'total_apps': len(all_files),
                'total_sheets': 0,
                'total_cells': 0,
                'total_formulas': 0,
                'total_findings': 0
            }
        }
        
        for idx, file_path in enumerate(all_files, 1):
            self.logger.info(f"\n{'='*60}")
            self.logger.info(f"Processing [{idx}/{len(all_files)}]: {file_path.name}")
            self.logger.info(f"{'='*60}")
            
            try:
                app_result = self.extract_single_application(file_path)
                batch_results['applications'][file_path.name] = app_result
                
                # Update batch totals
                summary = app_result.get('summary', {})
                batch_results['batch_summary']['total_sheets'] += summary.get('total_sheets', 0)
                batch_results['batch_summary']['total_cells'] += summary.get('total_cells_extracted', 0)
                batch_results['batch_summary']['total_formulas'] += summary.get('total_formulas', 0)
                batch_results['batch_summary']['total_findings'] += summary.get('intelligent_findings_count', 0)
                
            except Exception as e:
                self.logger.error(f"❌ Error processing {file_path.name}: {e}")
                batch_results['applications'][file_path.name] = {
                    'error': str(e),
                    'summary': {'extraction_failed': True}
                }
        
        return batch_results
    
    def extract_single_application(self, file_path: Path) -> dict:
        """Extract single application comprehensively"""
        result = {
            'file_info': {
                'name': file_path.name,
                'path': str(file_path),
                'timestamp': datetime.now().isoformat()
            },
            'sheets_data': {},
            'intelligent_findings': {},
            'summary': {}
        }
        
        # Load Excel file
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Process each sheet
        for sheet_name in wb_formulas.sheetnames:
            self.logger.info(f"  📋 Processing: {sheet_name}")
            
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]
            
            sheet_result = self._extract_sheet_data(ws_formulas, ws_values, sheet_name)
            result['sheets_data'][sheet_name] = sheet_result
            
            # Apply intelligent pattern matching for critical sheets
            if sheet_name in self.critical_patterns:
                findings = self._apply_intelligent_extraction(sheet_result, sheet_name)
                if findings:
                    result['intelligent_findings'][sheet_name] = findings
        
        wb_formulas.close()
        wb_values.close()
        
        # Generate summary
        result['summary'] = self._generate_app_summary(result)
        
        return result
    
    def _extract_sheet_data(self, ws_formulas, ws_values, sheet_name: str) -> dict:
        """Extract data from a single sheet"""
        sheet_data = {
            'metadata': {
                'sheet_name': sheet_name,
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            },
            'cells': {},
            'formulas': {},
            'merged_cells': []
        }
        
        # Get merged cells
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract meaningful cells
        for row in range(1, min(ws_formulas.max_row + 1, 300)):  # Limit for performance
            for col in range(1, min(ws_formulas.max_column + 1, 40)):
                
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__
                    }
                    
                    # Check for formula
                    if (hasattr(cell_formula, '_value') and 
                        isinstance(cell_formula._value, str) and 
                        cell_formula._value.startswith('=')):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                    
                    sheet_data['cells'][cell_ref] = cell_info
        
        return sheet_data
    
    def _apply_intelligent_extraction(self, sheet_data: dict, sheet_name: str) -> dict:
        """Apply intelligent pattern matching"""
        findings = {}
        patterns = self.critical_patterns[sheet_name]
        
        for field_name, search_terms in patterns.items():
            found_instances = []
            
            # Search through all cells for patterns
            for cell_ref, cell_info in sheet_data['cells'].items():
                if isinstance(cell_info['value'], str):
                    cell_text = str(cell_info['value']).lower()
                    
                    for term in search_terms:
                        if str(term).lower() in cell_text:
                            # Try to find associated value
                            associated_value = self._find_associated_value(cell_info, sheet_data)
                            
                            found_instances.append({
                                'label': cell_info['value'],
                                'location': cell_ref,
                                'row': cell_info['row'],
                                'col': cell_info['col'],
                                'associated_value': associated_value,
                                'search_term': term
                            })
                            break
            
            if found_instances:
                findings[field_name] = found_instances
        
        return findings
    
    def _find_associated_value(self, label_cell: dict, sheet_data: dict):
        """Find value associated with a label"""
        row = label_cell['row']
        col = label_cell['col']
        
        # Check common locations for values
        check_positions = [
            (row, col + 1),      # Right
            (row, col + 2),      # Two right
            (row + 1, col),      # Below
            (row - 1, col),      # Above
        ]
        
        for check_row, check_col in check_positions:
            check_ref = f"{openpyxl.utils.get_column_letter(check_col)}{check_row}"
            if check_ref in sheet_data['cells']:
                value = sheet_data['cells'][check_ref]['value']
                if value is not None and not isinstance(value, str):
                    return value
                elif isinstance(value, str) and value.strip() and len(value.strip()) < 100:
                    return value.strip()
        
        return None
    
    def _generate_app_summary(self, result: dict) -> dict:
        """Generate summary for single application"""
        summary = {
            'total_sheets': len(result['sheets_data']),
            'sheets_processed': list(result['sheets_data'].keys()),
            'total_cells_extracted': 0,
            'total_formulas': 0,
            'intelligent_findings_count': 0,
            'key_project_info': {}
        }
        
        # Count totals
        for sheet_data in result['sheets_data'].values():
            summary['total_cells_extracted'] += len(sheet_data['cells'])
            summary['total_formulas'] += len(sheet_data['formulas'])
        
        # Count intelligent findings
        for findings in result['intelligent_findings'].values():
            summary['intelligent_findings_count'] += len(findings)
        
        # Extract key project info from Application sheet
        if 'Application' in result['intelligent_findings']:
            app_findings = result['intelligent_findings']['Application']
            for field, instances in app_findings.items():
                if instances:
                    best_instance = instances[0]
                    summary['key_project_info'][field] = {
                        'label': best_instance['label'],
                        'value': best_instance['associated_value'],
                        'location': best_instance['location']
                    }
        
        return summary
    
    def create_batch_comparison_excel(self, batch_results: dict, output_path: Path):
        """Create comprehensive batch comparison Excel"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Batch Overview sheet
        self._create_batch_overview(wb, batch_results)
        
        # Create Comparison Dashboard
        self._create_comparison_dashboard(wb, batch_results)
        
        # Create detailed sheets for each application
        for app_name, app_data in batch_results['applications'].items():
            if 'error' not in app_data:
                safe_name = app_name.replace('.xlsx', '')[:20]
                self._create_app_detail_sheet(wb, safe_name, app_data)
        
        # Save the workbook
        wb.save(output_path)
        self.logger.info(f"💾 Batch comparison Excel saved: {output_path.name}")
        
        return output_path
    
    def _create_batch_overview(self, wb: Workbook, batch_results: dict):
        """Create batch overview sheet"""
        ws = wb.create_sheet(title="BATCH OVERVIEW", index=0)
        
        # Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        bold_font = Font(bold=True)
        
        # Title
        ws['A1'] = "COMPREHENSIVE BATCH EXTRACTION RESULTS"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:E1')
        
        row = 3
        
        # Batch Statistics
        ws[f'A{row}'] = "BATCH STATISTICS"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        batch_summary = batch_results['batch_summary']
        stats = [
            ("Total Applications", batch_summary['total_apps']),
            ("Total Sheets Processed", batch_summary['total_sheets']),
            ("Total Cells Extracted", batch_summary['total_cells']),
            ("Total Formulas Preserved", batch_summary['total_formulas']),
            ("Total Intelligent Findings", batch_summary['total_findings'])
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Per-Application Summary
        ws[f'A{row}'] = "PER-APPLICATION SUMMARY"
        ws[f'A{row}'].font = bold_font
        row += 1
        
        # Headers
        headers = ['Application', 'Sheets', 'Cells', 'Formulas', 'Findings', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        # Data rows
        for app_name, app_data in batch_results['applications'].items():
            ws.cell(row=row, column=1, value=app_name)
            
            if 'error' in app_data:
                ws.cell(row=row, column=6, value="ERROR")
            else:
                summary = app_data.get('summary', {})
                ws.cell(row=row, column=2, value=summary.get('total_sheets', 0))
                ws.cell(row=row, column=3, value=summary.get('total_cells_extracted', 0))
                ws.cell(row=row, column=4, value=summary.get('total_formulas', 0))
                ws.cell(row=row, column=5, value=summary.get('intelligent_findings_count', 0))
                ws.cell(row=row, column=6, value="SUCCESS")
            
            row += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_comparison_dashboard(self, wb: Workbook, batch_results: dict):
        """Create comparison dashboard"""
        ws = wb.create_sheet(title="COMPARISON DASHBOARD")
        
        # Title
        ws['A1'] = "V2.0 vs V1.0 EXTRACTION COMPARISON"
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        
        # Comparison metrics
        ws[f'A{row}'] = "ENHANCEMENT METRICS"
        ws[f'A{row}'].font = Font(bold=True, color="0066CC")
        row += 2
        
        # Calculate averages
        apps_with_data = [app for app in batch_results['applications'].values() if 'error' not in app]
        
        if apps_with_data:
            avg_sheets = sum(app['summary']['total_sheets'] for app in apps_with_data) / len(apps_with_data)
            avg_cells = sum(app['summary']['total_cells_extracted'] for app in apps_with_data) / len(apps_with_data)
            avg_formulas = sum(app['summary']['total_formulas'] for app in apps_with_data) / len(apps_with_data)
            avg_findings = sum(app['summary']['intelligent_findings_count'] for app in apps_with_data) / len(apps_with_data)
            
            metrics = [
                ("Average Sheets per App", f"{avg_sheets:.1f}"),
                ("Average Cells per App", f"{avg_cells:.0f}"),
                ("Average Formulas per App", f"{avg_formulas:.0f}"),
                ("Average Findings per App", f"{avg_findings:.1f}"),
                ("", ""),
                ("V1.0 Results (from previous run):", ""),
                ("- Files Processed", "3"),
                ("- Sheets Extracted", "~18"),
                ("- Cells Extracted", "~14,697"),
                ("- Formulas Preserved", "~5,119"),
                ("", ""),
                ("V2.0 IMPROVEMENTS:", ""),
                ("- Enhanced Pattern Matching", "✓"),
                ("- Comprehensive Sheet Coverage", "✓"),
                ("- Intelligent Field Detection", "✓"),
                ("- Professional Excel Output", "✓")
            ]
            
            for label, value in metrics:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
    
    def _create_app_detail_sheet(self, wb: Workbook, app_name: str, app_data: dict):
        """Create detailed sheet for specific application"""
        ws = wb.create_sheet(title=f"DETAIL_{app_name}")
        
        # Title
        ws['A1'] = f"DETAILED EXTRACTION: {app_name}"
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        
        # Application summary
        summary = app_data.get('summary', {})
        ws[f'A{row}'] = "APPLICATION SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, color="0066CC")
        row += 1
        
        summary_items = [
            ("Sheets Processed", summary.get('total_sheets', 0)),
            ("Cells Extracted", summary.get('total_cells_extracted', 0)),
            ("Formulas Preserved", summary.get('total_formulas', 0)),
            ("Intelligent Findings", summary.get('intelligent_findings_count', 0))
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Intelligent findings
        if app_data.get('intelligent_findings'):
            ws[f'A{row}'] = "INTELLIGENT FINDINGS"
            ws[f'A{row}'].font = Font(bold=True, color="006600")
            row += 1
            
            # Headers
            headers = ['Sheet', 'Field', 'Label', 'Value', 'Location']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1
            
            # Data
            for sheet_name, findings in app_data['intelligent_findings'].items():
                for field_name, instances in findings.items():
                    for instance in instances:
                        ws.cell(row=row, column=1, value=sheet_name)
                        ws.cell(row=row, column=2, value=field_name)
                        ws.cell(row=row, column=3, value=str(instance['label'])[:50])
                        ws.cell(row=row, column=4, value=str(instance['associated_value']) if instance['associated_value'] else 'N/A')
                        ws.cell(row=row, column=5, value=instance['location'])
                        row += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """Run batch comprehensive extraction"""
    print("="*80)
    print("🚀 BATCH COMPREHENSIVE EXTRACTION")
    print("   V2.0 Enhanced System vs Original V1.0 Results")
    print("="*80)
    
    # Paths
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/ctcac_extractions_v2")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Initialize extractor
    extractor = BatchComprehensiveExtractor()
    
    # Run batch extraction (same 3 apps as original)
    batch_results = extractor.extract_batch_applications(data_dir, sample_size=3)
    
    # Create comprehensive comparison Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f"BATCH_COMPARISON_V2_{timestamp}.xlsx"
    extractor.create_batch_comparison_excel(batch_results, output_file)
    
    # Print results
    print("\n" + "="*80)
    print("✅ BATCH COMPREHENSIVE EXTRACTION COMPLETE!")
    print("="*80)
    
    batch_summary = batch_results['batch_summary']
    print(f"📊 Applications Processed: {batch_summary['total_apps']}")
    print(f"📋 Total Sheets: {batch_summary['total_sheets']}")
    print(f"🔢 Total Cells: {batch_summary['total_cells']}")
    print(f"🧮 Total Formulas: {batch_summary['total_formulas']}")
    print(f"🎯 Total Findings: {batch_summary['total_findings']}")
    
    print("\n📈 V2.0 ENHANCEMENTS:")
    print("   - Comprehensive pattern matching")
    print("   - Enhanced field detection")
    print("   - Professional Excel dashboards")
    print("   - Detailed comparison analytics")
    
    print(f"\n💾 Comprehensive Comparison Excel: {output_file.name}")
    print("\nThis file contains V2.0 results with comparison to V1.0!")

if __name__ == "__main__":
    main()