#!/usr/bin/env python3
"""
Quick Sheet Analysis - Compare processed files to find 4,839 cell difference
"""

import openpyxl
from pathlib import Path

def quick_sheet_analysis():
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    
    # Use processed files (already have smart ranges applied)
    file_2024 = data_dir / "processed_2024_4pct_R1_24-409.xlsx"
    file_2025 = data_dir / "processed_2025_4pct_R1_25-464.xlsx"
    
    print(f"🔍 QUICK SHEET ANALYSIS - Processed Files")
    print(f"Finding the 4,839 extra cells in 2024 vs 2025")
    print("=" * 50)
    
    wb_2024 = openpyxl.load_workbook(file_2024, read_only=True)
    wb_2025 = openpyxl.load_workbook(file_2025, read_only=True)
    
    print(f"{'Sheet Name':<35} {'2024':<10} {'2025':<10} {'Diff':<10}")
    print("-" * 65)
    
    total_2024 = 0
    total_2025 = 0
    
    # Get all sheet names
    sheets_2024 = {sheet.title: sheet for sheet in wb_2024.worksheets}
    sheets_2025 = {sheet.title: sheet for sheet in wb_2025.worksheets}
    all_sheet_names = set(sheets_2024.keys()).union(set(sheets_2025.keys()))
    
    differences = []
    
    for sheet_name in sorted(all_sheet_names):
        cells_2024 = 0
        cells_2025 = 0
        
        if sheet_name in sheets_2024:
            sheet = sheets_2024[sheet_name]
            cells_2024 = (sheet.max_row or 0) * (sheet.max_column or 0)
            
        if sheet_name in sheets_2025:
            sheet = sheets_2025[sheet_name]
            cells_2025 = (sheet.max_row or 0) * (sheet.max_column or 0)
        
        diff = cells_2024 - cells_2025
        differences.append((sheet_name, cells_2024, cells_2025, diff))
        
        total_2024 += cells_2024
        total_2025 += cells_2025
        
        status = ""
        if abs(diff) > 1000:
            status = "🔴"
        elif abs(diff) > 100:
            status = "🟡"
            
        print(f"{sheet_name:<35} {cells_2024:<10,} {cells_2025:<10,} {diff:<10,} {status}")
    
    print("-" * 65)
    print(f"{'TOTALS':<35} {total_2024:<10,} {total_2025:<10,} {total_2024-total_2025:<10,}")
    
    print(f"\n🎯 BIGGEST CONTRIBUTORS:")
    sorted_diffs = sorted(differences, key=lambda x: abs(x[3]), reverse=True)
    for name, c2024, c2025, diff in sorted_diffs[:5]:
        if abs(diff) > 50:
            print(f"  {name}: {diff:+,} cells ({c2024:,} vs {c2025:,})")
    
    # Check for sheets that exist in one file but not the other
    only_2024 = set(sheets_2024.keys()) - set(sheets_2025.keys())
    only_2025 = set(sheets_2025.keys()) - set(sheets_2024.keys())
    
    if only_2024:
        print(f"\n📋 SHEETS ONLY IN 2024: {', '.join(only_2024)}")
    if only_2025:
        print(f"\n📋 SHEETS ONLY IN 2025: {', '.join(only_2025)}")
    
    wb_2024.close()
    wb_2025.close()

if __name__ == "__main__":
    quick_sheet_analysis()