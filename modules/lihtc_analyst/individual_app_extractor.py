#!/usr/bin/env python3
"""
INDIVIDUAL APPLICATION EXTRACTOR
Create separate comprehensive extraction files for each application
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime
import logging
import json

class IndividualAppExtractor:
    """Create individual comprehensive extractions for each CTCAC application"""
    
    def __init__(self):
        self.setup_logging()
        
        # Enhanced patterns for comprehensive extraction
        self.extraction_patterns = {
            'Application': {
                'project_name': ['Project Name', 'project name', 'Development Name'],
                'address': ['Site Address', 'Project Address'],
                'city': ['City'],
                'county': ['County'],  
                'zip_code': ['Zip', 'ZIP'],
                'total_units': ['Total Units', 'Unit Count', 'Number of units'],
                'applicant': ['CTCAC APPLICANT', 'Applicant Name'],
                'contact_person': ['Contact Person', 'Primary Contact'],
                'phone': ['Phone', 'Telephone'],
                'email': ['Email', 'E-mail'],
                'census_tract': ['Census Tract'],
                'annual_credits': ['annual Federal Credits', 'Annual Federal Tax Credit', 'Federal Credits'],
                'state_credits': ['State Credits', 'total State Credits'],
                'qct_status': ['Qualified Census Tract', 'QCT'],
                'dda_status': ['DDA', 'Difficult Development Area'],
                'set_aside': ['Set-Aside', 'Set Aside Election'],
                'geographic_area': ['Geographic Area', 'CTCAC Geographic'],
                'placed_in_service': ['Placed in Service', 'Service Date'],
                'application_date': ['Application Date'],
                'architect': ['Architect'],
                'contractor': ['General Contractor', 'Contractor'],
                'property_management': ['Property Management', 'Management Company']
            },
            
            'Sources and Uses Budget': {
                'tax_exempt_bond': ['Tax-Exempt', 'Bond Proceeds', 'Municipal Bond'],
                'tax_credit_equity': ['LIHTC Investor', 'Tax Credit Equity', 'Equity'],
                'taxable_loan': ['Taxable', 'Conventional Loan'],
                'seller_financing': ['Seller Financing', 'Seller Note'],
                'deferred_developer_fee': ['Deferred Developer Fee', 'Developer Fee Deferral'],
                'soft_loan': ['Soft Loan', 'Subordinate'],
                'grant_funds': ['Grant', 'Grant Funds'],
                'cash_equity': ['Cash Equity', 'Sponsor Equity'],
                'total_sources': ['Total Sources'],
                'acquisition_cost': ['Acquisition', 'Land Cost'],
                'construction_cost': ['Construction', 'Hard Cost'],
                'soft_costs': ['Soft Costs', 'Professional Fees'],
                'developer_fee': ['Developer Fee', 'Developer Overhead'],
                'reserves': ['Reserves', 'Reserve Funds'],
                'total_uses': ['Total Uses', 'Total Development Cost'],
                'cost_per_unit': ['Cost Per Unit', 'Per Unit Cost'],
                'lender_info': ['Lender', 'Financial Institution'],
                'interest_rate': ['Interest Rate', 'Rate'],
                'loan_term': ['Term', 'Maturity'],
                'loan_to_cost': ['Loan to Cost', 'LTC']
            },
            
            'Basis & Credits': {
                'eligible_basis': ['Eligible Basis', 'Total Eligible'],
                'qualified_basis': ['Qualified Basis', 'Total Qualified'],
                'applicable_percentage': ['Applicable Percentage', 'Credit Rate'],
                'annual_credits': ['Annual Credits', 'Annual Federal Credits'],
                'credit_period': ['Credit Period', '10-year'],
                'basis_adjustments': ['Basis Adjustment', 'Adjustment'],
                'acquisition_basis': ['Acquisition Basis'],
                'rehab_basis': ['Rehabilitation Basis', 'Rehab'],
                'new_construction_basis': ['New Construction Basis'],
                'boost_eligible': ['Boost Eligible', '30% Boost'],
                'federal_credits': ['Federal Credits'],
                'state_credits': ['State Credits']
            },
            
            'Points System': {
                'tie_breaker_points': ['Tie Breaker', 'Tie-Breaker Points'],
                'readiness_points': ['Readiness', 'Site Control'],
                'sustainability_points': ['Sustainability', 'Green Building', 'LEED'],
                'special_needs_points': ['Special Needs', 'Senior Housing'],
                'service_amenity_points': ['Service Amenity', 'Amenities'],
                'leveraging_points': ['Leveraging', 'Leverage'],
                'community_revitalization': ['Community Revitalization'],
                'public_transportation': ['Public Transportation', 'Transit'],
                'neighborhood_revitalization': ['Neighborhood Revitalization'],
                'total_points': ['Total Points', 'Point Total']
            }
        }
        
        self.logger.info("🎯 INDIVIDUAL APP EXTRACTOR Ready!")
    
    def setup_logging(self):
        """Setup logging"""
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def extract_individual_application(self, file_path: Path, output_dir: Path):
        """Create comprehensive individual extraction"""
        self.logger.info(f"🚀 INDIVIDUAL EXTRACTION: {file_path.name}")
        
        # Extract data
        app_data = self._extract_comprehensive_data(file_path)
        
        # Create individual Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = file_path.name.replace('.xlsx', '')
        excel_file = output_dir / f"INDIVIDUAL_{base_name}_{timestamp}.xlsx"
        
        self._create_individual_excel(app_data, excel_file)
        
        # Create JSON export for programmatic access
        json_file = output_dir / f"INDIVIDUAL_{base_name}_{timestamp}.json"
        self._create_json_export(app_data, json_file)
        
        # Print summary
        self._print_extraction_summary(app_data, excel_file, json_file)
        
        return excel_file, json_file
    
    def _extract_comprehensive_data(self, file_path: Path) -> dict:
        """Extract comprehensive data from application"""
        app_data = {
            'file_info': {
                'name': file_path.name,
                'path': str(file_path),
                'extraction_timestamp': datetime.now().isoformat()
            },
            'sheets_data': {},
            'intelligent_extractions': {},
            'raw_cell_dump': {},
            'formula_analysis': {},
            'summary': {}
        }
        
        # Load Excel file
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        # Process each sheet comprehensively
        for sheet_name in wb_formulas.sheetnames:
            self.logger.info(f"  📋 Extracting: {sheet_name}")
            
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]
            
            # Extract sheet data
            sheet_result = self._extract_sheet_comprehensive(ws_formulas, ws_values, sheet_name)
            app_data['sheets_data'][sheet_name] = sheet_result
            
            # Apply intelligent extraction for critical sheets
            if sheet_name in self.extraction_patterns:
                intelligent_results = self._apply_intelligent_patterns(sheet_result, sheet_name)
                if intelligent_results:
                    app_data['intelligent_extractions'][sheet_name] = intelligent_results
            
            # Create raw cell dump for first 100 cells with data
            app_data['raw_cell_dump'][sheet_name] = self._create_raw_dump(sheet_result)
            
            # Analyze formulas
            if sheet_result['formulas']:
                app_data['formula_analysis'][sheet_name] = self._analyze_formulas(sheet_result['formulas'])
        
        wb_formulas.close()
        wb_values.close()
        
        # Generate summary
        app_data['summary'] = self._generate_comprehensive_summary(app_data)
        
        return app_data
    
    def _extract_sheet_comprehensive(self, ws_formulas, ws_values, sheet_name: str) -> dict:
        """Extract everything from a sheet"""
        sheet_data = {
            'metadata': {
                'sheet_name': sheet_name,
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column,
                'merged_cells_count': len(list(ws_formulas.merged_cells.ranges))
            },
            'cells': {},
            'formulas': {},
            'merged_cells': [],
            'cell_styles': {},
            'data_types': {}
        }
        
        # Get merged cells
        for merged_range in ws_formulas.merged_cells.ranges:
            sheet_data['merged_cells'].append(str(merged_range))
        
        # Extract cells (expanded range for comprehensive coverage)
        for row in range(1, min(ws_formulas.max_row + 1, 500)):
            for col in range(1, min(ws_formulas.max_column + 1, 50)):
                
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                if cell_value.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    # Basic cell info
                    cell_info = {
                        'row': row,
                        'col': col,
                        'value': cell_value.value,
                        'display_value': str(cell_value.value)[:200] if cell_value.value else None,
                        'formula': None,
                        'data_type': type(cell_value.value).__name__
                    }
                    
                    # Check for formula
                    if (hasattr(cell_formula, '_value') and 
                        isinstance(cell_formula._value, str) and 
                        cell_formula._value.startswith('=')):
                        cell_info['formula'] = cell_formula._value
                        sheet_data['formulas'][cell_ref] = cell_formula._value
                    
                    # Track data types
                    data_type = type(cell_value.value).__name__
                    if data_type not in sheet_data['data_types']:
                        sheet_data['data_types'][data_type] = 0
                    sheet_data['data_types'][data_type] += 1
                    
                    sheet_data['cells'][cell_ref] = cell_info
        
        return sheet_data
    
    def _apply_intelligent_patterns(self, sheet_data: dict, sheet_name: str) -> dict:
        """Apply intelligent pattern matching"""
        results = {}
        patterns = self.extraction_patterns[sheet_name]
        
        for field_name, search_terms in patterns.items():
            matches = []
            
            # Search all cells for pattern matches
            for cell_ref, cell_info in sheet_data['cells'].items():
                if isinstance(cell_info['value'], str):
                    cell_text = str(cell_info['value']).lower()
                    
                    for term in search_terms:
                        if str(term).lower() in cell_text:
                            # Find associated value
                            associated_value = self._find_value_nearby(cell_info, sheet_data)
                            
                            matches.append({
                                'label_text': cell_info['value'],
                                'label_location': cell_ref,
                                'associated_value': associated_value,
                                'value_confidence': self._calculate_value_confidence(associated_value),
                                'search_term_matched': term,
                                'row': cell_info['row'],
                                'col': cell_info['col']
                            })
                            break
            
            if matches:
                results[field_name] = matches
        
        return results
    
    def _find_value_nearby(self, label_cell: dict, sheet_data: dict):
        """Find value near a label with multiple strategies"""
        row = label_cell['row']
        col = label_cell['col']
        
        # Strategy 1: Same cell (after colon, dash, etc.)
        if isinstance(label_cell['value'], str):
            text = str(label_cell['value'])
            for separator in [':', '-', '=', '|']:
                if separator in text:
                    parts = text.split(separator, 1)
                    if len(parts) > 1:
                        potential_value = parts[1].strip()
                        if potential_value and len(potential_value) < 100:
                            return potential_value
        
        # Strategy 2: Check nearby cells in priority order
        search_positions = [
            (row, col + 1),      # Right
            (row, col + 2),      # Two right
            (row + 1, col),      # Below
            (row + 1, col + 1),  # Below-right
            (row - 1, col + 1),  # Above-right
            (row, col + 3),      # Three right
            (row + 2, col),      # Two below
        ]
        
        for search_row, search_col in search_positions:
            search_ref = f"{openpyxl.utils.get_column_letter(search_col)}{search_row}"
            if search_ref in sheet_data['cells']:
                candidate = sheet_data['cells'][search_ref]['value']
                
                # Prefer numeric values
                if isinstance(candidate, (int, float)) and candidate != 0:
                    return candidate
                
                # Accept short, meaningful text
                if (isinstance(candidate, str) and 
                    candidate.strip() and 
                    len(candidate.strip()) < 100 and
                    candidate.strip().lower() not in ['yes', 'no', 'n/a', 'na']):
                    return candidate.strip()
        
        return None
    
    def _calculate_value_confidence(self, value) -> str:
        """Calculate confidence in extracted value"""
        if value is None:
            return "NONE"
        elif isinstance(value, (int, float)):
            return "HIGH"
        elif isinstance(value, str) and len(value.strip()) > 0:
            if any(char.isdigit() for char in value):
                return "MEDIUM"
            else:
                return "LOW"
        else:
            return "UNKNOWN"
    
    def _create_raw_dump(self, sheet_data: dict) -> list:
        """Create raw dump of first 100 meaningful cells"""
        raw_dump = []
        count = 0
        
        for cell_ref, cell_info in sheet_data['cells'].items():
            if count >= 100:
                break
            
            # Only include cells with meaningful data
            if (cell_info['value'] is not None and 
                str(cell_info['value']).strip() and 
                str(cell_info['value']).strip() not in ['0', '0.0']):
                
                raw_dump.append({
                    'location': cell_ref,
                    'value': str(cell_info['value'])[:100],
                    'data_type': cell_info['data_type'],
                    'has_formula': bool(cell_info.get('formula'))
                })
                count += 1
        
        return raw_dump
    
    def _analyze_formulas(self, formulas: dict) -> dict:
        """Analyze Excel formulas"""
        analysis = {
            'total_formulas': len(formulas),
            'formula_types': {},
            'complex_formulas': [],
            'sample_formulas': []
        }
        
        for cell_ref, formula in formulas.items():
            # Count formula types
            if 'SUM(' in formula:
                analysis['formula_types']['SUM'] = analysis['formula_types'].get('SUM', 0) + 1
            elif 'IF(' in formula:
                analysis['formula_types']['IF'] = analysis['formula_types'].get('IF', 0) + 1
            elif 'VLOOKUP(' in formula:
                analysis['formula_types']['VLOOKUP'] = analysis['formula_types'].get('VLOOKUP', 0) + 1
            else:
                analysis['formula_types']['OTHER'] = analysis['formula_types'].get('OTHER', 0) + 1
            
            # Identify complex formulas
            if len(formula) > 50 or formula.count('(') > 2:
                analysis['complex_formulas'].append({
                    'location': cell_ref,
                    'formula': formula[:100]
                })
            
            # Sample formulas
            if len(analysis['sample_formulas']) < 5:
                analysis['sample_formulas'].append({
                    'location': cell_ref,
                    'formula': formula[:100]
                })
        
        return analysis
    
    def _generate_comprehensive_summary(self, app_data: dict) -> dict:
        """Generate comprehensive summary"""
        summary = {
            'extraction_stats': {
                'total_sheets': len(app_data['sheets_data']),
                'total_cells': sum(len(sheet['cells']) for sheet in app_data['sheets_data'].values()),
                'total_formulas': sum(len(sheet['formulas']) for sheet in app_data['sheets_data'].values()),
                'total_intelligent_findings': sum(len(findings) for findings in app_data['intelligent_extractions'].values())
            },
            'sheet_breakdown': {},
            'key_findings': {},
            'data_quality': {}
        }
        
        # Sheet breakdown
        for sheet_name, sheet_data in app_data['sheets_data'].items():
            summary['sheet_breakdown'][sheet_name] = {
                'cells': len(sheet_data['cells']),
                'formulas': len(sheet_data['formulas']),
                'data_types': sheet_data['data_types'],
                'intelligent_findings': len(app_data['intelligent_extractions'].get(sheet_name, {}))
            }
        
        # Key findings from intelligent extractions
        for sheet_name, extractions in app_data['intelligent_extractions'].items():
            summary['key_findings'][sheet_name] = {}
            for field_name, matches in extractions.items():
                if matches:
                    best_match = matches[0]  # Take first match
                    summary['key_findings'][sheet_name][field_name] = {
                        'value': best_match['associated_value'],
                        'confidence': best_match['value_confidence'],
                        'location': best_match['label_location']
                    }
        
        # Data quality assessment
        total_findings_with_values = 0
        total_findings = 0
        
        for extractions in app_data['intelligent_extractions'].values():
            for matches in extractions.values():
                for match in matches:
                    total_findings += 1
                    if match['associated_value'] is not None:
                        total_findings_with_values += 1
        
        if total_findings > 0:
            summary['data_quality']['value_extraction_rate'] = (total_findings_with_values / total_findings) * 100
        else:
            summary['data_quality']['value_extraction_rate'] = 0
        
        summary['data_quality']['total_findings'] = total_findings
        summary['data_quality']['findings_with_values'] = total_findings_with_values
        
        return summary
    
    def _create_individual_excel(self, app_data: dict, output_path: Path):
        """Create comprehensive individual Excel file"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Executive Summary
        self._create_executive_summary(wb, app_data)
        
        # Create Intelligent Findings sheet
        self._create_intelligent_findings_sheet(wb, app_data)
        
        # Create Formula Analysis sheet
        self._create_formula_analysis_sheet(wb, app_data)
        
        # Create Raw Data Dump sheets for key sheets
        key_sheets = ['Application', 'Sources and Uses Budget', 'Basis & Credits', 'Points System']
        for sheet_name in key_sheets:
            if sheet_name in app_data['raw_cell_dump']:
                self._create_raw_dump_sheet(wb, sheet_name, app_data['raw_cell_dump'][sheet_name])
        
        # Save workbook
        wb.save(output_path)
        self.logger.info(f"💾 Individual Excel saved: {output_path.name}")
    
    def _create_executive_summary(self, wb: Workbook, app_data: dict):
        """Create executive summary sheet"""
        ws = wb.create_sheet(title="EXECUTIVE SUMMARY", index=0)
        
        # Styling
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        header_font = Font(bold=True, size=12, color="0066CC")
        
        # Title
        ws['A1'] = f"COMPREHENSIVE EXTRACTION: {app_data['file_info']['name']}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # Extraction Statistics
        ws[f'A{row}'] = "EXTRACTION STATISTICS"
        ws[f'A{row}'].font = header_font
        row += 1
        
        stats = app_data['summary']['extraction_stats']
        for label, value in stats.items():
            ws[f'A{row}'] = label.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Data Quality
        ws[f'A{row}'] = "DATA QUALITY METRICS"
        ws[f'A{row}'].font = header_font
        row += 1
        
        quality = app_data['summary']['data_quality']
        ws[f'A{row}'] = "Value Extraction Rate"
        ws[f'B{row}'] = f"{quality['value_extraction_rate']:.1f}%"
        row += 1
        ws[f'A{row}'] = "Total Field Matches"
        ws[f'B{row}'] = quality['total_findings']
        row += 1
        ws[f'A{row}'] = "Matches with Values"
        ws[f'B{row}'] = quality['findings_with_values']
        row += 2
        
        # Key Project Information
        ws[f'A{row}'] = "KEY PROJECT INFORMATION"
        ws[f'A{row}'].font = header_font
        row += 1
        
        # Extract top findings from Application sheet
        app_findings = app_data['summary']['key_findings'].get('Application', {})
        for field_name, finding_data in list(app_findings.items())[:10]:  # Top 10
            if finding_data['value']:
                ws[f'A{row}'] = field_name.replace('_', ' ').title()
                ws[f'B{row}'] = str(finding_data['value'])[:50]
                ws[f'C{row}'] = finding_data['confidence']
                row += 1
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _create_intelligent_findings_sheet(self, wb: Workbook, app_data: dict):
        """Create detailed intelligent findings sheet"""
        ws = wb.create_sheet(title="INTELLIGENT FINDINGS")
        
        # Headers
        headers = ['Sheet', 'Field', 'Label Found', 'Extracted Value', 'Confidence', 'Location', 'Search Term']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 2
        
        # Add all intelligent findings
        for sheet_name, extractions in app_data['intelligent_extractions'].items():
            for field_name, matches in extractions.items():
                for match in matches:
                    ws.cell(row=row, column=1, value=sheet_name)
                    ws.cell(row=row, column=2, value=field_name)
                    ws.cell(row=row, column=3, value=str(match['label_text'])[:100])
                    ws.cell(row=row, column=4, value=str(match['associated_value'])[:100] if match['associated_value'] else 'N/A')
                    ws.cell(row=row, column=5, value=match['value_confidence'])
                    ws.cell(row=row, column=6, value=match['label_location'])
                    ws.cell(row=row, column=7, value=match['search_term_matched'])
                    row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_formula_analysis_sheet(self, wb: Workbook, app_data: dict):
        """Create formula analysis sheet"""
        ws = wb.create_sheet(title="FORMULA ANALYSIS")
        
        # Title
        ws['A1'] = "EXCEL FORMULA ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Overall statistics
        total_formulas = sum(analysis['total_formulas'] for analysis in app_data['formula_analysis'].values())
        ws[f'A{row}'] = "Total Formulas Found:"
        ws[f'B{row}'] = total_formulas
        row += 2
        
        # Per-sheet breakdown
        ws[f'A{row}'] = "FORMULAS BY SHEET"
        ws[f'A{row}'].font = Font(bold=True, color="0066CC")
        row += 1
        
        for sheet_name, analysis in app_data['formula_analysis'].items():
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = analysis['total_formulas']
            row += 1
        
        row += 2
        
        # Sample complex formulas
        ws[f'A{row}'] = "SAMPLE COMPLEX FORMULAS"
        ws[f'A{row}'].font = Font(bold=True, color="006600")
        row += 1
        
        headers = ['Sheet', 'Location', 'Formula']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        for sheet_name, analysis in app_data['formula_analysis'].items():
            for complex_formula in analysis.get('complex_formulas', [])[:10]:
                ws.cell(row=row, column=1, value=sheet_name)
                ws.cell(row=row, column=2, value=complex_formula['location'])
                ws.cell(row=row, column=3, value=complex_formula['formula'])
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_raw_dump_sheet(self, wb: Workbook, sheet_name: str, raw_data: list):
        """Create raw data dump sheet"""
        safe_name = f"RAW_{sheet_name[:20]}"
        ws = wb.create_sheet(title=safe_name)
        
        # Title
        ws['A1'] = f"RAW DATA: {sheet_name}"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Headers
        headers = ['Location', 'Value', 'Data Type', 'Has Formula']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for idx, cell_data in enumerate(raw_data, 4):
            ws.cell(row=idx, column=1, value=cell_data['location'])
            ws.cell(row=idx, column=2, value=cell_data['value'])
            ws.cell(row=idx, column=3, value=cell_data['data_type'])
            ws.cell(row=idx, column=4, value=cell_data['has_formula'])
        
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for col in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter and max_length > 0:
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_json_export(self, app_data: dict, json_path: Path):
        """Create JSON export for programmatic access"""
        with open(json_path, 'w') as f:
            json.dump(app_data, f, indent=2, default=str)
        self.logger.info(f"📄 JSON export saved: {json_path.name}")
    
    def _print_extraction_summary(self, app_data: dict, excel_file: Path, json_file: Path):
        """Print extraction summary"""
        stats = app_data['summary']['extraction_stats']
        quality = app_data['summary']['data_quality']
        
        print("\n" + "="*80)
        print(f"✅ INDIVIDUAL EXTRACTION COMPLETE: {app_data['file_info']['name']}")
        print("="*80)
        print(f"📊 Sheets Processed: {stats['total_sheets']}")
        print(f"🔢 Cells Extracted: {stats['total_cells']}")
        print(f"🧮 Formulas Preserved: {stats['total_formulas']}")
        print(f"🎯 Intelligent Findings: {stats['total_intelligent_findings']}")
        print(f"📈 Value Extraction Rate: {quality['value_extraction_rate']:.1f}%")
        
        print(f"\n💾 Excel File: {excel_file.name}")
        print(f"📄 JSON File: {json_file.name}")


def main():
    """Extract individual comprehensive files for each of the 3 applications"""
    print("="*80)
    print("🎯 INDIVIDUAL APPLICATION EXTRACTIONS")
    print("   Creating separate comprehensive files for each app")
    print("="*80)
    
    # Paths
    data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/individual_extractions")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # The same 3 applications
    applications = [
        "2025_4pct_R1_25-464.xlsx",
        "2025_4pct_R1_25-433.xlsx", 
        "2025_4pct_R1_25-425.xlsx"
    ]
    
    # Initialize extractor
    extractor = IndividualAppExtractor()
    
    # Process each application individually
    for app_name in applications:
        app_path = data_dir / app_name
        if app_path.exists():
            print(f"\n{'='*60}")
            print(f"Processing: {app_name}")
            print(f"{'='*60}")
            
            excel_file, json_file = extractor.extract_individual_application(app_path, output_dir)
        else:
            print(f"❌ File not found: {app_name}")
    
    print("\n" + "="*80)
    print("✅ ALL INDIVIDUAL EXTRACTIONS COMPLETE!")
    print("="*80)
    print(f"📁 Output Directory: {output_dir}")
    print("\nEach application now has its own comprehensive Excel and JSON files!")
    print("You can see exactly what data was scraped from each application.")

if __name__ == "__main__":
    main()