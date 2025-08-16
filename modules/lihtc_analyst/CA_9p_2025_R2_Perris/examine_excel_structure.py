#!/usr/bin/env python3
"""
Script to examine the structure of the Excel file and display its contents,
particularly focusing on column A and the area around cell B7.
"""

import pandas as pd
import openpyxl
from pathlib import Path

# Define the Excel file path
excel_file = Path("/Users/williamrice/HERR Dropbox/Bill Rice/Structured Consultants/AI Projects/CA_9p_2025_R2_Perris/9_Percent_Applications/TAB_08__Rehabilitation_Credit_Applications/Exhbit-B-RAWFILE_Att-8-Rehab Summary.xlsx")

print(f"Examining Excel file: {excel_file.name}")
print("=" * 80)

# Load the workbook to examine its structure
wb = openpyxl.load_workbook(excel_file)

# List all sheet names
print(f"\nWorkbook contains {len(wb.sheetnames)} sheet(s):")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Examine the first sheet
sheet = wb.active
print(f"\nExamining active sheet: '{sheet.title}'")
print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")

# Display column A contents and the area around B7
print("\n" + "=" * 80)
print("COLUMN A CONTENTS (first 30 rows):")
print("=" * 80)

for row in range(1, min(31, sheet.max_row + 1)):
    cell_a = sheet[f'A{row}']
    if cell_a.value is not None:
        print(f"A{row}: {cell_a.value}")

# Display the area around B7 (rows 5-15, columns A-E)
print("\n" + "=" * 80)
print("AREA AROUND B7 (rows 5-15, columns A-E):")
print("=" * 80)

for row in range(5, min(16, sheet.max_row + 1)):
    row_data = []
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = sheet[f'{col}{row}']
        value = cell.value if cell.value is not None else "[empty]"
        row_data.append(f"{col}{row}: {str(value)[:20]}")
    print(" | ".join(row_data))

# Check if there are any merged cells around B7
print("\n" + "=" * 80)
print("MERGED CELLS (if any):")
print("=" * 80)
merged_ranges = sheet.merged_cells.ranges
if merged_ranges:
    for merged_range in merged_ranges:
        print(f"  Merged range: {merged_range}")
else:
    print("  No merged cells found")

# Use pandas to read the Excel file and display a broader view
print("\n" + "=" * 80)
print("PANDAS VIEW OF THE DATA:")
print("=" * 80)

try:
    df = pd.read_excel(excel_file, sheet_name=0, header=None)
    print(f"\nDataFrame shape: {df.shape}")
    print("\nFirst 20 rows and 6 columns:")
    print(df.iloc[:20, :6])
except Exception as e:
    print(f"Error reading with pandas: {e}")

print("\n" + "=" * 80)
print("Script completed.")