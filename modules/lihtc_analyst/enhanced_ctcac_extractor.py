#!/usr/bin/env python3
"""
Enhanced CTCAC Application Data Extractor
Incorporates lessons from Strike Leader's docling research and WINGMAN's extraction reports
Supports both 4p and 9p applications from 2023-2025
"""

import pandas as pd
import json
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import hashlib
import numpy as np

class EnhancedCTCACExtractor:
    """
    Advanced extraction system for California CTCAC LIHTC applications
    Implements hybrid approach: Docling for complex PDFs, direct extraction for Excel
    """
    
    def __init__(self):
        # Key sheets based on CTCAC application structure
        self.critical_sheets = [
            'Application',
            'Sources and Uses Budget', 
            'Basis & Credits',
            'Sources and Basis Breakdown',
            'Points System',
            'Tie Breaker'
        ]
        
        # LIHTC-specific terms from WINGMAN's validated benchmark
        self.lihtc_terms = {
            'financial': [
                'eligible basis', 'qualified basis', 'applicable percentage',
                'developer fee', 'acquisition cost', 'construction cost',
                'hard costs', 'soft costs', 'total development cost'
            ],
            'regulatory': [
                'section 42', 'qct', 'dda', 'extended use agreement',
                'regulatory agreement', 'placed in service', 'carryover allocation'
            ],
            'scoring': [
                'tie breaker', 'readiness points', 'sustainable building',
                'special needs', 'rural', 'at-risk', 'preservation'
            ]
        }
        
        # Field mapping for consistent extraction
        self.field_mappings = {
            'project_name': ['Project Name', 'Development Name', 'Property Name'],
            'total_units': ['Total Units', 'Unit Count', 'Number of Units'],
            'tax_credit_requested': ['Annual Federal Tax Credit Requested', 'Tax Credits Requested'],
            'total_dev_cost': ['Total Development Cost', 'Total Project Cost', 'TDC'],
            'developer_fee': ['Developer Fee', 'Developer Overhead', 'Developer Compensation']
        }
        
    def extract_excel_application(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract data from CTCAC Excel application with enhanced validation
        """
        print(f"\n{'='*80}")
        print(f"📊 ENHANCED CTCAC EXTRACTION: {file_path.name}")
        print(f"{'='*80}")
        
        extraction_result = {
            'file': file_path.name,
            'timestamp': datetime.now().isoformat(),
            'type': self._detect_application_type(file_path),
            'sheets': {},
            'key_metrics': {},
            'validation': {},
            'errors': []
        }
        
        try:
            # Load Excel file
            excel_file = pd.ExcelFile(file_path)
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            print(f"📋 Found {len(excel_file.sheet_names)} sheets")
            
            # Extract each critical sheet
            for sheet_name in self.critical_sheets:
                if sheet_name in excel_file.sheet_names:
                    print(f"\n  Processing: {sheet_name}")
                    sheet_data = self._extract_sheet_intelligent(
                        excel_file, wb, sheet_name, file_path
                    )
                    extraction_result['sheets'][sheet_name] = sheet_data
                    
                    # Extract key metrics from specific sheets
                    if sheet_name == 'Application':
                        self._extract_application_metrics(sheet_data, extraction_result)
                    elif sheet_name == 'Sources and Uses Budget':
                        self._extract_financial_metrics(sheet_data, extraction_result)
                    elif sheet_name == 'Basis & Credits':
                        self._extract_basis_metrics(sheet_data, extraction_result)
                        
            # Validate extraction quality
            extraction_result['validation'] = self._validate_extraction(extraction_result)
            
            # Generate summary
            extraction_result['summary'] = self._generate_summary(extraction_result)
            
        except Exception as e:
            extraction_result['errors'].append(f"Extraction error: {str(e)}")
            print(f"❌ Error: {e}")
            
        return extraction_result
    
    def _detect_application_type(self, file_path: Path) -> str:
        """Detect if this is a 4% or 9% application"""
        if '9p' in str(file_path).lower() or '9_percent' in str(file_path).lower():
            return '9_percent'
        elif '4p' in str(file_path).lower() or '4_percent' in str(file_path).lower():
            return '4_percent'
        return 'unknown'
    
    def _extract_sheet_intelligent(self, excel_file: pd.ExcelFile, 
                                   wb: openpyxl.Workbook, 
                                   sheet_name: str,
                                   file_path: Path) -> Dict:
        """
        Intelligent extraction using multiple strategies
        Inspired by Strike Leader's hybrid approach
        """
        sheet_data = {
            'raw_data': {},
            'structured_data': {},
            'formulas': {},
            'validation_rules': {},
            'extracted_values': {}
        }
        
        # Strategy 1: Pandas extraction for tabular data
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        sheet_data['raw_data'] = df.to_dict('records')
        
        # Strategy 2: Openpyxl for formulas and specific cells
        ws = wb[sheet_name]
        
        # Extract key cells with formulas
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    # Check for formulas
                    if hasattr(cell, 'formula') and cell.formula:
                        sheet_data['formulas'][cell_ref] = {
                            'formula': cell.formula,
                            'value': cell.value
                        }
                    
                    # Look for LIHTC terms
                    if isinstance(cell.value, str):
                        for category, terms in self.lihtc_terms.items():
                            for term in terms:
                                if term.lower() in str(cell.value).lower():
                                    if category not in sheet_data['extracted_values']:
                                        sheet_data['extracted_values'][category] = {}
                                    sheet_data['extracted_values'][category][cell_ref] = cell.value
        
        # Strategy 3: Pattern-based extraction for specific fields
        sheet_data['structured_data'] = self._extract_structured_fields(df, ws)
        
        return sheet_data
    
    def _extract_structured_fields(self, df: pd.DataFrame, ws) -> Dict:
        """Extract specific structured fields using patterns"""
        structured = {}
        
        # Look for labeled values (Label: Value pattern)
        for idx, row in df.iterrows():
            for col_name, value in row.items():
                if pd.notna(value) and isinstance(value, str):
                    # Check if this looks like a label
                    for field_key, patterns in self.field_mappings.items():
                        for pattern in patterns:
                            if pattern.lower() in str(value).lower():
                                # Look for the value in adjacent cells
                                col_idx = df.columns.get_loc(col_name)
                                if col_idx < len(df.columns) - 1:
                                    next_val = row.iloc[col_idx + 1]
                                    if pd.notna(next_val):
                                        structured[field_key] = {
                                            'label': value,
                                            'value': next_val,
                                            'location': f"Row {idx+1}"
                                        }
        
        return structured
    
    def _extract_application_metrics(self, sheet_data: Dict, result: Dict):
        """Extract key metrics from Application sheet"""
        metrics = result['key_metrics']
        
        # Try to extract from structured data first
        if 'structured_data' in sheet_data:
            for field in ['project_name', 'total_units', 'tax_credit_requested']:
                if field in sheet_data['structured_data']:
                    metrics[field] = sheet_data['structured_data'][field]['value']
        
        # Extract from raw data patterns
        if 'raw_data' in sheet_data and sheet_data['raw_data']:
            for row in sheet_data['raw_data']:
                for key, value in row.items():
                    if pd.notna(value):
                        # Look for specific patterns
                        if isinstance(key, str):
                            if 'project' in key.lower() and 'name' in key.lower():
                                metrics['project_name'] = value
                            elif 'units' in key.lower() and 'total' in key.lower():
                                try:
                                    metrics['total_units'] = int(value)
                                except:
                                    pass
    
    def _extract_financial_metrics(self, sheet_data: Dict, result: Dict):
        """Extract financial metrics from Sources and Uses"""
        metrics = result['key_metrics']
        
        # Look for totals
        if 'raw_data' in sheet_data:
            for row in sheet_data['raw_data']:
                for key, value in row.items():
                    if pd.notna(value) and pd.notna(key):
                        key_str = str(key).lower()
                        if 'total' in key_str and 'development' in key_str:
                            try:
                                metrics['total_development_cost'] = float(value)
                            except:
                                pass
                        elif 'developer fee' in key_str:
                            try:
                                metrics['developer_fee'] = float(value)
                            except:
                                pass
    
    def _extract_basis_metrics(self, sheet_data: Dict, result: Dict):
        """Extract basis and credit calculations"""
        metrics = result['key_metrics']
        
        # Look for basis-related values
        if 'extracted_values' in sheet_data and 'financial' in sheet_data['extracted_values']:
            for cell_ref, value in sheet_data['extracted_values']['financial'].items():
                if 'eligible basis' in str(value).lower():
                    # Try to find the actual value
                    metrics['eligible_basis_label'] = value
                elif 'qualified basis' in str(value).lower():
                    metrics['qualified_basis_label'] = value
    
    def _validate_extraction(self, result: Dict) -> Dict:
        """Validate extraction quality using WINGMAN's QA protocols"""
        validation = {
            'completeness': 0,
            'lihtc_relevance': 0,
            'data_quality': 0,
            'issues': []
        }
        
        # Check completeness
        expected_fields = ['project_name', 'total_units', 'total_development_cost']
        found_fields = sum(1 for f in expected_fields if f in result['key_metrics'])
        validation['completeness'] = (found_fields / len(expected_fields)) * 100
        
        # Check LIHTC relevance
        lihtc_term_count = 0
        for sheet_data in result['sheets'].values():
            if 'extracted_values' in sheet_data:
                for category_values in sheet_data['extracted_values'].values():
                    lihtc_term_count += len(category_values)
        
        validation['lihtc_relevance'] = min(100, lihtc_term_count * 5)  # 5 points per term found
        
        # Data quality checks
        if 'total_units' in result['key_metrics']:
            units = result['key_metrics']['total_units']
            if isinstance(units, (int, float)) and 1 <= units <= 500:
                validation['data_quality'] += 33
            else:
                validation['issues'].append("Total units value seems incorrect")
        
        if 'total_development_cost' in result['key_metrics']:
            tdc = result['key_metrics']['total_development_cost']
            if isinstance(tdc, (int, float)) and tdc > 1000000:
                validation['data_quality'] += 33
            else:
                validation['issues'].append("TDC value seems too low")
                
        if 'developer_fee' in result['key_metrics']:
            validation['data_quality'] += 34
            
        return validation
    
    def _generate_summary(self, result: Dict) -> Dict:
        """Generate extraction summary"""
        summary = {
            'extraction_quality': 'UNKNOWN',
            'sheets_processed': len(result['sheets']),
            'key_metrics_found': len(result['key_metrics']),
            'validation_score': 0
        }
        
        # Calculate overall score
        if result['validation']:
            scores = [
                result['validation']['completeness'],
                result['validation']['lihtc_relevance'],
                result['validation']['data_quality']
            ]
            summary['validation_score'] = sum(scores) / len(scores)
            
            if summary['validation_score'] >= 80:
                summary['extraction_quality'] = 'HIGH'
            elif summary['validation_score'] >= 60:
                summary['extraction_quality'] = 'MEDIUM'
            else:
                summary['extraction_quality'] = 'LOW'
                
        return summary
    
    def extract_multiple_applications(self, directory: Path) -> List[Dict]:
        """Extract data from multiple CTCAC applications"""
        results = []
        
        # Find all Excel files
        excel_files = list(directory.glob("**/*.xlsx"))
        
        print(f"\n🔍 Found {len(excel_files)} Excel files to process")
        
        for file_path in excel_files:
            if not file_path.name.startswith('~'):  # Skip temp files
                result = self.extract_excel_application(file_path)
                results.append(result)
                
        return results
    
    def save_extraction_results(self, results: List[Dict], output_dir: Path):
        """Save extraction results to JSON and markdown report"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save JSON
        json_file = output_dir / f"ctcac_extraction_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_file, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        print(f"\n💾 Saved JSON results to: {json_file}")
        
        # Generate markdown report
        md_content = self._generate_markdown_report(results)
        md_file = output_dir / f"ctcac_extraction_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
        with open(md_file, 'w') as f:
            f.write(md_content)
        print(f"📄 Saved report to: {md_file}")
        
        return json_file, md_file
    
    def _generate_markdown_report(self, results: List[Dict]) -> str:
        """Generate comprehensive markdown report"""
        md = f"""# ENHANCED CTCAC EXTRACTION REPORT
**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M')}
**Files Processed**: {len(results)}

## EXTRACTION SUMMARY

| File | Type | Quality | Metrics Found | Validation Score |
|------|------|---------|---------------|------------------|
"""
        
        for result in results:
            file_name = result['file'][:30] + '...' if len(result['file']) > 30 else result['file']
            md += f"| {file_name} | {result.get('type', 'N/A')} | "
            md += f"{result.get('summary', {}).get('extraction_quality', 'N/A')} | "
            md += f"{result.get('summary', {}).get('key_metrics_found', 0)} | "
            md += f"{result.get('summary', {}).get('validation_score', 0):.1f}% |\n"
        
        md += "\n## KEY METRICS EXTRACTED\n\n"
        
        for result in results:
            if result.get('key_metrics'):
                md += f"\n### {result['file']}\n"
                for key, value in result['key_metrics'].items():
                    md += f"- **{key}**: {value}\n"
        
        md += "\n## VALIDATION ISSUES\n\n"
        
        has_issues = False
        for result in results:
            if result.get('validation', {}).get('issues'):
                has_issues = True
                md += f"\n### {result['file']}\n"
                for issue in result['validation']['issues']:
                    md += f"- ⚠️ {issue}\n"
        
        if not has_issues:
            md += "✅ No validation issues found\n"
        
        md += f"""
## TECHNICAL NOTES

This extraction used the enhanced hybrid approach incorporating:
- **Docling-inspired**: Multi-strategy extraction (pandas + openpyxl)
- **LIHTC Term Validation**: {len(self.lihtc_terms)} validated terms
- **QA Protocols**: WINGMAN's mandatory validation framework
- **Pattern Recognition**: Field mapping for common CTCAC structures

## RECOMMENDATIONS

1. **High Quality Extractions**: Focus on files with >80% validation scores
2. **Manual Review**: Files with <60% scores need manual verification  
3. **Missing Data**: Check formulas and merged cells for hidden values
4. **Next Steps**: Consider using GPT-OSS for semantic understanding of complex fields
"""
        
        return md


def demo_extraction():
    """Demonstrate the enhanced extraction system"""
    
    # Use the CA 9p application we have
    test_file = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/CA_9p_2025_R2_Perris/9_Percent_Applications/TAB_00__APPLICATION/2025_9_Percent_Competitive_Tax_Credit_Application.xlsx")
    
    if not test_file.exists():
        print(f"❌ Test file not found: {test_file}")
        return
    
    # Initialize extractor
    extractor = EnhancedCTCACExtractor()
    
    # Extract single file
    result = extractor.extract_excel_application(test_file)
    
    # Save results
    output_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/lihtc_analyst/extraction_results")
    extractor.save_extraction_results([result], output_dir)
    
    # Print summary
    print("\n" + "="*80)
    print("📊 EXTRACTION COMPLETE")
    print("="*80)
    print(f"Quality: {result['summary']['extraction_quality']}")
    print(f"Validation Score: {result['summary']['validation_score']:.1f}%")
    print(f"Key Metrics Found: {result['summary']['key_metrics_found']}")
    
    if result['key_metrics']:
        print("\n🎯 Key Metrics:")
        for key, value in result['key_metrics'].items():
            print(f"  - {key}: {value}")


if __name__ == "__main__":
    demo_extraction()