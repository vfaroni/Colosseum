#!/usr/bin/env python3
"""
🦁 M4 BEAST CTCAC CRUSHER - MASSIVELY PARALLEL PROCESSING
Mission: Demolish 570 CTCAC applications at maximum speed with quality
Hardware: M4 Max - 16 cores, 128GB RAM
Strategy: Multi-process parallel processing with real-time monitoring
"""

import os
import json
import time
import psutil
import openpyxl  
from pathlib import Path
from datetime import datetime
from multiprocessing import Pool, Manager, cpu_count
from concurrent.futures import ProcessPoolExecutor, as_completed
import signal
import sys
from typing import List, Dict, Tuple

# Import our proven V1.9 processor components
class SmartRangeDetector:
    """Intelligent range detection for CTCAC applications"""
    
    def __init__(self):
        self.SAFE_MAX_ROWS = 500     
        self.SAFE_MAX_COLS = 50      
    
    def detect_smart_range(self, sheet, sheet_name: str) -> Tuple[int, int]:
        """Detect actual CTCAC data boundaries vs Excel used range pollution"""
        excel_max_row = sheet.max_row or 1
        excel_max_col = sheet.max_column or 1
        
        if excel_max_col > 1000 or excel_max_row > 5000:
            return self._scan_for_actual_boundaries(sheet, sheet_name)
        
        return min(excel_max_row, self.SAFE_MAX_ROWS), min(excel_max_col, self.SAFE_MAX_COLS)
    
    def _scan_for_actual_boundaries(self, sheet, sheet_name: str) -> Tuple[int, int]:
        """Quick scan for actual data limits"""
        max_search_row = min(sheet.max_row or self.SAFE_MAX_ROWS, 2000)
        max_search_col = min(sheet.max_column or self.SAFE_MAX_COLS, 100)
        
        last_data_row = 1
        last_data_col = 1
        
        # Quick reverse scan
        for row_num in range(max_search_row, 0, -1):
            found_data = False
            for col_num in range(1, min(max_search_col + 1, 51)):
                try:
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        cell_text = str(cell.value).strip()
                        if cell_text and len(cell_text) > 0:
                            last_data_row = row_num
                            found_data = True
                            break
                except:
                    continue
            if found_data:
                break
        
        # Quick forward scan
        for col_num in range(1, min(max_search_col + 1, 51)):
            found_data = False
            for row_num in range(1, min(last_data_row + 1, 501)):
                try:
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        cell_text = str(cell.value).strip()
                        if cell_text and len(cell_text) > 0:
                            last_data_col = col_num
                            found_data = True
                            break
                except:
                    continue
            if not found_data and col_num > 20:
                break
        
        safe_row = min(last_data_row + 10, self.SAFE_MAX_ROWS)
        safe_col = min(last_data_col + 5, self.SAFE_MAX_COLS)
        
        return safe_row, safe_col

def process_single_ctcac_file(file_path_str: str) -> Dict:
    """Single file processor for parallel execution"""
    file_path = Path(file_path_str)
    
    try:
        start_time = time.time()
        range_detector = SmartRangeDetector()
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path, read_only=False)
        
        # Remove unnecessary sheets
        sheets_removed = []
        sheets_to_remove = []
        for sheet in wb.worksheets:
            if any(term in sheet.title.lower() for term in ["checklist items", "application checklist"]):
                sheets_to_remove.append(sheet.title)
        
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
                sheets_removed.append(sheet_name)
        
        file_result = {
            'file_name': file_path.name,
            'processing_time': None,
            'sheets_processed': 0,
            'total_cells_processed': 0,
            'sheets_removed': sheets_removed,
            'anomaly_status': 'NORMAL',
            'year': _extract_year(file_path.name),
            'process_id': os.getpid()
        }
        
        total_cells = 0
        sheets_processed = 0
        
        # Process each remaining sheet with smart range detection
        for sheet in wb.worksheets:
            max_row, max_col = range_detector.detect_smart_range(sheet, sheet.title)
            sheet_cells = max_row * max_col
            total_cells += sheet_cells
            sheets_processed += 1
            
            # Quick cell processing with datetime fix
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if hasattr(cell.value, 'isoformat'):
                        cell.value = cell.value.isoformat()
        
        # Anomaly detection
        if total_cells > 500000:
            file_result['anomaly_status'] = 'CRITICAL_ANOMALY'
            wb.close()
            return file_result
        
        # Save processed file
        output_path = file_path.parent / f"processed_{file_path.name}"
        wb.save(output_path)
        wb.close()
        
        # Finalize results
        processing_time = time.time() - start_time
        file_result.update({
            'processing_time': round(processing_time, 2),
            'sheets_processed': sheets_processed,
            'total_cells_processed': total_cells,
            'status': 'SUCCESS'
        })
        
        return file_result
        
    except Exception as e:
        return {
            'file_name': file_path.name,
            'error': str(e),
            'status': 'ERROR',
            'process_id': os.getpid()
        }

def _extract_year(filename: str) -> int:
    """Extract year from filename"""
    if '2025' in filename:
        return 2025
    elif '2024' in filename:
        return 2024
    elif '2023' in filename:
        return 2023
    else:
        return 0

class M4BeastCTCACCrusher:
    """M4 Max powered parallel CTCAC processing beast"""
    
    def __init__(self):
        self.data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
        self.max_workers = 14  # Leave 2 cores for system/monitoring
        self.max_memory_gb = 80  # Use up to 80GB, leave buffer
        self.batch_size = 50  # Process in batches to manage memory
        
        # Performance tracking
        self.start_time = None
        self.stats = {
            'files_processed': 0,
            'files_succeeded': 0,
            'files_failed': 0,
            'total_processing_time': 0,
            'total_cells_processed': 0,
            'anomalies_detected': 0,
            'sheets_removed': 0
        }
        
    def get_remaining_files(self) -> List[Path]:
        """Get list of unprocessed 2024 and 2023 files"""
        print("🔍 Scanning for remaining files to process...")
        
        all_files = []
        processed_files = set()
        
        # Find all CTCAC files
        for pattern in ["*2024*.xlsx", "*2023*.xlsx"]:
            files = list(self.data_dir.glob(pattern))
            all_files.extend(files)
        
        # Find already processed files
        for processed in self.data_dir.glob("processed_*.xlsx"):
            original_name = processed.name.replace("processed_", "")
            processed_files.add(original_name)
        
        # Filter to unprocessed files only
        remaining_files = [f for f in all_files if f.name not in processed_files]
        
        # Sort by year and number for organized processing
        remaining_files.sort(key=lambda x: (self._extract_year_from_path(x), x.name))
        
        print(f"📊 Found {len(remaining_files)} files to process:")
        years = {}
        for f in remaining_files:
            year = self._extract_year_from_path(f)
            years[year] = years.get(year, 0) + 1
        
        for year, count in sorted(years.items()):
            print(f"  {year}: {count} files")
        
        return remaining_files
    
    def _extract_year_from_path(self, path: Path) -> int:
        """Extract year from file path"""
        if '2025' in path.name:
            return 2025
        elif '2024' in path.name:
            return 2024
        elif '2023' in path.name:
            return 2023
        else:
            return 0
    
    def monitor_system_resources(self):
        """Real-time system monitoring"""
        memory = psutil.virtual_memory()
        cpu = psutil.cpu_percent(interval=0.1)
        
        memory_used_gb = memory.used / (1024**3)
        memory_percent = memory.percent
        
        return {
            'cpu_percent': cpu,
            'memory_used_gb': round(memory_used_gb, 1),
            'memory_percent': round(memory_percent, 1),
            'available_gb': round(memory.available / (1024**3), 1)
        }
    
    def process_batch_parallel(self, file_batch: List[Path]) -> List[Dict]:
        """Process a batch of files in parallel"""
        batch_size = len(file_batch)
        print(f"\n🚀 Processing batch of {batch_size} files with {self.max_workers} workers")
        
        # Convert paths to strings for multiprocessing
        file_paths = [str(f) for f in file_batch]
        
        batch_results = []
        batch_start_time = time.time()
        
        # Use ProcessPoolExecutor for robust parallel processing
        with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all files for processing
            future_to_file = {
                executor.submit(process_single_ctcac_file, file_path): file_path 
                for file_path in file_paths
            }
            
            completed_count = 0
            
            # Process completed futures as they finish
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                completed_count += 1
                
                try:
                    result = future.result()
                    batch_results.append(result)
                    
                    # Quick progress update
                    if result.get('status') == 'SUCCESS':
                        cells = result.get('total_cells_processed', 0)
                        time_taken = result.get('processing_time', 0)
                        print(f"  ✅ {completed_count}/{batch_size}: {Path(file_path).name} - {cells:,} cells in {time_taken}s")
                    else:
                        print(f"  ❌ {completed_count}/{batch_size}: {Path(file_path).name} - ERROR")
                        
                    # System monitoring every 10 files
                    if completed_count % 10 == 0:
                        resources = self.monitor_system_resources()
                        print(f"    💻 System: {resources['cpu_percent']:.1f}% CPU, {resources['memory_used_gb']}GB RAM ({resources['memory_percent']:.1f}%)")
                        
                        # Memory protection
                        if resources['memory_percent'] > 85:
                            print(f"    ⚠️  High memory usage - may reduce batch size")
                
                except Exception as e:
                    print(f"  ❌ {completed_count}/{batch_size}: {Path(file_path).name} - Exception: {e}")
                    batch_results.append({
                        'file_name': Path(file_path).name,
                        'error': str(e),
                        'status': 'ERROR'
                    })
        
        batch_time = time.time() - batch_start_time
        print(f"🏁 Batch completed in {batch_time:.1f}s ({batch_time/batch_size:.2f}s per file)")
        
        return batch_results
    
    def crush_all_files(self):
        """Main execution - crush all remaining CTCAC files"""
        print("🦁 M4 BEAST CTCAC CRUSHER - INITIALIZING")
        print("=" * 60)
        
        # System info
        system_resources = self.monitor_system_resources()
        print(f"🖥️  M4 Max: 16 cores, 128GB RAM")
        print(f"💪 Workers: {self.max_workers} parallel processes")
        print(f"💾 Available RAM: {system_resources['available_gb']}GB")
        print(f"🎯 Target: All remaining 2024/2023 CTCAC applications")
        
        # Get files to process
        remaining_files = self.get_remaining_files()
        total_files = len(remaining_files)
        
        if total_files == 0:
            print("✅ No remaining files to process!")
            return
        
        print(f"\n🚀 COMMENCING PARALLEL PROCESSING OF {total_files} FILES")
        print("=" * 60)
        
        self.start_time = time.time()
        all_results = []
        
        # Process in batches to manage memory
        for batch_start in range(0, total_files, self.batch_size):
            batch_end = min(batch_start + self.batch_size, total_files)
            batch = remaining_files[batch_start:batch_end]
            
            batch_num = (batch_start // self.batch_size) + 1
            total_batches = (total_files + self.batch_size - 1) // self.batch_size
            
            print(f"\n📦 BATCH {batch_num}/{total_batches} ({len(batch)} files)")
            print("-" * 40)
            
            batch_results = self.process_batch_parallel(batch)
            all_results.extend(batch_results)
            
            # Update stats
            self._update_stats(batch_results)
            
            # Progress report
            self._print_progress_report()
            
            # Small break between batches for system stability
            if batch_num < total_batches:
                print(f"⏸️  Brief pause for system optimization...")
                time.sleep(2)
        
        # Final report
        self._generate_final_report(all_results)
    
    def _update_stats(self, results: List[Dict]):
        """Update processing statistics"""
        for result in results:
            self.stats['files_processed'] += 1
            
            if result.get('status') == 'SUCCESS':
                self.stats['files_succeeded'] += 1
                self.stats['total_cells_processed'] += result.get('total_cells_processed', 0)
                self.stats['total_processing_time'] += result.get('processing_time', 0)
                self.stats['sheets_removed'] += len(result.get('sheets_removed', []))
                
                if result.get('anomaly_status') != 'NORMAL':
                    self.stats['anomalies_detected'] += 1
            else:
                self.stats['files_failed'] += 1
    
    def _print_progress_report(self):
        """Print current progress statistics"""
        elapsed = time.time() - self.start_time
        files_per_second = self.stats['files_processed'] / elapsed if elapsed > 0 else 0
        
        print(f"\n📊 PROGRESS REPORT:")
        print(f"  ✅ Succeeded: {self.stats['files_succeeded']}")
        print(f"  ❌ Failed: {self.stats['files_failed']}")  
        print(f"  🔢 Total Cells: {self.stats['total_cells_processed']:,}")
        print(f"  🗑️  Sheets Removed: {self.stats['sheets_removed']}")
        print(f"  ⚡ Speed: {files_per_second:.2f} files/sec")
        print(f"  ⏱️  Elapsed: {elapsed:.1f}s")
    
    def _generate_final_report(self, all_results: List[Dict]):
        """Generate comprehensive final processing report"""
        total_time = time.time() - self.start_time
        
        print(f"\n🏛️ M4 BEAST CTCAC CRUSHER - MISSION COMPLETE!")
        print("=" * 60)
        print(f"🎯 FINAL STATISTICS:")
        print(f"  📁 Files Processed: {self.stats['files_processed']}")
        print(f"  ✅ Success Rate: {(self.stats['files_succeeded']/max(self.stats['files_processed'],1)*100):.1f}%")
        print(f"  ⏱️  Total Time: {total_time:.1f}s")
        print(f"  ⚡ Average Speed: {self.stats['files_processed']/total_time:.2f} files/sec")
        print(f"  🔢 Total Cells: {self.stats['total_cells_processed']:,}")
        print(f"  🗑️  Sheets Cleaned: {self.stats['sheets_removed']}")
        print(f"  🚨 Anomalies: {self.stats['anomalies_detected']}")
        
        # Save detailed results
        results_file = f"m4_beast_ctcac_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        final_report = {
            'system_info': {
                'processor': 'M4 Max',
                'cores_used': self.max_workers,
                'total_ram_gb': 128,
                'processing_date': datetime.now().isoformat()
            },
            'processing_stats': self.stats,
            'performance_metrics': {
                'total_time_seconds': round(total_time, 2),
                'files_per_second': round(self.stats['files_processed']/total_time, 2),
                'cells_per_second': round(self.stats['total_cells_processed']/total_time, 0),
                'average_file_time': round(self.stats['total_processing_time']/max(self.stats['files_succeeded'],1), 2)
            },
            'detailed_results': all_results
        }
        
        with open(results_file, 'w') as f:
            json.dump(final_report, f, indent=2)
        
        print(f"  📄 Report Saved: {results_file}")
        
        if self.stats['files_failed'] == 0:
            print(f"\n🏆 FLAWLESS VICTORY - ALL FILES PROCESSED SUCCESSFULLY!")
        else:
            print(f"\n✅ MISSION SUCCESS - {self.stats['files_succeeded']} files ready for RAG integration")

def main():
    """Launch the M4 Beast CTCAC Crusher"""
    
    # Signal handling for clean shutdown
    def signal_handler(sig, frame):
        print('\n🛑 Graceful shutdown requested...')
        sys.exit(0)
    
    signal.signal(signal.SIGINT, signal_handler)
    
    # Initialize and run the crusher
    crusher = M4BeastCTCACCrusher()
    crusher.crush_all_files()

if __name__ == "__main__":
    main()