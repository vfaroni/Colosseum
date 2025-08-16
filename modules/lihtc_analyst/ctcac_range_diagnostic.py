#!/usr/bin/env python3
"""
CTCAC Range Diagnostic Tool
Investigate Excel range differences between 2024 vs 2025 files
Mission: CTCAC_PROCESSING_FIX_20250810
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime

class CTCACRangeDiagnostic:
    def __init__(self):
        self.data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
        self.results = {}
        
    def analyze_excel_structure(self, file_path):
        """Analyze Excel file structure and ranges"""
        print(f"🔍 Analyzing: {file_path.name}")
        
        try:
            # Load with openpyxl for detailed structure analysis
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            file_analysis = {
                'file_name': file_path.name,
                'file_size': file_path.stat().st_size,
                'sheet_count': len(wb.worksheets),
                'sheets': {}
            }
            
            total_used_cells = 0
            
            for sheet in wb.worksheets:
                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                used_range_size = max_row * max_col if max_row and max_col else 0
                
                # Count actual data cells
                actual_data_cells = 0
                if max_row and max_col and max_row < 1000 and max_col < 100:  # Safety limit
                    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                        for cell in row:
                            if cell.value is not None and str(cell.value).strip():
                                actual_data_cells += 1
                
                sheet_info = {
                    'max_row': max_row,
                    'max_column': max_col, 
                    'used_range_size': used_range_size,
                    'actual_data_cells': actual_data_cells,
                    'range_ratio': used_range_size / max(actual_data_cells, 1)
                }
                
                file_analysis['sheets'][sheet.title] = sheet_info
                total_used_cells += used_range_size
                
                print(f"  📊 {sheet.title}: {max_row}x{max_col} = {used_range_size:,} cells, {actual_data_cells:,} with data")
                
            file_analysis['total_used_range'] = total_used_cells
            wb.close()
            
            return file_analysis
            
        except Exception as e:
            print(f"❌ Error analyzing {file_path.name}: {e}")
            return None
    
    def compare_files(self, problem_file, good_file):
        """Compare problem vs good file structures"""
        print(f"\n🆚 COMPARING: {problem_file.name} vs {good_file.name}")
        
        problem_analysis = self.analyze_excel_structure(problem_file)
        good_analysis = self.analyze_excel_structure(good_file)
        
        if not problem_analysis or not good_analysis:
            return
            
        print(f"\n📊 COMPARISON RESULTS:")
        print(f"Problem File Total Used Range: {problem_analysis['total_used_range']:,} cells")
        print(f"Good File Total Used Range: {good_analysis['total_used_range']:,} cells")
        print(f"Inflation Ratio: {problem_analysis['total_used_range'] / good_analysis['total_used_range']:.1f}x")
        
        # Sheet-by-sheet comparison
        print(f"\n📋 SHEET-BY-SHEET ANALYSIS:")
        for sheet_name in problem_analysis['sheets']:
            if sheet_name in good_analysis['sheets']:
                prob = problem_analysis['sheets'][sheet_name]
                good = good_analysis['sheets'][sheet_name]
                
                range_diff = prob['used_range_size'] / max(good['used_range_size'], 1)
                print(f"  {sheet_name}:")
                print(f"    Problem: {prob['max_row']}x{prob['max_column']} = {prob['used_range_size']:,}")  
                print(f"    Good:    {good['max_row']}x{good['max_column']} = {good['used_range_size']:,}")
                print(f"    Ratio:   {range_diff:.1f}x")
                
        return problem_analysis, good_analysis
    
    def run_diagnostic(self):
        """Run complete diagnostic on problem files"""
        print("🚨 CTCAC RANGE DIAGNOSTIC - PHASE 1 ROOT CAUSE ANALYSIS")
        print("=" * 60)
        
        # Test files from mission plan
        problem_files = [
            "2024_4pct_R1_24-409.xlsx",  # Known problem file
            "2024_4pct_R1_24-408.xlsx",  # Another 2024 sample
        ]
        
        good_files = [
            "2025_4pct_R1_25-464.xlsx",  # Known good baseline
            "2025_4pct_R1_25-405.xlsx",  # Another 2025 sample  
        ]
        
        # Run comparisons
        for problem_name in problem_files:
            problem_path = self.data_dir / problem_name
            if problem_path.exists():
                for good_name in good_files:
                    good_path = self.data_dir / good_name
                    if good_path.exists():
                        analysis = self.compare_files(problem_path, good_path)
                        if analysis:
                            self.results[f"{problem_name}_vs_{good_name}"] = analysis
                        break
                print("-" * 40)
        
        # Summary diagnosis
        self.print_diagnosis_summary()
        
    def print_diagnosis_summary(self):
        """Print diagnostic summary and root cause hypothesis"""
        print(f"\n🎯 ROOT CAUSE DIAGNOSIS SUMMARY:")
        print("=" * 50)
        
        if self.results:
            for comparison, (problem, good) in self.results.items():
                ratio = problem['total_used_range'] / good['total_used_range']
                print(f"📈 {comparison}: {ratio:.1f}x inflation detected")
        
        print(f"\n💡 HYPOTHESIS:")
        print("1. 2024 Excel files have expanded 'used range' due to formatting")
        print("2. V1.7 iter_rows() processes entire used range, not just data")
        print("3. Smart range detection needed to limit to actual data boundaries")
        print("4. CTCAC forms should not exceed ~500 rows x 50 columns")
        
        print(f"\n🔧 RECOMMENDED FIX:")
        print("- Implement get_smart_data_range() function")
        print("- Limit processing to actual data boundaries")  
        print("- Add safety limits: max 500 rows, 50 columns")
        print("- Add real-time anomaly detection")

if __name__ == "__main__":
    diagnostic = CTCACRangeDiagnostic()
    diagnostic.run_diagnostic()