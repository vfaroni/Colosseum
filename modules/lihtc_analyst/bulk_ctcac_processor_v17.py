#!/usr/bin/env python3
"""
CTCAC BULK PROCESSOR V1.7 - YEAR-AWARE PRODUCTION SYSTEM
Bulk process all 804 4% CTCAC applications (2023-2025) with format awareness
Built on proven V1.6 conservative cleanup engine
"""

import os
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

class YearAwareCTCACProcessor:
    """Production bulk processor for all CTCAC 4% applications with year format awareness"""
    
    def __init__(self, max_workers: int = 4):
        self.setup_directories()
        self.setup_logging()
        self.max_workers = max_workers
        
        # Year-specific configurations
        self.year_configs = {
            2025: {
                'expected_sheets': [
                    'Application', 'Sources and Uses Budget', 'Basis & Credits',
                    'Sources and Basis Breakdown', 'Points System', 'Tie Breaker',
                    'CalHFA Addendum', 'Subsidy Contract Calculation', '15 Year Pro Forma'
                ],
                'seismic_tiebreaker': True,  # 2025 has seismic tiebreaker changes
                'format_notes': '2025 format with seismic tiebreaker modifications'
            },
            2024: {
                'expected_sheets': [
                    'Application', 'Sources and Uses Budget', 'Basis & Credits',
                    'Sources and Basis Breakdown', 'Points System', 'Tie Breaker',
                    'CalHFA Addendum', 'Subsidy Contract Calculation', '15 Year Pro Forma'
                ],
                'seismic_tiebreaker': False,  # Standard format
                'format_notes': '2024 standard format'
            },
            2023: {
                'expected_sheets': [
                    'Application', 'Sources and Uses Budget', 'Basis & Credits',
                    'Sources and Basis Breakdown', 'Points System', 'Tie Breaker',
                    'CalHFA Addendum', 'Subsidy Contract Calculation', '15 Year Pro Forma'
                ],
                'seismic_tiebreaker': False,  # Standard format
                'format_notes': '2023 standard format'
            }
        }
        
        # V1.6 Conservative cleanup patterns (proven system)
        self.preserve_patterns = [
            r'\$[\d,]+',  # Dollar amounts
            r'\d+\.?\d*%',  # Percentages
            r'\d{4}-\d{2}-\d{2}',  # Dates
            r'Section\s+\d+',  # Section references
            r'Tab\s+\d+',  # Tab references
            r'^[A-Z]\d+$',  # Cell references
            r'^\d+$',  # Pure numbers
            r'^[IVX]+\.',  # Roman numerals
        ]
        
        self.safe_removal_patterns = [
            r'^.{100,}$',  # Long text (>100 chars) - V1.6 proven pattern
        ]
        
        self.stats = {
            'total_files': 0,
            'processed_files': 0,
            'skipped_files': 0,
            'error_files': 0,
            'total_sheets': 0,
            'total_cells': 0,
            'total_cleaned': 0,
            'processing_time': 0,
            'errors': [],
            'year_stats': {2023: {}, 2024: {}, 2025: {}}
        }
    
    def setup_directories(self):
        """Setup organized directory structure for processed files"""
        self.raw_data_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Data_Sets/california/CA_LIHTC_Applications/raw_data")
        self.output_base = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/processed_ctcac_data")
        
        # Create year/round subdirectories
        for year in [2023, 2024, 2025]:
            for round_num in ['R1', 'R2', 'R3']:
                output_dir = self.output_base / str(year) / round_num
                output_dir.mkdir(parents=True, exist_ok=True)
        
        # Logs directory
        self.logs_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/bulk_processing_logs")
        self.logs_dir.mkdir(parents=True, exist_ok=True)
    
    def setup_logging(self):
        """Setup comprehensive logging system"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        log_file = self.logs_dir / f"bulk_processing_v17_{timestamp}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def parse_filename(self, filename: str) -> Dict[str, Any]:
        """Parse CTCAC filename to extract year, round, and application number"""
        patterns = [
            # 2025 R1: 2025_4pct_R1_25-XXX.xlsx
            r'2025_4pct_R1_25-(\d+)\.xlsx',
            # 2025 R2: 2025_4pct_R2_25-XXX.xlsx  
            r'2025_4pct_R2_25-(\d+)\.xlsx',
            # 2024: 2024_4pct_RX_24-XXX.xlsx
            r'2024_4pct_R([123])_24-(\d+)\.xlsx',
            # 2023: 2023_4pct_RX_23-XXX.xlsx
            r'2023_4pct_R([123])_23-(\d+)\.xlsx',
        ]
        
        for i, pattern in enumerate(patterns):
            match = re.match(pattern, filename)
            if match:
                if i == 0:  # 2025 R1
                    return {
                        'year': 2025,
                        'round': 'R1',
                        'app_num': match.group(1),
                        'full_app_id': f"25-{match.group(1)}",
                        'valid': True
                    }
                elif i == 1:  # 2025 R2
                    return {
                        'year': 2025,
                        'round': 'R2', 
                        'app_num': match.group(1),
                        'full_app_id': f"25-{match.group(1)}",
                        'valid': True
                    }
                elif i == 2:  # 2024
                    return {
                        'year': 2024,
                        'round': f"R{match.group(1)}",
                        'app_num': match.group(2),
                        'full_app_id': f"24-{match.group(2)}",
                        'valid': True
                    }
                elif i == 3:  # 2023
                    return {
                        'year': 2023,
                        'round': f"R{match.group(1)}",
                        'app_num': match.group(2),
                        'full_app_id': f"23-{match.group(2)}",
                        'valid': True
                    }
        
        return {'valid': False, 'filename': filename}
    
    def conservative_cleanup_cell(self, cell_value: Any, sheet_name: str) -> Tuple[Any, bool, str]:
        """V1.6 proven conservative cleanup logic with JSON serialization fix"""
        if cell_value is None:
            return cell_value, False, "Blank cell preserved"
        
        # Handle datetime objects for JSON serialization
        if hasattr(cell_value, 'isoformat'):  # datetime objects
            cell_value = cell_value.isoformat()
        
        text = str(cell_value).strip()
        if not text:
            return cell_value, False, "Empty cell preserved"
        
        # Always preserve certain patterns (V1.6 proven)
        for pattern in self.preserve_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return cell_value, False, f"Preserved due to pattern: {pattern}"
        
        # Check for safe removal patterns (V1.6 conservative threshold)
        for pattern in self.safe_removal_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                # Double-check for variable data
                if any(preserve_pattern in text.lower() for preserve_pattern in ['$', '%', '20', 'section', 'tab']):
                    return cell_value, False, "Contains variable data - preserved"
                return None, True, f"Safe removal: >100 char instructional text"
        
        return cell_value, False, "Conservative default - preserved"
    
    def process_single_file(self, file_path: Path) -> Dict[str, Any]:
        """Process single CTCAC application using V1.6 proven methods"""
        start_time = time.time()
        filename = file_path.name
        
        try:
            # Parse filename for metadata
            file_info = self.parse_filename(filename)
            if not file_info['valid']:
                return {
                    'success': False,
                    'error': f"Invalid filename format: {filename}",
                    'filename': filename
                }
            
            year = file_info['year']
            round_type = file_info['round']
            app_id = file_info['full_app_id']
            
            self.logger.info(f"📊 Processing {app_id} ({year} {round_type}): {filename}")
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Get year configuration
            config = self.year_configs.get(year, self.year_configs[2025])
            expected_sheets = config['expected_sheets']
            
            # Process each sheet
            extraction_data = {}
            cleanup_stats = {'removed': 0, 'preserved': 0}
            sheets_processed = 0
            
            for sheet_name in wb.sheetnames:
                if sheet_name in expected_sheets:
                    sheet = wb[sheet_name]
                    sheet_data = []
                    
                    # Extract all cells with conservative cleanup
                    for row in sheet.iter_rows():
                        row_data = []
                        for cell in row:
                            # V1.6 conservative cleanup
                            cleaned_value, was_removed, reason = self.conservative_cleanup_cell(
                                cell.value, sheet_name
                            )
                            
                            if was_removed:
                                cleanup_stats['removed'] += 1
                            else:
                                cleanup_stats['preserved'] += 1
                            
                            row_data.append({
                                'value': cleaned_value,
                                'coordinate': cell.coordinate,
                                'removed': was_removed,
                                'cleanup_reason': reason
                            })
                        sheet_data.append(row_data)
                    
                    extraction_data[sheet_name] = sheet_data
                    sheets_processed += 1
            
            # Create output filename and path
            output_dir = self.output_base / str(year) / round_type
            output_filename = f"{app_id}_processed_v17.json"
            output_path = output_dir / output_filename
            
            # Save processed data
            processing_report = {
                'filename': filename,
                'app_id': app_id,
                'year': year,
                'round': round_type,
                'processing_timestamp': datetime.now().isoformat(),
                'sheets_processed': sheets_processed,
                'cleanup_stats': cleanup_stats,
                'preservation_rate': f"{(cleanup_stats['preserved'] / (cleanup_stats['preserved'] + cleanup_stats['removed']) * 100):.1f}%",
                'format_config': config['format_notes'],
                'extraction_data': extraction_data
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(processing_report, f, indent=2, ensure_ascii=False)
            
            processing_time = time.time() - start_time
            
            return {
                'success': True,
                'filename': filename,
                'app_id': app_id,
                'year': year,
                'round': round_type,
                'sheets_processed': sheets_processed,
                'cleanup_stats': cleanup_stats,
                'processing_time': processing_time,
                'output_path': str(output_path)
            }
        
        except Exception as e:
            error_msg = f"Error processing {filename}: {str(e)}"
            self.logger.error(error_msg)
            return {
                'success': False,
                'filename': filename,
                'error': error_msg,
                'processing_time': time.time() - start_time
            }
    
    def get_file_inventory(self) -> Dict[str, List[Path]]:
        """Get organized inventory of all CTCAC files by year"""
        inventory = {2023: [], 2024: [], 2025: []}
        
        for file_path in self.raw_data_dir.glob("*.xlsx"):
            if file_path.name.startswith(('.', '~')):  # Skip system files
                continue
                
            file_info = self.parse_filename(file_path.name)
            if file_info['valid']:
                year = file_info['year']
                inventory[year].append(file_path)
        
        return inventory
    
    def process_year_batch(self, year: int, file_list: List[Path]) -> Dict[str, Any]:
        """Process all files for a specific year"""
        self.logger.info(f"🚀 Processing {year} applications: {len(file_list)} files")
        
        year_start = time.time()
        year_stats = {
            'files_attempted': len(file_list),
            'files_successful': 0,
            'files_failed': 0,
            'total_sheets': 0,
            'total_cells_preserved': 0,
            'total_cells_removed': 0,
            'errors': []
        }
        
        # Process files with thread pool
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_file = {
                executor.submit(self.process_single_file, file_path): file_path 
                for file_path in file_list
            }
            
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    
                    if result['success']:
                        year_stats['files_successful'] += 1
                        year_stats['total_sheets'] += result['sheets_processed']
                        year_stats['total_cells_preserved'] += result['cleanup_stats']['preserved']
                        year_stats['total_cells_removed'] += result['cleanup_stats']['removed']
                        
                        self.logger.info(f"✅ {result['app_id']}: {result['sheets_processed']} sheets, "
                                       f"{result['cleanup_stats']['preserved']} preserved, "
                                       f"{result['cleanup_stats']['removed']} removed")
                    else:
                        year_stats['files_failed'] += 1
                        year_stats['errors'].append(result.get('error', 'Unknown error'))
                        self.logger.error(f"❌ Failed: {file_path.name}")
                
                except Exception as e:
                    year_stats['files_failed'] += 1
                    error_msg = f"Thread execution error for {file_path.name}: {str(e)}"
                    year_stats['errors'].append(error_msg)
                    self.logger.error(error_msg)
        
        year_processing_time = time.time() - year_start
        year_stats['processing_time'] = year_processing_time
        
        # Calculate preservation rate
        total_cells = year_stats['total_cells_preserved'] + year_stats['total_cells_removed']
        preservation_rate = (year_stats['total_cells_preserved'] / total_cells * 100) if total_cells > 0 else 0
        year_stats['preservation_rate'] = f"{preservation_rate:.1f}%"
        
        self.logger.info(f"🏁 {year} Complete: {year_stats['files_successful']}/{year_stats['files_attempted']} files, "
                        f"{preservation_rate:.1f}% preservation, {year_processing_time:.1f}s")
        
        return year_stats
    
    def run_bulk_processing(self):
        """Main bulk processing orchestrator"""
        self.logger.info("🏛️ CTCAC BULK PROCESSOR V1.7 - PRODUCTION RUN")
        self.logger.info("=" * 80)
        
        overall_start = time.time()
        
        # Get file inventory
        inventory = self.get_file_inventory()
        total_files = sum(len(files) for files in inventory.values())
        
        self.logger.info(f"📊 INVENTORY SUMMARY:")
        self.logger.info(f"   2023: {len(inventory[2023])} files")
        self.logger.info(f"   2024: {len(inventory[2024])} files") 
        self.logger.info(f"   2025: {len(inventory[2025])} files")
        self.logger.info(f"   TOTAL: {total_files} files")
        self.logger.info("=" * 80)
        
        # Process in order: 2025 → 2024 → 2023 (newest first)
        processing_order = [2025, 2024, 2023]
        
        for year in processing_order:
            if inventory[year]:
                year_stats = self.process_year_batch(year, inventory[year])
                self.stats['year_stats'][year] = year_stats
                
                # Update overall stats
                self.stats['processed_files'] += year_stats['files_successful']
                self.stats['error_files'] += year_stats['files_failed']
                self.stats['total_sheets'] += year_stats['total_sheets']
                self.stats['total_cells'] += year_stats['total_cells_preserved'] + year_stats['total_cells_removed']
                self.stats['total_cleaned'] += year_stats['total_cells_removed']
                self.stats['errors'].extend(year_stats['errors'])
        
        overall_time = time.time() - overall_start
        self.stats['total_files'] = total_files
        self.stats['processing_time'] = overall_time
        
        # Generate final report
        self.generate_final_report()
    
    def generate_final_report(self):
        """Generate comprehensive final processing report"""
        self.logger.info("=" * 80)
        self.logger.info("🏁 BULK PROCESSING COMPLETE!")
        self.logger.info("=" * 80)
        
        # Overall statistics
        total_preserved = sum(year['total_cells_preserved'] for year in self.stats['year_stats'].values())
        overall_preservation = (total_preserved / self.stats['total_cells'] * 100) if self.stats['total_cells'] > 0 else 0
        
        self.logger.info(f"📊 OVERALL RESULTS:")
        self.logger.info(f"   Files Processed: {self.stats['processed_files']}/{self.stats['total_files']}")
        self.logger.info(f"   Success Rate: {(self.stats['processed_files']/self.stats['total_files']*100):.1f}%")
        self.logger.info(f"   Total Sheets: {self.stats['total_sheets']}")
        self.logger.info(f"   Total Cells: {self.stats['total_cells']:,}")
        self.logger.info(f"   Cells Cleaned: {self.stats['total_cleaned']:,}")
        self.logger.info(f"   Preservation Rate: {overall_preservation:.1f}%")
        self.logger.info(f"   Processing Time: {self.stats['processing_time']:.1f} seconds")
        
        # Year-by-year breakdown
        self.logger.info(f"\n📅 YEAR-BY-YEAR BREAKDOWN:")
        for year in [2025, 2024, 2023]:
            if year in self.stats['year_stats'] and self.stats['year_stats'][year]:
                stats = self.stats['year_stats'][year]
                self.logger.info(f"   {year}: {stats['files_successful']}/{stats['files_attempted']} files, "
                               f"{stats['preservation_rate']} preservation, "
                               f"{stats['processing_time']:.1f}s")
        
        # Save detailed report
        report_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_file = self.logs_dir / f"bulk_processing_report_v17_{report_timestamp}.json"
        
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(self.stats, f, indent=2, ensure_ascii=False, default=str)
        
        self.logger.info(f"\n📋 Detailed report saved: {report_file}")
        
        if self.stats['errors']:
            self.logger.info(f"\n⚠️  Errors encountered: {len(self.stats['errors'])}")
            for error in self.stats['errors'][:5]:  # Show first 5 errors
                self.logger.info(f"   {error}")
            if len(self.stats['errors']) > 5:
                self.logger.info(f"   ... and {len(self.stats['errors']) - 5} more errors")


def main():
    """Run the V1.7 bulk processor"""
    print("🏛️ CTCAC BULK PROCESSOR V1.7")
    print("Year-Aware Production System")
    print("Processing all 804 4% applications (2023-2025)")
    print("Built on proven V1.6 conservative cleanup engine")
    print("=" * 60)
    
    processor = YearAwareCTCACProcessor(max_workers=4)
    processor.run_bulk_processing()
    
    print("\n" + "=" * 60)
    print("🎯 V1.7 BULK PROCESSING COMPLETE!")
    print("All processed files saved with year/round organization")
    print("Ready for JSON → CTCAC RAG integration")
    print("=" * 60)


if __name__ == "__main__":
    main()