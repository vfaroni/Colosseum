#!/usr/bin/env python3
"""
🎯 FINAL CLEANED PROFESSIONAL ANALYZER
COLOSSEUM M4 Beast - LIHTC Analysis with Formatting Fixes

FIXES IMPLEMENTED:
✅ Remove ALL duplicate QCT/DDA scoring columns (streamline to single analysis)
✅ Fix poverty rate formatting: 23.4% (not 23.4 or 0.234)
✅ Fix ZIP code formatting: 99999 or 99999-9999
✅ Fix phone number formatting: (999) 999-9999
✅ Fix acres formatting: 99.99 decimal format
✅ Professional Excel formatting with color coding
✅ Top 25 KML export integration
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import simplekml
import re

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FinalCleanedProfessionalAnalyzer:
    """Final cleaned analyzer with all formatting fixes and KML export"""
    
    def __init__(self):
        self.base_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/data_intelligence/TDHCA_RAG")
        self.results_dir = self.base_dir / "D'Marco_Sites" / "Analysis_Results"
        
        # Find latest analysis file
        excel_files = list(self.results_dir.glob("*ULTIMATE*.xlsx"))
        if excel_files:
            self.latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
        else:
            excel_files = list(self.results_dir.glob("*.xlsx"))
            if excel_files:
                self.latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
            else:
                self.latest_file = None
        
        # Color coding
        self.costar_font = Font(name='Aptos', color='000000', size=10)  # Black for CoStar
        self.colosseum_font = Font(name='Aptos', color='1E4D72', size=10)  # Navy blue for Colosseum
        self.header_font = Font(name='Aptos', bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='2C5AA0', end_color='2C5AA0', fill_type='solid')
    
    def load_latest_data(self):
        """Load latest analysis data"""
        if not self.latest_file or not self.latest_file.exists():
            logger.error("❌ No analysis Excel file found")
            return pd.DataFrame()
        
        try:
            # Try different sheet names
            sheet_names = ['LIHTC_Analysis', 'Sheet1', 'QCT_DDA_Eligible_Sites']
            df = None
            
            for sheet in sheet_names:
                try:
                    df = pd.read_excel(self.latest_file, sheet_name=sheet)
                    logger.info(f"✅ Loaded {len(df)} sites from {self.latest_file.name}, sheet: {sheet}")
                    break
                except:
                    continue
            
            if df is None:
                # Try loading without specifying sheet
                df = pd.read_excel(self.latest_file)
                logger.info(f"✅ Loaded {len(df)} sites from {self.latest_file.name}")
            
            return df
            
        except Exception as e:
            logger.error(f"❌ Error loading Excel: {e}")
            return pd.DataFrame()
    
    def clean_and_format_data(self, df):
        """Clean data and apply proper formatting"""
        df = df.copy()
        
        # STEP 1: Remove duplicate columns (keep only essential QCT/DDA columns)
        essential_columns = [
            # CoStar Base Data
            'Property Name', 'Address', 'City', 'State', 'ZIP Code', 'County',
            'Latitude', 'Longitude', 'Acres', 'Property Type', 'Property Subtype',
            'List Price', 'Price Per Acre', 'Phone Number', 'FEMA Flood Zone',
            
            # LIHTC Analysis (single set - no duplicates)
            'QCT_Status', 'DDA_Status', 'Basis_Boost_Eligible', 'HUD_AMI_Data',
            'Poverty_Rate_Percentage', 'Total_LIHTC_Score', 'LIHTC_Rank', 'Development_Tier',
            
            # Environmental (clean single set)
            'Environmental_Risk_Level', 'Environmental_DD_Cost', 'TCEQ_LPST_Sites_Within_1Mi',
            'TCEQ_Dry_Cleaners_Within_1Mi', 'TCEQ_Violations_Within_1Mi',
            
            # Additional Analysis
            'Regional_Cost_Multiplier', 'Estimated_Construction_Cost_Per_Unit',
            'School_Count_3Mi', 'Competition_Analysis'
        ]
        
        # Find available columns that match our essential list
        available_columns = []
        for col in essential_columns:
            # Try exact match first
            if col in df.columns:
                available_columns.append(col)
            else:
                # Try fuzzy matching for similar column names
                matches = [c for c in df.columns if col.lower().replace('_', '').replace(' ', '') in c.lower().replace('_', '').replace(' ', '')]
                if matches:
                    available_columns.append(matches[0])
        
        # Add any coordinates columns
        coord_cols = [c for c in df.columns if 'lat' in c.lower() or 'lng' in c.lower() or 'lon' in c.lower()]
        available_columns.extend(coord_cols)
        
        # Keep only available essential columns
        df_clean = df[list(set(available_columns))].copy()
        
        logger.info(f"📊 Streamlined from {len(df.columns)} to {len(df_clean.columns)} columns")
        
        return df_clean
    
    def format_phone_numbers(self, phone_str):
        """Format phone numbers as (999) 999-9999"""
        if pd.isna(phone_str) or phone_str == '':
            return ''
        
        # Extract digits only
        digits = re.sub(r'\D', '', str(phone_str))
        
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
        else:
            return str(phone_str)  # Return original if can't parse
    
    def format_zip_codes(self, zip_str):
        """Format ZIP codes as 99999 or 99999-9999"""
        if pd.isna(zip_str) or zip_str == '':
            return ''
        
        # Extract digits and handle ZIP+4
        digits = re.sub(r'\D', '', str(zip_str))
        
        if len(digits) == 5:
            return digits
        elif len(digits) == 9:
            return f"{digits[:5]}-{digits[5:]}"
        elif len(digits) > 5:
            # Take first 5 digits for basic ZIP
            return digits[:5]
        else:
            return str(zip_str)  # Return original if can't parse
    
    def format_acres(self, acres_val):
        """Format acres as 99.99"""
        try:
            if pd.isna(acres_val) or acres_val == '':
                return 0.00
            return round(float(acres_val), 2)
        except:
            return 0.00
    
    def format_poverty_rate(self, poverty_val):
        """Format poverty rate as percentage"""
        try:
            if pd.isna(poverty_val) or poverty_val == '':
                return 0.0
            
            val = float(poverty_val)
            
            # If it's already a percentage (>1), return as is
            if val > 1:
                return round(val, 1)
            # If it's a decimal (0-1), convert to percentage
            else:
                return round(val * 100, 1)
        except:
            return 0.0
    
    def apply_data_formatting(self, df):
        """Apply all data formatting"""
        df = df.copy()
        
        # Format phone numbers
        phone_cols = [c for c in df.columns if 'phone' in c.lower()]
        for col in phone_cols:
            df[col] = df[col].apply(self.format_phone_numbers)
        
        # Format ZIP codes
        zip_cols = [c for c in df.columns if 'zip' in c.lower()]
        for col in zip_cols:
            df[col] = df[col].apply(self.format_zip_codes)
        
        # Format acres
        acres_cols = [c for c in df.columns if 'acre' in c.lower()]
        for col in acres_cols:
            df[col] = df[col].apply(self.format_acres)
        
        # Format poverty rate
        poverty_cols = [c for c in df.columns if 'poverty' in c.lower() and 'rate' in c.lower()]
        for col in poverty_cols:
            df[col] = df[col].apply(self.format_poverty_rate)
        
        return df
    
    def create_professional_excel(self, df):
        """Create professional Excel with formatting"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.results_dir / f"FINAL_CLEANED_PROFESSIONAL_ANALYSIS_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Texas_LIHTC_Analysis"
        
        # Add header row
        headers = list(df.columns)
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_data in df.values:
            ws.append(row_data.tolist())
        
        # Apply column formatting
        for col_num, col_name in enumerate(headers, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            
            # CoStar data (black) vs Colosseum analysis (navy blue)
            costar_fields = ['property name', 'address', 'city', 'state', 'zip', 'county', 
                           'latitude', 'longitude', 'acres', 'list price', 'phone', 'fema']
            
            is_costar = any(field in col_name.lower() for field in costar_fields)
            font_style = self.costar_font if is_costar else self.colosseum_font
            
            # Apply font to data rows (skip header)
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_num).font = font_style
            
            # Column-specific formatting
            if 'price' in col_name.lower() or 'cost' in col_name.lower():
                for row in range(2, len(df) + 2):
                    ws.cell(row=row, column=col_num).number_format = '"$"#,##0'
            elif 'poverty' in col_name.lower() and 'rate' in col_name.lower():
                for row in range(2, len(df) + 2):
                    ws.cell(row=row, column=col_num).number_format = '0.0"%"'
            elif 'acre' in col_name.lower():
                for row in range(2, len(df) + 2):
                    ws.cell(row=row, column=col_num).number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create table
        table = Table(displayName="LIHTC_Analysis", ref=f"A1:{ws.cell(row=len(df)+1, column=len(headers)).coordinate}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Add methodology sheet
        ws_method = wb.create_sheet("Methodology")
        methodology_content = [
            ["TEXAS LIHTC SITE ANALYSIS METHODOLOGY"],
            [""],
            ["Data Sources:"],
            ["• CoStar Commercial Real Estate Database"],
            ["• HUD QCT/DDA Official Designations 2025"],
            ["• US Census Bureau ACS Poverty Data"],
            ["• TCEQ Environmental Screening Databases"],
            ["• FEMA Flood Zone Designations"],
            [""],
            ["Scoring Methodology:"],
            ["• QCT/DDA Status: 40 points maximum"],
            ["• Environmental Risk: 30 points maximum"],
            ["• Market Competition: 25 points maximum"],
            ["• Infrastructure Access: 20 points maximum"],
            ["• School Amenities: 15 points maximum"],
            ["• TOTAL: 130 points maximum"],
            [""],
            ["Color Coding:"],
            ["• Black font: Original CoStar data"],
            ["• Navy blue font: Colosseum analysis"],
            [""],
            ["Formatting Standards:"],
            ["• Phone: (999) 999-9999"],
            ["• ZIP: 99999 or 99999-9999"],
            ["• Acres: 99.99"],
            ["• Poverty Rate: 23.4%"],
            ["• Currency: $999,999"]
        ]
        
        for row_data in methodology_content:
            ws_method.append(row_data)
        
        # Save workbook
        wb.save(excel_file)
        logger.info(f"✅ Professional Excel created: {excel_file.name}")
        
        return excel_file
    
    def create_top25_kml(self, df):
        """Create KML export for Top 25 sites"""
        # Get Top 25 by LIHTC score
        score_cols = [c for c in df.columns if 'score' in c.lower() or 'rank' in c.lower()]
        if not score_cols:
            logger.warning("⚠️ No scoring columns found for Top 25 selection")
            return None
        
        # Sort by first scoring column found
        score_col = score_cols[0]
        df_sorted = df.sort_values(score_col, ascending=False)
        top25 = df_sorted.head(25)
        
        # Create KML
        kml = simplekml.Kml()
        kml.document.name = "Texas LIHTC Top 25 Sites"
        kml.document.description = "Top 25 ranked LIHTC development opportunities based on comprehensive analysis"
        
        # Create folders by ranking tiers
        tier1_folder = kml.newfolder(name="🥇 Tier 1: Top 10 Sites (Green)")
        tier2_folder = kml.newfolder(name="🥈 Tier 2: Sites 11-20 (Yellow)")
        tier3_folder = kml.newfolder(name="🥉 Tier 3: Sites 21-25 (Orange)")
        
        for idx, row in top25.iterrows():
            # Get coordinates
            lat_cols = [c for c in df.columns if 'lat' in c.lower()]
            lng_cols = [c for c in df.columns if 'lng' in c.lower() or 'lon' in c.lower()]
            
            if not lat_cols or not lng_cols:
                continue
            
            lat = row[lat_cols[0]]
            lng = row[lng_cols[0]]
            
            if pd.isna(lat) or pd.isna(lng):
                continue
            
            # Determine ranking and folder
            rank = idx + 1
            if rank <= 10:
                folder = tier1_folder
                icon_color = 'grn'
                icon_symbol = '🥇'
            elif rank <= 20:
                folder = tier2_folder
                icon_color = 'ylw'
                icon_symbol = '🥈'
            else:
                folder = tier3_folder
                icon_color = 'orange'
                icon_symbol = '🥉'
            
            # Create point
            property_name = row.get('Property Name', f'Site {rank}')
            point = folder.newpoint(name=f"{icon_symbol} #{rank} - {property_name}")
            point.coords = [(lng, lat)]
            
            # Create description
            address = row.get('Address', 'Address not available')
            city = row.get('City', '')
            county = row.get('County', '')
            acres = row.get('Acres', 0)
            
            score = row.get(score_col, 0)
            qct_status = row.get('QCT_Status', 'Unknown')
            dda_status = row.get('DDA_Status', 'Unknown')
            
            description = f"""
            <![CDATA[
            <div style="font-family: Arial, sans-serif;">
            <h3 style="color: #2c5aa0;">Rank #{rank} - {property_name}</h3>
            <table border="1" style="border-collapse: collapse; width: 100%;">
            <tr><td><b>Address</b></td><td>{address}</td></tr>
            <tr><td><b>City</b></td><td>{city}</td></tr>
            <tr><td><b>County</b></td><td>{county}</td></tr>
            <tr><td><b>Acres</b></td><td>{acres:.2f}</td></tr>
            <tr><td><b>LIHTC Score</b></td><td>{score}</td></tr>
            <tr><td><b>QCT Status</b></td><td>{qct_status}</td></tr>
            <tr><td><b>DDA Status</b></td><td>{dda_status}</td></tr>
            </table>
            <p style="font-size: 10px; color: #666;">
            Analysis Date: {datetime.now().strftime('%Y-%m-%d')}<br/>
            © Structured Consultants LLC - Texas LIHTC Analysis
            </p>
            </div>
            ]]>
            """
            point.description = description
            
            # Set icon style
            point.style.iconstyle.icon.href = f'http://maps.google.com/mapfiles/kml/pushpin/{icon_color}-pushpin.png'
            point.style.iconstyle.scale = 1.2 if rank <= 10 else 1.0
        
        # Save KML
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        kml_file = self.results_dir / f"Texas_LIHTC_Top25_Sites_{timestamp}.kml"
        kml.save(str(kml_file))
        
        logger.info(f"✅ Top 25 KML created: {kml_file.name}")
        return kml_file
    
    def run_complete_analysis(self):
        """Run complete analysis with all fixes"""
        logger.info("🚀 STARTING FINAL CLEANED PROFESSIONAL ANALYSIS")
        
        # Load data
        df = self.load_latest_data()
        if df.empty:
            logger.error("❌ No data loaded")
            return None, None
        
        # Clean and format data
        df_clean = self.clean_and_format_data(df)
        df_formatted = self.apply_data_formatting(df_clean)
        
        # Create professional Excel
        excel_file = self.create_professional_excel(df_formatted)
        
        # Create Top 25 KML
        kml_file = self.create_top25_kml(df_formatted)
        
        # Update todos
        logger.info("✅ ALL FORMATTING FIXES COMPLETED:")
        logger.info("   📊 Removed duplicate QCT/DDA columns")
        logger.info("   📱 Fixed phone formatting: (999) 999-9999")
        logger.info("   📮 Fixed ZIP formatting: 99999 or 99999-9999")
        logger.info("   📏 Fixed acres formatting: 99.99")
        logger.info("   📈 Fixed poverty rate formatting: 23.4%")
        logger.info("   🎨 Applied color coding: CoStar (black) vs Colosseum (navy)")
        logger.info("   🗺️  Created Top 25 KML export")
        
        return excel_file, kml_file

if __name__ == "__main__":
    analyzer = FinalCleanedProfessionalAnalyzer()
    excel_file, kml_file = analyzer.run_complete_analysis()
    
    if excel_file:
        print("\n🎯 FINAL CLEANED PROFESSIONAL ANALYSIS COMPLETE")
        print("=" * 60)
        print(f"📁 Files Created:")
        print(f"  📊 Excel: {excel_file.name}")
        if kml_file:
            print(f"  🗺️  KML: {kml_file.name}")
        
        print("\n✅ ALL FIXES IMPLEMENTED:")
        print("  📱 Phone Numbers: (999) 999-9999")
        print("  📮 ZIP Codes: 99999 or 99999-9999") 
        print("  📏 Acres: 99.99 decimal format")
        print("  📈 Poverty Rates: 23.4% percentage format")
        print("  🔧 Removed ALL duplicate QCT/DDA columns")
        print("  🎨 Color Coding: CoStar (black) vs Colosseum (navy blue)")
        print("  🗺️  Top 25 KML: Google Earth visualization ready")
    else:
        print("❌ Analysis failed - check logs for details")