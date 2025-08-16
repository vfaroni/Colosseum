#!/usr/bin/env python3
"""
🎯 ENHANCED COLOSSEUM ANALYZER WITH CENSUS COUNTY LOOKUP
Final comprehensive LIHTC analysis with high-accuracy county data

ENHANCEMENTS IMPLEMENTED:
✅ US Census API county lookup (100% free, high accuracy)
✅ TDHCA Regions integration from Colosseum directory
✅ Professional color coding (blue headers, CoStar=black, Colosseum=navy)
✅ Proper ranking hierarchy (flood zones bottom, <240 units middle, 240+ units top)
✅ Remove environmental columns (technically inaccurate for large parcels)
✅ Formatting fixes: phone (999) 999-9999, ZIP 99999/99999-9999, acres 99.99, poverty 23.4%
✅ Preserve all 5 professional worksheets
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

# Import our Census county lookup
from census_county_lookup import CensusCountyLookup

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EnhancedColosseumAnalyzerWithCensus:
    """Enhanced analyzer with Census API county lookup and TDHCA regions"""
    
    def __init__(self):
        self.base_dir = Path("/Users/williamrice/Library/CloudStorage/Dropbox-HERR/Bill Rice/Colosseum/modules/data_intelligence/TDHCA_RAG")
        self.results_dir = self.base_dir / "D'Marco_Sites" / "Analysis_Results"
        
        # Initialize Census county lookup
        self.county_lookup = CensusCountyLookup()
        
        # Load TDHCA Regions data
        self.regions_file = self.base_dir / "TDHCA_Regions" / "TDHCA_Regions.xlsx"
        self.regions_df = self.load_tdhca_regions()
        
        # Target source file (surgically cleaned version)
        self.source_file = self.results_dir / "SURGICALLY_CLEANED_ANALYSIS_20250801_222520.xlsx"
        
        # Professional color coding
        self.costar_font = Font(name='Aptos', color='000000', size=10)  # Black for CoStar
        self.colosseum_font = Font(name='Aptos', color='1E4D72', size=10)  # Navy blue for Colosseum
        self.header_font = Font(name='Aptos', bold=True, color='FFFFFF', size=11)
        self.header_fill = PatternFill(start_color='2C5AA0', end_color='2C5AA0', fill_type='solid')
        
        # Environmental columns to remove (technically inaccurate for large parcels)
        self.environmental_columns = [
            'Environmental_Risk_Level', 'Environmental_Risk_Score', 'Environmental_DD_Cost',
            'LPST_Sites_Within_Mile', 'Dry_Cleaners_Within_Mile', 'Environmental_Details',
            'Flood_Risk_Level', 'Competition_Risk_Level', 'Flood Risk'
        ]
    
    def load_tdhca_regions(self):
        """Load TDHCA regions data for county mapping"""
        try:
            regions_df = pd.read_excel(self.regions_file)
            # Clean column names (remove trailing spaces)
            regions_df.columns = regions_df.columns.str.strip()
            logger.info(f"✅ Loaded TDHCA regions: {len(regions_df)} counties across {regions_df['Region'].nunique()} regions")
            return regions_df
        except Exception as e:
            logger.error(f"❌ Error loading TDHCA regions: {e}")
            return pd.DataFrame()
    
    def format_phone_numbers(self, phone_str):
        """Format phone numbers as (999) 999-9999"""
        if pd.isna(phone_str) or phone_str == '' or phone_str == 'Not Available':
            return 'Not Available'
        
        digits = re.sub(r'\D', '', str(phone_str))
        
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
        else:
            return str(phone_str)
    
    def format_zip_codes(self, zip_str):
        """Format ZIP codes as 99999 or 99999-9999"""
        if pd.isna(zip_str) or zip_str == '':
            return ''
        
        digits = re.sub(r'\D', '', str(zip_str))
        
        if len(digits) == 5:
            return digits
        elif len(digits) == 9:
            return f"{digits[:5]}-{digits[5:]}"
        elif len(digits) > 5:
            return digits[:5]
        else:
            return str(zip_str)
    
    def format_acres(self, acres_val):
        """Format acres as 99.99"""
        try:
            if pd.isna(acres_val) or acres_val == '':
                return 0.00
            return round(float(acres_val), 2)
        except:
            return 0.00
    
    def format_poverty_rate(self, poverty_val):
        """Format poverty rate as percentage (23.4%)"""
        try:
            if pd.isna(poverty_val) or poverty_val == '':
                return 0.0
            
            val = float(poverty_val)
            
            if val > 1:
                return round(val, 1)
            else:
                return round(val * 100, 1)
        except:
            return 0.0
    
    def add_census_counties_and_regions(self, df):
        """Add county data via Census API and map to TDHCA regions"""
        logger.info("🗺️ ADDING CENSUS COUNTIES AND TDHCA REGIONS")
        
        # Find coordinate columns
        lat_cols = [c for c in df.columns if 'lat' in c.lower()]
        lng_cols = [c for c in df.columns if 'lng' in c.lower() or 'lon' in c.lower()]
        
        if not lat_cols or not lng_cols:
            logger.error("❌ No coordinate columns found for county lookup")
            df['County'] = 'No Coordinates'
            df['TDHCA_Region'] = 'Unknown'
            return df
        
        lat_col = lat_cols[0]
        lng_col = lng_cols[0]
        
        logger.info(f"📍 Using coordinates: {lat_col}, {lng_col}")
        
        # Perform batch county lookup using Census API
        df = self.county_lookup.batch_county_lookup(df, lat_col, lng_col)
        
        # Map counties to TDHCA regions
        if not self.regions_df.empty:
            # Create mapping dictionary - handle both "County" and "County Name County" formats
            region_map = {}
            for _, row in self.regions_df.iterrows():
                county_name = row['County'].strip()
                region = row['Region'].strip()
                
                # Add multiple mapping formats to handle variations
                region_map[county_name] = region
                
                # If it doesn't already end with "County", add that version too
                if not county_name.lower().endswith('county'):
                    region_map[f"{county_name} County"] = region
                
                # Also add version without "County" if it has it
                if county_name.lower().endswith('county'):
                    base_name = county_name.replace(' County', '').replace(' county', '')
                    region_map[base_name] = region
            
            # Map TDHCA regions
            df['TDHCA_Region'] = df['County'].map(region_map).fillna('Unknown Region')
            
            # Log mapping results
            mapped_count = len(df[df['TDHCA_Region'] != 'Unknown Region'])
            logger.info(f"✅ TDHCA regions mapped: {mapped_count}/{len(df)} sites")
            
            region_summary = df['TDHCA_Region'].value_counts()
            logger.info(f"📊 Region distribution: {region_summary.head().to_dict()}")
            
        else:
            df['TDHCA_Region'] = 'No Region Data'
        
        return df
    
    def calculate_ranking_hierarchy(self, df):
        """Implement proper ranking hierarchy: flood zones bottom, <240 units middle, 240+ top"""
        df = df.copy()
        
        # Find relevant columns
        unit_cols = [c for c in df.columns if 'max_units' in c.lower() or 'development_units' in c.lower()]
        flood_cols = [c for c in df.columns if 'fema' in c.lower() and 'flood' in c.lower()]
        score_cols = [c for c in df.columns if 'total' in c.lower() and 'score' in c.lower()]
        
        # Use first available columns
        unit_col = unit_cols[0] if unit_cols else None
        flood_col = flood_cols[0] if flood_cols else None
        score_col = score_cols[0] if score_cols else None
        
        logger.info(f"🎯 Ranking using: Units={unit_col}, Flood={flood_col}, Score={score_col}")
        
        # Calculate ranking tiers
        df['Ranking_Tier'] = 'TIER_3_PREMIUM'  # Default
        df['Ranking_Score'] = 0.0
        df['Unit_Capacity'] = 0.0
        
        if unit_col and unit_col in df.columns:
            # Calculate unit capacity
            df['Unit_Capacity'] = pd.to_numeric(df[unit_col], errors='coerce').fillna(0)
            
            # Tier assignment based on unit capacity and flood risk
            for idx, row in df.iterrows():
                units = row['Unit_Capacity']
                is_flood_zone = False
                
                # Check flood zone risk
                if flood_col and flood_col in df.columns:
                    flood_value = str(row[flood_col]).upper()
                    is_flood_zone = any(zone in flood_value for zone in ['A', 'AE', 'AO', 'AH', 'V', 'VE'])
                
                # Assign tiers based on hierarchy
                if is_flood_zone:
                    df.at[idx, 'Ranking_Tier'] = 'TIER_5_FLOOD_RISK'
                    df.at[idx, 'Ranking_Score'] = 10.0  # Lowest priority
                elif units < 240:
                    df.at[idx, 'Ranking_Tier'] = 'TIER_4_UNDERSIZED'
                    df.at[idx, 'Ranking_Score'] = 20.0 + min(units / 240 * 30, 30)  # 20-50 points
                elif units >= 400:
                    df.at[idx, 'Ranking_Tier'] = 'TIER_1_EXCELLENT'
                    df.at[idx, 'Ranking_Score'] = 90.0 + min((units - 400) / 200 * 10, 10)  # 90-100 points
                elif units >= 300:
                    df.at[idx, 'Ranking_Tier'] = 'TIER_2_STRONG'
                    df.at[idx, 'Ranking_Score'] = 70.0 + (units - 300) / 100 * 20  # 70-90 points
                else:  # 240-299 units
                    df.at[idx, 'Ranking_Tier'] = 'TIER_3_VIABLE'
                    df.at[idx, 'Ranking_Score'] = 50.0 + (units - 240) / 60 * 20  # 50-70 points
        
        # Add original LIHTC score bonus if available
        if score_col and score_col in df.columns:
            original_scores = pd.to_numeric(df[score_col], errors='coerce').fillna(0)
            # Add normalized LIHTC score (0-20 points) to ranking score
            max_score = original_scores.max() if original_scores.max() > 0 else 100
            df['Ranking_Score'] += (original_scores / max_score) * 20
        
        # Final ranking
        df = df.sort_values(['Ranking_Score'], ascending=False).reset_index(drop=True)
        df['Final_LIHTC_Rank'] = range(1, len(df) + 1)
        
        # Log tier distribution
        tier_summary = df['Ranking_Tier'].value_counts()
        logger.info("🏆 RANKING TIER DISTRIBUTION:")
        for tier, count in tier_summary.items():
            logger.info(f"   {tier}: {count} sites")
        
        return df
    
    def remove_environmental_columns(self, df):
        """Remove environmental analysis columns (technically inaccurate for large parcels)"""
        df = df.copy()
        
        # Find environmental columns that actually exist
        existing_env_cols = [col for col in self.environmental_columns if col in df.columns]
        
        if existing_env_cols:
            df = df.drop(columns=existing_env_cols)
            logger.info(f"🗑️ Removed {len(existing_env_cols)} environmental columns (technically inaccurate for large parcels)")
            for col in existing_env_cols:
                logger.info(f"   - Removed: {col}")
        else:
            logger.info("ℹ️ No environmental columns found to remove")
        
        return df
    
    def apply_data_formatting(self, df):
        """Apply all data formatting fixes"""
        df = df.copy()
        
        # Format phone numbers
        phone_cols = [c for c in df.columns if 'phone' in c.lower()]
        for col in phone_cols:
            if col in df.columns:
                df[col] = df[col].apply(self.format_phone_numbers)
                logger.info(f"📱 Fixed phone formatting: {col}")
        
        # Format ZIP codes
        zip_cols = [c for c in df.columns if 'zip' in c.lower()]
        for col in zip_cols:
            if col in df.columns:
                df[col] = df[col].apply(self.format_zip_codes)
                logger.info(f"📮 Fixed ZIP formatting: {col}")
        
        # Format acres
        acres_cols = [c for c in df.columns if 'acre' in c.lower()]
        for col in acres_cols:
            if col in df.columns:
                df[col] = df[col].apply(self.format_acres)
                logger.info(f"📏 Fixed acres formatting: {col}")
        
        # Format poverty rates
        poverty_cols = [c for c in df.columns if 'poverty' in c.lower() and 'rate' in c.lower()]
        for col in poverty_cols:
            if col in df.columns:
                df[col] = df[col].apply(self.format_poverty_rate)
                logger.info(f"📈 Fixed poverty rate formatting: {col}")
        
        return df
    
    def apply_professional_formatting(self, workbook):
        """Apply professional Excel formatting with color coding"""
        
        # Define CoStar vs Colosseum column categories
        costar_fields = [
            'address', 'city', 'state', 'zip', 'county', 'latitude', 'longitude', 
            'acres', 'property', 'list', 'price', 'phone', 'fema', 'size', 'rating',
            'sale', 'name', 'listing', 'broker'
        ]
        
        for sheet_name in workbook.sheetnames:
            if any(term in sheet_name.lower() for term in ['analysis', 'ready']):
                ws = workbook[sheet_name]
                
                # Style headers
                for col_num in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=1, column=col_num)
                    header_cell.font = self.header_font
                    header_cell.fill = self.header_fill
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Style data rows with color coding
                for row_num in range(2, ws.max_row + 1):
                    for col_num in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        header_text = ws.cell(row=1, column=col_num).value or ''
                        
                        # Determine font color based on column type
                        is_costar = any(field in header_text.lower() for field in costar_fields)
                        
                        if any(term in header_text.lower() for term in ['tdhca_region', 'county', 'ranking', 'tier']):
                            # Census/TDHCA/Ranking columns are Colosseum analysis (navy)
                            cell.font = self.colosseum_font
                        elif is_costar:
                            # CoStar original data (black)
                            cell.font = self.costar_font
                        else:
                            # Other analysis columns (navy)
                            cell.font = self.colosseum_font
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        return workbook
    
    def create_enhanced_analysis(self):
        """Create enhanced analysis with Census county lookup"""
        logger.info("🚀 STARTING ENHANCED COLOSSEUM ANALYSIS WITH CENSUS COUNTY LOOKUP")
        
        if not self.source_file.exists():
            logger.error(f"❌ Source file not found: {self.source_file}")
            return None
        
        # Load all worksheets
        xl_file = pd.ExcelFile(self.source_file)
        logger.info(f"📋 Processing {len(xl_file.sheet_names)} worksheets: {xl_file.sheet_names}")
        
        # Create output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.results_dir / f"ENHANCED_CENSUS_ANALYSIS_{timestamp}.xlsx"
        
        # Process each worksheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            for sheet_name in xl_file.sheet_names:
                logger.info(f"🔧 Processing worksheet: {sheet_name}")
                
                # Load sheet
                df = pd.read_excel(self.source_file, sheet_name=sheet_name)
                original_shape = df.shape
                
                # Apply enhancements to analysis sheets
                if any(term in sheet_name.lower() for term in ['analysis', 'ready']):
                    # 1. Remove environmental columns
                    df = self.remove_environmental_columns(df)
                    
                    # 2. Add Census counties and TDHCA regions (this is the key enhancement!)
                    df = self.add_census_counties_and_regions(df)
                    
                    # 3. Calculate proper ranking hierarchy
                    df = self.calculate_ranking_hierarchy(df)
                    
                    # 4. Apply data formatting
                    df = self.apply_data_formatting(df)
                
                # Write to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                logger.info(f"✅ {sheet_name}: {original_shape} → {df.shape}")
        
        # Apply professional formatting
        workbook = load_workbook(output_file)
        workbook = self.apply_professional_formatting(workbook)
        workbook.save(output_file)
        
        logger.info(f"💾 Enhanced analysis with Census counties saved: {output_file.name}")
        return output_file
    
    def run_enhanced_analysis(self):
        """Run complete enhanced analysis with Census county lookup"""
        logger.info("🎯 ENHANCED COLOSSEUM ANALYZER WITH CENSUS API INTEGRATION")
        
        output_file = self.create_enhanced_analysis()
        
        if output_file:
            logger.info("✅ ENHANCED ANALYSIS WITH CENSUS COUNTIES COMPLETE:")
            logger.info("   🗺️ Census API county lookup (100% free, high accuracy)")
            logger.info("   🏛️ TDHCA regions mapped from official county data")
            logger.info("   🏆 Proper ranking hierarchy implemented")
            logger.info("   🗑️ Environmental columns removed (technical accuracy)")
            logger.info("   📱 Phone formatting: (999) 999-9999")
            logger.info("   📮 ZIP formatting: 99999/99999-9999")
            logger.info("   📏 Acres formatting: 99.99")
            logger.info("   📈 Poverty formatting: 23.4%")
            logger.info("   🎨 Color coding: CoStar (black) vs Colosseum (navy)")
            logger.info("   💙 Blue headers restored")
            
            return output_file
        else:
            logger.error("❌ Enhanced analysis failed")
            return None

if __name__ == "__main__":
    analyzer = EnhancedColosseumAnalyzerWithCensus()
    output_file = analyzer.run_enhanced_analysis()
    
    if output_file:
        print(f"\n🎯 ENHANCED COLOSSEUM ANALYSIS WITH CENSUS COUNTIES COMPLETE")
        print("=" * 70)
        print(f"📁 Output: {output_file.name}")
        print("\n✅ ALL ENHANCEMENTS IMPLEMENTED:")
        print("  🗺️ Census API: Official US government county lookup")
        print("  🏛️ TDHCA Regions: Accurate mapping from county data") 
        print("  🏆 Ranking Hierarchy: Flood zones → <240 units → 240+ units")
        print("  🗑️ Environmental Removed: Technically accurate approach")
        print("  📱 Phone: (999) 999-9999")
        print("  📮 ZIP: 99999 or 99999-9999")
        print("  📏 Acres: 99.99 format")
        print("  📈 Poverty: 23.4% format")
        print("  🎨 Color Coding: CoStar (black) vs Colosseum (navy)")
        print("  💙 Blue Headers: Professional Excel formatting")
        print("  📋 All 5 Worksheets: Preserved and enhanced")
        print("\n🚀 Ready for LIHTC analysis and client presentation!")
    else:
        print("❌ Enhanced analysis failed")