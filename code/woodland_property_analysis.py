#!/usr/bin/env python3
"""
Complete property analysis for the Woodland-related property
"""

import pandas as pd
import openpyxl
import os
from pathlib import Path
from openpyxl.utils import get_column_letter

def get_property_details(file_path):
    """Get complete property details for the Woodland property"""
    filename = Path(file_path).name
    
    print(f"\n{'='*100}")
    print(f"PROPERTY ANALYSIS: {filename}")
    print(f"{'='*100}")
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Find the worksheet with raw data
        data_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'raw' in sheet_name.lower() or 'data' in sheet_name.lower():
                data_sheet = workbook[sheet_name]
                break
        
        if not data_sheet:
            data_sheet = workbook[workbook.sheetnames[-1]]  # Use last sheet as fallback
            
        print(f"Using worksheet: '{data_sheet.title}'")
        
        # Find the row with "The Woodlands" city (row 63)
        target_row = 63
        
        # Get all column headers
        headers = {}
        for col in range(1, 100):
            cell = data_sheet.cell(row=1, column=col)
            if cell.value:
                col_letter = get_column_letter(col)
                headers[col_letter] = str(cell.value)
        
        # Get all data for the target row
        property_data = {}
        for col in range(1, 100):
            cell = data_sheet.cell(row=target_row, column=col)
            if cell.value:
                col_letter = get_column_letter(col)
                header = headers.get(col_letter, f"Column_{col_letter}")
                property_data[header] = str(cell.value)
        
        print(f"\nPROPERTY DETAILS FOR THE WOODLANDS LISTING:")
        print("=" * 80)
        
        # Group and display the data logically
        property_info = {}
        broker_info = {}
        financial_info = {}
        location_info = {}
        zoning_info = {}
        other_info = {}
        
        for header, value in property_data.items():
            header_lower = header.lower()
            if any(word in header_lower for word in ['address', 'city', 'state', 'zip', 'street', 'latitude', 'longitude', 'legal']):
                location_info[header] = value
            elif any(word in header_lower for word in ['broker', 'agent', 'phone', 'company']):
                broker_info[header] = value
            elif any(word in header_lower for word in ['price', 'acre', 'cost', 'rent', 'value', 'income', 'tax']):
                financial_info[header] = value
            elif any(word in header_lower for word in ['zoning', 'flood', 'qct', 'dda', 'basis', 'zone']):
                zoning_info[header] = value
            elif any(word in header_lower for word in ['property', 'land', 'type', 'status', 'sale', 'listing']):
                property_info[header] = value
            else:
                other_info[header] = value
        
        # Display organized information
        if property_info:
            print("\n🏠 PROPERTY INFORMATION:")
            print("-" * 40)
            for header, value in property_info.items():
                print(f"  {header}: {value}")
        
        if location_info:
            print("\n📍 LOCATION INFORMATION:")
            print("-" * 40)
            for header, value in location_info.items():
                print(f"  {header}: {value}")
        
        if broker_info:
            print("\n🏢 BROKER/AGENT INFORMATION:")
            print("-" * 40)
            for header, value in broker_info.items():
                print(f"  {header}: {value}")
        
        if financial_info:
            print("\n💰 FINANCIAL INFORMATION:")
            print("-" * 40)
            for header, value in financial_info.items():
                print(f"  {header}: {value}")
        
        if zoning_info:
            print("\n🏛️ ZONING & REGULATORY:")
            print("-" * 40)
            for header, value in zoning_info.items():
                print(f"  {header}: {value}")
        
        if other_info:
            print("\n📋 ADDITIONAL INFORMATION:")
            print("-" * 40)
            for header, value in other_info.items():
                print(f"  {header}: {value}")
                    
    except Exception as e:
        print(f"Error analyzing {filename}: {str(e)}")

def main():
    # Check all files for property details
    file_paths = [
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Complete_20250616_233933.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Corrected_20250616_234759.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Enhanced_Excel_20250616_231632.xlsx"
    ]
    
    print("COMPLETE WOODLAND PROPERTY ANALYSIS")
    print("=" * 100)
    print("Property: 81 FM 3454 Rd, Huntsville, TX")
    print("Listed by: Compass RE Texas, LLC in The Woodlands, TX")
    print("Contact: Sherri Barrett, 832-899-4788")
    
    for file_path in file_paths:
        if os.path.exists(file_path):
            get_property_details(file_path)
        else:
            print(f"File not found: {Path(file_path).name}")

if __name__ == "__main__":
    main()