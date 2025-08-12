#!/usr/bin/env python3
"""
Focused search for "Woodland" mentions in CoStar export Excel files
Focus specifically on hotel/development project context
"""

import pandas as pd
import openpyxl
import os
import re
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

def search_excel_for_woodland_focused(file_path):
    """Search all worksheets in an Excel file for mentions of 'Woodland' with focus on hotels/development"""
    results = []
    
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    
    try:
        # Load workbook to get sheet names
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        print(f"\nSearching file: {Path(file_path).name}")
        
        for sheet_name in sheet_names:
            try:
                # Read each sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Search for 'Woodland' (case-insensitive) in all columns
                woodland_matches = []
                
                for col_idx, col_name in enumerate(df.columns):
                    for row_idx, cell_value in enumerate(df[col_name]):
                        if pd.notna(cell_value) and isinstance(cell_value, str):
                            if re.search(r'woodland', str(cell_value), re.IGNORECASE):
                                # Excel row numbers start at 1, plus 1 for header
                                excel_row = row_idx + 2
                                excel_col = openpyxl.utils.get_column_letter(col_idx + 1)
                                cell_ref = f"{excel_col}{excel_row}"
                                
                                # Get full row context
                                row_data = df.iloc[row_idx].to_dict()
                                
                                # Check if this is hotel/development related
                                is_hotel_related = check_hotel_context(cell_value, row_data)
                                is_development_related = check_development_context(cell_value, row_data)
                                is_legal_entity = check_legal_entity_context(cell_value, row_data)
                                
                                woodland_matches.append({
                                    'cell_reference': cell_ref,
                                    'column_name': col_name,
                                    'row_index': excel_row,
                                    'cell_value': str(cell_value),
                                    'is_hotel_related': is_hotel_related,
                                    'is_development_related': is_development_related,
                                    'is_legal_entity': is_legal_entity,
                                    'context_summary': get_context_summary(row_data)
                                })
                
                if woodland_matches:
                    results.append({
                        'sheet_name': sheet_name,
                        'matches': woodland_matches,
                        'total_matches': len(woodland_matches)
                    })
                    
            except Exception as e:
                results.append({
                    'sheet_name': sheet_name,
                    'error': f"Error reading sheet: {str(e)}"
                })
                
    except Exception as e:
        return {"error": f"Error opening file: {str(e)}"}
    
    return results

def check_hotel_context(cell_value, row_data):
    """Check if the woodland mention is in hotel context"""
    hotel_keywords = ['hotel', 'inn', 'resort', 'lodge', 'motel', 'suites', 'hospitality']
    
    # Check the cell value itself
    if any(keyword in cell_value.lower() for keyword in hotel_keywords):
        return True
    
    # Check other fields in the row
    for key, value in row_data.items():
        if pd.notna(value) and isinstance(value, str):
            if any(keyword in str(value).lower() for keyword in hotel_keywords):
                return True
    
    return False

def check_development_context(cell_value, row_data):
    """Check if the woodland mention is in development context"""
    dev_keywords = ['development', 'project', 'construction', 'build', 'developer', 'planned', 'proposed']
    
    # Check the cell value itself
    if any(keyword in cell_value.lower() for keyword in dev_keywords):
        return True
    
    # Check other fields in the row
    for key, value in row_data.items():
        if pd.notna(value) and isinstance(value, str):
            if any(keyword in str(value).lower() for keyword in dev_keywords):
                return True
    
    return False

def check_legal_entity_context(cell_value, row_data):
    """Check if the woodland mention is related to legal entities"""
    entity_keywords = ['llc', 'lp', 'corp', 'inc', 'corporation', 'partnership', 'ltd', 'limited']
    
    # Check the cell value itself
    if any(keyword in cell_value.lower() for keyword in entity_keywords):
        return True
    
    # Check other fields in the row
    for key, value in row_data.items():
        if pd.notna(value) and isinstance(value, str):
            if any(keyword in str(value).lower() for keyword in entity_keywords):
                return True
    
    return False

def get_context_summary(row_data):
    """Get a summary of key context information"""
    context = {}
    
    # Priority fields to include in summary
    priority_fields = [
        'name', 'property_name', 'project_name', 'address', 'city', 'state',
        'property_type', 'property type', 'secondary_type', 'secondary type',
        'owner', 'developer', 'listing_broker_company', 'listing broker company',
        'sale_price', 'sale price', 'price'
    ]
    
    for key, value in row_data.items():
        if pd.notna(value):
            key_lower = str(key).lower()
            if any(field in key_lower for field in priority_fields):
                context[key] = str(value)
    
    return context

def main():
    # Define the file paths
    file_paths = [
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/CoStar_Texas_PreFiltered_20250616_194859.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/CoStar_Land_Analysis_With_Counties_20250617_223737.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/TDHCA_Analysis_Results/CoStar_TX_Land_TDHCA_Analysis_20250602_203319.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-8.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-9.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-10.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-11.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-12.xlsx"
    ]
    
    # Check which files exist
    existing_files = []
    missing_files = []
    
    print("=== FOCUSED WOODLAND SEARCH IN COSTAR EXCEL FILES ===\n")
    print("Target: Hotel/Development/Legal Entity contexts")
    print("Checking file availability:")
    
    for file_path in file_paths:
        if os.path.exists(file_path):
            existing_files.append(file_path)
            print(f"✓ Found: {Path(file_path).name}")
        else:
            missing_files.append(file_path)
            print(f"✗ Missing: {Path(file_path).name}")

    print(f"\nTotal files found: {len(existing_files)}")
    
    # Search existing files
    all_results = {}
    total_matches = 0
    hotel_matches = 0
    development_matches = 0
    legal_entity_matches = 0
    
    print(f"\n{'='*60}")
    print("BEGINNING FOCUSED SEARCH")
    print(f"{'='*60}")
    
    for file_path in existing_files:
        file_name = Path(file_path).name
        print(f"\nProcessing: {file_name}")
        
        results = search_excel_for_woodland_focused(file_path)
        all_results[file_name] = results
        
        if isinstance(results, dict) and 'error' in results:
            print(f"ERROR: {results['error']}")
        elif isinstance(results, list):
            file_matches = sum(sheet.get('total_matches', 0) for sheet in results if 'total_matches' in sheet)
            total_matches += file_matches
            print(f"  Total matches: {file_matches}")
    
    # Generate focused report
    print(f"\n{'='*60}")
    print("FOCUSED WOODLAND SEARCH REPORT")
    print(f"{'='*60}")
    
    print(f"\nEXECUTIVE SUMMARY:")
    print(f"- Files searched: {len(existing_files)}")
    print(f"- Total 'Woodland' matches found: {total_matches}")
    
    # Analyze matches by category
    hotel_related_matches = []
    development_related_matches = []
    legal_entity_matches_list = []
    geographic_matches = []
    
    for file_name, results in all_results.items():
        if isinstance(results, list):
            for sheet_result in results:
                if 'matches' in sheet_result:
                    for match in sheet_result['matches']:
                        match_info = {
                            'file': file_name,
                            'sheet': sheet_result['sheet_name'],
                            'cell_ref': match['cell_reference'],
                            'column': match['column_name'],
                            'value': match['cell_value'],
                            'context': match['context_summary']
                        }
                        
                        if match['is_hotel_related']:
                            hotel_related_matches.append(match_info)
                        if match['is_development_related']:
                            development_related_matches.append(match_info)
                        if match['is_legal_entity']:
                            legal_entity_matches_list.append(match_info)
                        
                        # Check if it's geographic (like "The Woodlands" city/area)
                        if 'woodlands' in match['cell_value'].lower() and 'the woodlands' in match['cell_value'].lower():
                            geographic_matches.append(match_info)
    
    print(f"\nCATEGORY ANALYSIS:")
    print(f"- Hotel-related matches: {len(hotel_related_matches)}")
    print(f"- Development-related matches: {len(development_related_matches)}")
    print(f"- Legal entity matches: {len(legal_entity_matches_list)}")
    print(f"- Geographic references (The Woodlands): {len(geographic_matches)}")
    
    # Detailed findings for each category
    if hotel_related_matches:
        print(f"\n{'='*50}")
        print("HOTEL-RELATED WOODLAND MENTIONS")
        print(f"{'='*50}")
        for match in hotel_related_matches:
            print(f"\nFile: {match['file']}")
            print(f"Sheet: {match['sheet']}")
            print(f"Cell: {match['cell_ref']} ({match['column']})")
            print(f"Value: {match['value']}")
            if match['context']:
                print(f"Context:")
                for key, value in match['context'].items():
                    print(f"  {key}: {value}")
    
    if development_related_matches:
        print(f"\n{'='*50}")
        print("DEVELOPMENT-RELATED WOODLAND MENTIONS")
        print(f"{'='*50}")
        for match in development_related_matches:
            print(f"\nFile: {match['file']}")
            print(f"Sheet: {match['sheet']}")
            print(f"Cell: {match['cell_ref']} ({match['column']})")
            print(f"Value: {match['value']}")
            if match['context']:
                print(f"Context:")
                for key, value in match['context'].items():
                    print(f"  {key}: {value}")
    
    if legal_entity_matches_list:
        print(f"\n{'='*50}")
        print("LEGAL ENTITY WOODLAND MENTIONS")
        print(f"{'='*50}")
        for match in legal_entity_matches_list:
            print(f"\nFile: {match['file']}")
            print(f"Sheet: {match['sheet']}")
            print(f"Cell: {match['cell_ref']} ({match['column']})")
            print(f"Value: {match['value']}")
            if match['context']:
                print(f"Context:")
                for key, value in match['context'].items():
                    print(f"  {key}: {value}")
    
    # Show all unique Woodland values found
    print(f"\n{'='*50}")
    print("ALL UNIQUE WOODLAND VALUES FOUND")
    print(f"{'='*50}")
    unique_values = set()
    for file_name, results in all_results.items():
        if isinstance(results, list):
            for sheet_result in results:
                if 'matches' in sheet_result:
                    for match in sheet_result['matches']:
                        unique_values.add(match['cell_value'])
    
    for value in sorted(unique_values):
        print(f"- {value}")
    
    print(f"\n{'='*60}")
    print("SEARCH COMPLETE")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()