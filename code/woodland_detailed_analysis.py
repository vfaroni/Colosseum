#!/usr/bin/env python3
"""
Detailed analysis of Woodland mentions in D'Marco Excel files
Focus on the real estate company in The Woodlands, TX
"""

import pandas as pd
import openpyxl
import os
from pathlib import Path
from openpyxl.utils import get_column_letter

def analyze_woodland_company(file_path):
    """Analyze the specific real estate company in The Woodlands"""
    filename = Path(file_path).name
    
    print(f"\n{'='*80}")
    print(f"DETAILED ANALYSIS: {filename}")
    print(f"{'='*80}")
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook['Raw Data']
        
        # Find the row with "The Woodlands" city (row 63)
        target_row = 63
        
        print(f"\nDetailed analysis of Row {target_row} (The Woodlands real estate company):")
        print("-" * 60)
        
        # Get comprehensive context for this row (columns A through BZ)
        row_data = {}
        for col in range(1, 80):  # Columns A through BZ
            cell = worksheet.cell(row=target_row, column=col)
            if cell.value:
                col_letter = get_column_letter(col)
                row_data[col_letter] = str(cell.value)
        
        # Print all available data for this row
        for col_letter in sorted(row_data.keys()):
            value = row_data[col_letter]
            print(f"  {col_letter}{target_row}: {value}")
            
        # Also check a few rows above and below for additional context
        print(f"\nContext rows around The Woodlands entry:")
        print("-" * 60)
        
        for context_row in range(target_row - 3, target_row + 4):
            if context_row == target_row:
                continue
                
            print(f"\nRow {context_row}:")
            context_data = {}
            for col in range(1, 80):
                cell = worksheet.cell(row=context_row, column=col)
                if cell.value:
                    col_letter = get_column_letter(col)
                    context_data[col_letter] = str(cell.value)
            
            # Show key columns for context
            key_columns = ['AD', 'AE', 'AF', 'AH', 'AI', 'AJ', 'AK']  # Company, lat, lon, address, city, phone, state
            for col in key_columns:
                if col in context_data:
                    print(f"    {col}{context_row}: {context_data[col]}")
                    
        # Let's also check if there are column headers
        print(f"\nColumn Headers (Row 1):")
        print("-" * 60)
        header_row = 1
        for col in range(25, 50):  # Focus on columns around AD-AX
            cell = worksheet.cell(row=header_row, column=col)
            if cell.value:
                col_letter = get_column_letter(col)
                print(f"  {col_letter}1: {cell.value}")
                
    except Exception as e:
        print(f"Error analyzing {filename}: {str(e)}")

def main():
    # Focus on the files that have The Woodlands company
    target_files = [
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Complete_20250616_233933.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Corrected_20250616_234759.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/DMarco_Enhanced_Excel_20250616_231632.xlsx"
    ]
    
    print("DETAILED WOODLAND COMPANY ANALYSIS")
    print("=" * 80)
    print("Focusing on: Compass RE Texas, LLC in The Woodlands, TX")
    print("Address: 1800 Hughes Landing Blvd, 725")
    print("Phone: 8328994788")
    
    for file_path in target_files:
        if os.path.exists(file_path):
            analyze_woodland_company(file_path)
        else:
            print(f"File not found: {Path(file_path).name}")

if __name__ == "__main__":
    main()