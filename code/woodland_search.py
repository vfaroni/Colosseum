#!/usr/bin/env python3
"""
Systematic search for "Woodland" mentions in CoStar export Excel files
Focus on hotel/development project context
"""

import pandas as pd
import openpyxl
import os
import re
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

def search_excel_for_woodland(file_path):
    """Search all worksheets in an Excel file for mentions of 'Woodland'"""
    results = []
    
    if not os.path.exists(file_path):
        return {"error": f"File not found: {file_path}"}
    
    try:
        # Load workbook to get sheet names
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        print(f"\nSearching file: {Path(file_path).name}")
        print(f"Worksheets found: {sheet_names}")
        
        for sheet_name in sheet_names:
            try:
                # Read each sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                print(f"  Searching sheet '{sheet_name}' - Shape: {df.shape}")
                
                # Search for 'Woodland' (case-insensitive) in all columns
                woodland_matches = []
                
                for col_idx, col_name in enumerate(df.columns):
                    for row_idx, cell_value in enumerate(df[col_name]):
                        if pd.notna(cell_value) and isinstance(cell_value, str):
                            if re.search(r'woodland', str(cell_value), re.IGNORECASE):
                                # Excel row numbers start at 1, plus 1 for header
                                excel_row = row_idx + 2
                                excel_col = openpyxl.utils.get_column_letter(col_idx + 1)
                                cell_ref = f"{excel_col}{excel_row}"
                                
                                woodland_matches.append({
                                    'cell_reference': cell_ref,
                                    'column_name': col_name,
                                    'row_index': excel_row,
                                    'cell_value': str(cell_value),
                                    'context_data': get_row_context(df, row_idx)
                                })
                
                if woodland_matches:
                    results.append({
                        'sheet_name': sheet_name,
                        'matches': woodland_matches,
                        'total_matches': len(woodland_matches)
                    })
                    print(f"    Found {len(woodland_matches)} 'Woodland' matches in sheet '{sheet_name}'")
                else:
                    print(f"    No 'Woodland' matches found in sheet '{sheet_name}'")
                    
            except Exception as e:
                results.append({
                    'sheet_name': sheet_name,
                    'error': f"Error reading sheet: {str(e)}"
                })
                print(f"    Error reading sheet '{sheet_name}': {str(e)}")
                
    except Exception as e:
        return {"error": f"Error opening file: {str(e)}"}
    
    return results

def get_row_context(df, row_idx):
    """Get context information from the entire row"""
    context = {}
    row_data = df.iloc[row_idx]
    
    # Look for key columns that might contain relevant context
    key_columns = [
        'property_name', 'property name', 'name', 'project_name', 'project name',
        'owner', 'developer', 'entity', 'legal_entity', 'legal entity',
        'address', 'street_address', 'street address', 'location',
        'property_type', 'property type', 'type', 'use_type', 'use type',
        'hotel', 'development', 'project',
        'city', 'state', 'zip', 'county',
        'price', 'value', 'amount', 'sale_price', 'sale price',
        'llc', 'lp', 'corp', 'corporation', 'partnership', 'inc'
    ]
    
    for col_name in df.columns:
        col_lower = str(col_name).lower()
        cell_value = row_data[col_name]
        
        # Include if column name contains key terms or if cell value is not null
        if any(key in col_lower for key in key_columns) or pd.notna(cell_value):
            if pd.notna(cell_value):
                context[col_name] = str(cell_value)
    
    return context

def main():
    # Define the file paths
    file_paths = [
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/CoStar_Texas_PreFiltered_20250616_194859.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/CoStar_Land_Analysis_With_Counties_20250617_223737.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/priorcode/TDHCA_Analysis_Results/CoStar_TX_Land_TDHCA_Analysis_20250602_203319.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-8.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-9.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-10.xlsx",
        "/Users/vitorfaroni/Library/ CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-11.xlsx",
        "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Colosseum/modules/lihtc_analyst/botn_engine/Sites/CostarExport-12.xlsx"
    ]
    
    # Check which files exist
    existing_files = []
    missing_files = []
    
    print("=== WOODLAND SEARCH IN COSTAR EXCEL FILES ===\n")
    print("Checking file availability:")
    
    for file_path in file_paths:
        if os.path.exists(file_path):
            existing_files.append(file_path)
            print(f"✓ Found: {Path(file_path).name}")
        else:
            missing_files.append(file_path)
            print(f"✗ Missing: {Path(file_path).name}")

    print(f"\nTotal files found: {len(existing_files)}")
    print(f"Total files missing: {len(missing_files)}")
    
    if missing_files:
        print(f"\nMissing files:")
        for file_path in missing_files:
            print(f"  - {file_path}")
    
    # Search existing files
    all_results = {}
    total_matches = 0
    
    print(f"\n{'='*60}")
    print("BEGINNING SYSTEMATIC SEARCH")
    print(f"{'='*60}")
    
    for file_path in existing_files:
        file_name = Path(file_path).name
        print(f"\n{'='*40}")
        print(f"PROCESSING: {file_name}")
        print(f"{'='*40}")
        
        results = search_excel_for_woodland(file_path)
        all_results[file_name] = results
        
        if isinstance(results, dict) and 'error' in results:
            print(f"ERROR: {results['error']}")
        elif isinstance(results, list):
            file_matches = sum(sheet.get('total_matches', 0) for sheet in results if 'total_matches' in sheet)
            total_matches += file_matches
            print(f"SUMMARY: {file_matches} total matches found in {file_name}")
    
    # Generate comprehensive report
    print(f"\n{'='*60}")
    print("COMPREHENSIVE WOODLAND SEARCH REPORT")
    print(f"{'='*60}")
    
    print(f"\nEXECUTIVE SUMMARY:")
    print(f"- Files searched: {len(existing_files)}")
    print(f"- Files missing: {len(missing_files)}")
    print(f"- Total 'Woodland' matches found: {total_matches}")
    
    # Detailed findings
    for file_name, results in all_results.items():
        print(f"\n{'='*50}")
        print(f"FILE: {file_name}")
        print(f"{'='*50}")
        
        if isinstance(results, dict) and 'error' in results:
            print(f"ERROR: {results['error']}")
            continue
        
        if not results:
            print("No matches found in this file.")
            continue
        
        for sheet_result in results:
            if 'error' in sheet_result:
                print(f"SHEET ERROR: {sheet_result['sheet_name']} - {sheet_result['error']}")
                continue
                
            sheet_name = sheet_result['sheet_name']
            matches = sheet_result.get('matches', [])
            
            print(f"\nWORKSHEET: {sheet_name}")
            print(f"Matches found: {len(matches)}")
            
            for i, match in enumerate(matches, 1):
                print(f"\n  MATCH #{i}:")
                print(f"    Cell Reference: {match['cell_reference']}")
                print(f"    Column: {match['column_name']}")
                print(f"    Value: {match['cell_value']}")
                
                # Show context data
                context = match.get('context_data', {})
                if context:
                    print(f"    Row Context:")
                    for key, value in context.items():
                        if len(str(value)) > 100:
                            value = str(value)[:100] + "..."
                        print(f"      {key}: {value}")
    
    # Look for patterns
    print(f"\n{'='*60}")
    print("PATTERN ANALYSIS")
    print(f"{'='*60}")
    
    hotel_related = []
    development_related = []
    legal_entity_related = []
    
    for file_name, results in all_results.items():
        if isinstance(results, list):
            for sheet_result in results:
                if 'matches' in sheet_result:
                    for match in sheet_result['matches']:
                        match_text = match['cell_value'].lower()
                        context = match.get('context_data', {})
                        
                        # Check for hotel context
                        if any(term in match_text for term in ['hotel', 'inn', 'resort', 'lodge']):
                            hotel_related.append({
                                'file': file_name,
                                'sheet': sheet_result['sheet_name'],
                                'match': match
                            })
                        
                        # Check for development context
                        if any(term in match_text for term in ['development', 'project', 'construction', 'build']):
                            development_related.append({
                                'file': file_name,
                                'sheet': sheet_result['sheet_name'],
                                'match': match
                            })
                        
                        # Check for legal entity context
                        if any(term in match_text for term in ['llc', 'lp', 'corp', 'inc', 'partnership']):
                            legal_entity_related.append({
                                'file': file_name,
                                'sheet': sheet_result['sheet_name'],
                                'match': match
                            })
    
    print(f"\nHOTEL-RELATED WOODLAND MENTIONS: {len(hotel_related)}")
    for item in hotel_related:
        print(f"  - {item['file']} > {item['sheet']} > {item['match']['cell_reference']}: {item['match']['cell_value']}")
    
    print(f"\nDEVELOPMENT-RELATED WOODLAND MENTIONS: {len(development_related)}")
    for item in development_related:
        print(f"  - {item['file']} > {item['sheet']} > {item['match']['cell_reference']}: {item['match']['cell_value']}")
    
    print(f"\nLEGAL ENTITY-RELATED WOODLAND MENTIONS: {len(legal_entity_related)}")
    for item in legal_entity_related:
        print(f"  - {item['file']} > {item['sheet']} > {item['match']['cell_reference']}: {item['match']['cell_value']}")
    
    print(f"\n{'='*60}")
    print("SEARCH COMPLETE")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()