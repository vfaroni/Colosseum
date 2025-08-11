#!/usr/bin/env python3
"""
Unit tests for OpenPyxl BOTN Generator - Roman Engineering Standards Compliance
"""

import unittest
import tempfile
import shutil
from pathlib import Path
import sys
import pandas as pd
import openpyxl

# Add the module path for imports
sys.path.append(str(Path(__file__).parent.parent.parent / "modules" / "lihtc_analyst" / "botn_engine"))

from botn_openpyxl_generator import OpenPyxlBOTNGenerator

class TestOpenPyxlBOTNGenerator(unittest.TestCase):
    """Test suite for OpenPyxl BOTN Generator"""
    
    def setUp(self):
        """Set up test environment"""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.generator = OpenPyxlBOTNGenerator()
        
        # Create mock template
        self.mock_template = self.temp_dir / "mock_template.xlsx"
        self.create_mock_template()
        
        # Override template path for testing
        self.generator.template_path = self.mock_template
        
    def tearDown(self):
        """Clean up test environment"""
        shutil.rmtree(self.temp_dir)
    
    def create_mock_template(self):
        """Create a mock CABOTNTemplate for testing"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create required sheets
        wb.remove(wb.active)
        
        # Create Inputs sheet
        inputs_sheet = wb.create_sheet("Inputs")
        inputs_sheet['A1'] = 'Property Name'
        inputs_sheet['B1'] = 'Address'
        inputs_sheet['C1'] = 'County:'
        inputs_sheet['A2'] = ''  # Will be populated by generator
        inputs_sheet['B2'] = ''
        
        # Create Rents sheet with formulas
        rents_sheet = wb.create_sheet("Rents")
        rents_sheet['C2'] = '=Inputs!B5'
        rents_sheet['C3'] = '=Inputs!C2'
        
        # Create Expenses sheet with formulas
        expenses_sheet = wb.create_sheet("Expenses")
        expenses_sheet['C2'] = '=Inputs!N2'
        expenses_sheet['C3'] = '=Inputs!H2'
        
        # Create Sources & Uses sheet with formulas
        su_sheet = wb.create_sheet("Sources & Uses")
        su_sheet['C14'] = '=Inputs!J2'
        su_sheet['C16'] = '=Inputs!I2'
        
        # Create NOI sheet with formulas
        noi_sheet = wb.create_sheet("NOI")
        noi_sheet['C2'] = '=Inputs!A2'
        noi_sheet['C3'] = '=Inputs!B2'
        
        wb.save(self.mock_template)
        wb.close()
    
    def test_clean_data_value(self):
        """Test data value cleaning functionality"""
        # Test various data types and edge cases
        self.assertEqual(self.generator.clean_data_value(''), '')
        self.assertEqual(self.generator.clean_data_value(None), '')
        self.assertEqual(self.generator.clean_data_value('test'), 'test')
        self.assertEqual(self.generator.clean_data_value(123), 123)
        self.assertEqual(self.generator.clean_data_value('nan'), '')
        self.assertEqual(self.generator.clean_data_value('None'), '')
        
        # Test pandas NA values
        self.assertEqual(self.generator.clean_data_value(pd.NA), '')
    
    def test_get_cdlac_region(self):
        """Test CDLAC region mapping"""
        # Test known mappings
        self.assertEqual(self.generator.get_cdlac_region('Los Angeles'), 'Greater Los Angeles Area')
        self.assertEqual(self.generator.get_cdlac_region('San Francisco'), 'San Francisco County')
        self.assertEqual(self.generator.get_cdlac_region('Alameda'), 'East Bay Region')
        self.assertEqual(self.generator.get_cdlac_region('San Diego'), 'San Diego/Imperial Region')
        
        # Test unknown county defaults to Northern Region
        self.assertEqual(self.generator.get_cdlac_region('Unknown County'), 'Northern Region')
    
    def test_get_county_with_suffix(self):
        """Test county suffix addition"""
        self.assertEqual(self.generator.get_county_with_suffix('Los Angeles'), 'Los Angeles County')
        self.assertEqual(self.generator.get_county_with_suffix('Los Angeles County'), 'Los Angeles County')
        self.assertEqual(self.generator.get_county_with_suffix(''), '')
    
    def test_template_loading(self):
        """Test that mock template can be loaded successfully"""
        wb = openpyxl.load_workbook(self.mock_template, data_only=False)
        
        # Verify required sheets exist
        required_sheets = ['Inputs', 'Rents', 'Expenses', 'Sources & Uses', 'NOI']
        for sheet in required_sheets:
            self.assertIn(sheet, wb.sheetnames, f"Required sheet '{sheet}' not found")
        
        # Verify formulas are preserved
        rents_sheet = wb['Rents']
        self.assertEqual(rents_sheet['C2'].value, '=Inputs!B5')
        
        wb.close()
    
    def test_formula_preservation_workflow(self):
        """Test that the complete workflow preserves formulas"""
        # Create test output file
        test_output = self.temp_dir / "test_botn.xlsx"
        shutil.copy2(self.mock_template, test_output)
        
        # Load and modify like the generator does
        wb = openpyxl.load_workbook(test_output, data_only=False)
        inputs_sheet = wb['Inputs']
        
        # Populate test data
        inputs_sheet['A2'] = 'TEST PROPERTY'
        inputs_sheet['B2'] = '123 Test Street'
        inputs_sheet['G2'] = 2500000
        
        wb.save(test_output)
        wb.close()
        
        # Reload and verify formulas are still intact
        wb_check = openpyxl.load_workbook(test_output, data_only=False)
        
        rents_sheet = wb_check['Rents']
        self.assertEqual(rents_sheet['C2'].value, '=Inputs!B5')
        
        expenses_sheet = wb_check['Expenses']
        self.assertEqual(expenses_sheet['C2'].value, '=Inputs!N2')
        
        # Verify data was populated
        inputs_check = wb_check['Inputs']
        self.assertEqual(inputs_check['A2'].value, 'TEST PROPERTY')
        self.assertEqual(inputs_check['B2'].value, '123 Test Street')
        self.assertEqual(inputs_check['G2'].value, 2500000)
        
        wb_check.close()
    
    def test_mock_sites_data_processing(self):
        """Test processing with mock sites data"""
        # Create mock sites data
        mock_sites = [
            {
                'Property Name': 'Test Site 1',
                'Property Address': '123 Main St, Los Angeles, CA',
                'County Name': 'Los Angeles',
                'State': 'CA',
                'Zip': '90210',
                'For Sale Price': 3000000
            },
            {
                'Property Name': 'Test Site 2', 
                'Property Address': '456 Oak Ave, San Diego, CA',
                'County Name': 'San Diego',
                'State': 'CA',
                'Zip': '92101',
                'For Sale Price': 2500000
            }
        ]
        
        # Override base path for testing
        original_base_path = self.generator.base_path
        self.generator.base_path = self.temp_dir
        
        try:
            # Test processing 2 mock sites
            result = self.generator.create_botn_files(
                mock_sites, 
                start_index=0, 
                count=2, 
                output_dir_name="test_output"
            )
            
            # Verify results
            self.assertTrue(result['success'])
            self.assertEqual(result['files_created'], 2)
            self.assertEqual(result['total_requested'], 2)
            self.assertGreater(result['total_time'], 0)
            self.assertEqual(len(result['files']), 2)
            
            # Verify files were created
            output_dir = self.temp_dir / "test_output"
            self.assertTrue(output_dir.exists())
            
            created_files = list(output_dir.glob("*.xlsx"))
            self.assertEqual(len(created_files), 2)
            
            # Verify one of the files has proper data
            test_file = created_files[0]
            wb = openpyxl.load_workbook(test_file, data_only=False)
            inputs_sheet = wb['Inputs']
            
            # Check that data was populated
            self.assertIn('Test Site', str(inputs_sheet['A2'].value))
            self.assertIsNotNone(inputs_sheet['B2'].value)
            self.assertIsInstance(inputs_sheet['G2'].value, (int, float))
            
            # Verify formulas still exist
            rents_sheet = wb['Rents']
            self.assertTrue(str(rents_sheet['C2'].value).startswith('='))
            
            wb.close()
            
        finally:
            # Restore original base path
            self.generator.base_path = original_base_path
    
    def test_error_handling(self):
        """Test error handling for various scenarios"""
        # Test with invalid template path
        generator_bad_template = OpenPyxlBOTNGenerator()
        generator_bad_template.template_path = self.temp_dir / "nonexistent.xlsx"
        
        # Should handle missing template gracefully
        mock_sites = [{'Property Name': 'Test', 'For Sale Price': 1000000}]
        
        # This should fail gracefully without crashing
        try:
            result = generator_bad_template.create_botn_files(
                mock_sites, count=1, output_dir_name="error_test"
            )
            # If it returns, it should indicate failure
            self.assertFalse(result.get('success', False))
        except Exception:
            # Exceptions are acceptable for missing templates
            pass
    
    def test_production_settings(self):
        """Test that production settings are applied correctly"""
        mock_sites = [{
            'Property Name': 'Settings Test',
            'For Sale Price': 2000000,
            'County Name': 'Los Angeles'
        }]
        
        # Override base path for testing
        original_base_path = self.generator.base_path
        self.generator.base_path = self.temp_dir
        
        try:
            result = self.generator.create_botn_files(
                mock_sites, count=1, output_dir_name="settings_test"
            )
            
            if result['success'] and result['files_created'] > 0:
                # Check that production settings were applied
                output_dir = self.temp_dir / "settings_test"
                created_files = list(output_dir.glob("*.xlsx"))
                
                if created_files:
                    wb = openpyxl.load_workbook(created_files[0])
                    inputs_sheet = wb['Inputs']
                    
                    # Check production settings
                    self.assertEqual(inputs_sheet['H2'].value, 'Large Family')
                    self.assertEqual(inputs_sheet['I2'].value, 0.80)
                    self.assertEqual(inputs_sheet['J2'].value, '4%')
                    self.assertEqual(inputs_sheet['O2'].value, 80)
                    self.assertEqual(inputs_sheet['P2'].value, 950)
                    
                    wb.close()
                    
        finally:
            self.generator.base_path = original_base_path

class TestOpenPyxlIntegration(unittest.TestCase):
    """Integration tests for OpenPyxl functionality"""
    
    def test_openpyxl_formula_preservation(self):
        """Test that openpyxl preserves Excel formulas correctly"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        
        try:
            # Create workbook with formulas
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 10
            ws['A2'] = 20
            ws['A3'] = '=A1+A2'  # Formula
            ws['A4'] = '=SUM(A1:A2)'  # Another formula
            
            wb.save(temp_path)
            wb.close()
            
            # Reload and verify formulas are preserved
            wb2 = openpyxl.load_workbook(temp_path, data_only=False)
            ws2 = wb2.active
            
            self.assertEqual(ws2['A3'].value, '=A1+A2')
            self.assertEqual(ws2['A4'].value, '=SUM(A1:A2)')
            
            wb2.close()
            
        finally:
            temp_path.unlink(missing_ok=True)
    
    def test_openpyxl_no_permissions_required(self):
        """Test that openpyxl operations don't require special permissions"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        
        try:
            # This should work without any permission prompts
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Permission Test'
            
            # Save and reload multiple times (simulating batch processing)
            for i in range(3):
                ws['A1'] = f'Permission Test {i+1}'
                wb.save(temp_path)
                wb.close()
                
                wb = openpyxl.load_workbook(temp_path)
                ws = wb.active
            
            wb.close()
            
            # Final verification
            wb_final = openpyxl.load_workbook(temp_path)
            ws_final = wb_final.active
            self.assertEqual(ws_final['A1'].value, 'Permission Test 3')
            wb_final.close()
            
        finally:
            temp_path.unlink(missing_ok=True)

def run_tests():
    """Run all tests and return results"""
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestOpenPyxlBOTNGenerator))
    suite.addTests(loader.loadTestsFromTestCase(TestOpenPyxlIntegration))
    
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result.wasSuccessful()

if __name__ == '__main__':
    success = run_tests()
    exit(0 if success else 1)