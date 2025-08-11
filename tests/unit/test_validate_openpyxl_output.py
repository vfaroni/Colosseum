#!/usr/bin/env python3
"""
Unit tests for OpenPyxl Output Validator - Roman Engineering Standards Compliance
"""

import unittest
import tempfile
import shutil
from pathlib import Path
import sys
import openpyxl

# Add the module path for imports
sys.path.append(str(Path(__file__).parent.parent.parent / "modules" / "lihtc_analyst" / "botn_engine"))

from validate_openpyxl_output import OpenPyxlOutputValidator

class TestOpenPyxlOutputValidator(unittest.TestCase):
    """Test suite for OpenPyxl Output Validator"""
    
    def setUp(self):
        """Set up test environment"""
        self.temp_dir = Path(tempfile.mkdtemp())
        self.validator = OpenPyxlOutputValidator()
        
        # Create mock template and test directory
        self.create_test_environment()
        
        # Override paths for testing
        self.validator.base_path = self.temp_dir
        self.validator.template_path = self.temp_dir / "botntemplate" / "CABOTNTemplate.xlsx"
        self.validator.test_output_dir = self.temp_dir / "test_output"
        
    def tearDown(self):
        """Clean up test environment"""
        shutil.rmtree(self.temp_dir)
    
    def create_test_environment(self):
        """Create test environment with mock template and files"""
        # Create template directory
        template_dir = self.temp_dir / "botntemplate"
        template_dir.mkdir()
        
        # Create mock template
        template_path = template_dir / "CABOTNTemplate.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create required sheets
        sheets_data = {
            'Inputs': [('A1', 'Property Name'), ('B1', 'Address'), ('A2', ''), ('B2', '')],
            'Rents': [('C2', '=Inputs!B5'), ('C3', '=Inputs!C2')],
            'Expenses': [('C2', '=Inputs!N2'), ('C3', '=Inputs!H2')],
            'Sources & Uses': [('C14', '=Inputs!J2'), ('C16', '=Inputs!I2')],
            'NOI': [('C2', '=Inputs!A2'), ('C3', '=Inputs!B2')],
            'Data>>': [],
            'Developer Fee Max': [],
            'Section8-FY24': [],
            '2025 FMR': [],
            '2025 SAFMR': []
        }
        
        for sheet_name, data in sheets_data.items():
            sheet = wb.create_sheet(sheet_name)
            for cell_ref, value in data:
                sheet[cell_ref] = value
        
        wb.save(template_path)
        wb.close()
        
        # Create test output directory with mock BOTN files
        test_output_dir = self.temp_dir / "test_output"
        test_output_dir.mkdir()
        
        # Create mock BOTN files
        for i in range(2):
            botn_file = test_output_dir / f"Test_Site_{i+1}_BOTN.xlsx"
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Create sheets with populated data and formulas
            inputs_sheet = wb.create_sheet("Inputs")
            inputs_sheet['A2'] = f'Test Property {i+1}'
            inputs_sheet['B2'] = f'123 Test St {i+1}'
            inputs_sheet['G2'] = 2500000 + (i * 100000)
            inputs_sheet['H2'] = 'Large Family'
            inputs_sheet['I2'] = 0.80
            inputs_sheet['O2'] = 80
            inputs_sheet['P2'] = 950
            
            # Create formula sheets
            rents_sheet = wb.create_sheet("Rents")
            rents_sheet['C2'] = '=Inputs!B5'
            rents_sheet['C3'] = '=Inputs!C2'
            
            expenses_sheet = wb.create_sheet("Expenses") 
            expenses_sheet['C2'] = '=Inputs!N2'
            expenses_sheet['C3'] = '=Inputs!H2'
            
            su_sheet = wb.create_sheet("Sources & Uses")
            su_sheet['C14'] = '=Inputs!J2'
            su_sheet['C16'] = '=Inputs!I2'
            
            noi_sheet = wb.create_sheet("NOI")
            noi_sheet['C2'] = '=Inputs!A2'
            noi_sheet['C3'] = '=Inputs!B2'
            
            wb.save(botn_file)
            wb.close()
    
    def test_validator_initialization(self):
        """Test that validator initializes correctly"""
        validator = OpenPyxlOutputValidator()
        self.assertIsInstance(validator.base_path, Path)
        self.assertTrue(str(validator.template_path).endswith('CABOTNTemplate.xlsx'))
    
    def test_validate_single_file_success(self):
        """Test successful validation of a single BOTN file"""
        # Get the first test file
        botn_files = list(self.validator.test_output_dir.glob("*_BOTN.xlsx"))
        self.assertGreater(len(botn_files), 0, "Test BOTN files should exist")
        
        result = self.validator.validate_single_file(botn_files[0])
        
        # Check validation results structure
        self.assertIn('file_name', result)
        self.assertIn('worksheets_present', result) 
        self.assertIn('formulas_found', result)
        self.assertIn('data_populated', result)
        self.assertIn('errors', result)
        
        # Check that expected sheets are found
        expected_sheets = ['Inputs', 'Rents', 'Expenses', 'Sources & Uses', 'NOI']
        for sheet in expected_sheets:
            self.assertIn(sheet, result['worksheets_present'])
        
        # Check that data was populated
        self.assertIn('Property Name', result['data_populated'])
        self.assertIn('Address', result['data_populated'])
        
        # Check that formulas were found
        self.assertTrue(len(result['formulas_found']) > 0)
        for sheet_name in ['Rents', 'Expenses', 'Sources & Uses', 'NOI']:
            if sheet_name in result['formulas_found']:
                self.assertTrue(len(result['formulas_found'][sheet_name]) > 0)
    
    def test_validate_single_file_missing_file(self):
        """Test validation with missing file"""
        fake_path = self.temp_dir / "nonexistent_BOTN.xlsx"
        result = self.validator.validate_single_file(fake_path)
        
        self.assertIn('validation_failed', result)
        self.assertTrue(result['validation_failed'])
        self.assertIn('error', result)
    
    def test_validate_all_generated_files(self):
        """Test validation of all generated files"""
        results = self.validator.validate_all_generated_files()
        
        self.assertIsInstance(results, list)
        self.assertGreater(len(results), 0, "Should find test files")
        
        # Check that all results have required structure
        for result in results:
            if 'validation_failed' not in result:
                self.assertIn('file_name', result)
                self.assertIn('formulas_found', result)
                self.assertIn('data_populated', result)
    
    def test_compare_with_template(self):
        """Test template comparison functionality"""
        # This should not crash and should complete comparison
        try:
            self.validator.compare_with_template()
            # If we get here without exception, the test passes
            self.assertTrue(True)
        except Exception as e:
            self.fail(f"Template comparison should not fail: {str(e)}")
    
    def test_formula_detection(self):
        """Test that formulas are properly detected"""
        # Create a simple test file with known formulas
        test_file = self.temp_dir / "formula_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add known formulas
        ws['A1'] = '=B1+C1'
        ws['A2'] = '=SUM(B1:C1)'
        ws['B1'] = 10
        ws['C1'] = 20
        
        wb.save(test_file)
        wb.close()
        
        # Load and check for formulas
        wb_check = openpyxl.load_workbook(test_file, data_only=False)
        ws_check = wb_check.active
        
        # Verify formulas are preserved
        self.assertEqual(ws_check['A1'].value, '=B1+C1')
        self.assertEqual(ws_check['A2'].value, '=SUM(B1:C1)')
        
        wb_check.close()
    
    def test_data_population_detection(self):
        """Test that populated data is properly detected"""
        botn_files = list(self.validator.test_output_dir.glob("*_BOTN.xlsx"))
        if botn_files:
            result = self.validator.validate_single_file(botn_files[0])
            
            if 'validation_failed' not in result:
                # Check specific data fields
                data_pop = result['data_populated']
                self.assertIsInstance(data_pop.get('Property Name'), str)
                self.assertIsInstance(data_pop.get('Purchase Price'), (int, float))
                self.assertEqual(data_pop.get('Housing Type'), 'Large Family')
                self.assertEqual(data_pop.get('Credit Pricing'), 0.80)

class TestValidatorIntegration(unittest.TestCase):
    """Integration tests for validator functionality"""
    
    def test_validator_with_real_openpyxl_workflow(self):
        """Test validator works with actual OpenPyxl operations"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        
        try:
            # Create a workbook similar to BOTN structure
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Create Inputs sheet
            inputs_sheet = wb.create_sheet("Inputs")
            inputs_sheet['A2'] = 'Integration Test Property'
            inputs_sheet['B2'] = '999 Integration Ave'
            inputs_sheet['G2'] = 3500000
            
            # Create sheet with formula
            calc_sheet = wb.create_sheet("Calculations")
            calc_sheet['C1'] = '=Inputs!A2'
            calc_sheet['C2'] = '=Inputs!G2*1.1'
            
            wb.save(temp_path)
            wb.close()
            
            # Verify the file can be validated
            validator = OpenPyxlOutputValidator()
            result = validator.validate_single_file(temp_path)
            
            # Should not have validation_failed
            self.assertNotIn('validation_failed', result)
            self.assertIn('file_name', result)
            
        finally:
            temp_path.unlink(missing_ok=True)
    
    def test_performance_with_multiple_files(self):
        """Test validator performance with multiple files"""
        import time
        
        temp_dir = Path(tempfile.mkdtemp())
        try:
            # Create multiple test files
            file_count = 5
            for i in range(file_count):
                test_file = temp_dir / f"perf_test_{i}.xlsx"
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = f'Performance Test {i}'
                ws['A2'] = f'=A1 & " - Formula"'
                wb.save(test_file)
                wb.close()
            
            # Create validator and override test directory
            validator = OpenPyxlOutputValidator()
            validator.test_output_dir = temp_dir
            
            # Time the validation
            start_time = time.time()
            
            # Validate all files (rename them to match expected pattern)
            for i, file_path in enumerate(temp_dir.glob("perf_test_*.xlsx")):
                new_name = temp_dir / f"Perf_Test_{i}_BOTN.xlsx"
                file_path.rename(new_name)
            
            results = validator.validate_all_generated_files()
            end_time = time.time()
            
            # Verify results
            self.assertEqual(len(results), file_count)
            self.assertLess(end_time - start_time, 10.0)  # Should complete in under 10 seconds
            
        finally:
            shutil.rmtree(temp_dir)

def run_tests():
    """Run all validator tests"""
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestOpenPyxlOutputValidator))
    suite.addTests(loader.loadTestsFromTestCase(TestValidatorIntegration))
    
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result.wasSuccessful()

if __name__ == '__main__':
    success = run_tests()
    exit(0 if success else 1)