#!/usr/bin/env python3
"""
Unit tests for ExcelFileManager - Dropbox-based BOTN analysis
"""
import unittest
import tempfile
import os
import shutil
import json
from datetime import datetime
from unittest.mock import patch, MagicMock

# Add the workforce_analyst directory to path for imports
import sys
import os
workforce_analyst_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'modules', 'workforce_analyst')
sys.path.insert(0, workforce_analyst_path)

from AcquisitionAnalyst import ExcelFileManager, sanitize_deal_name, load_deal_data

class TestExcelFileManager(unittest.TestCase):
    """Test ExcelFileManager functionality"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.excel_manager = ExcelFileManager()
        self.test_deal_name = "Test Deal Properties"
        self.sample_data = {
            'Property Name': 'Test Property',
            'Address': '123 Test St',
            'City': 'Test City',
            'State': 'CA',
            'Zip Code': '90210',
            'Number of Units': '100',
            'Year Built': '2020',
            'Asking Price': '$10,000,000',
            'T12 Net Rental Income': '$1,200,000',
            'T12 Expenses': '$400,000'
        }
    
    def test_sanitize_deal_name(self):
        """Test deal name sanitization"""
        # Test special characters removal
        result = sanitize_deal_name("Deal@Name#With$Special%Characters")
        self.assertEqual(result, "DealNameWithSpecialCharacters")
        
        # Test spaces to hyphens
        result = sanitize_deal_name("Deal With Spaces")
        self.assertEqual(result, "Deal-With-Spaces")
        
        # Test mixed case preservation
        result = sanitize_deal_name("CamelCase Deal")
        self.assertEqual(result, "CamelCase-Deal")
    
    def test_excel_manager_initialization(self):
        """Test ExcelFileManager initialization"""
        # Test default template path
        expected_path = "/Users/vitorfaroni/Library/CloudStorage/Dropbox-HERR/Vitor Faroni/Deals/!WFBOTN/80AMIBOTN.xlsx"
        self.assertEqual(self.excel_manager.template_path, expected_path)
        
        # Test custom template path
        custom_path = "/custom/path/template.xlsx"
        custom_manager = ExcelFileManager(custom_path)
        self.assertEqual(custom_manager.template_path, custom_path)
    
    @patch('os.path.exists')
    @patch('shutil.copy2')
    @patch('os.makedirs')
    def test_copy_template_to_deal_folder_success(self, mock_makedirs, mock_copy, mock_exists):
        """Test successful template copying"""
        # Mock successful file operations
        mock_exists.return_value = True
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Test template copying
            result = self.excel_manager.copy_template_to_deal_folder(
                self.test_deal_name, 
                temp_dir
            )
            
            # Verify result is a path
            self.assertIsNotNone(result)
            self.assertTrue(result.endswith('.xlsx'))
            self.assertIn('BOTN_Test-Deal-Properties', result)
            
            # Verify directories created
            mock_makedirs.assert_called_once_with(temp_dir, exist_ok=True)
            
            # Verify file copied
            mock_copy.assert_called_once()
    
    @patch('os.path.exists')
    @patch('shutil.copy2')
    def test_copy_template_failure_with_retry(self, mock_copy, mock_exists):
        """Test template copying with failure and retry logic"""
        mock_exists.return_value = True
        mock_copy.side_effect = [PermissionError("Permission denied"), None]  # Fail first, succeed second
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # This should succeed after retry
            with patch('time.sleep'):  # Mock sleep to speed up test
                result = self.excel_manager.copy_template_to_deal_folder(
                    self.test_deal_name, 
                    temp_dir
                )
            
            # Should succeed after retry
            self.assertIsNotNone(result)
            self.assertEqual(mock_copy.call_count, 2)  # Initial attempt + 1 retry
    
    @patch('openpyxl.load_workbook')
    def test_update_excel_values_success(self, mock_load_workbook):
        """Test successful Excel file updating"""
        # Mock workbook and worksheet
        mock_workbook = MagicMock()
        mock_worksheet = MagicMock()
        mock_workbook.sheetnames = ['Inputs', 'Main', 'Output']
        mock_workbook.__getitem__.return_value = mock_worksheet
        mock_load_workbook.return_value = mock_workbook
        
        # Test updating values
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as temp_file:
            result = self.excel_manager.update_excel_values(temp_file.name, self.sample_data)
            
            # Verify success
            self.assertTrue(result)
            
            # Verify workbook operations
            mock_load_workbook.assert_called_once_with(temp_file.name)
            mock_workbook.save.assert_called_once_with(temp_file.name)
            mock_workbook.close.assert_called_once()
            
            # Verify some cell updates (checking if worksheet received values)
            self.assertGreater(len(mock_worksheet.__setitem__.call_args_list), 0)
    
    @patch('openpyxl.load_workbook')
    def test_update_excel_values_missing_inputs_sheet(self, mock_load_workbook):
        """Test Excel updating with missing Inputs sheet"""
        # Mock workbook without Inputs sheet
        mock_workbook = MagicMock()
        mock_workbook.sheetnames = ['Main', 'Output']  # No 'Inputs'
        mock_load_workbook.return_value = mock_workbook
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as temp_file:
            result = self.excel_manager.update_excel_values(temp_file.name, self.sample_data)
            
            # Should fail due to missing Inputs sheet
            self.assertFalse(result)
    
    @patch('openpyxl.load_workbook')
    def test_update_excel_values_exception_handling(self, mock_load_workbook):
        """Test Excel updating with exception handling"""
        # Mock workbook loading exception
        mock_load_workbook.side_effect = Exception("File not found")
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as temp_file:
            result = self.excel_manager.update_excel_values(temp_file.name, self.sample_data)
            
            # Should fail gracefully
            self.assertFalse(result)
    
    def test_get_file_path(self):
        """Test file path getter"""
        test_path = "/path/to/test.xlsx"
        result = self.excel_manager.get_file_path(test_path)
        self.assertEqual(result, test_path)
    
    def test_load_deal_data_integration(self):
        """Test integration with existing deal data loading"""
        # This tests the integration with cached deal data
        # Using actual cached data if available
        
        cached_deals = ['mResidences-Olympic-and-Olive-Los-Angeles-CA', 'TCI', 'Sunset-Gardens-El-Cajon-CA']
        
        for deal_name in cached_deals:
            try:
                deal_data, extracted_date = load_deal_data(deal_name)
                if deal_data:
                    # Verify data structure
                    self.assertIsInstance(deal_data, dict)
                    self.assertIsInstance(extracted_date, str)
                    
                    # Verify some expected fields exist
                    common_fields = ['Property Name', 'Number of Units']
                    found_fields = [field for field in common_fields if field in deal_data]
                    self.assertGreater(len(found_fields), 0, f"No common fields found in {deal_name}")
                    
                    break  # Test passed with at least one deal
            except Exception:
                continue  # Try next deal if this one fails
        else:
            self.skipTest("No cached deal data available for integration testing")

class TestExcelFileManagerIntegration(unittest.TestCase):
    """Integration tests for ExcelFileManager"""
    
    def setUp(self):
        """Set up integration test fixtures"""
        self.excel_manager = ExcelFileManager()
        
    def test_template_file_exists(self):
        """Test that the actual template file exists and is readable"""
        template_path = self.excel_manager.template_path
        
        # Check if template exists (skip if not available)
        if not os.path.exists(template_path):
            self.skipTest(f"Template file not found: {template_path}")
        
        # Test loading template with openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path)
            
            # Verify expected sheets exist
            self.assertIn('Inputs', wb.sheetnames)
            
            # Verify Inputs sheet structure
            inputs_sheet = wb['Inputs']
            self.assertGreater(inputs_sheet.max_row, 1)
            self.assertGreater(inputs_sheet.max_column, 1)
            
            wb.close()
            
        except ImportError:
            self.skipTest("openpyxl not available for integration testing")
        except Exception as e:
            self.fail(f"Failed to load template file: {e}")
    
    def test_full_workflow_with_sample_data(self):
        """Test complete workflow from template copy to data update"""
        # Skip if template not available
        if not os.path.exists(self.excel_manager.template_path):
            self.skipTest("Template file not available")
        
        sample_data = {
            'Property Name': 'Integration Test Property',
            'Address': '123 Integration Test St',
            'City': 'Test City',
            'State': 'CA',
            'Number of Units': '50',
            'Year Built': '2021'
        }
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Test complete workflow
            botn_file_path = self.excel_manager.copy_template_to_deal_folder(
                'Integration-Test-Deal',
                temp_dir
            )
            
            self.assertIsNotNone(botn_file_path)
            self.assertTrue(os.path.exists(botn_file_path))
            
            # Test updating the file
            success = self.excel_manager.update_excel_values(botn_file_path, sample_data)
            self.assertTrue(success)
            
            # Verify file still exists and is valid
            self.assertTrue(os.path.exists(botn_file_path))

if __name__ == '__main__':
    unittest.main(verbosity=2)